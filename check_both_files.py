import os
import sys
from openpyxl import load_workbook

output = []

# Ariza_Listesi_KAYSERI.xlsx check
ariza_file = 'logs/kayseri/ariza_listesi/Ariza_Listesi_KAYSERI.xlsx'
output.append(f"\n=== {ariza_file} ===")
if os.path.exists(ariza_file):
    try:
        wb = load_workbook(ariza_file, data_only=True)
        ws = wb.active
        # A sütundan FF'leri bul
        ff_a = []
        for row in range(5, min(ws.max_row + 1, 1000)):
            val = ws.cell(row=row, column=1).value
            if val and 'FF-' in str(val):
                ff_a.append((row, val))
        if ff_a:
            output.append(f"Last 3 in Column A:")
            for row, val in ff_a[-3:]:
                output.append(f"  Row {row}: {val}")
        wb.close()
    except Exception as e:
        output.append(f"Error: {e}")
else:
    output.append("File not found")

# Fracas_KAYSERI.xlsx check
fracas_file = 'logs/kayseri/ariza_listesi/Fracas_KAYSERI.xlsx'
output.append(f"\n=== {fracas_file} ===")
if os.path.exists(fracas_file):
    try:
        wb = load_workbook(fracas_file, data_only=True)
        if 'FRACAS' in wb.sheetnames:
            ws = wb['FRACAS']
            # E sütundan FF'leri bul
            ff_e = []
            for row in range(5, min(ws.max_row + 1, 1000)):
                val = ws.cell(row=row, column=5).value
                if val and 'FF-' in str(val):
                    ff_e.append((row, val))
            if ff_e:
                output.append(f"Last 3 in Column E (FRACAS sheet):")
                for row, val in ff_e[-3:]:
                    output.append(f"  Row {row}: {val}")
            else:
                output.append("No FF in Column E")
            wb.close()
    except Exception as e:
        output.append(f"Error: {e}")
else:
    output.append("File not found")

output.append("\n✓ Done")

# Dosyaya yaz
with open('check_result.txt', 'w', encoding='utf-8') as f:
    f.write('\n'.join(output))

# Stdout'a da yaz
print('\n'.join(output))
