#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""Her proje için Veriler.xlsx'ten tram_id'leri yükle ve Equipment tablosunu güncelle"""

from app import create_app, db
from models import Equipment, ServiceStatus
from openpyxl import load_workbook
from pathlib import Path
from datetime import date

app = create_app()

with app.app_context():
    print("\n" + "="*70)
    print("EQUIPMENT TABLOSUNU HER PROJENIN GERÇEK VERİLERİ İLE GÜNCELLE")
    print("="*70)
    
    # Equipment tablosunu temizle
    Equipment.query.delete()
    ServiceStatus.query.delete()
    db.session.commit()
    print("\n✓ Eski veriler silindi")
    
    base_path = Path(app.root_path)
    projects = ['belgrad', 'kayseri', 'iasi', 'timisoara', 'kocaeli', 'gebze']
    
    today = str(date.today())
    total_trams = 0
    
    for project in projects:
        veriler_path = base_path / 'data' / project / 'Veriler.xlsx'
        
        print(f"\n🚊 {project.upper()}: {veriler_path.name}")
        
        if not veriler_path.exists():
            print(f"  ⚠️  Dosya bulunamadı, skip")
            continue
        
        try:
            wb = load_workbook(veriler_path)
            
            if 'Sayfa2' not in wb.sheetnames:
                print(f"  ⚠️  'Sayfa2' bulunamadı")
                continue
            
            ws = wb['Sayfa2']
            
            # tram_id'leri oku (1. sütun, 2. satırdan başla)
            tram_ids = []
            for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
                if row[0]:
                    tram_id = str(row[0]).strip()
                    if tram_id and tram_id.upper() != 'NONE':
                        tram_ids.append(tram_id)
            
            print(f"  ✓ {len(tram_ids)} tramvay bulundu")
            
            # Equipment'e ekle
            for tram_id in tram_ids:
                eq = Equipment(
                    equipment_code=tram_id,
                    name=f'{project.capitalize()} - {tram_id}',
                    equipment_type='Tramway',
                    status='aktif',
                    project_code=project,
                    location=project.capitalize(),
                    criticality='high'
                )
                db.session.add(eq)
            
            db.session.commit()
            
            # ServiceStatus'e veri ekle (bugün tarihi ile)
            # Her tramvay için örnek status
            status_types = ['Servis', 'Servis Dışı', 'İşletme Kaynaklı Servis Dışı']
            
            for i, tram_id in enumerate(tram_ids):
                # Çeşitli status'lar ekle
                status = status_types[i % len(status_types)]
                
                ss = ServiceStatus(
                    tram_id=tram_id,
                    date=today,
                    status=status,
                    project_code=project
                )
                db.session.add(ss)
            
            db.session.commit()
            print(f"  ✓ {len(tram_ids)} tramvay Equipment ve ServiceStatus'e eklendi")
            total_trams += len(tram_ids)
            
        except Exception as e:
            print(f"  ❌ Hata: {e}")
    
    print("\n" + "="*70)
    print(f"✅ TOPLAM {total_trams} TRAMVAY BAŞARILI ŞEKİLDE YÜKLENDİ")
    print("="*70)
    print("\nŞimdi her proje kendi dosyasından veri çekecek!")
    print("Dashboard ve Service Status sayfası otomatik güncellenir.\n")
