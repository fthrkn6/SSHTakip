#!/usr/bin/env python
# -*- coding: utf-8 -*-

import os
from openpyxl import load_workbook

# Template'i kontrol et
template_path = r"data\kayseri\FR_010_R06_SSH HBR.xlsx"

if os.path.exists(template_path):
    print(f"Template found: {template_path}")
    wb = load_workbook(template_path)
    
    # Fracas sheet'ini aç
    if 'FRACAS' in wb.sheetnames:
        ws = wb['FRACAS']
        print(f"FRACAS sheet found")
        print(f"A1 value: {ws['A1'].value}")
        print(f"A2 value: {ws['A2'].value}")
        print(f"A3 value: {ws['A3'].value}")
        print(f"A4 value: {ws['A4'].value}")
        print(f"A5 value: {ws['A5'].value}")
        
        # İlk birkaç veri satırını kontrol et
        for row in range(1, 10):
            val = ws[f'A{row}'].value
            print(f"A{row}: {val}")
    else:
        print(f"FRACAS sheet not found. Sheets: {wb.sheetnames}")
    
    wb.close()
else:
    print(f"Template not found: {template_path}")
