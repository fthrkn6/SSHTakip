import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

# Load workbook
wb = openpyxl.load_workbook('data/belgrad/FR_010_R06_SSH HBR.xlsx')
ws = wb.active

print('='*80)
print('EXCEL CHECKBOX/FORM CONTROL ANALIZI')
print('='*80)

print('\n🔍 Template icerisinde Form Controls var mi?')
print(f"  Max Row: {ws.max_row}, Max Col: {ws.max_column}")

# Check for drawing shapes (checkbox controls often stored as shapes)
if hasattr(ws, '_drawing'):
    print(f"  Drawing bulundu: {ws._drawing}")
else:
    print(f"  OK Drawing yok (checkbox control yok)")

# Check for data validation
print(f"\n✓ Data Validation Rules ({len(ws.data_validations.dataValidation)} adet):")
for dv in ws.data_validations.dataValidation:
    print(f"  - {dv}")

# Check all shapes/drawing objects
if ws._images:
    print(f"\n✓ Resimler bulundu: {len(ws._images)}")
    for img in ws._images:
        print(f"  - {img}")

print("\n\n✅ ÇÖZÜM:")
print("openpyxl kütüphanesi ile checkbox kontrolü eklemek sınırlı.")
print("Alternatif yöntemler:")
print("  1️⃣  Hücreye ✓ karakteri yazmak (basit, etkili)")
print("  2️⃣  Hücre rengini değiştirmek (yeşil = işaretli)")
print("  3️⃣  Form control eklemek (Python-UNO ile VBA - kompleks)")
print("\nÖNERİ: ✓ karakteri yazmak en pratik!")
