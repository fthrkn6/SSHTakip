"""
HBR (Hata Bildirim Raporu) Yönetim Sistemi
Yeni arıza bildirmelerinden HBR Excel dosyaları oluşturma
"""

import os
import logging
from datetime import datetime
from copy import copy
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import shutil

logger = logging.getLogger(__name__)


class HBRManager:
    """HBR (Hata Bildirim Raporu) Excel dosyalarını yönet"""
    
    # HBR dosya eşlemesi
    HBR_CELLS = {
        'malzeme_no': 'B6',           # Material number
        'malzeme_adi': 'D6',          # Material name
        'rapor_tarihi': 'E6',         # Report date
        'ariza_tarihi': 'G6',         # Failure date
        'ncr_no': 'I6',               # SSH NCR No
        'ariza_km': 'G7',             # Failure km
        'tedarikci': 'J7',            # Supplier
        'musteri': 'E8',              # Customer
        'tespit_yontemi': 'F8',       # Detection method
        'musteri_bildirimi': 'H8',    # Customer notification
        'ariza_sinifi_a': 'G9',       # Failure class A
        'ariza_sinifi_b': 'G10',      # Failure class B
        'ariza_sinifi_c': 'G11',      # Failure class C
        'ilk_defa': 'H9',             # First time
        'tekrarlayan_arac': 'A12',    # Repeated in same vehicle
        'tekrarlayan_farkli': 'E12',  # Repeated in different vehicle
        'ariza_tanimi': 'B17',        # Failure description
        'arac_modulu': 'D19',         # Vehicle module
        'parca_seri_no': 'G19',       # Part serial number
        'fotograf': 'B20',            # Photo
        'ssh_sorumlusu': 'B22'        # SSH responsible (calculated from user)
    }
    
    @staticmethod
    def create_ncr_number(project_code, counter):
        """
        NCR numarası oluştur: {PROJECT_CODE}-NCR-{COUNTER:03d} formatında
        Örnek: BEL25-NCR-001, GDM7-NCR-002, etc.
        
        Args:
            project_code: Proje kodu (BEL25, GDM7, IASI16+18, etc)
            counter: NCR sayıcısı
            
        Returns:
            NCR numarası string'i
        """
        # Proje kodunu güvenli şekilde hazırla (büyük harf)
        code = str(project_code).strip().upper()
        return f"{code}-NCR-{counter:03d}"
    
    @staticmethod
    def create_hbr_file(project_name, app_root, session_user, failure_data, project_code=None, counter=1):
        """
        HBR Excel dosyası oluştur
        
        Args:
            project_name: Proje adı (belgrad, iasi, etc)
            app_root: Flask app root path
            session_user: Login yapan kullanıcı nesnesi
            failure_data: Arıza bildiri verisi (dict)
            project_code: Proje kodu (opt) - None ise Veriler.xlsx'ten okunur
            counter: NCR sayıcısı
            
        Returns:
            Oluşturulan dosyanın yolu veya None
        """
        
        try:
            # Proje kodunu oku (sağlanmamışsa)
            if not project_code:
                try:
                    from utils_hbr_numbering import get_project_code_from_veriler
                    project_code = get_project_code_from_veriler(project_name)
                except:
                    project_code = project_name.upper()[:3] + "25"  # Fallback
            
            # HBR klasörü oluştur (logs/{project_name}/HBR/)
            hbr_dir = os.path.join(app_root, 'logs', project_name, 'HBR')
            os.makedirs(hbr_dir, exist_ok=True)
            
            # Template dosyasını bul
            template_path = os.path.join(app_root, 'data', project_name, 'FR_010_R06_SSH HBR.xlsx')
            if not os.path.exists(template_path):
                # Alternatif: belgrad'dan al
                template_path = os.path.join(app_root, 'data', 'belgrad', 'FR_010_R06_SSH HBR.xlsx')
            
            if not os.path.exists(template_path):
                logger.error(f"HBR template bulunamadı: {template_path}")
                return None
            
            # NCR numarası oluştur (project_code kullan)
            ncr_number = HBRManager.create_ncr_number(project_code, counter)
            
            # Dosya adı: {PROJECT_CODE}-NCR-001_TARİH_KOD
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"{ncr_number}_{timestamp}.xlsx"
            output_path = os.path.join(hbr_dir, filename)
            
            logger.info(f"[HBR] Dosya oluşturuluyor: {filename} (Proje: {project_name} → {project_code})")
            
            # Template'i kopyala
            shutil.copy2(template_path, output_path)
            
            # Excel dosyasını aç ve doldur
            wb = load_workbook(output_path)
            ws = wb.active
            
            # Verileri hücrelere yazma
            hbr_data = HBRManager._prepare_hbr_data(failure_data, session_user, project_code)
            
            for key, cell_ref in HBRManager.HBR_CELLS.items():
                if key in hbr_data and hbr_data[key] is not None:
                    cell = ws[cell_ref]
                    
                    if key == 'fotograf':
                        # Fotoğraf ayrıca işlenecek
                        continue
                    elif key in ['ariza_sinifi_a', 'ariza_sinifi_b', 'ariza_sinifi_c',
                                'ilk_defa', 'tekrarlayan_arac', 'tekrarlayan_farkli',
                                'tespit_yontemi', 'musteri_bildirimi']:
                        # Checkbox hücreler - işaretlemek için 'X' yaz
                        if hbr_data[key]:
                            cell.value = '✓'
                            cell.font = Font(bold=True, size=12)
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                    else:
                        cell.value = hbr_data[key]
                        # Datetime'ı string'e dönüştür
                        if isinstance(hbr_data[key], datetime):
                            cell.value = hbr_data[key].strftime('%Y-%m-%d')
            
            # Fotoğraf varsa ekle
            if 'fotograf' in failure_data and failure_data['fotograf']:
                HBRManager._add_image_to_hbr(ws, failure_data['fotograf'], 'B20')
            
            wb.save(output_path)
            logger.info(f"HBR dosyası oluşturuldu: {output_path}")
            
            return output_path
            
        except Exception as e:
            logger.error(f"HBR oluşturma hatası: {e}")
            return None
    
    @staticmethod
    def _prepare_hbr_data(failure_data, session_user, project_code):
        """Arıza verilerinden HBR verilerini hazırla"""
        
        hbr_data = {
            'malzeme_no': failure_data.get('malzeme_no', ''),
            'malzeme_adi': failure_data.get('malzeme_adi', ''),
            'rapor_tarihi': datetime.now(),
            'ariza_tarihi': failure_data.get('ariza_tarihi'),
            'ncr_no': failure_data.get('ncr_no', ''),
            'ariza_km': failure_data.get('ariza_km', ''),
            'tedarikci': failure_data.get('tedarikci', ''),
            'musteri': failure_data.get('musteri', ''),
            'tespit_yontemi': failure_data.get('tespit_yontemi') == 'Bozankaya',  # checkbox
            'musteri_bildirimi': failure_data.get('musteri_bildirimi', False),  # checkbox
            
            # Arıza sınıfı - sadece seçileni işaretle
            'ariza_sinifi_a': failure_data.get('ariza_sinifi') == 'A',
            'ariza_sinifi_b': failure_data.get('ariza_sinifi') == 'B',
            'ariza_sinifi_c': failure_data.get('ariza_sinifi') == 'C',
            
            # Arıza tipi - sadece seçileni işaretle
            'ilk_defa': failure_data.get('ariza_tipi') == 'ilk_defa',
            'tekrarlayan_arac': failure_data.get('ariza_tipi') == 'tekrarlayan_arac',
            'tekrarlayan_farkli': failure_data.get('ariza_tipi') == 'tekrarlayan_farkli',
            
            # Açıklamalar
            'ariza_tanimi': failure_data.get('ariza_tanimi', ''),
            'arac_modulu': failure_data.get('arac_modulu', ''),
            'parca_seri_no': failure_data.get('parca_seri_no', ''),
            
            # SSH sorumlusu - login yapan kullanıcı
            'ssh_sorumlusu': session_user.username if session_user else '',
            
            'fotograf': failure_data.get('fotograf')
        }
        
        return hbr_data
    
    @staticmethod
    def _add_image_to_hbr(ws, image_path, cell_ref='B20'):
        """Excel dosyasına fotoğraf ekle"""
        try:
            if not os.path.exists(image_path):
                logger.warning(f"Fotoğraf bulunamadı: {image_path}")
                return
            
            # Fotoğraf boyutlandır (max 4x4 cm)
            img = XLImage(image_path)
            img.width = 120
            img.height = 100
            
            ws.add_image(img, cell_ref)
            logger.info(f"Fotoğraf eklendi: {image_path}")
            
        except Exception as e:
            logger.warning(f"Fotoğraf ekleme hatası: {e}")
    
    @staticmethod
    def get_next_ncr_counter(project_name, app_root):
        """
        Bir sonraki NCR sayıcısını al
        
        HBR klasöründen mevcut dosyaları tara,
        en yüksek numarayı bul ve bir sonrakini döndür.
        Proje kodunu Veriler.xlsx'ten otomatik okuyor.
        
        Args:
            project_name: Proje adı (belgrad, iasi, etc)
            app_root: Flask app root path
            
        Returns:
            Tuple: (project_code, next_counter)
                project_code: Veriler.xlsx'ten okunan proje kodu (BEL25, GDM7, etc)
                next_counter: Sonraki NCR sayıcısı (integer)
        """
        import re
        
        try:
            # Proje kodunu Veriler.xlsx'ten oku
            from utils_hbr_numbering import get_project_code_from_veriler
            project_code = get_project_code_from_veriler(project_name)
        except:
            project_code = project_name.upper()[:3] + "25"  # Fallback
        
        hbr_dir = os.path.join(app_root, 'logs', project_name, 'HBR')
        
        if not os.path.exists(hbr_dir):
            logger.info(f"[HBR] Yeni klasör: {project_name}/{project_code}")
            return project_code, 1
        
        # Proje kodunu pattern'e dönüştür
        pattern = re.compile(rf'^{re.escape(project_code)}-NCR-(\d+)', re.IGNORECASE)
        
        # Var olan HBR dosyalarını tara
        max_num = 0
        try:
            for f in os.listdir(hbr_dir):
                match = pattern.match(f)
                if match:
                    num = int(match.group(1))
                    max_num = max(max_num, num)
                    logger.debug(f"[HBR] Bulunan: {f} → {num}")
        except Exception as e:
            logger.warning(f"HBR dosyaları tarama hatası: {e}")
        
        next_counter = max_num + 1
        logger.info(f"[HBR] Proje: {project_name} / {project_code}, Son: {max_num}, Sonraki: {next_counter}")
        
        return project_code, next_counter
