"""
Raporlama ve Log Yönetimi Sistemi
Tüm sayfalar için tutarlı rapor formatı (Arıza Listesi stili)
JSON, Excel ve Log dosyaları
"""

import os
import json
from datetime import datetime, timedelta
from pathlib import Path
import logging
from typing import List, Dict, Any
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Configure logger
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    encoding='utf-8'
)
logger = logging.getLogger(__name__)


class ReportSystem:
    """Merkezi Raporlama Sistemi"""
    
    BASE_LOG_DIR = Path(__file__).parent / 'logs'
    REPORTS_DIR = BASE_LOG_DIR / 'reports'
    HISTORY_DIR = BASE_LOG_DIR / 'history'
    
    # Klasör yapısını oluştur
    REPORT_TYPES = {
        'dashboard': 'dashboard_reports',
        'maintenance': 'maintenance_reports',
        'km': 'km_reports',
        'scenarios': 'scenarios',
        'service': 'service_reports',
        'system': 'system_logs'
    }
    
    @staticmethod
    def init_directories():
        """Gerekli log klasörlerini oluştur"""
        ReportSystem.REPORTS_DIR.mkdir(parents=True, exist_ok=True)
        ReportSystem.HISTORY_DIR.mkdir(parents=True, exist_ok=True)
        
        for report_type in ReportSystem.REPORT_TYPES.values():
            (ReportSystem.REPORTS_DIR / report_type).mkdir(parents=True, exist_ok=True)
        
        logger.info("✅ Report directories initialized")
    
    @staticmethod
    def get_report_path(report_type: str, project: str = 'belgrad') -> Path:
        """Rapor dosyasının yolunu döndür"""
        if report_type not in ReportSystem.REPORT_TYPES:
            raise ValueError(f"Unknown report type: {report_type}")
        
        type_dir = ReportSystem.REPORT_TYPES[report_type]
        report_dir = ReportSystem.REPORTS_DIR / type_dir / project
        report_dir.mkdir(parents=True, exist_ok=True)
        return report_dir
    
    @staticmethod
    def log_action(action: str, details: Dict[str, Any], severity: str = 'INFO'):
        """Sistem eylemini history'ye kaydet"""
        ReportSystem.init_directories()
        
        log_file = ReportSystem.HISTORY_DIR / 'changes.log'
        
        log_entry = {
            'timestamp': datetime.now().isoformat(),
            'action': action,
            'severity': severity,
            'details': details
        }
        
        with open(log_file, 'a', encoding='utf-8') as f:
            f.write(json.dumps(log_entry, ensure_ascii=False) + '\n')
        
        if severity == 'ERROR':
            logger.error(f"❌ {action}: {details}")
        elif severity == 'WARNING':
            logger.warning(f"⚠️  {action}: {details}")
        else:
            logger.info(f"✅ {action}")
    
    @staticmethod
    def export_to_excel(
        data: List[Dict],
        report_type: str,
        title: str,
        columns: List[str],
        project: str = 'belgrad',
        color_map: Dict = None,
        filename: str = None
    ) -> Path:
        """
        Excel'e rapor dışa aktar (Arıza Listesi stili)
        
        Args:
            data: Rapor verileri
            report_type: Rapor tipi (dashboard, maintenance, etc.)
            title: Rapor başlığı
            columns: Sütun başlıkları
            project: Proje adı
            color_map: {column_name: {value: color_hex}} format renk haritası
            filename: Özel dosya adı (opsiyonel)
        """
        ReportSystem.init_directories()
        
        if not filename:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"{title.replace(' ', '_')}_{timestamp}.xlsx"
        
        report_path = ReportSystem.get_report_path(report_type, project)
        file_path = report_path / filename
        
        # Workbook oluştur
        wb = Workbook()
        ws = wb.active
        ws.title = "Rapor"
        
        # Başlık
        ws.merge_cells('A1:' + get_column_letter(len(columns)) + '1')
        header_cell = ws['A1']
        header_cell.value = title
        header_cell.font = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
        header_cell.fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
        header_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 25
        
        # Tarih ve saat
        ws.merge_cells('A2:' + get_column_letter(len(columns)) + '2')
        date_cell = ws['A2']
        date_cell.value = f"Rapor Tarihi: {datetime.now().strftime('%d.%m.%Y %H:%M:%S')}"
        date_cell.font = Font(name='Calibri', size=10, italic=True)
        date_cell.alignment = Alignment(horizontal='right')
        
        # Sütun başlıkları
        for col_idx, col_name in enumerate(columns, 1):
            cell = ws.cell(row=4, column=col_idx)
            cell.value = col_name
            cell.font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
            cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        ws.row_dimensions[4].height = 30
        
        # Veriler
        borders = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row_idx, row_data in enumerate(data, 5):
            for col_idx, col_name in enumerate(columns, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = row_data.get(col_name, '-')
                cell.border = borders
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                
                # Renk haritası uygula
                if color_map and col_name in color_map:
                    cell_value = str(row_data.get(col_name, ''))
                    if cell_value in color_map[col_name]:
                        color = color_map[col_name][cell_value]
                        cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
        
        # Sütun genişliklerini ayarla
        for col_idx, col_name in enumerate(columns, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 18
        
        # Dosyayı kaydet
        wb.save(file_path)
        logger.info(f"✅ Report exported: {file_path}")
        
        ReportSystem.log_action(
            f"Report exported: {report_type}",
            {'file': str(file_path), 'rows': len(data), 'type': report_type}
        )
        
        return file_path
    
    @staticmethod
    def get_recent_reports(report_type: str, project: str = 'belgrad', days: int = 30) -> List[Dict]:
        """Son N gündeki raporları listele"""
        report_path = ReportSystem.get_report_path(report_type, project)
        cutoff_date = datetime.now() - timedelta(days=days)
        
        reports = []
        for file in sorted(report_path.glob('*.xlsx'), reverse=True):
            if file.stat().st_mtime > cutoff_date.timestamp():
                reports.append({
                    'filename': file.name,
                    'size': file.stat().st_size,
                    'created': datetime.fromtimestamp(file.stat().st_mtime),
                    'path': str(file)
                })
        
        return reports
    
    @staticmethod
    def export_to_json(
        data: List[Dict],
        report_type: str,
        title: str,
        project: str = 'belgrad',
        filename: str = None
    ) -> Path:
        """JSON formatında rapor dışa aktar"""
        ReportSystem.init_directories()
        
        if not filename:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"{title.replace(' ', '_')}_{timestamp}.json"
        
        report_path = ReportSystem.get_report_path(report_type, project)
        file_path = report_path / filename
        
        report_data = {
            'title': title,
            'generated_at': datetime.now().isoformat(),
            'report_type': report_type,
            'data': data,
            'total_records': len(data)
        }
        
        with open(file_path, 'w', encoding='utf-8') as f:
            json.dump(report_data, f, ensure_ascii=False, indent=2)
        
        logger.info(f"✅ JSON report exported: {file_path}")
        return file_path
    
    @staticmethod
    def get_system_logs(hours: int = 24) -> List[Dict]:
        """Son N saatlik sistem loglarını döndür"""
        log_file = ReportSystem.HISTORY_DIR / 'changes.log'
        
        if not log_file.exists():
            return []
        
        logs = []
        cutoff_time = datetime.now() - timedelta(hours=hours)
        
        with open(log_file, 'r', encoding='utf-8') as f:
            for line in f:
                try:
                    entry = json.loads(line)
                    entry_time = datetime.fromisoformat(entry['timestamp'])
                    if entry_time > cutoff_time:
                        logs.append(entry)
                except:
                    pass
        
        return sorted(logs, key=lambda x: x['timestamp'], reverse=True)
    
    @staticmethod
    def cleanup_old_reports(days: int = 90):
        """Eski raporları sil (N gün öncesinden)"""
        cutoff_date = datetime.now() - timedelta(days=days)
        deleted_count = 0
        
        for report_type_dir in ReportSystem.REPORTS_DIR.rglob('*.xlsx'):
            if report_type_dir.stat().st_mtime < cutoff_date.timestamp():
                report_type_dir.unlink()
                deleted_count += 1
        
        ReportSystem.log_action(
            f"Cleanup: Deleted old reports",
            {'deleted_count': deleted_count, 'older_than_days': days}
        )
        
        return deleted_count


# Dashboard Rapor Formatı
class DashboardReport:
    """Dashboard Raporlama"""
    
    @staticmethod
    def generate(tramvaylar: List[Dict], project: str = 'belgrad') -> Path:
        """Dashboard snapshot raporunu oluştur"""
        ReportSystem.init_directories()
        
        # Verileri hazırla
        report_data = []
        for i, tram in enumerate(tramvaylar, 1):
            report_data.append({
                '#': i,
                'Araç No': tram.get('equipment_code', '-'),
                'Durum': tram.get('status', 'Bilinmiyor'),
                'Güncel KM': tram.get('current_km', 0),
                'Toplam KM': tram.get('total_km', 0),
                'Son Bakım': tram.get('last_maintenance', '-'),
                'Bakım Durumu': tram.get('maintenance_status', '-'),
                'Arıza Sayısı': tram.get('failure_count', 0)
            })
        
        columns = ['#', 'Araç No', 'Durum', 'Güncel KM', 'Toplam KM', 'Son Bakım', 'Bakım Durumu', 'Arıza Sayısı']
        
        # Renk haritası (Durum bazlı)
        color_map = {
            'Durum': {
                'Aktif': 'C6EFCE',  # Yeşil
                'Pasif': 'FFC7CE',  # Kırmızı
                'Bakımda': 'FFFFEB', # Sarı
            }
        }
        
        return ReportSystem.export_to_excel(
            report_data,
            'dashboard',
            f'Dashboard Raporu - {datetime.now().strftime("%d.%m.%Y")}',
            columns,
            project,
            color_map
        )


# Bakım Planları Rapor Formatı
class MaintenanceReport:
    """Bakım Planları Raporlama"""
    
    @staticmethod
    def generate(maintenance_data: List[Dict], project: str = 'belgrad') -> Path:
        """Bakım Planları raporunu oluştur"""
        ReportSystem.init_directories()
        
        report_data = []
        for i, item in enumerate(maintenance_data, 1):
            report_data.append({
                '#': i,
                'Araç No': item.get('tram_id', '-'),
                'Araç Adı': item.get('name', '-'),
                'Güncel KM': item.get('current_km', 0),
                'Toplam KM': item.get('total_km', 0),
                'Sonraki Bakım': item.get('next_maintenance_km', '-'),
                'Bakım Tipi': item.get('maintenance_type', '-'),
                'Aciliyet': item.get('urgency', 'Normal'),
                'Son İşlem': item.get('last_operation', '-')
            })
        
        columns = ['#', 'Araç No', 'Araç Adı', 'Güncel KM', 'Toplam KM', 'Sonraki Bakım', 'Bakım Tipi', 'Aciliyet', 'Son İşlem']
        
        # Aciliyet renkleri (Arıza sınıfları gibi)
        color_map = {
            'Aciliyet': {
                'Kritik': 'FF0000',      # Kırmızı (A)
                'Yüksek': 'FF8C00',      # Turuncu (B)
                'Normal': 'FFD700',      # Sarı (C)
                'Düşük': '28A745'        # Yeşil (D)
            }
        }
        
        return ReportSystem.export_to_excel(
            report_data,
            'maintenance',
            f'Bakım Planları Raporu - {datetime.now().strftime("%d.%m.%Y")}',
            columns,
            project,
            color_map
        )


# KM Rapor Formatı
class KMReport:
    """Tramvay KM Raporlama"""
    
    @staticmethod
    def generate(km_data: List[Dict], project: str = 'belgrad') -> Path:
        """KM verilerinin raporunu oluştur"""
        ReportSystem.init_directories()
        
        report_data = []
        for i, item in enumerate(km_data, 1):
            report_data.append({
                '#': i,
                'Araç No': item.get('equipment_code', '-'),
                'Güncel KM': item.get('current_km', 0),
                'Toplam KM': item.get('total_km', 0),
                'Aylık KM': item.get('monthly_km', 0),
                'Son Güncelleme': item.get('last_update', '-'),
                'Güncelleme Yapmış': item.get('updated_by', '-'),
                'Durum': item.get('status', '-')
            })
        
        columns = ['#', 'Araç No', 'Güncel KM', 'Toplam KM', 'Aylık KM', 'Son Güncelleme', 'Güncelleme Yapmış', 'Durum']
        
        return ReportSystem.export_to_excel(
            report_data,
            'km',
            f'KM Raporu - {datetime.now().strftime("%d.%m.%Y")}',
            columns,
            project
        )


# Senaryolar Analiz
class ScenarioAnalysis:
    """Senaryo Bazlı Analiz"""
    
    @staticmethod
    def get_high_failure_trams(all_trams: List[Dict], threshold: int = 10) -> List[Dict]:
        """Son 30 günde N+ arıza olan tramvaylar"""
        high_failure = [t for t in all_trams if t.get('failure_count', 0) >= threshold]
        return sorted(high_failure, key=lambda x: x.get('failure_count', 0), reverse=True)
    
    @staticmethod
    def get_overdue_maintenance(maintenance_data: List[Dict]) -> List[Dict]:
        """Bakımı gecikmiş tramvaylar"""
        overdue = [m for m in maintenance_data if m.get('urgency') == 'Kritik']
        return sorted(overdue, key=lambda x: x.get('current_km', 0), reverse=True)
    
    @staticmethod
    def get_high_km_trams(all_trams: List[Dict], percentile: int = 90) -> List[Dict]:
        """Yüksek KM'li tramvaylar (90. persentil üstü)"""
        kms = [int(t.get('current_km', 0)) for t in all_trams]
        threshold = sorted(kms)[int(len(kms) * percentile / 100)]
        
        high_km = [t for t in all_trams if int(t.get('current_km', 0)) >= threshold]
        return sorted(high_km, key=lambda x: int(x.get('current_km', 0)), reverse=True)
    
    @staticmethod
    def generate_scenario_report(scenario_name: str, scenario_data: List[Dict], project: str = 'belgrad') -> Path:
        """Senaryo raporunu oluştur"""
        ReportSystem.init_directories()
        
        report_data = []
        for i, item in enumerate(scenario_data, 1):
            report_data.append({
                '#': i,
                'Araç No': item.get('equipment_code', item.get('tram_id', '-')),
                'Detay 1': item.get('failure_count', item.get('urgency', '-')),
                'Detay 2': item.get('current_km', '--'),
                'Gözlem Tarihi': datetime.now().strftime('%d.%m.%Y'),
                'Öneri': f"İnceleme yapılmalı"
            })
        
        columns = ['#', 'Araç No', 'Detay 1', 'Detay 2', 'Gözlem Tarihi', 'Öneri']
        
        return ReportSystem.export_to_excel(
            report_data,
            'scenarios',
            f'Senaryo: {scenario_name} - {datetime.now().strftime("%d.%m.%Y")}',
            columns,
            project
        )


# Başlatma
def init_reporting_system():
    """Raporlama sistemini başlat"""
    ReportSystem.init_directories()
    logger.info("✅ Reporting system initialized successfully")


if __name__ == '__main__':
    init_reporting_system()
    print("✅ Reporting system ready")
