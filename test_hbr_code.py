#!/usr/bin/env python
# -*- coding: utf-8 -*-

import os
from openpyxl import load_workbook

# Test veriler.xlsx
veriler_path = os.path.join(os.path.dirname(__file__), 'data', 'kayseri', 'Veriler.xlsx')
print(f"Veriler.xlsx yolu: {veriler_path}")
print(f"Dosya var mı?: {os.path.exists(veriler_path)}")

if os.path.exists(veriler_path):
    try:
        wb = load_workbook(veriler_path)
        sheets = wb.sheetnames
        print(f"Sheet'ler: {sheets}")
        
        if 'Sayfa2' in sheets:
            ws = wb['Sayfa2']
            b2_val = ws['B2'].value
            print(f"Sayfa2 B2: {b2_val} (type: {type(b2_val)})")
        else:
            print(f"Sayfa2 yok!")
        
        wb.close()
    except Exception as e:
        import traceback
        print(f"Error: {e}")
        traceback.print_exc()
else:
    print("File not found!")
