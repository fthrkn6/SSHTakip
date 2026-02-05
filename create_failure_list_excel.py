"""
Arıza Listesi - Yeni Arıza Bildirmeleri İçin Excel Dosyası Oluşturucu
Güzel Styling ile
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import os

def create_ariza_listesi_workbook():
    """
    Yeni Arıza Listesi Excel dosyası oluştur
    Güzel styling ile
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Ariza Listesi"
    
    # Başlık stili
    title_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    title_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Başlık satırı
    ws['A1'] = "ARIZA LİSTESİ - BELGRAD PROJESİ"
    ws.merge_cells('A1:R1')
    ws['A1'].font = title_font
    ws['A1'].fill = title_fill
    ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 25
    
    # Alt başlık - tarih
    ws['A2'] = f"Oluşturma Tarihi: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    ws.merge_cells('A2:R2')
    ws['A2'].font = Font(italic=True, size=10)
    ws['A2'].alignment = Alignment(horizontal="right")
    
    # Sütun başlıkları (Row 4)
    headers = [
        'FRACAS ID',          # A
        'Araç No',            # B
        'Araç Modül',         # C
        'Kilometre',          # D
        'Tarih',              # E
        'Saat',               # F
        'Sistem',             # G
        'Alt Sistem',         # H
        'Tedarikçi',          # I
        'Arıza Sınıfı',       # J
        'Arıza Kaynağı',      # K
        'Garanti Kapsamı',    # L
        'Arıza Tanımı',       # M
        'Yapılan İşlem',      # N
        'Aksiyon',            # O
        'Parça Kodu',         # P
        'Parça Adı',          # Q
        'Durum'               # R
    ]
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    
    ws.row_dimensions[4].height = 30
    
    # Sütun genişlikleri
    column_widths = {
        'A': 15,  # FRACAS ID
        'B': 12,  # Araç No
        'C': 14,  # Araç Modül
        'D': 12,  # Kilometre
        'E': 12,  # Tarih
        'F': 10,  # Saat
        'G': 14,  # Sistem
        'H': 14,  # Alt Sistem
        'I': 14,  # Tedarikçi
        'J': 14,  # Arıza Sınıfı
        'K': 14,  # Arıza Kaynağı
        'L': 14,  # Garanti Kapsamı
        'M': 20,  # Arıza Tanımı
        'N': 18,  # Yapılan İşlem
        'O': 18,  # Aksiyon
        'P': 14,  # Parça Kodu
        'Q': 20,  # Parça Adı
        'R': 12,  # Durum
    }
    
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width
    
    # Örnek veri stili (row 5+)
    light_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    data_font = Font(size=10)
    
    # 10 boş satır oluştur (örnek için)
    for row_idx in range(5, 15):
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = border
            cell.font = data_font
            if row_idx % 2 == 0:
                cell.fill = light_fill
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    
    ws.row_dimensions[5].height = 20
    
    # Freeze panes (başlıkları dondur)
    ws.freeze_panes = "A5"
    
    return wb


# Test: Dosya oluştur
if __name__ == "__main__":
    # logs/ariza_listesi klasörü oluştur
    os.makedirs('logs/ariza_listesi', exist_ok=True)
    
    # Dosya adı: Ariza_Listesi_BELGRAD_{TARIH}.xlsx
    filename = f"Ariza_Listesi_BELGRAD_{datetime.now().strftime('%Y%m%d')}.xlsx"
    filepath = os.path.join('logs/ariza_listesi', filename)
    
    # Workbook oluştur ve kaydet
    wb = create_ariza_listesi_workbook()
    wb.save(filepath)
    
    print(f"✓ Arıza Listesi oluşturuldu: {filepath}")
    print(f"  - Güzel styling ile")
    print(f"  - 18 sütun")
    print(f"  - Örnek satırlar dahil")
