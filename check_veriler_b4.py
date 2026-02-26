#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
veriler.xlsx B4 hücresini kontrol et
"""

from openpyxl import load_workbook
import os

project = 'BELGRAD'
veriler_path = os.path.join('data', project, 'veriler.xlsx')

if os.path.exists(veriler_path):
    wb = load_workbook(veriler_path)
    ws = wb.active
    
    print("📊 Veriler.xlsx - İçerik Kontrolü")
    print("=" * 50)
    print(f"B1: {ws['B1'].value}")
    print(f"B2: {ws['B2'].value}")
    print(f"B3: {ws['B3'].value}")
    print(f"B4: {ws['B4'].value} ✓ (Müşteri)")
    print(f"B5: {ws['B5'].value}")
    print(f"C3: {ws['C3'].value}")
    print("=" * 50)
    
    wb.close()
else:
    print(f"❌ Dosya yok: {veriler_path}")
