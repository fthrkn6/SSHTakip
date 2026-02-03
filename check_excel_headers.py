#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Excel Sütunları Kontrol Scripti
"""

import pandas as pd
from openpyxl import load_workbook

excel_path = r"c:\Users\ferki\Desktop\bozankaya_ssh_takip\data\belgrad\BEL25_FRACAS.xlsx"

print("="*70)
print("EXCEL HEADER SATIRI - İlk Satırdaki Başlıklar")
print("="*70)

wb = load_workbook(excel_path)
ws = wb['FRACAS']

print("\nHeader satırından tüm hücreler:")
for col_num, cell in enumerate(ws[1], 1):
    if cell.value:
        print(f"  Sütun {col_num:2d}: '{cell.value}'")

wb.close()

print("\n" + "="*70)
print("PANDAS İLE OKUNAN BAŞLIKLAR")
print("="*70)

df = pd.read_excel(excel_path, sheet_name='FRACAS', header=0, engine='openpyxl')
print("\nPandas tarafından okunan sütunlar:")
for i, col in enumerate(df.columns, 1):
    print(f"  Sütun {i:2d}: '{col}'")
