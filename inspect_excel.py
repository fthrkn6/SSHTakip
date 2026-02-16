#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""En yeni Excel dosyasını ayrıntılı kontrol"""
from pathlib import Path
from openpyxl import load_workbook

# En yeni Excel'i bul
rca_dir = Path('logs/root_cause_analysis')
excel_files = sorted(rca_dir.glob('*.xlsx'), key=lambda x: x.stat().st_mtime)

if not excel_files:
    print("Excel dosyasi yok!")
    exit(1)

latest = excel_files[-1]
print(f"Dosya: {latest.name}")
print(f"Boyut: {latest.stat().st_size} bytes\n")

# Excel'i aç
wb = load_workbook(latest)
ws = wb.active

print(f"Sayfa: {ws.title}")
print(f"Boyut: {ws.dimensions}")
print(f"Satir: {ws.max_row}, Sutun: {ws.max_column}\n")

# Hücrelerden veri topla
print("EXCEL VERİLERİ:")
print("=" * 80)

for row_idx in range(1, min(ws.max_row + 1, 45)):
    row_data = []
    for col_idx in range(1, 9):
        cell = ws.cell(row_idx, col_idx)
        val = cell.value
        if val is None:
            row_data.append('')
        else:
            row_data.append(str(val)[:20])
    
    # Eğer hatalı satırın içinde veri varsa print et
    if any(row_data):
        line = f"R{row_idx:2}: "
        for i, val in enumerate(row_data, 1):
            if val:
                line += f"[C{i}: {val}] "
        print(line)

print("\n" + "=" * 80)

# Merge cells kontrol
if ws.merged_cells:
    print(f"\nMerge Cells: {len(ws.merged_cells)} adet")
    for merged in list(ws.merged_cells.ranges)[:10]:
        print(f"  - {merged}")

# Tables kontrol
if ws.tables:
    print(f"\nTables: {len(ws.tables)} adet")
    for table_name, table in ws.tables.items():
        print(f"  - {table_name}: {table.ref}")
else:
    print(f"\nTables: YOKU (muhtemelen bu nedenini Excel bos goriyor)")

print("\n" + "=" * 80)
