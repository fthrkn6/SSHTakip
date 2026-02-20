"""
Arıza Listesi'nden tüm verileri çekip Excel'e aktarma
"""

import os
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import shutil

# Klasörleri tanımla
ariza_listesi_dir = os.path.join(os.path.dirname(__file__), 'logs', 'ariza_listesi')

if not os.path.exists(ariza_listesi_dir):
    print(f"❌ Arıza Listesi klasörü bulunamadı: {ariza_listesi_dir}")
    exit(1)

# Bugünün dosyasını bul
today_date = datetime.now().strftime('%Y%m%d')
today_file = os.path.join(ariza_listesi_dir, f"Ariza_Listesi_BELGRAD_{today_date}.xlsx")

if not os.path.exists(today_file):
    print(f"❌ Bugünün Arıza Listesi dosyası bulunamadı: {today_file}")
    print(f"\n📁 Klasördeki dosyalar:")
    for f in os.listdir(ariza_listesi_dir):
        print(f"   - {f}")
    exit(1)

print(f"✓ Dosya bulundu: {today_file}")

# Dosyayı aç ve oku
try:
    wb = load_workbook(today_file, data_only=False)
    ws = wb.active
    
    print(f"\n📊 Arıza Listesi'nden veriler okunuyor...")
    print(f"   Max satır: {ws.max_row}")
    print(f"   Max sütun: {ws.max_column}")
    
    # Başlık bilgileri
    print(f"\n📝 Başlık Bilgileri:")
    print(f"   Row 1: {ws['A1'].value}")
    print(f"   Row 2: {ws['A2'].value}")
    
    # Header satırı (Row 4)
    print(f"\n📋 Sütun Başlıkları (Row 4):")
    headers = []
    for col in range(1, ws.max_column + 1):
        cell_value = ws.cell(row=4, column=col).value
        headers.append(cell_value)
        print(f"   Sütun {col}: {cell_value}")
    
    # Verileri oku (Row 5+)
    print(f"\n📈 Veri satırları (Row 5+):")
    data_rows = []
    row_count = 0
    
    for row in range(5, ws.max_row + 1):
        # İlk sütunu kontrol et (FRACAS ID)
        fracas_id = ws.cell(row=row, column=1).value
        
        if fracas_id:  # Boş olmayan satır
            row_data = []
            for col in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=row, column=col).value
                row_data.append(cell_value)
            data_rows.append(row_data)
            row_count += 1
            
            if row_count <= 5:  # İlk 5 satırı göster
                print(f"   Row {row}: {fracas_id} | {row_data[1]} | {row_data[6]} (Sistem)")
    
    print(f"\n✅ Toplam {row_count} satır veri bulundu")
    
    # Export Excel'i oluştur
    export_file = os.path.join(os.path.dirname(__file__), 'Ariza_Listesi_BELGRAD_EXPORT.xlsx')
    
    from openpyxl import Workbook
    wb_export = Workbook()
    ws_export = wb_export.active
    ws_export.title = "Ariza Listesi"
    
    # Başlık
    ws_export['A1'] = "ARIZA LİSTESİ - BELGRAD PROJESİ (Export)"
    ws_export.merge_cells('A1:R1')
    
    title_font = Font(bold=True, size=12, color="FFFFFF")
    title_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    ws_export['A1'].font = title_font
    ws_export['A1'].fill = title_fill
    ws_export['A1'].alignment = Alignment(horizontal="center", vertical="center")
    ws_export.row_dimensions[1].height = 25
    
    # Tarih
    ws_export['A2'] = f"Export Tarihi: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    ws_export.merge_cells('A2:R2')
    ws_export['A2'].font = Font(italic=True, size=10)
    ws_export['A2'].alignment = Alignment(horizontal="right")
    
    # Sütun başlıkları
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                   top=Side(style='thin'), bottom=Side(style='thin'))
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws_export.cell(row=4, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    
    ws_export.row_dimensions[4].height = 30
    
    # Veri satırları
    data_fill = PatternFill(start_color="D9E8F5", end_color="D9E8F5", fill_type="solid")
    
    for row_idx, row_data in enumerate(data_rows, 5):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_export.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.border = border
            cell.font = Font(size=10)
            
            if row_idx % 2 == 0:
                cell.fill = data_fill
            
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        
        ws_export.row_dimensions[row_idx].height = 20
    
    # Kolon genişlikleri
    ws_export.column_dimensions['A'].width = 15  # FRACAS ID
    ws_export.column_dimensions['B'].width = 12  # Araç No
    ws_export.column_dimensions['C'].width = 12  # Araç Modül
    ws_export.column_dimensions['D'].width = 10  # Kilometre
    ws_export.column_dimensions['E'].width = 12  # Tarih
    ws_export.column_dimensions['F'].width = 10  # Saat
    ws_export.column_dimensions['G'].width = 12  # Sistem
    ws_export.column_dimensions['H'].width = 12  # Alt Sistem
    
    for col in range(9, 19):
        ws_export.column_dimensions[chr(64 + col)].width = 15
    
    # Frozen panes
    ws_export.freeze_panes = "A5"
    
    # Kaydet
    wb_export.save(export_file)
    
    print(f"\n✅ Export tamamlandı!")
    print(f"📁 Dosya: {export_file}")
    print(f"📊 Satır: {row_count}")
    print(f"📋 Sütun: {len(headers)}")
    
except Exception as e:
    print(f"❌ Hata: {e}")
    import traceback
    traceback.print_exc()
