import openpyxl
import os
import shutil

excel_file = r'c:\Users\ferki\Desktop\bozankaya_ssh_takip\logs\ariza_listesi\Ariza_Listesi_BELGRAD.xlsx'
temp_file = r'c:\Users\ferki\Desktop\bozankaya_ssh_takip\logs\ariza_listesi\temp_ariza.xlsx'

# Kopyala
shutil.copy(excel_file, temp_file)

# Temp dosyayı açıp düzenle
wb = openpyxl.load_workbook(temp_file)
ws = wb['Ariza Listesi']

# Personel sayılarını yaz (Row 5-12, Column 30)
personel_sayilari = [2, 1, 3, 2, 1, 2, 1, 3]

print('Personel Sayılarını Excel\'e Yazılıyor:')
print('=' * 50)

for idx, personel_sayı in enumerate(personel_sayilari):
    row_idx = 5 + idx
    araç = ws.cell(row_idx, 2).value
    ws.cell(row_idx, 30).value = personel_sayı
    print(f'Row {row_idx}: Araç={araç}, Personel Sayısı={personel_sayı}')

# Temp dosyayı kaydet
wb.save(temp_file)
wb.close()

# Orijinal dosyayı sil ve temp'yi kopyala
os.remove(excel_file)
shutil.copy(temp_file, excel_file)
os.remove(temp_file)

print('\n✅ Tüm veriler kaydedildi!')

