#!/usr/bin/env python
from app import create_app
from openpyxl import load_workbook
import tempfile
import os

app = create_app()
with app.test_client() as client:
    # Login
    client.post('/login', data={'username': 'testuser', 'password': 'password'})
    
    # Get Excel export
    response = client.get('/servis/excel/daily-table')
    print(f'Response Status: {response.status_code}')
    
    if response.status_code == 200:
        print(f'File Size: {len(response.data)} bytes')
        
        # Save and read
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as f:
            f.write(response.data)
            temp_path = f.name
        
        try:
            wb = load_workbook(temp_path)
            info_sheet = wb['Bilgi']
            data_sheet = wb['Günlük Durum']
            
            print(f'A5 (Toplam Kayıt): {info_sheet["A5"].value}')
            print(f'A6 (Tramvay Sayısı): {info_sheet["A6"].value}')
            print(f'Data Sheet Rows: {data_sheet.max_row}')
        finally:
            os.unlink(temp_path)
