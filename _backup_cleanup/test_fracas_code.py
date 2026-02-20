from openpyxl import load_workbook
import os

# Belgrad
belgrad_file = r'c:\Users\fatiherkin\Desktop\bozankaya_ssh_takip\data\belgrad\Veriler.xlsx'
if os.path.exists(belgrad_file):
    try:
        wb = load_workbook(belgrad_file)
        ws = wb['Sayfa2']
        print('=== BELGRAD ===')
        print(f'B2 değeri: {ws["B2"].value}')
        print(f'A1: {ws["A1"].value}, B1: {ws["B1"].value}')
        print(f'A2: {ws["A2"].value}, B2: {ws["B2"].value}')
        print(f'A3: {ws["A3"].value}, B3: {ws["B3"].value}')
    except Exception as e:
        print(f'Belgrad hata: {e}')
else:
    print('Belgrad Veriler.xlsx bulunamadı')

print()

# Kayseri
kayseri_file = r'c:\Users\fatiherkin\Desktop\bozankaya_ssh_takip\data\kayseri\Veriler.xlsx'
if os.path.exists(kayseri_file):
    try:
        wb = load_workbook(kayseri_file)
        ws = wb['Sayfa2']
        print('=== KAYSERI ===')
        print(f'B2 değeri: {ws["B2"].value}')
        print(f'A1: {ws["A1"].value}, B1: {ws["B1"].value}')
        print(f'A2: {ws["A2"].value}, B2: {ws["B2"].value}')
        print(f'A3: {ws["A3"].value}, B3: {ws["B3"].value}')
    except Exception as e:
        print(f'Kayseri hata: {e}')
else:
    print('Kayseri Veriler.xlsx bulunamadı')
