#!/usr/bin/env python
from app import create_app
from openpyxl import load_workbook
import tempfile
import os

app = create_app()
with app.test_client() as client:
    client.post('/login', data={'username': 'testuser', 'password': 'password'})
    
    response = client.get('/servis/excel/daily-table')
    print(f'Status: {response.status_code}')
    
    if response.status_code == 200:
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as f:
            f.write(response.data)
            temp_path = f.name
        
        try:
            wb = load_workbook(temp_path)
            ws = wb['Servis Durumu']
            
            print(f'Rows: {ws.max_row}, Cols: {ws.max_column}')
            print(f'A1: {ws["A1"].value}')
            print(f'B1: {ws["B1"].value}')
            print(f'A2: {ws["A2"].value}')
            print(f'B2 value: {ws["B2"].value}')
            
            # Check colors
            if ws["B2"].fill:
                print(f'B2 color: {ws["B2"].fill.start_color.rgb if ws["B2"].fill.start_color else "None"}')
        finally:
            os.unlink(temp_path)
    else:
        print(response.get_json())
