import openpyxl
import os

excel_file = r'logs\ariza_listesi\Ariza_Listesi_BELGRAD.xlsx'
if os.path.exists(excel_file):
    wb = openpyxl.load_workbook(excel_file)
    ws = wb['Ariza Listesi']
    
    # Check header row 4 (openpyxl row 4)
    print('Header Row (Row 4) - Last 5 columns:')
    headers = [ws.cell(4, i).value for i in range(26, 31)]
    for i, h in enumerate(headers, start=26):
        print(f'Col {i}: {h}')
    
    print('\nFirst Data Row (Row 5) - Last 5 columns:')
    for i in range(26, 31):
        cell = ws.cell(5, i)
        print(f'Col {i}: Value={cell.value}, Format={cell.number_format}')
    
    print('\nAll rows - Personnel Count column (Col 30):')
    for row_idx in range(5, 10):  # Check first 5 data rows
        cell = ws.cell(row_idx, 30)
        print(f'Row {row_idx}: Value={cell.value}, Type={type(cell.value).__name__}')
    
    wb.close()
else:
    print('File not found')
