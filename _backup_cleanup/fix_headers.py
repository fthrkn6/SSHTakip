#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Mevcut Excel dosyalarının başlıklarını güncelle (standalone)
"""
import os
import sys
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Doğru başlıklar (29 sütun)
headers = [
    'FRACAS ID', 'Araç No', 'Araç Modül', 'Kilometre', 'Tarih', 'Saat',
    'Sistem', 'Alt Sistem', 'Tedarikçi', 'Arıza Sınıfı', 'Arıza Kaynağı', 'Arıza Tipi',
    'Garanti Kapsamı', 'Arıza Tanımı', 'Yapılan İşlem', 'Aksiyon', 'Parça Kodu', 'Parça Adı', 'Parça Seri No', 'Adedi',
    'Tamir Başlama Tarihi', 'Tamir Başlama Saati', 'Tamir Bitişi Tarihi', 'Tamir Bitişi Saati', 'Tamir Süresi', 'MTTR (dk)',
    'Servise Veriliş Tarihi', 'Servise Veriliş Saati', 'Durum'
]

# Sütun genişlikleri
column_widths = [13, 10, 12, 10, 12, 10, 12, 12, 12, 14, 14, 18, 12, 20, 14, 10, 12, 12, 12, 10, 15, 14, 14, 14, 14, 12, 14, 14, 10]

ariza_listesi_dir = os.path.join(os.path.dirname(__file__), 'logs', 'ariza_listesi')
updated_count = 0
error_count = 0

if os.path.exists(ariza_listesi_dir):
    print(f"📁 Klasör: {ariza_listesi_dir}\n")
    
    for file in os.listdir(ariza_listesi_dir):
        if 'Ariza_Listesi' in file and file.endswith('.xlsx'):
            file_path = os.path.join(ariza_listesi_dir, file)
            print(f"📄 Dosya: {file}")
            
            try:
                wb = load_workbook(file_path)
                ws = wb.active
                
                # Başlık formatı
                header_font = Font(bold=True, color="FFFFFF", size=10)
                header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
                border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                               top=Side(style='thin'), bottom=Side(style='thin'))
                
                # 4. satıra başlıkları yaz
                for col_idx, header in enumerate(headers, 1):
                    cell = ws.cell(row=4, column=col_idx)
                    cell.value = header
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    cell.border = border
                
                # Sütun genişliklerini ayarla
                for idx, width in enumerate(column_widths, 1):
                    ws.column_dimensions[get_column_letter(idx)].width = width
                
                ws.row_dimensions[4].height = 30
                
                # Dosyayı kaydet
                wb.save(file_path)
                print(f"   ✅ Başlıklar güncellendi ({len(headers)} sütun)")
                print(f"   ✅ Sütun genişlikleri ayarlandı\n")
                updated_count += 1
                
            except Exception as e:
                print(f"   ❌ Hata: {e}\n")
                error_count += 1
else:
    print(f"❌ Klasör bulunamadı: {ariza_listesi_dir}")
    sys.exit(1)

print(f"\n{'='*60}")
print(f"📊 Özet:")
print(f"   ✅ Güncellenen dosya: {updated_count}")
print(f"   ❌ Hata: {error_count}")
print(f"{'='*60}")
