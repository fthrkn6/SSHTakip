#!/usr/bin/env python
"""Test verilerini Excel'e ekle"""
import pandas as pd
import os
from openpyxl import load_workbook
from datetime import datetime, timedelta

file = 'data/belgrad/BEL25_FRACAS(Hata Raporlama Analizi ve Düzeltici Aksiyon Sitemi) Formu - 2025.12.3 S.xlsx'

# Başlama değeri
test_data = [
    {
        'Project': 'Belgrad',
        'Araç Numarası\n\nVehicle Number': '1531',
        'Araç Module\n\nVehicle Module': 'MC',
        'Araç Kilometresi\n\nVehicle Kilometer': '150000',
        'FRACAS ID': 'BEL25-001',
        'Hata Tarih Saat': datetime.now() - timedelta(days=5),
        'Sistem': 'Elektrik',
        'Alt Sistemler': 'Şarj Sistemi',
        'İlgili Tedarikçi\nRelevant Supplier': 'Bosch',
        'Arıza Tanımı\nFailure Description': 'Batarya şarj olmuyor',
        'Arıza Sınıfı\n\n\nFailure Class': 'Kritik',
        'Arıza Kaynağı\n\nSource of Failure': 'Fabrika Hatası',
        'Tamir Başlama Tarih Saat': datetime.now() - timedelta(days=5),
        'Tamir Bitiş Tarih Saat': datetime.now() - timedelta(days=4),
        'Tamir Süresi (dakika)\n\n\n\nRepair Time (dakika)': '480',
        'Servise Veriliş Tarih Saat': datetime.now() - timedelta(days=4),
        'Aksiyon\n\nAction': 'Değişim',
        'Garanti Kapsamı\n\nWarranty Coverage': 'Evet',
        'Parça Kodu\n\nPart Code': 'BOS-12345',
                'Parça Adı\n\nPart Name': 'Oto Şarj',
        'Adeti\nQuantity': 1,
        'İşçilik Maliyeti\n\n\nLabor Cost': 11,
    },
    {
        'Project': 'Belgrad',
        'Araç Numarası\n\nVehicle Number': '1532',
        'Araç Module\n\nVehicle Module': 'SA',
        'Araç Kilometresi\n\nVehicle Kilometer': '120000',
        'FRACAS ID': 'BEL25-002',
        'Hata Tarih Saat': datetime.now() - timedelta(days=3),
        'Sistem': 'Mekanik',
        'Alt Sistemler': 'Fren Sistemi',
        'İlgili Tedarikçi\nRelevant Supplier': 'ZF',
        'Arıza Tanımı\nFailure Description': 'Fren pedi aşınmış',
        'Arıza Sınıfı\n\n\nFailure Class': 'Yüksek',
        'Arıza Kaynağı\n\nSource of Failure': 'Yıpranma',
        'Tamir Başlama Tarih Saat': datetime.now() - timedelta(days=3),
        'Tamir Bitiş Tarih Saat': datetime.now() - timedelta(days=2),
        'Tamir Süresi (dakika)\n\n\n\nRepair Time (dakika)': '240',
        'Servise Veriliş Tarih Saat': datetime.now() - timedelta(days=2),
        'Aksiyon\n\nAction': 'Değişim',
        'Garanti Kapsamı\n\nWarranty Coverage': 'Hayır',
        'Parça Kodu\n\nPart Code': 'ZF-67890',
        'Parça Adı\n\nPart Name': 'Fren Pedi Seti',
        'Adeti\nQuantity': 4,
        'İşçilik Maliyeti\n\n\nLabor Cost': 11,
    },
    {
        'Project': 'Belgrad',
        'Araç Numarası\n\nVehicle Number': '1533',
        'Araç Module\n\nVehicle Module': 'T',
        'Araç Kilometresi\n\nVehicle Kilometer': '180000',
        'FRACAS ID': 'BEL25-003',
        'Hata Tarih Saat': datetime.now() - timedelta(days=1),
        'Sistem': 'Yazılım',
        'Alt Sistemler': 'Kontrol Paneli',
        'İlgili Tedarikçi\nRelevant Supplier': 'Alstom',
        'Arıza Tanımı\nFailure Description': 'Ekran donuş sorunu',
        'Arıza Sınıfı\n\n\nFailure Class': 'Orta',
        'Arıza Kaynağı\n\nSource of Failure': 'Yazılım Hatası',
        'Tamir Başlama Tarih Saat': datetime.now() - timedelta(hours=24),
        'Tamir Bitiş Tarih Saat': datetime.now(),
        'Tamir Süresi (dakika)\n\n\n\nRepair Time (dakika)': '120',
        'Servise Veriliş Tarih Saat': datetime.now(),
        'Aksiyon\n\nAction': 'Reset',
        'Garanti Kapsamı\n\nWarranty Coverage': 'Evet',
        'Parça Kodu\n\nPart Code': 'ALI-99999',
        'Parça Adı\n\nPart Name': 'Kontrol Modülü',
        'Adeti\nQuantity': 1,
        'İşçilik Maliyeti\n\n\nLabor Cost': 11,
    }
]

try:
    # Excel açıkla
    wb = load_workbook(file)
    ws = wb['FRACAS']
    
    # Header satırı (satır 4 - row 4)
    # Satır 5'ten başla (index başından itibaren)
    start_row = 5
    
    for idx, row_data in enumerate(test_data):
        current_row = start_row + idx
        col = 1
        for key, val in row_data.items():
            cell = ws.cell(row=current_row, column=col)
            cell.value = val
            col += 1
    
    # Kaydet
    wb.save(file)
    print(f"✓ {len(test_data)} test verisi eklendi")
    print(f"✓ Dosya kaydedildi: {file}")
    
except Exception as e:
    print(f"❌ Hata: {e}")
    import traceback
    traceback.print_exc()
