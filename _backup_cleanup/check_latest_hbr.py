import openpyxl
import os
from datetime import datetime

# En son oluşturulan HBR dosyasını bul
hbr_dir = "c:\\Users\\fatiherkin\\Desktop\\bozankaya_ssh_takip\\logs\\belgrad\\HBR"
files = [f for f in os.listdir(hbr_dir) if f.startswith('BOZ-NCR-') and f.endswith('.xlsx')]
files.sort()
latest_file = files[-1] if files else None

if not latest_file:
    print("❌ HBR dosyası bulunamadı")
else:
    file_path = os.path.join(hbr_dir, latest_file)
    print(f"📄 Son Dosya: {latest_file}")
    print(f"{'='*80}\n")
    
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    
    print("📊 KRİTİK HÜCRELERDE NE YAZILI?")
    print("-" * 80)
    
    critical_cells = {
        'G9': 'Kritik → A (checkbox)',
        'G10': 'Yüksek → B (checkbox)',
        'G11': 'Orta/Düşük → C (checkbox)',
        'H9': 'İlk defa (merged H9:K11)',
        'A12': 'Tekrarlayan aynı (merged A12:C14)',
        'E12': 'Tekrarlayan farklı (merged E12:F14)',
        'B22': 'İsim yazışı alanı',
        'D23': 'SSH Sorumlusu'
    }
    
    for cell_ref, description in critical_cells.items():
        cell = ws[cell_ref]
        print(f"\n✓ {cell_ref} ({description}):")
        print(f"    Değer: '{cell.value}'")
        print(f"    Tipir: {type(cell.value).__name__}")
        
        # Check if merged
        for merged_range in ws.merged_cells.ranges:
            if cell_ref in merged_range:
                print(f"    ⚠️  MERGED: {merged_range}")
                break
    
    print(f"\n\n📋 İLK 15 SATIRDA TÜM DOLU HÜCRELER:")
    print("-" * 80)
    for row_num in range(1, 16):
        for col_num in range(1, 13):  # A-L
            cell = ws.cell(row=row_num, column=col_num)
            if cell.value:
                col_letter = openpyxl.utils.get_column_letter(col_num)
                print(f"  {col_letter}{row_num:2}: {str(cell.value)[:60]}")
