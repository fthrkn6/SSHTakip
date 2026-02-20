#!/usr/bin/env python
from app import create_app
from openpyxl import load_workbook
import tempfile
import os

app = create_app()
with app.test_client() as client:
    client.post('/login', data={'username': 'testuser', 'password': 'password'})
    
    response = client.get('/servis/excel/daily-table')
    print(f'Status: {response.status_code}')
    
    if response.status_code == 200:
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as f:
            f.write(response.data)
            temp_path = f.name
        
        try:
            wb = load_workbook(temp_path)
            ws = wb['Servis Durumu']
            
            print(f'Rows: {ws.max_row}, Cols: {ws.max_column}')
            print(f'Header (Row 1) first 5: {[ws.cell(1, i).value for i in range(1, min(6, ws.max_column+1))]}')
            
            # Show first 5 rows
            for row in range(2, min(7, ws.max_row+1)):
                row_data = [ws.cell(row, i).value for i in range(1, min(6, ws.max_column+1))]
                print(f'Row {row}: {row_data}')
        finally:
            os.unlink(temp_path)
