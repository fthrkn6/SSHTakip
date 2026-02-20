#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""RCA raporu manuel olarak oluştur ve kontrol et"""
import sys
from datetime import datetime
from pathlib import Path
sys.path.insert(0, '.')

from app import create_app
from utils_root_cause_analysis import RootCauseAnalyzer

app = create_app()

with app.app_context():
    print("=" * 70)
    print("RCA RAPOR MANUEL OLUSTURMA")
    print("=" * 70)
    
    # Analiz yap
    print("\n1. Veriler analiz ediliyor...")
    analysis = RootCauseAnalyzer.analyze_service_disruptions(
        start_date='2026-02-01',
        end_date='2026-02-16'
    )
    
    print(f"   Kayit: {analysis['total_records']}")
    print(f"   Sistemler: {len(analysis['sistem_analysis'])}")
    
    if analysis['total_records'] == 0:
        print("\nHATA: Veri yok!")
        print("Sistem ve Alt Sistem Analizi:")
        sys.exit(1)
    
    # Sistem verilerini göster
    print("\n2. Sistem verileri:")
    for sistem, data in analysis['sistem_analysis'].items():
        print(f"   - {sistem}: {data['days']} gun, {data['count']} olay, {data['affected_tram_count']} arac")
        for alt_sistem, alt_data in data['alt_sistemler'].items():
            print(f"     * {alt_sistem}: {alt_data['days']} gun / {alt_data['count']} olay")
    
    # Excel oluştur
    print("\n3. Excel dosyasi olusturuluyor...")
    filepath = RootCauseAnalyzer.generate_rca_excel(analysis)
    print(f"   Dosya: {filepath}")
    
    # Excel dosyasını aç ve kontrol et
    from openpyxl import load_workbook
    
    wb = load_workbook(filepath)
    ws = wb.active
    
    print(f"\n4. Excel dosya kontrolu:")
    print(f"   Sayfalar: {wb.sheetnames}")
    print(f"   Aktif sayfa: {ws.title}")
    print(f"   Boyut (dimensions): {ws.dimensions}")
    print(f"   Satir sayisi: {ws.max_row}")
    print(f"   Sutun sayisi: {ws.max_column}")
    
    # İç veriyi göster
    print(f"\n5. Tablo verileri (ilk 15 satir):")
    count = 0
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=15), 1):
        has_data = False
        line = f"   R{row_idx}: "
        for col_idx, cell in enumerate(row[:8], 1):
            if cell.value:
                has_data = True
                val = str(cell.value)[:25]
                line += f"[{val}] "
            else:
                line += "[.] "
        if has_data:
            print(line)
            count += 1
    
    print(f"\n   Toplam dolu satir (ilk 15'te): {count}")
    
    print("\n" + "=" * 70)
    
    # Özel tablo kontrol
    print("\n6. Tablo kaynakları:")
    print(f"   Tablolar: {list(ws.tables.keys())}")
    for table_name, table in ws.tables.items():
        print(f"   - {table_name}: {table.ref}")
    
    print("\n" + "=" * 70)
