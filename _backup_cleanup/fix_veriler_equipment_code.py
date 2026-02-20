#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Veriler.xlsx'e equipment_code sütunu ekle
"""

import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

print("\n" + "="*80)
print("VERILER.XLSX'E EQUIPMENT_CODE SÜTUNU EKLE")
print("="*80)

# BELGRAD - 25 tramvay (1531-1555 -> BEL-01 to BEL-25)
belgrad_file = 'data/belgrad/Veriler.xlsx'
if os.path.exists(belgrad_file):
    print(f"\n📍 BELGRAD: {belgrad_file}")
    
    # Excel dosyasını aç
    wb = load_workbook(belgrad_file)
    ws = wb['Sayfa2']
    
    # Yeni sütun ekle (tram_id'den sonra, index 1'de equipment_code)
    # Önce başlığı kontrol et
    if ws.cell(1, 1).value and 'tram_id' in str(ws.cell(1, 1).value):
        print("   ✓ tram_id sütunu bulundu (column A)")
    
    # B sütununa equipment_code başlığı ekle
    ws.insert_cols(2)  # tram_id'den sonra yeni sütun ekle
    ws.cell(1, 2).value = 'equipment_code'
    print("   ✓ equipment_code sütunu eklendi (column B)")
    
    # Veriler ekle: 1531 -> BEL-01, 1532 -> BEL-02, ..., 1555 -> BEL-25
    for row in range(2, ws.max_row + 1):
        tram_id = ws.cell(row, 1).value
        if tram_id and isinstance(tram_id, (int, float)):
            # Sayıdan sıra numarasını çıkar
            number = int(tram_id) - 1530  # 1531 -> 1, 1532 -> 2, ..., 1555 -> 25
            if 1 <= number <= 25:
                equipment_code = f"BEL-{number:02d}"
                ws.cell(row, 2).value = equipment_code
    
    # Dosyayı kaydet
    wb.save(belgrad_file)
    print(f"   ✓ Dosya kaydedildi: {belgrad_file}")
    
    # Kontrol et
    df = pd.read_excel(belgrad_file, sheet_name='Sayfa2', header=0)
    print(f"   Columns: {list(df.columns)}")
    print(f"   First 5 rows:")
    for idx, row in df.head().iterrows():
        print(f"     {row['tram_id']} -> {row['equipment_code']}")

# KAYSERI - 11 tramvay (3872-3882 -> K001, K002, K003, ...)
kayseri_file = 'data/kayseri/Veriler.xlsx'
if os.path.exists(kayseri_file):
    print(f"\n📍 KAYSERI: {kayseri_file}")
    
    try:
        # Excel dosyasını aç
        wb = load_workbook(kayseri_file)
        ws = wb['Sayfa2']
        
        # B sütununa equipment_code başlığı ekle (varsa değiştir, yoksa yeni ekle)
        # Önce kontrol et - sütun adı ne?
        col_headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 2)]
        print(f"   Current columns: {col_headers}")
        
        # equipment_code sütunu varsa kullanıcının belirlediği format
        # Kayseri'de K001, K002, K003 olmalı
        
        # B sütununa ekle
        ws.insert_cols(2)
        ws.cell(1, 2).value = 'equipment_code'
        print("   ✓ equipment_code sütunu eklendi (column B)")
        
        # Veriler ekle: K001, K002, K003 (sadece 3 tane)
        kayseri_codes = ['K001', 'K002', 'K003']
        for row in range(2, min(ws.max_row + 1, 5)):  # Max 3 tramvay
            tram_id = ws.cell(row, 1).value
            if tram_id and row - 1 < len(kayseri_codes) + 1:
                equipment_code = kayseri_codes[row - 2] if row - 2 < len(kayseri_codes) else f"K{row-1:03d}"
                ws.cell(row, 2).value = equipment_code
        
        # Dosyayı kaydet
        wb.save(kayseri_file)
        print(f"   ✓ Dosya kaydedildi: {kayseri_file}")
        
        # Kontrol et
        df = pd.read_excel(kayseri_file, sheet_name='Sayfa2', header=0)
        print(f"   Columns: {list(df.columns)}")
        print(f"   First 5 rows:")
        for idx, row in df.head().iterrows():
            print(f"     {row['tram_id']} -> {row['equipment_code']}")
    except Exception as e:
        print(f"   ❌ Error: {e}")

print("\n" + "="*80)
