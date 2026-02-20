#!/usr/bin/env python
from app import create_app
from openpyxl import load_workbook
import tempfile
import os

app = create_app()
with app.test_client() as client:
    client.post('/login', data={'username': 'testuser', 'password': 'password'})
    
    response = client.get('/servis/excel/daily-table')
    
    if response.status_code == 200:
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as f:
            f.write(response.data)
            temp_path = f.name
        
        try:
            wb = load_workbook(temp_path)
            print(f'Sheet names: {wb.sheetnames}')
            print(f'Sheet count: {len(wb.sheetnames)}')
        finally:
            os.unlink(temp_path)
