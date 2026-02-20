#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Fracas_BELGRAD.xlsx başlıklarını kontrol et"""
from openpyxl import load_workbook
from pathlib import Path

excel_path = Path('logs/belgrad/ariza_listesi/Fracas_BELGRAD.xlsx')

if not excel_path.exists():
    print(f"Dosya bulunamadi: {excel_path}")
    exit(1)

wb = load_workbook(excel_path)
ws = wb.active

print("=" * 80)
print(f"Excel Dosyasi: {excel_path.name}")
print(f"Sayfa: {ws.title}")
print("=" * 80)

# 4. satır başlıkları
print("\n4. SATIR BASLIKLAR (Header):")
print("-" * 80)
headers = []
for col_idx in range(1, 50):
    cell = ws.cell(row=4, column=col_idx)
    if cell.value:
        headers.append((col_idx, cell.value))
        print(f"  Sutun {col_idx:2} (Col {chr(64+col_idx)}): {cell.value}")
    else:
        if col_idx > 20:  # 20'den sonra boş ise dur
            break

print(f"\nToplam Sutun (4. satirda): {len(headers)}")

# İlk birkaç satır veri göster
print("\n" + "=" * 80)
print("ILK VERI SATIRI (5. satir):")
print("-" * 80)
for col_idx, header in headers[:10]:
    cell_value = ws.cell(row=5, column=col_idx).value
    print(f"  {header}: {cell_value}")

print("\n" + "=" * 80)
