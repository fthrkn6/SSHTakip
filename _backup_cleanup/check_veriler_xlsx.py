#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""Veriler.xlsx Sayfa2 tram_id sütununu kontrol et"""

import os
from pathlib import Path
from openpyxl import load_workbook

base_path = Path('c:\\Users\\fatiherkin\\Desktop\\bozankaya_ssh_takip')
projects = ['belgrad', 'kayseri', 'iasi', 'timisoara', 'kocaeli', 'gebze']

print("\n" + "="*70)
print("📋 VERİLER.XLSX SAYFA2 - TRAM_ID SÜTUNU KONTROL")
print("="*70)

for project in projects:
    veriler_path = base_path / 'data' / project / 'Veriler.xlsx'
    
    print(f"\n🚊 {project.upper()}: {veriler_path}")
    
    if not veriler_path.exists():
        print(f"  ⚠️  DOSYA BULUNAMADI")
        continue
    
    try:
        wb = load_workbook(veriler_path)
        
        # Sayfa2'yi kontrol et
        if 'Sayfa2' in wb.sheetnames:
            ws = wb['Sayfa2']
            
            # Sütun başlıklarını kontrol et (ilk satır)
            headers = []
            for cell in ws[1]:
                if cell.value:
                    headers.append(cell.value)
            
            print(f"  📊 Sütun Başlıkları: {headers}")
            
            # tram_id sütununu bul
            tram_id_col = None
            for i, header in enumerate(headers, 1):
                if 'tram' in str(header).lower() or 'araç' in str(header).lower():
                    tram_id_col = i
                    print(f"  ✓ Sütun {i}: {header}")
                    break
            
            if tram_id_col:
                # tram_id'leri oku (satır 2'den başla)
                tram_ids = []
                for row in ws.iter_rows(min_row=2, max_col=tram_id_col, values_only=True):
                    if row and row[tram_id_col-1]:
                        tram_id = str(row[tram_id_col-1]).strip()
                        if tram_id and tram_id.upper() != 'NONE':
                            tram_ids.append(tram_id)
                
                print(f"  📍 Bulunan Tramvaylar ({len(tram_ids)}):")
                for i, tid in enumerate(tram_ids[:15], 1):  # İlk 15'i göster
                    print(f"     {i}. {tid}")
                if len(tram_ids) > 15:
                    print(f"     ... ve {len(tram_ids)-15} daha")
            else:
                print(f"  ❌ tram_id sütunu bulunamadı")
                print(f"     Mevcut sütunlar: {headers}")
        else:
            print(f"  ❌ 'Sayfa2' sayfası bulunamadı")
            print(f"     Mevcut sayfalar: {wb.sheetnames}")
    
    except Exception as e:
        print(f"  ❌ Hata: {e}")

print("\n" + "="*70)
print("\n✅ Kontrol tamamlandı. Doğru mu? Evet ise 'y' yaz, değişiklik istiyorsan 'n' yaz.")
print("="*70 + "\n")
