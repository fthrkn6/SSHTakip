import openpyxl

# Load workbook
wb = openpyxl.load_workbook('data/belgrad/FR_010_R06_SSH HBR.xlsx')
ws = wb.active

print('='*80)
print('HBR EXCEL DOSYASI YAPISAL ANALİZİ')
print('='*80)

print('\n📊 MERGED HÜCRELERİ:')
for merged_range in ws.merged_cells.ranges:
    print(f'  {merged_range}')

print('\n\n📋 İLK 15 SATIR - HÜCRELERİN İÇERİĞİ:')
print('Row | Cell       | Value')
print('-'*80)
for row_num in range(1, 16):
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
        cell = ws[f'{col}{row_num}']
        if cell.value:
            value_str = str(cell.value)[:45]
            print(f'{row_num:3} | {col}{row_num:6} | {value_str}')

print('\n\n🎯 KONTROL ALANLAR (G9, G10, G11, H9, A12, E12):')
cells_to_check = ['G9', 'G10', 'G11', 'H9', 'A12', 'E12']
for cell_ref in cells_to_check:
    cell = ws[cell_ref]
    is_merged = False
    merged_parent = None
    for merged_range in ws.merged_cells.ranges:
        if cell_ref in merged_range:
            is_merged = True
            merged_parent = str(merged_range)
            break
    
    print(f'  {cell_ref}: Değer="{cell.value}" | Merged={is_merged}', end='')
    if is_merged:
        print(f' [{merged_parent}]')
    else:
        print()

print('\n\n📊 TÜÜM HÜCRE YAPISALI:')
print(f'Max Row: {ws.max_row}')
print(f'Max Column: {ws.max_column}')

print('\n✅ DEŞ LU HÜCRE LİSTESİ:')
for row in ws.iter_rows():
    for cell in row:
        if cell.value:
            # Kontrol et merged mi
            is_merged = False
            for merged_range in ws.merged_cells.ranges:
                if cell.coordinate in merged_range:
                    is_merged = True
                    break
            merged_info = ' [MERGED]' if is_merged else ''
            print(f'  {cell.coordinate:4} = {str(cell.value)[:50]}{merged_info}')
