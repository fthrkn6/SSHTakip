from openpyxl import load_workbook

wb = load_workbook('data/belgrad/Veriler_KATEGORIZE.xlsx')
ws = wb['Sayfa2']

print('\n' + '='*100)
print('KATEGORIZE EDILMIŞ EXCEL YAPISI')
print('='*100 + '\n')

# Header göster
print(f"{'Tram':<8} | {'L1: ANA SİSTEM':<30} | {'L2: ALT SİS':<20} | {'L3: BİLEŞEN':<20} | {'L4: PARÇA':<20}")
print('-'*100)

# İlk 8 veri satırını göster (header'ı atla)
for row_idx in range(3, min(11, ws.max_row + 1)):
    tram = ws.cell(row=row_idx, column=1).value
    l1 = ws.cell(row=row_idx, column=6).value or '---'
    l2 = ws.cell(row=row_idx, column=7).value or '---'
    l3 = ws.cell(row=row_idx, column=8).value or '---'
    l4 = ws.cell(row=row_idx, column=9).value or '---'
    
    print(f"{str(tram):<8} | {str(l1)[:28]:<30} | {str(l2)[:18]:<20} | {str(l3)[:18]:<20} | {str(l4)[:18]:<20}")

print('\n' + '='*100)
