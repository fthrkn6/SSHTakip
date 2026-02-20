#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""Gerçek tramvay verilerini Veriler.xlsx'ten oku ve ekle"""

from app import create_app, db
from models import Equipment
import os
from openpyxl import load_workbook
from pathlib import Path

app = create_app()

with app.app_context():
    print("\n" + "="*60)
    print("📊 GERÇEK TRAMVAY VERİSİ EKLENIYOR...")
    print("="*60)
    
    base_path = Path(app.root_path)
    projects = ['belgrad', 'kayseri', 'iasi', 'timisoara', 'kocaeli', 'gebze']
    
    total_added = 0
    
    for project in projects:
        veriler_path = base_path / 'data' / project / 'Veriler.xlsx'
        
        if not veriler_path.exists():
            print(f"\n⚠️  {project}: {veriler_path} bulunamadı")
            continue
        
        print(f"\n🚊 {project.upper()} - {veriler_path.name} okunuyor...")
        
        try:
            wb = load_workbook(veriler_path)
            
            # Sayfa2'yi kontrol et
            if 'Sayfa2' in wb.sheetnames:
                ws = wb['Sayfa2']
                
                # Tram ID'lerini oku (ilk sütun)
                tram_ids = []
                for row in ws.iter_rows(min_row=3, values_only=True):  # Başlık satırını geç
                    if row[0]:  # İlk sütun boş değilse
                        tram_id = str(row[0]).strip()
                        if tram_id and tram_id.upper() != 'NONE':
                            tram_ids.append(tram_id)
                
                print(f"  Bulundu: {len(tram_ids)} tramvay")
                
                # Database'e ekle
                added = 0
                for tram_id in tram_ids:
                    # Zaten varsa skip et
                    existing = Equipment.query.filter_by(equipment_code=tram_id).first()
                    if existing:
                        continue
                    
                    eq = Equipment(
                        equipment_code=tram_id,
                        name=f'{project.capitalize()} - {tram_id}',
                        equipment_type='Tramway',
                        status='aktif',  # Default olarak aktif
                        project_code=project,
                        location=project.capitalize(),
                        criticality='high'
                    )
                    db.session.add(eq)
                    added += 1
                
                db.session.commit()
                print(f"  ✅ {added} tramvay eklendi")
                total_added += added
            else:
                print(f"  ⚠️  'Sayfa2' sayfası bulunamadı (Sayfalar: {wb.sheetnames})")
                
        except Exception as e:
            print(f"  ❌ Hata: {e}")
    
    print("\n" + "="*60)
    print(f"✅ Toplam {total_added} gerçek tramvay eklendi!")
    print("="*60 + "\n")
