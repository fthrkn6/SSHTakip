#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""Test script to verify Detaylı Bilgi field handling in FRACAS template"""

import sys
import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime

# Add current directory to path
sys.path.insert(0, os.path.dirname(__file__))

from utils_fracas_writer import FracasWriter

def test_detayli_bilgi():
    """Test that detayli_bilgi field is correctly written to Column Z"""
    
    print("\n" + "="*60)
    print("TEST: Detaylı Bilgi Field Handling")
    print("="*60)
    
    # Prepare test data
    test_data = {
        'arac_numarasi': 'TEST-001',    # Required
        'sistem': 'Motor',
        'alt_sistem': 'Elektrik',
        'tedarikci': 'Test Supplier',
        'hata_tarih': '2026-02-17',     # Required
        'hata_saat': '14:30',           # Required
        'ariza_tanimi': 'Test failure',  # Required
        'ariza_sinifi': 'B',            # Required
        'ariza_kaynagi': 'Design',
        'yapilan_islem': 'Test action',
        'aksiyon': 'Tamir',
        'garanti_kapsami': 'Evet',
        'ariza_tespit_yontemi': 'Visual',
        'detayli_bilgi': 'Servise Engel',  # NEW FIELD - maps to Column Z
        'parca_kodu': 'TP-001',
        'parca_adi': 'Test Part',
        'parca_seri_no': 'SN-12345',
        'adet': 1,
        'tamir_baslama_tarih': '2026-02-17',
        'tamir_baslama_saati': '15:00',
        'tamir_bitisi_tarih': '2026-02-17',
        'tamir_bitisi_saati': '16:00',
        'personel_sayisi': 2,
    }
    
    # Initialize FracasWriter
    print("\n1. Initializing FracasWriter...")
    writer = FracasWriter('belgrad')
    
    # Write test data
    print("2. Writing test data to FRACAS template...")
    try:
        result = writer.write_failure_data(test_data)
        if result.get('success'):
            print(f"✅ Data written successfully")
            print(f"   Row: {result['row']}")
            print(f"   File: {result['file_path']}")
            fracas_file = result['file_path']
        else:
            print(f"❌ Write failed: {result.get('error')}")
            return False
    except Exception as e:
        print(f"❌ Exception during write: {str(e)}")
        import traceback
        traceback.print_exc()
        return False
    
    # Verify data was written to correct columns
    print("\n3. Verifying data in FRACAS template...")
    try:
        # Load the FRACAS file
        xl_book = load_workbook(fracas_file)
        ws = xl_book['FRACAS']
        
        row_num = result['row']
        
        # Check key columns
        checks = [
            ('B (Araç No)', 'B', 'TEST-001'),
            ('Z (Detaylı Bilgi)', 'Z', 'Servise Engel'),  # NEW FIELD CHECK - PRIMARY TEST
            ('Q (Personel)', 'Q', 2),
        ]
        
        print(f"\n   Checking row {row_num} in FRACAS sheet:")
        all_passed = True
        
        for label, col, expected_value in checks:
            cell_value = ws[f'{col}{row_num}'].value
            
            # Handle different types of comparisons
            if isinstance(expected_value, int):
                try:
                    passed = int(cell_value) == expected_value if cell_value else False
                except (ValueError, TypeError):
                    passed = False
            else:
                passed = str(cell_value) == str(expected_value) if cell_value else False
            
            status = "✅" if passed else "❌"
            print(f"   {status} {label}: {cell_value} (expected: {expected_value})")
            
            if not passed:
                all_passed = False
        
        if all_passed:
            print("\n✅ ALL CHECKS PASSED!")
            print("   Detaylı Bilgi field correctly written to Column Z")
            return True
        else:
            print("\n❌ SOME CHECKS FAILED!")
            return False
            
    except Exception as e:
        print(f"❌ Verification failed: {str(e)}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == '__main__':
    success = test_detayli_bilgi()
    sys.exit(0 if success else 1)
