from openpyxl import load_workbook
import json

excel_file = 'data/belgrad/Belgrad-Bakım.xlsx'
wb = load_workbook(excel_file)

# Bakım seviyeleri sırası
maintenance_levels = ['6K', '18K', '24K', '36K', '60K', '70K', '100K', '140K', '210K', '300K']

# Sayısal part'ı çıkar
def get_km_value(level_str):
    return int(level_str.replace('K', '')) * 1000

all_maintenance = {}

for level in maintenance_levels:
    sheet = wb[level]
    km_value = get_km_value(level)
    
    # WORKS satırını bul
    works_row = None
    for row in sheet.iter_rows(min_row=1, max_row=50):
        for cell in row:
            if cell.value and str(cell.value).strip() == 'WORKS':
                works_row = cell.row
                works_col = cell.column
                break
        if works_row:
            break
    
    if works_row:
        works = []
        # WORKS satırından sonra boş olmayan hücreleri topla
        for i in range(works_row + 1, sheet.max_row + 1):
            cell = sheet.cell(row=i, column=works_col)
            if cell.value and str(cell.value).strip():
                works.append(str(cell.value).strip())
            elif works:  # Boş satır sonra dene
                if cell.value is None:
                    continue
                else:
                    break
        
        all_maintenance[level] = {
            'km': km_value,
            'label': f'{level} KM Bakımı',
            'works': works[:30]  # Max 30 iş
        }

# JSON olarak kaydet
with open('data/belgrad/maintenance.json', 'w', encoding='utf-8') as f:
    json.dump(all_maintenance, f, ensure_ascii=False, indent=2)

print("✅ Bakım verileri başarıyla çıkarıldı!")
print(f"\nToplamda {len(all_maintenance)} bakım seviyesi:")
for level, data in all_maintenance.items():
    print(f"  {level}: {len(data['works'])} iş")
