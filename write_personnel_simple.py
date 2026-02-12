import openpyxl

excel_file = r'logs\ariza_listesi\Ariza_Listesi_BELGRAD.xlsx'

wb = openpyxl.load_workbook(excel_file)
ws = wb['Ariza Listesi']

# Personel sayılarını yaz (Row 5-12, Column 30)
personel_sayilari = [2, 1, 3, 2, 1, 2, 1, 3]

print('Personel Sayılarını Excel\'e Yazılıyor:')
print('=' * 50)

for idx, personel_sayı in enumerate(personel_sayilari):
    row_idx = 5 + idx
    araç = ws.cell(row_idx, 2).value
    ws.cell(row_idx, 30).value = personel_sayı
    print(f'Row {row_idx}: Araç={araç}, Personel Sayısı={personel_sayı}')

# Dosyayı kaydet
wb.save(excel_file)
wb.close()

print('\n✅ Tüm veriler kaydedildi!')

# Doğrula
print('\nDoğrulanıyor:')
wb2 = openpyxl.load_workbook(excel_file)
ws2 = wb2['Ariza Listesi']
for idx in range(8):
    val = ws2.cell(5 + idx, 30).value
    print(f'Row {5+idx}, Col 30: {val}')
wb2.close()
