from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
import os
import shutil
import time

project = 'kayseri'
base_path = r'c:\Users\fatiherkin\Desktop\bozankaya_ssh_takip\data'
grid_path = os.path.join(base_path, project, 'service_status_grid.xlsx')
backup_path = grid_path + '.backup'

print(f"\n{'='*80}")
print(f"KAYSERI GRID - SEMBOL NORMALİZASYONU")
print(f"{'='*80}\n")

# Dosya kilitliyse, birkaç saniye bekle
max_retries = 3
retry_count = 0

while retry_count < max_retries:
    try:
        # Backup yap
        if not os.path.exists(backup_path):
            shutil.copy2(grid_path, backup_path)
            print(f"✓ Backup oluşturuldu: {backup_path}")
        
        # Dosyayı aç ve düzenle
        wb = load_workbook(grid_path)
ws = wb.active

# Symbol replacement mapping
replace_count = 0
symbol_map = {
    '√': '✓',      # SQUARE ROOT → CHECK MARK
    '⚠': '⚠',      # WARNING (keep as-is)
    '✗': '✗',      # BALLOT X (keep as-is)
}

# Status codes
status_codes = {
    '✓': 'aktif',
    '⚠': 'isletme_kaynakli',
    '✗': 'servis_disi'
}

# Color mapping
STATUS_COLORS = {
    'aktif': 'FF00B050',           # Yeşil
    'servis_disi': 'FFFF0000',     # Kırmızı
    'isletme_kaynakli': 'FFFFC000' # Turuncu
}

# Tüm veri hücrelerini kontrol et ve değiştir
replaced = 0
for row_idx in range(2, ws.max_row + 1):
    for col_idx in range(2, ws.max_column + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell_value = cell.value
        
        if cell_value:
            str_value = str(cell_value).strip()
            
            # Sembol değiştir
            if str_value in symbol_map:
                new_symbol = symbol_map[str_value]
                if str_value != new_symbol:
                    cell.value = new_symbol
                    replaced += 1
                
                # Renk ve stil uygula
                status = status_codes.get(new_symbol)
                if status and status in STATUS_COLORS:
                    cell.fill = PatternFill(
                        start_color=STATUS_COLORS[status],
                        end_color=STATUS_COLORS[status],
                        fill_type='solid'
                    )
                
                cell.font = Font(bold=True, size=14)
                cell.alignment = Alignment(horizontal='center', vertical='center')

wb.save(grid_path)
print(f"✓ Normalizasyon tamamlandı!")
print(f"  Değiştirilen sembol: √ → ✓ ({replaced} hücre)")
print(f"  Dosya kaydedildi: {grid_path}")
print(f"\n{'='*80}\n")
