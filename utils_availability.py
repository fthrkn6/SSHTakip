"""
Availability Metrikleri ve Raporlama Utility
Araç kullanılabilirlik hesaplamaları ve raporlar
"""

from datetime import datetime, timedelta, date
from sqlalchemy import func
from models import db, ServiceLog, AvailabilityMetrics, RootCauseAnalysis, Failure, Equipment
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
import logging

# Logging konfigürasyonu
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('logs/availability.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)


class AvailabilityCalculator:
    """Availability metriklerini hesapla"""
    
    @staticmethod
    def calculate_daily_availability(tram_id, target_date=None):
        """Günlük availability hesapla"""
        if target_date is None:
            target_date = date.today()
        
        logger.info(f"Calculating daily availability for {tram_id} on {target_date}")
        
        # Mevcut metrikleri kontrol et
        existing = AvailabilityMetrics.query.filter_by(
            tram_id=tram_id,
            metric_date=target_date,
            report_period='daily'
        ).first()
        
        if existing:
            return existing
        
        # Log verilerinden hesapla
        start_time = datetime.combine(target_date, datetime.min.time())
        end_time = datetime.combine(target_date, datetime.max.time())
        
        logs = ServiceLog.query.filter(
            ServiceLog.tram_id == tram_id,
            ServiceLog.log_date >= start_time,
            ServiceLog.log_date <= end_time
        ).order_by(ServiceLog.log_date).all()
        
        total_downtime = sum(log.duration_hours or 0 for log in logs 
                            if log.new_status in ['Servis Dışı', 'Bakımda', 'Arızalı'])
        
        operational_hours = 24 - total_downtime
        availability_pct = (operational_hours / 24) * 100 if 24 > 0 else 0
        
        # Metrikleri kaydet
        metric = AvailabilityMetrics(
            tram_id=tram_id,
            metric_date=target_date,
            total_hours=24,
            operational_hours=operational_hours,
            downtime_hours=total_downtime,
            availability_percentage=round(availability_pct, 2),
            report_period='daily'
        )
        
        db.session.add(metric)
        db.session.commit()
        logger.info(f"Daily availability calculated: {availability_pct}%")
        return metric
    
    @staticmethod
    def calculate_period_availability(tram_id, start_date, end_date, period_type='weekly'):
        """Belirtilen dönem için availability hesapla"""
        
        logger.info(f"Calculating {period_type} availability for {tram_id} from {start_date} to {end_date}")
        
        # Mevcut metrikleri kontrol et
        existing = AvailabilityMetrics.query.filter_by(
            tram_id=tram_id,
            metric_date=start_date,
            report_period=period_type
        ).first()
        
        if existing:
            return existing
        
        logs = ServiceLog.query.filter(
            ServiceLog.tram_id == tram_id,
            ServiceLog.log_date >= start_date,
            ServiceLog.log_date <= end_date
        ).all()
        
        # Dönem saatlerini hesapla
        period_hours = (end_date - start_date).total_seconds() / 3600
        total_downtime = sum(log.duration_hours or 0 for log in logs)
        operational_hours = period_hours - total_downtime
        availability_pct = (operational_hours / period_hours) * 100 if period_hours > 0 else 0
        
        # Sistem bazında kategorize
        systems_downtime = {}
        for log in logs:
            sistem = log.sistem or 'Bilinmiyor'
            if sistem not in systems_downtime:
                systems_downtime[sistem] = 0
            systems_downtime[sistem] += log.duration_hours or 0
        
        # Failure sayısı
        failures = Failure.query.filter(
            Failure.tram_id == tram_id,
            Failure.failure_date >= start_date,
            Failure.failure_date <= end_date
        ).count()
        
        metric = AvailabilityMetrics(
            tram_id=tram_id,
            metric_date=start_date,
            total_hours=period_hours,
            operational_hours=operational_hours,
            downtime_hours=total_downtime,
            availability_percentage=round(availability_pct, 2),
            report_period=period_type,
            failure_count=failures
        )
        
        db.session.add(metric)
        db.session.commit()
        logger.info(f"Period availability calculated: {availability_pct}%")
        return metric
    
    @staticmethod
    def get_availability_report(tram_id, start_date, end_date):
        """Detaylı availability raporu getir"""
        
        logger.info(f"Generating availability report for {tram_id}")
        
        report = {
            'tram_id': tram_id,
            'period': f"{start_date.strftime('%d.%m.%Y')} - {end_date.strftime('%d.%m.%Y')}",
            'daily_metrics': [],
            'system_breakdown': {},
            'total_availability': 0
        }
        
        # Günlük metrikler
        current_date = start_date
        total_hours = 0
        total_operational = 0
        
        while current_date <= end_date:
            daily_metric = AvailabilityCalculator.calculate_daily_availability(tram_id, current_date)
            report['daily_metrics'].append({
                'date': current_date.strftime('%d.%m.%Y'),
                'availability': daily_metric.availability_percentage,
                'downtime': daily_metric.downtime_hours
            })
            
            total_hours += daily_metric.total_hours
            total_operational += daily_metric.operational_hours
            current_date += timedelta(days=1)
        
        # Toplam availability
        report['total_availability'] = round((total_operational / total_hours * 100) if total_hours > 0 else 0, 2)
        
        # Sistem bazında breakdown
        logs = ServiceLog.query.filter(
            ServiceLog.tram_id == tram_id,
            ServiceLog.log_date >= datetime.combine(start_date, datetime.min.time()),
            ServiceLog.log_date <= datetime.combine(end_date, datetime.max.time())
        ).all()
        
        for log in logs:
            sistem = log.sistem or 'Bilinmiyor'
            if sistem not in report['system_breakdown']:
                report['system_breakdown'][sistem] = {
                    'downtime': 0,
                    'incidents': 0,
                    'subsystems': {}
                }
            
            report['system_breakdown'][sistem]['downtime'] += log.duration_hours or 0
            report['system_breakdown'][sistem]['incidents'] += 1
            
            if log.alt_sistem:
                if log.alt_sistem not in report['system_breakdown'][sistem]['subsystems']:
                    report['system_breakdown'][sistem]['subsystems'][log.alt_sistem] = {
                        'downtime': 0,
                        'incidents': 0
                    }
                report['system_breakdown'][sistem]['subsystems'][log.alt_sistem]['downtime'] += log.duration_hours or 0
                report['system_breakdown'][sistem]['subsystems'][log.alt_sistem]['incidents'] += 1
        
        return report


class ExcelReportGenerator:
    """Excel raporları oluştur"""
    
    @staticmethod
    def create_availability_report_excel(tram_ids, start_date, end_date, output_path):
        """Availability raporu Excel'e aktar"""
        
        logger.info(f"Creating availability Excel report for {len(tram_ids)} vehicles")
        
        wb = Workbook()
        wb.remove(wb.active)
        
        # Stilleri tanımla
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        subheader_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        subheader_font = Font(bold=True, size=10)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 1. ÖZET SAYFASI
        ws_summary = wb.create_sheet('Özet', 0)
        ws_summary['A1'] = 'ARAÇ KULLANILABILIRLIK RAPORU - ÖZET'
        ws_summary['A1'].font = Font(bold=True, size=14, color="FFFFFF")
        ws_summary['A1'].fill = header_fill
        ws_summary.merge_cells('A1:G1')
        
        ws_summary['A2'] = f"Dönem: {start_date.strftime('%d.%m.%Y')} - {end_date.strftime('%d.%m.%Y')}"
        ws_summary['A2'].font = Font(size=11)
        ws_summary['A3'] = f"Rapor Tarihi: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
        
        row = 5
        headers = ['Araç ID', 'Toplam Kullanılabilirlik %', 'Downtime Saatleri', 
                   'Operasyon Saatleri', 'Toplam Saatler', 'Arıza Sayısı']
        
        for col, header in enumerate(headers, 1):
            cell = ws_summary.cell(row=row, column=col)
            cell.value = header
            cell.fill = subheader_fill
            cell.font = subheader_font
            cell.border = border
        
        row = 6
        for tram_id in tram_ids:
            report = AvailabilityCalculator.get_availability_report(tram_id, start_date, end_date)
            
            ws_summary.cell(row=row, column=1).value = tram_id
            ws_summary.cell(row=row, column=2).value = report['total_availability']
            
            # Downtime topla
            total_downtime = sum(m['downtime'] for m in report['daily_metrics'])
            ws_summary.cell(row=row, column=3).value = round(total_downtime, 2)
            
            # Operasyon saatleri
            period_hours = (end_date - start_date).days * 24
            operational = period_hours - total_downtime
            ws_summary.cell(row=row, column=4).value = round(operational, 2)
            ws_summary.cell(row=row, column=5).value = period_hours
            
            # Arıza sayısı
            failure_count = Failure.query.filter(
                Failure.tram_id == tram_id,
                Failure.failure_date >= datetime.combine(start_date, datetime.min.time()),
                Failure.failure_date <= datetime.combine(end_date, datetime.max.time())
            ).count()
            ws_summary.cell(row=row, column=6).value = failure_count
            
            row += 1
        
        # Sütun genişliğini ayarla
        ws_summary.column_dimensions['A'].width = 15
        ws_summary.column_dimensions['B'].width = 18
        ws_summary.column_dimensions['C'].width = 15
        ws_summary.column_dimensions['D'].width = 18
        ws_summary.column_dimensions['E'].width = 15
        ws_summary.column_dimensions['F'].width = 12
        
        # 2. GÜNLÜK METRİKLER SAYFASI
        ws_daily = wb.create_sheet('Günlük Metrikler', 1)
        ws_daily['A1'] = 'GÜNLÜK KULLANILABILIRLIK METRİKLERİ'
        ws_daily['A1'].font = Font(bold=True, size=12, color="FFFFFF")
        ws_daily['A1'].fill = header_fill
        ws_daily.merge_cells('A1:D1')
        
        row = 3
        for tram_id in tram_ids:
            ws_daily.cell(row=row, column=1).value = f"Araç: {tram_id}"
            ws_daily.cell(row=row, column=1).font = Font(bold=True, size=11)
            
            row += 1
            
            daily_headers = ['Tarih', 'Kullanılabilirlik %', 'Downtime (Saat)', 'Operasyon (Saat)']
            for col, header in enumerate(daily_headers, 1):
                cell = ws_daily.cell(row=row, column=col)
                cell.value = header
                cell.fill = subheader_fill
                cell.font = subheader_font
                cell.border = border
            
            row += 1
            
            report = AvailabilityCalculator.get_availability_report(tram_id, start_date, end_date)
            for daily_metric in report['daily_metrics']:
                ws_daily.cell(row=row, column=1).value = daily_metric['date']
                ws_daily.cell(row=row, column=2).value = daily_metric['availability']
                ws_daily.cell(row=row, column=3).value = daily_metric['downtime']
                ws_daily.cell(row=row, column=4).value = 24 - daily_metric['downtime']
                row += 1
            
            row += 2
        
        ws_daily.column_dimensions['A'].width = 15
        ws_daily.column_dimensions['B'].width = 18
        ws_daily.column_dimensions['C'].width = 15
        ws_daily.column_dimensions['D'].width = 15
        
        # 3. SİSTEM BAZINDA ANALIZ
        ws_system = wb.create_sheet('Sistem Analizi', 2)
        ws_system['A1'] = 'SİSTEM BAZINDA DOWNTIME ANALİZİ'
        ws_system['A1'].font = Font(bold=True, size=12, color="FFFFFF")
        ws_system['A1'].fill = header_fill
        ws_system.merge_cells('A1:E1')
        
        row = 3
        system_headers = ['Araç ID', 'Sistem', 'Alt Sistem', 'Downtime (Saat)', 'Olay Sayısı']
        for col, header in enumerate(system_headers, 1):
            cell = ws_system.cell(row=row, column=col)
            cell.value = header
            cell.fill = subheader_fill
            cell.font = subheader_font
            cell.border = border
        
        row = 4
        for tram_id in tram_ids:
            report = AvailabilityCalculator.get_availability_report(tram_id, start_date, end_date)
            
            for sistem, data in report['system_breakdown'].items():
                ws_system.cell(row=row, column=1).value = tram_id
                ws_system.cell(row=row, column=2).value = sistem
                ws_system.cell(row=row, column=3).value = '-'
                ws_system.cell(row=row, column=4).value = round(data['downtime'], 2)
                ws_system.cell(row=row, column=5).value = data['incidents']
                row += 1
                
                # Alt sistemler
                for alt_sistem, subsys_data in data['subsystems'].items():
                    ws_system.cell(row=row, column=1).value = tram_id
                    ws_system.cell(row=row, column=2).value = sistem
                    ws_system.cell(row=row, column=3).value = alt_sistem
                    ws_system.cell(row=row, column=4).value = round(subsys_data['downtime'], 2)
                    ws_system.cell(row=row, column=5).value = subsys_data['incidents']
                    row += 1
        
        for col in range(1, 6):
            ws_system.column_dimensions[get_column_letter(col)].width = 15
        
        # Dosyayı kaydet
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        wb.save(output_path)
        logger.info(f"Availability Excel report saved to {output_path}")
        
        return output_path
    
    @staticmethod
    def create_root_cause_analysis_excel(tram_ids, output_path):
        """Root Cause Analysis Excel raporu oluştur"""
        
        logger.info(f"Creating Root Cause Analysis Excel report")
        
        wb = Workbook()
        ws = wb.active
        ws.title = 'Root Cause Analizi'
        
        # Header
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        ws['A1'] = 'ROOT CAUSE ANALYSIS (KÖK NEDEN ANALİZİ) RAPORU'
        ws['A1'].font = Font(bold=True, size=14, color="FFFFFF")
        ws['A1'].fill = header_fill
        ws.merge_cells('A1:H1')
        
        ws['A2'] = f"Oluşturulma Tarihi: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
        
        row = 4
        headers = [
            'Araç ID', 'Sistem', 'Alt Sistem', 'Arıza Açıklaması', 
            'Kök Neden', 'Şiddet Seviyesi', 'Sıklık', 'Durum'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
        
        row = 5
        
        for tram_id in tram_ids:
            analyses = RootCauseAnalysis.query.filter_by(tram_id=tram_id).all()
            
            for analysis in analyses:
                ws.cell(row=row, column=1).value = tram_id
                ws.cell(row=row, column=2).value = analysis.sistem
                ws.cell(row=row, column=3).value = analysis.alt_sistem or '-'
                ws.cell(row=row, column=4).value = analysis.failure_description[:50]
                ws.cell(row=row, column=5).value = analysis.root_cause[:50]
                ws.cell(row=row, column=6).value = analysis.severity_level
                ws.cell(row=row, column=7).value = analysis.frequency
                ws.cell(row=row, column=8).value = analysis.status
                
                row += 1
        
        # Sütun genişliğini ayarla
        column_widths = [12, 15, 15, 25, 25, 15, 10, 12]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width
        
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        wb.save(output_path)
        logger.info(f"Root Cause Analysis Excel report saved to {output_path}")
        
        return output_path


def log_service_status_change(tram_id, new_status, sistem=None, alt_sistem=None, reason=None, duration_hours=0, user_id=None):
    """Servis durumu değişikliğini log'la"""
    
    # Son durumu bul
    last_log = ServiceLog.query.filter_by(tram_id=tram_id).order_by(ServiceLog.log_date.desc()).first()
    previous_status = last_log.new_status if last_log else None
    
    log = ServiceLog(
        tram_id=tram_id,
        log_date=datetime.utcnow(),
        previous_status=previous_status,
        new_status=new_status,
        sistem=sistem,
        alt_sistem=alt_sistem,
        reason=reason,
        duration_hours=duration_hours,
        created_by=user_id,
        notes=f"Status changed from {previous_status} to {new_status}"
    )
    
    db.session.add(log)
    db.session.commit()
    
    logger.info(f"Service status logged for {tram_id}: {previous_status} -> {new_status}")
    
    return log
