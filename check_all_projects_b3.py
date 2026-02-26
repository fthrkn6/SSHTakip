#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Her projede veriler.xlsx B3 hücresini kontrol et
"""

from openpyxl import load_workbook
import os
from pathlib import Path

projects = [
    'belgrad', 'gebze', 'iasi', 'kayseri', 'kocaeli', 'timisoara'
]

print("📊 Her Proje - veriler.xlsx B3 Hücre İçeriği")
print("=" * 60)

for project in projects:
    veriler_path = os.path.join('data', project, 'veriler.xlsx')
    
    if os.path.exists(veriler_path):
        try:
            wb = load_workbook(veriler_path)
            ws = wb.active
            
            b3_value = ws['B3'].value or '(boş)'
            print(f"✓ {project.upper():12} → B3: {b3_value}")
            
            wb.close()
        except Exception as e:
            print(f"❌ {project.upper():12} → Hata: {e}")
    else:
        print(f"❌ {project.upper():12} → Dosya yok")

print("=" * 60)
