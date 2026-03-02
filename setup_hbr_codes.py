#!/usr/bin/env python
# -*- coding: utf-8 -*-

import os
import sys
from pathlib import Path

# Change to script directory
script_dir = os.path.dirname(os.path.abspath(__file__))
os.chdir(script_dir)

try:
    from openpyxl import load_workbook
    
    # Setup paths for all projects
    projects_setup = {
        'kayseri': 'KAY5+6',
        'belgrad': 'BEL25',
        'iasi': 'IAS25',
        'timisoara': 'TIM25',
        'gebze': 'GEB25',
        'kocaeli': 'KOC25'
    }
    
    for project, code in projects_setup.items():
        data_dir = os.path.join('data', project)
        
        # Find Veriler.xlsx (case-insensitive)
        veriler_file = None
        if os.path.exists(data_dir):
            for f in os.listdir(data_dir):
                if f.lower() == 'veriler.xlsx':
                    veriler_file = os.path.join(data_dir, f)
                    break
        
        if veriler_file:
            print(f"\n[{project}] {veriler_file}")
            try:
                wb = load_workbook(veriler_file)
                print(f"  Sheets: {wb.sheetnames}")
                
                if 'Sayfa2' in wb.sheetnames:
                    ws = wb['Sayfa2']
                    current_val = ws['B2'].value
                    print(f"  B2 (before): {current_val}")
                    
                    ws['B2'].value = code
                    wb.save(veriler_file)
                    print(f"  B2 (after): {code}")
                    print(f"  ✓ Saved!")
                else:
                    print(f"  ! Sayfa2 sheet not found. Available: {wb.sheetnames}")
                
                wb.close()
            except Exception as e:
                print(f"  ✗ Error: {e}")
        else:
            print(f"\n[{project}] ! veriler.xlsx not found in {data_dir}")
    
    print("\nDone!")

except Exception as e:
    import traceback
    print(f"Fatal error: {e}")
    traceback.print_exc()
    sys.exit(1)
