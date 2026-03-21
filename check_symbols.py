from openpyxl import load_workbook
import os

projects = ['belgrad', 'istanbul', 'ankara', 'iasi', 'timisoara', 'kayseri', 'kocaeli', 'gebze', 'samsun']
base_path = r'c:\Users\fatiherkin\Desktop\bozankaya_ssh_takip\data'

print("\n" + "="*80)
print("TÜÜM PROJELERİN SİMBOL KONTROL")
print("="*80 + "\n")

for project in projects:
    grid_path = os.path.join(base_path, project, 'service_status_grid.xlsx')
    
    if not os.path.exists(grid_path):
        print(f"❌ {project:12} | File not found")
        continue
    
    wb = load_workbook(grid_path, data_only=True)
    ws = wb.active
    
    # Örnek bir veri hücresi bul
    example_symbol = None
    for row_idx in range(2, min(10, ws.max_row + 1)):
        for col_idx in range(2, ws.max_column + 1):
            cell_val = ws.cell(row=row_idx, column=col_idx).value
            if cell_val:
                example_symbol = cell_val
                break
        if example_symbol:
            break
    
    if example_symbol:
        # Sembolü analiz et
        print(f"✓ {project:12} | Symbol: '{example_symbol}'")
        print(f"  Hex: {example_symbol.encode('unicode-escape').decode()}")
        print(f"  Unicode: U+{ord(example_symbol):04X}")
        if example_symbol == '✓':
            print(f"  ✓ CORRECT (CHECK MARK - U+2713)")
        elif example_symbol == '√':
            print(f"  ✗ WRONG (SQUARE ROOT - U+221A) - Should be U+2713!")
        elif example_symbol == '⚠':
            print(f"  ✓ CORRECT (WARNING - U+26A0)")
        elif example_symbol == '✗':
            print(f"  ✓ CORRECT (BALLOT X - U+2717)")
        print()
    
    wb.close()

print("="*80)
print("UYARI: Eğer Kayseri '√' simbolü gösteriyorsa, get_availability_data() '✓' arıyor!")
print("FIX: Simbol normalizasyonu yapmalı...")
print("="*80 + "\n")
