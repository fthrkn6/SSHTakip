"""
FRACAS ve KPI için Filtreleme ve Raporlama Modülü
"""

from flask import Blueprint, render_template, request, jsonify, session, send_file
from flask_login import login_required, current_user
from datetime import datetime, timedelta
import json
import pandas as pd
import os
from pathlib import Path
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment

analytics_bp = Blueprint('analytics', __name__, url_prefix='/analytics')


class AnalyticsFilter:
    """FRACAS ve KPI verilerini filtreleme"""
    
    def __init__(self, df):
        self.original_df = df.copy()
        self.df = df.copy()
    
    def filter_by_month(self, month: int, year: int):
        """Belirli ay ve yıla göre filtrele"""
        date_cols = ['tarih', 'date', 'hata tarih', 'failure date']
        date_col = None
        
        for col in self.df.columns:
            if col.lower() in date_cols or any(d in col.lower() for d in date_cols):
                date_col = col
                break
        
        if date_col:
            self.df[date_col] = pd.to_datetime(self.df[date_col], errors='coerce')
            self.df = self.df[
                (self.df[date_col].dt.month == month) &
                (self.df[date_col].dt.year == year)
            ]
        
        return self
    
    def filter_by_date_range(self, start_date: str, end_date: str):
        """Tarih aralığına göre filtrele (YYYY-MM-DD)"""
        date_cols = ['tarih', 'date', 'hata tarih', 'failure date']
        date_col = None
        
        for col in self.df.columns:
            if col.lower() in date_cols or any(d in col.lower() for d in date_cols):
                date_col = col
                break
        
        if date_col:
            try:
                start = pd.to_datetime(start_date)
                end = pd.to_datetime(end_date)
                self.df[date_col] = pd.to_datetime(self.df[date_col], errors='coerce')
                self.df = self.df[
                    (self.df[date_col] >= start) &
                    (self.df[date_col] <= end)
                ]
            except Exception as e:
                print(f"Tarih filtreleme hatası: {e}")
        
        return self
    
    def filter_by_vehicle(self, vehicle_id: str):
        """Araç koduna göre filtrele"""
        vehicle_cols = ['araç', 'araç no', 'tram', 'vehicle', 'araç numarası']
        vehicle_col = None
        
        for col in self.df.columns:
            if col.lower() in vehicle_cols or any(v in col.lower() for v in vehicle_cols):
                vehicle_col = col
                break
        
        if vehicle_col:
            self.df = self.df[self.df[vehicle_col].astype(str) == str(vehicle_id)]
        
        return self
    
    def filter_by_module(self, module: str):
        """Sistem/Modüle göre filtrele"""
        module_cols = ['modül', 'sistem', 'module', 'system', 'alt sistem']
        module_col = None
        
        for col in self.df.columns:
            if col.lower() in module_cols or any(m in col.lower() for m in module_cols):
                module_col = col
                break
        
        if module_col:
            self.df = self.df[self.df[module_col].astype(str).str.contains(module, case=False, na=False)]
        
        return self
    
    def filter_by_equipment(self, equipment: str):
        """Ekipmana göre filtrele"""
        equipment_cols = ['ekipman', 'equipment', 'cihaz', 'device']
        equipment_col = None
        
        for col in self.df.columns:
            if col.lower() in equipment_cols or any(e in col.lower() for e in equipment_cols):
                equipment_col = col
                break
        
        if equipment_col:
            self.df = self.df[self.df[equipment_col].astype(str).str.contains(equipment, case=False, na=False)]
        
        return self
    
    def get_filtered_data(self):
        """Filtrelenmiş veriyi döndür"""
        return self.df
    
    def get_stats(self):
        """Filtrelenmiş verinin istatistiklerini döndür"""
        stats = {
            'total_records': len(self.df),
            'filtered_percentage': (len(self.df) / len(self.original_df) * 100) if len(self.original_df) > 0 else 0,
            'original_records': len(self.original_df)
        }
        return stats


def export_to_excel(df: pd.DataFrame, title: str, filters: dict = None) -> BytesIO:
    """DataFrame'i Excel'e dışa aktar"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Veriler"
    
    # Başlık
    ws.merge_cells('A1:' + chr(64 + len(df.columns)) + '1')
    header_cell = ws['A1']
    header_cell.value = title
    header_cell.font = Font(bold=True, size=14, color='FFFFFF')
    header_cell.fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
    header_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 25
    
    # Filtre bilgileri
    if filters:
        filter_text = " | ".join([f"{k}: {v}" for k, v in filters.items() if v])
        ws.merge_cells('A2:' + chr(64 + len(df.columns)) + '2')
        filter_cell = ws['A2']
        filter_cell.value = f"Filtreler: {filter_text}"
        filter_cell.font = Font(italic=True, size=10)
        filter_cell.alignment = Alignment(horizontal='left')
    
    # Rapor tarihi
    ws.merge_cells('A3:' + chr(64 + len(df.columns)) + '3')
    date_cell = ws['A3']
    date_cell.value = f"Rapor Tarihi: {datetime.now().strftime('%d.%m.%Y %H:%M:%S')}"
    date_cell.font = Font(italic=True, size=10)
    date_cell.alignment = Alignment(horizontal='right')
    
    # Sütun başlıkları
    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=5, column=col_idx)
        cell.value = col_name
        cell.font = Font(bold=True, color='FFFFFF', size=11)
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    ws.row_dimensions[5].height = 25
    
    # Veriler
    borders = None
    for row_idx, row_data in enumerate(df.values, 6):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    # Sütun genişlikleri
    for col_idx, col_name in enumerate(df.columns, 1):
        ws.column_dimensions[chr(64 + col_idx)].width = 18
    
    # BytesIO'ya yazı
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output


# ==================== FRACAS FİLTRELEME API'ları ====================
@analytics_bp.route('/fracas-filter', methods=['POST'])
@login_required
def fracas_filter():
    """FRACAS verilerini filtrele ve istatistikleri döndür"""
    try:
        from routes.fracas import load_ariza_listesi_data, load_fracas_data
        
        filters = request.json
        
        # Verileri yükle
        df = load_ariza_listesi_data() or load_fracas_data()
        if df is None:
            return jsonify({'success': False, 'message': 'Veri yüklenemedi'}), 500
        
        # Filtreleri uygula
        analyzer = AnalyticsFilter(df)
        
        if filters.get('month') and filters.get('year'):
            analyzer.filter_by_month(int(filters['month']), int(filters['year']))
        
        if filters.get('start_date') and filters.get('end_date'):
            analyzer.filter_by_date_range(filters['start_date'], filters['end_date'])
        
        if filters.get('vehicle_id'):
            analyzer.filter_by_vehicle(filters['vehicle_id'])
        
        if filters.get('module'):
            analyzer.filter_by_module(filters['module'])
        
        if filters.get('equipment'):
            analyzer.filter_by_equipment(filters['equipment'])
        
        # İstatistikleri hesapla
        filtered_df = analyzer.get_filtered_data()
        stats = analyzer.get_stats()
        total_records = len(analyzer.original_df)
        
        return jsonify({
            'success': True,
            'record_count': len(filtered_df),
            'total_records': total_records,
            'stats': stats,
            'sample_data': filtered_df.head(10).to_dict('records')
        })
    
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@analytics_bp.route('/fracas-export', methods=['POST'])
@login_required
def fracas_export():
    """Filtrelenmiş FRACAS verilerini Excel'e dışa aktar"""
    try:
        from routes.fracas import load_ariza_listesi_data, load_fracas_data
        
        filters = request.json
        
        # Verileri yükle
        df = load_ariza_listesi_data() or load_fracas_data()
        if df is None:
            return jsonify({'success': False, 'message': 'Veri yüklenemedi'}), 500
        
        # Filtreleri uygula
        analyzer = AnalyticsFilter(df)
        
        if filters.get('month') and filters.get('year'):
            analyzer.filter_by_month(int(filters['month']), int(filters['year']))
        
        if filters.get('start_date') and filters.get('end_date'):
            analyzer.filter_by_date_range(filters['start_date'], filters['end_date'])
        
        if filters.get('vehicle_id'):
            analyzer.filter_by_vehicle(filters['vehicle_id'])
        
        if filters.get('module'):
            analyzer.filter_by_module(filters['module'])
        
        if filters.get('equipment'):
            analyzer.filter_by_equipment(filters['equipment'])
        
        filtered_df = analyzer.get_filtered_data()
        
        # Excel'e dışa aktar
        excel_file = export_to_excel(filtered_df, 'FRACAS Raporu', dict(filters))
        
        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f"FRACAS_Raporu_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
    
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


# ==================== KPI FİLTRELEME API'ları ====================
@analytics_bp.route('/kpi-export', methods=['POST'])
@login_required
def kpi_export():
    """KPI verilerini Excel'e dışa aktar"""
    try:
        filters = request.json
        
        # KPI verisi oluştur
        kpi_data = {
            'Metrik': ['MTBF', 'MTTR', 'Hazırlık', 'Maliyet', 'Güvenilirlik'],
            'Değer': [1200, 4.5, 98.5, 15000, 96.0],
            'Hedef': [1500, 3.0, 99.0, 12000, 98.0]
        }
        df = pd.DataFrame(kpi_data)
        
        # Excel'e dışa aktar
        excel_file = export_to_excel(df, 'KPI Dashboard Raporu', dict(filters))
        
        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f"KPI_Raporu_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
    
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500
