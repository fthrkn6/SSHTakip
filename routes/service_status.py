"""
Servis Durumu ve Availability Route'ları
Araç servis durumunu yönet, raporla ve analiz et
"""

from flask import Blueprint, render_template, request, jsonify, session, flash, redirect, url_for, send_file, current_app
from flask_login import login_required, current_user
from models import db, ServiceStatus, AvailabilityMetrics, RootCauseAnalysis, Failure, Equipment
from datetime import datetime, timedelta, date
from sqlalchemy import desc
import logging
import sys

logger = logging.getLogger(__name__)
from utils.utils_availability import (
    log_service_status_change,
    AvailabilityCalculator,
    ExcelReportGenerator as AvailabilityExcelGenerator
)
from utils.utils_service_status import AvailabilityAnalyzer, ExcelReportGenerator as EnhancedExcelGenerator
from utils.utils_daily_service_logger import log_service_status as log_service_to_file
from utils.utils_service_status_excel_logger import log_service_status_to_excel
from utils.utils_excel_grid_manager import ExcelGridManager, RCAExcelManager
from utils.utils_project_excel_store import upsert_service_status
from openpyxl import load_workbook
import os
import json
import logging
import pandas as pd

logger = logging.getLogger(__name__)

bp = Blueprint('service_status', __name__, url_prefix='/servis')


def get_tram_ids_from_veriler(project_code=None):
    """Veriler.xlsx Sayfa2'den equipment_code'leri yükle - tüm sayfalarda tek kaynak"""
    if project_code is None:
        project_code = session.get('current_project', 'belgrad')
    
    veriler_path = os.path.join(current_app.root_path, 'data', project_code, 'Veriler.xlsx')
    
    if not os.path.exists(veriler_path):
        # Fallback: Equipment tablosundan çek (project_code ile filtrele)
        return [eq.equipment_code for eq in Equipment.query.filter_by(parent_id=None, project_code=project_code).all()]
    
    try:
        df = pd.read_excel(veriler_path, sheet_name='Sayfa2', header=0)
        # equipment_code sütununu kullan (eğer varsa), yoksa tram_id'leri string'e çevir
        if 'equipment_code' in df.columns:
            equipment_codes = df['equipment_code'].dropna().unique().tolist()
            return [str(c) for c in equipment_codes]
        elif 'tram_id' in df.columns:
            tram_ids = df['tram_id'].dropna().unique().tolist()
            return [str(t) for t in tram_ids]
        # Fallback: Equipment tablosundan çek
        return [eq.equipment_code for eq in Equipment.query.filter_by(parent_id=None, project_code=project_code).all()]
    except Exception as e:
        logger.error(f'Veriler.xlsx okuma hatasi ({project_code}): {e}')
        # Fallback: Equipment tablosundan çek (project_code ile filtrele)
        return [eq.equipment_code for eq in Equipment.query.filter_by(parent_id=None, project_code=project_code).all()]

bp = Blueprint('service_status', __name__, url_prefix='/servis')


@bp.route('/durumu', methods=['GET'])
@login_required
def service_status_page():
    """Servis durumu dashboard sayfası"""
    try:
        current_project = session.get('current_project', 'belgrad')
        
        # Veriler.xlsx Sayfa2'den equipment_code'leri al
        tram_ids = get_tram_ids_from_veriler(current_project)
        
        # Equipment Code'lara göre Equipment'i filtrele (DB'den status, name vb al)
        if tram_ids:
            equipment_list = Equipment.query.filter(
                Equipment.equipment_code.in_(tram_ids),
                Equipment.parent_id == None,
                Equipment.project_code == current_project
            ).all()
        else:
            # Fallback: Equipment tablosundan direkt çek (proje-spesifik)
            equipment_list = Equipment.query.filter_by(parent_id=None, project_code=current_project).all()
        
        # Bugünün tarihi
        today_date = str(date.today())
        
        # Her tramvay için durumunu getir
        tram_status_data = []
        for equipment in equipment_list:
            # ServiceStatus'ten bugünün kaydını getir
            status_record = ServiceStatus.query.filter_by(
                tram_id=equipment.equipment_code,
                date=today_date,
                project_code=current_project
            ).first()
            
            # Durum belirle
            status_color = 'success'  # Default yeşil
            status_display = 'aktif'
            status_badge = 'Aktif'
            
            if status_record:
                status_value = status_record.status if status_record.status else ''
                aciklama = status_record.aciklama if status_record.aciklama else ''
                
                # CRITICAL: Check "İşletme Kaynaklı" FIRST before "Servis Dışı" 
                # because "İşletme Kaynaklı Servis Dışı" contains both patterns!
                # Use case-insensitive Turkish pattern matching (no lower() to avoid Unicode issues)
                if 'İşletme' in status_value or 'işletme' in status_value:
                    status_color = 'warning'
                    status_display = 'isletme'  # İşletme kaynaklı = special status for counting
                    status_badge = 'İşletme Kaynaklı'
                elif 'Dışı' in status_value or 'dışı' in status_value:
                    status_color = 'danger'
                    status_display = 'ariza'
                    status_badge = 'Servis Dışı'
                elif 'Servis' in status_value or 'servis' in status_value:
                    status_color = 'success'
                    status_display = 'aktif'
                    status_badge = 'Aktif'
                else:
                    status_color = 'success'
                    status_display = 'aktif'
                    status_badge = 'Aktif'
            else:
                # ServiceStatus yoksa default 'aktif'
                status_color = 'success'
                status_display = 'aktif'
                status_badge = 'Aktif'
                aciklama = ''
            
            # Availability metrikleri getir (veya ServiceStatus'ten hesapla)
            latest_metric = AvailabilityMetrics.query.filter_by(
                tram_id=equipment.equipment_code
            ).order_by(AvailabilityMetrics.metric_date.desc()).first()
            
            # Eğer metric yoksa ServiceStatus'ten hesapla
            if latest_metric:
                availability = latest_metric.availability_percentage
                downtime = latest_metric.downtime_hours
                operational = latest_metric.operational_hours
            else:
                # ServiceStatus'ten hesapla
                if status_display == 'ariza':
                    availability = 0
                    operational = 0
                    downtime = 24
                elif status_display == 'isletme':
                    availability = 50
                    operational = 12
                    downtime = 12
                else:  # aktif
                    availability = 100
                    operational = 24
                    downtime = 0
            
            tram_status_data.append({
                'equipment_code': equipment.equipment_code,
                'name': equipment.name,
                'location': equipment.location if hasattr(equipment, 'location') else '',
                'total_km': equipment.total_km if hasattr(equipment, 'total_km') else 0,
                'status': status_display,
                'status_color': status_color,
                'status_badge': status_badge,
                'status_record': status_record,
                'latest_metric': latest_metric,
                'availability': availability,
                'downtime': downtime,
                'operational': operational
            })
        
        # Son servis durumu kayıtlarını getir (tüm araçlar - sadece aktif proje)
        recent_statuses = ServiceStatus.query.filter_by(
            project_code=current_project
        ).order_by(
            desc(ServiceStatus.updated_at)
        ).limit(100).all()
        
        # DYNAMIC STATS HESAPLAMA - ServiceStatus'ten
        servis_disi_count = 0
        isletme_count = 0
        aktif_count = 0
        toplam = len(tram_status_data)
        
        for tram in tram_status_data:
            if tram['status'] == 'ariza':
                servis_disi_count += 1
            elif tram['status'] == 'isletme':
                isletme_count += 1
            elif tram['status'] == 'aktif':
                aktif_count += 1
        
        # Availability = (Aktif + İşletme Kaynaklı) / Toplam * 100
        # İşletme Kaynaklı olanlar çalışabilir durumda olsa da tamir bekliyordur
        # Sadece Aktif olanlar tam kullanılabilir
        availability = ((aktif_count) / toplam * 100) if toplam > 0 else 0
        
        # Seçili dönem
        period = request.args.get('period', 'monthly')
        
        # STATS dict'i oluştur
        stats = {
            'servis_disi': servis_disi_count,
            'isletme': isletme_count,
            'aktif': aktif_count,
            'toplam': toplam,
            'availability': round(availability, 1)
        }
        
        return render_template('servis_durumu.html', 
                             equipment_list=tram_status_data,
                             tram_ids=[eq['equipment_code'] for eq in tram_status_data],
                             recent_logs=recent_statuses,
                             stats=stats,
                             period=period,
                             current_date=datetime.now())
    
    except Exception as e:
        import traceback
        error_trace = traceback.format_exc()
        logger.error(f"Error loading service status page: {str(e)}\n{error_trace}")
        print(f"[ERROR] Service Status Page: {str(e)}")
        print(error_trace)
        flash(f'Servis durumu sayfası yüklenirken hata oluştu: {str(e)}', 'danger')
        return redirect(url_for('dashboard.index'))


@bp.route('/durumu/tablo', methods=['GET'])
@login_required
def service_status_table():
    """Servis durumu tablosunu getir - DASHBOARD LOGIC COPY"""
    try:
        current_project = session.get('current_project', 'belgrad')
        today_date = str(date.today())
        
        # Veriler.xlsx'ten tram_id'leri al (proje-spesifik)
        tram_ids = get_tram_ids_from_veriler(current_project)
        
        # Equipment'leri Project filtresiyle al - Aynı dashboard gibi
        equipment_list = Equipment.query.filter(
            Equipment.equipment_code.in_(tram_ids) if tram_ids else True,
            Equipment.parent_id == None,
            Equipment.project_code == current_project
        ).order_by(Equipment.equipment_code).all()
        
        # Her tramvay için ServiceStatus'ten durum al - DASHBOARD LOGİĞİ
        table_data = []
        aktif_count = 0
        isletme_count = 0  # İşletme Kaynaklı Servis Dışı
        ariza_count = 0
        
        for equipment in equipment_list:
            # PRIMARY: ServiceStatus'ten bugünün kaydını al
            service_record = ServiceStatus.query.filter_by(
                tram_id=equipment.equipment_code,
                date=today_date,
                project_code=current_project
            ).first()
            
            # Durum belirle - DASHBOARD LOGİĞİ
            status_display = 'Aktif'
            status_type = 'aktif'
            
            # ServiceStatus varsa onu kullan
            if service_record and service_record.status:
                service_status = service_record.status
                
                # İşletme Kaynaklı check FIRST (çünkü "Dışı" yazı içeriyor)
                if 'İşletme' in service_status or 'işletme' in service_status.lower():
                    status_display = 'İşletme Kaynaklı Servis Dışı'
                    status_type = 'işletme'
                    isletme_count += 1
                elif 'Dışı' in service_status or 'dışı' in service_status.lower() or 'ariza' in service_status.lower():
                    status_display = 'Servis Dışı'
                    status_type = 'ariza'
                    ariza_count += 1
                else:
                    status_display = 'Aktif'
                    status_type = 'aktif'
                    aktif_count += 1
            else:
                # ServiceStatus yoksa default 'aktif'
                status_display = 'Aktif'
                status_type = 'aktif'
                aktif_count += 1
            
            # Availability metrikleri getir
            latest_metric = AvailabilityMetrics.query.filter_by(
                tram_id=equipment.equipment_code
            ).order_by(AvailabilityMetrics.metric_date.desc()).first()
            
            table_data.append({
                'tram_id': equipment.equipment_code,
                'name': equipment.name,
                'status': status_display,
                'sistem': service_record.sistem if service_record and service_record.sistem else '-',
                'alt_sistem': service_record.alt_sistem if service_record and service_record.alt_sistem else '-',
                'last_status_change': service_record.updated_at.strftime('%d.%m.%Y %H:%M') if service_record and service_record.updated_at else '-',
                'availability': latest_metric.availability_percentage if latest_metric else 0,
                'downtime': latest_metric.downtime_hours if latest_metric else 0
            })
        
        # TOPLAM HESAPLAMA - DASHBOARD LOGİĞİ
        # Fleet Kullanılabilirlik Oranı = (Aktif + İşletme Kaynaklı) / Toplam * 100
        total_tram = len(equipment_list)
        kullanilabilir = aktif_count + isletme_count
        availability_percent = round((kullanilabilir / total_tram * 100), 1) if total_tram > 0 else 0
        
        # Response'ta stats'ı da geri gönder
        return jsonify({
            'table_data': table_data,
            'stats': {
                'operational': aktif_count,
                'maintenance': isletme_count,
                'outofservice': ariza_count,
                'total': total_tram,
                'availability': availability_percent
            }
        })
    
    except Exception as e:
        logger.error(f"Error getting service status table: {str(e)}")
        return jsonify({'error': str(e)}), 400


@bp.route('/durumu/log', methods=['POST'])
@login_required
def log_service_status():
    """Servis durumu değişikliğini log'la"""
    try:
        data = request.get_json()
        
        tram_id = data.get('tram_id')
        new_status = data.get('status')
        sistem = data.get('sistem')
        alt_sistem = data.get('alt_sistem')
        reason = data.get('reason')
        duration_hours = float(data.get('duration_hours', 0))
        tarih = data.get('tarih')  # Seçili tarih parametresi
        
        if not tram_id or not new_status:
            return jsonify({'error': 'Tram ID ve durum gerekli'}), 400
        
        # Proje kodunu al (lowercase tutarlılık için)
        project_code = session.get('current_project', 'belgrad').lower()
        
        # Tramvay adını al
        equipment = Equipment.query.filter_by(equipment_code=str(tram_id), project_code=project_code).first()
        tram_name = equipment.name if equipment else f'Tramvay {tram_id}'
        
        # **YENİ LOGGER: Günlük log klasörüne kaydet**
        log_service_to_file(
            tram_id=tram_id,
            status=new_status,
            sistem=sistem or '',
            alt_sistem=alt_sistem or '',
            aciklama=reason or '',
            user=current_user.username if current_user else 'system',
            project_code=project_code.lower()
        )
        
        # **EXCEL LOGGER: Excel dosyasına kaydet**
        log_service_status_to_excel(
            tram_id=tram_id,
            tram_name=tram_name,
            status=new_status,
            sistem=sistem or '',
            alt_sistem=alt_sistem or '',
            aciklama=reason or '',
            user=current_user.username if current_user else 'system',
            action='GÜNCELLE',
            project_code=project_code.lower()
        )
        
        # **EXCEL GRID: Grid dosyasına kaydı aktar (durum tablosu)**
        try:
            grid_manager = ExcelGridManager(project_code.lower())
            # Durum kodu oluştur (Turkish character-safe)
            status_code = 'aktif'
            if 'İşletme' in new_status or 'işletme' in new_status.lower():
                status_code = 'isletme_kaynakli'
            elif 'Dışı' in new_status or 'dışı' in new_status.lower() or 'ariza' in new_status.lower():
                status_code = 'servis_disi'
            
            # Grid'e yaz - seçili tarih veya bugün
            grid_date = tarih if tarih else date.today().strftime('%Y-%m-%d')
            grid_manager.update_status(current_app.root_path, tram_id, grid_date, status_code)
            
            # RCA kaydı ekle (sadece servis dışı ise)
            if status_code == 'servis_disi' and sistem:
                rca_manager = RCAExcelManager(project_code.lower())
                rca_manager.add_rca_record(
                    current_app.root_path,
                    tram_id,
                    sistem,
                    alt_sistem,
                    status_code,
                    reason or ''
                )
        except Exception as excel_error:
            logger.warning(f'Excel grid yazma hatası (devam et): {excel_error}')
        
        # Log'u kaydet (eski sistem)
        log_entry = log_service_status_change(
            tram_id=tram_id,
            new_status=new_status,
            sistem=sistem,
            alt_sistem=alt_sistem,
            reason=reason,
            duration_hours=duration_hours,
            user_id=current_user.id,
            date_param=tarih,  # Seçili tarih parametrisini geçir
            project_code=project_code  # Proje kodunu geçir
        )
        
        # **SERVICE EXCEL SYNC: Ana servis durum Excel'ini güncelle (sync_service_excel_to_db ile uyumlu)**
        try:
            status_date = tarih if tarih else date.today().strftime('%Y-%m-%d')
            upsert_service_status(
                project_code=project_code,
                status_date=status_date,
                tram_id=str(tram_id),
                status=new_status,
                sistem=sistem or '',
                alt_sistem=alt_sistem or '',
                aciklama=reason or '',
                updated_by=current_user.username if current_user else 'system'
            )
        except Exception as excel_sync_error:
            logger.warning(f'Service Excel sync hatası (devam et): {excel_sync_error}')
        
        # Günlük availability'i güncelle
        AvailabilityCalculator.calculate_daily_availability(tram_id)
        
        # Excel Grid'den son availability'i oku
        try:
            grid_manager = ExcelGridManager(project_code)
            availability_data = grid_manager.get_availability_data(current_app.root_path)
            current_availability = availability_data.get(str(tram_id), 0)
        except:
            current_availability = 0
        
        logger.info(f"Service status logged for {tram_id}: {new_status}")
        
        return jsonify({
            'success': True,
            'message': 'Servis durumu başarıyla kaydedildi',
            'log_id': log_entry.id,
            'tram_id': tram_id,
            'tram_name': tram_name,
            'status': new_status,
            'availability': current_availability,
            'timestamp': datetime.now().strftime('%d.%m.%Y %H:%M:%S')
        })
    
    except Exception as e:
        logger.error(f"Error logging service status: {str(e)}")
        return jsonify({'error': str(e)}), 400


@bp.route('/rapor/gunluk', methods=['GET'])
@login_required
def daily_report():
    """Günlük availability raporu"""
    try:
        tram_id = request.args.get('tram_id')
        target_date = request.args.get('date', date.today().strftime('%Y-%m-%d'))
        target_date = datetime.strptime(target_date, '%Y-%m-%d').date()
        
        if tram_id:
            metric = AvailabilityCalculator.calculate_daily_availability(tram_id, target_date)
            data = {
                'tram_id': tram_id,
                'date': target_date.strftime('%d.%m.%Y'),
                'availability': metric.availability_percentage,
                'operational_hours': metric.operational_hours,
                'downtime_hours': metric.downtime_hours
            }
            return jsonify(data)
        else:
            return jsonify({'error': 'Tram ID gerekli'}), 400
    
    except Exception as e:
        logger.error(f"Error generating daily report: {str(e)}")
        return jsonify({'error': str(e)}), 400


@bp.route('/rapor/haftalik', methods=['GET'])
@login_required
def weekly_report():
    """Haftalık availability raporu"""
    try:
        tram_id = request.args.get('tram_id')
        start_date = request.args.get('start_date')
        
        if not start_date:
            start_date = date.today() - timedelta(days=7)
        else:
            start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
        
        end_date = start_date + timedelta(days=6)
        
        if tram_id:
            metric = AvailabilityCalculator.calculate_period_availability(
                tram_id, start_date, end_date, 'weekly'
            )
            data = {
                'tram_id': tram_id,
                'period': f"{start_date.strftime('%d.%m.%Y')} - {end_date.strftime('%d.%m.%Y')}",
                'availability': metric.availability_percentage,
                'operational_hours': metric.operational_hours,
                'downtime_hours': metric.downtime_hours,
                'failure_count': metric.failure_count
            }
            return jsonify(data)
        else:
            return jsonify({'error': 'Tram ID gerekli'}), 400
    
    except Exception as e:
        logger.error(f"Error generating weekly report: {str(e)}")
        return jsonify({'error': str(e)}), 400


@bp.route('/rapor/aylik', methods=['GET'])
@login_required
def monthly_report():
    """Aylık availability raporu"""
    try:
        tram_id = request.args.get('tram_id')
        start_date = request.args.get('start_date')
        
        if not start_date:
            today = date.today()
            start_date = date(today.year, today.month, 1)
        else:
            start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
        
        # Ayın son günü
        if start_date.month == 12:
            end_date = date(start_date.year + 1, 1, 1) - timedelta(days=1)
        else:
            end_date = date(start_date.year, start_date.month + 1, 1) - timedelta(days=1)
        
        if tram_id:
            metric = AvailabilityCalculator.calculate_period_availability(
                tram_id, start_date, end_date, 'monthly'
            )
            data = {
                'tram_id': tram_id,
                'period': f"{start_date.strftime('%d.%m.%Y')} - {end_date.strftime('%d.%m.%Y')}",
                'availability': metric.availability_percentage,
                'operational_hours': metric.operational_hours,
                'downtime_hours': metric.downtime_hours,
                'failure_count': metric.failure_count
            }
            return jsonify(data)
        else:
            return jsonify({'error': 'Tram ID gerekli'}), 400
    
    except Exception as e:
        logger.error(f"Error generating monthly report: {str(e)}")
        return jsonify({'error': str(e)}), 400


@bp.route('/rapor/3aylik', methods=['GET'])
@login_required
def quarterly_report():
    """3 aylık availability raporu"""
    try:
        tram_id = request.args.get('tram_id')
        
        if tram_id:
            today = date.today()
            start_date = today - timedelta(days=90)
            end_date = today
            
            metric = AvailabilityCalculator.calculate_period_availability(
                tram_id, start_date, end_date, 'quarterly'
            )
            data = {
                'tram_id': tram_id,
                'period': f"{start_date.strftime('%d.%m.%Y')} - {end_date.strftime('%d.%m.%Y')}",
                'availability': metric.availability_percentage,
                'operational_hours': metric.operational_hours,
                'downtime_hours': metric.downtime_hours,
                'failure_count': metric.failure_count
            }
            return jsonify(data)
        else:
            return jsonify({'error': 'Tram ID gerekli'}), 400
    
    except Exception as e:
        logger.error(f"Error generating quarterly report: {str(e)}")
        return jsonify({'error': str(e)}), 400


@bp.route('/rapor/6aylik', methods=['GET'])
@login_required
def biannual_report():
    """6 aylık availability raporu"""
    try:
        tram_id = request.args.get('tram_id')
        
        if tram_id:
            today = date.today()
            start_date = today - timedelta(days=180)
            end_date = today
            
            metric = AvailabilityCalculator.calculate_period_availability(
                tram_id, start_date, end_date, 'biannual'
            )
            data = {
                'tram_id': tram_id,
                'period': f"{start_date.strftime('%d.%m.%Y')} - {end_date.strftime('%d.%m.%Y')}",
                'availability': metric.availability_percentage,
                'operational_hours': metric.operational_hours,
                'downtime_hours': metric.downtime_hours,
                'failure_count': metric.failure_count
            }
            return jsonify(data)
        else:
            return jsonify({'error': 'Tram ID gerekli'}), 400
    
    except Exception as e:
        logger.error(f"Error generating biannual report: {str(e)}")
        return jsonify({'error': str(e)}), 400


@bp.route('/rapor/yillik', methods=['GET'])
@login_required
def yearly_report():
    """Yıllık availability raporu"""
    try:
        tram_id = request.args.get('tram_id')
        year = request.args.get('year', date.today().year)
        
        if tram_id:
            start_date = date(int(year), 1, 1)
            end_date = date(int(year), 12, 31)
            
            metric = AvailabilityCalculator.calculate_period_availability(
                tram_id, start_date, end_date, 'yearly'
            )
            data = {
                'tram_id': tram_id,
                'period': f"01.01.{year} - 31.12.{year}",
                'availability': metric.availability_percentage,
                'operational_hours': metric.operational_hours,
                'downtime_hours': metric.downtime_hours,
                'failure_count': metric.failure_count
            }
            return jsonify(data)
        else:
            return jsonify({'error': 'Tram ID gerekli'}), 400
    
    except Exception as e:
        logger.error(f"Error generating yearly report: {str(e)}")
        return jsonify({'error': str(e)}), 400


@bp.route('/rapor/total', methods=['GET'])
@login_required
def total_report():
    """Toplam availability raporu"""
    try:
        tram_id = request.args.get('tram_id')
        
        if tram_id:
            # İlk log'dan bugüne kadar
            first_log = ServiceLog.query.filter_by(tram_id=tram_id).order_by(
                ServiceLog.log_date.asc()
            ).first()
            
            if first_log:
                start_date = first_log.log_date.date()
            else:
                start_date = date.today() - timedelta(days=365)
            
            end_date = date.today()
            
            metric = AvailabilityCalculator.calculate_period_availability(
                tram_id, start_date, end_date, 'total'
            )
            data = {
                'tram_id': tram_id,
                'period': f"{start_date.strftime('%d.%m.%Y')} - {end_date.strftime('%d.%m.%Y')}",
                'availability': metric.availability_percentage,
                'operational_hours': metric.operational_hours,
                'downtime_hours': metric.downtime_hours,
                'failure_count': metric.failure_count
            }
            return jsonify(data)
        else:
            return jsonify({'error': 'Tram ID gerekli'}), 400
    
    except Exception as e:
        logger.error(f"Error generating total report: {str(e)}")
        return jsonify({'error': str(e)}), 400


@bp.route('/excel/availability', methods=['GET', 'POST'])
@login_required
def export_availability_excel():
    """Availability raporunu Excel'e aktar"""
    try:
        tram_ids = [eq.id_tram for eq in Equipment.query.all()]
        
        start_date_str = request.args.get('start_date', str(date.today() - timedelta(days=30)))
        end_date_str = request.args.get('end_date', str(date.today()))
        
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        
        # Klasör oluştur
        output_dir = 'availability_analiz'
        os.makedirs(output_dir, exist_ok=True)
        
        # Dosya adı
        filename = f"Availability_Raporu_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.xlsx"
        filepath = os.path.join(output_dir, filename)
        
        if not tram_ids:
            return jsonify({'error': 'Araç bulunamadı'}), 400
        
        # Excel oluştur (utils_availability.ExcelReportGenerator)
        AvailabilityExcelGenerator.create_availability_report_excel(tram_ids, start_date, end_date, filepath)
        
        logger.info(f"Availability report exported to {filepath}")
        
        return send_file(filepath, as_attachment=True, download_name=filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    
    except Exception as e:
        logger.error(f"Error exporting availability report: {str(e)}")
        return jsonify({'error': str(e)}), 400


@bp.route('/excel/rootcause', methods=['GET', 'POST'])
@login_required
def export_rootcause_excel():
    """Root Cause Analysis raporunu RCA Excel'den indir"""
    try:
        project = session.get('current_project', 'belgrad')
        rca_manager = RCAExcelManager(project)
        rca_path = rca_manager.get_rca_path(current_app.root_path)
        
        if not os.path.exists(rca_path):
            rca_manager.init_rca(current_app.root_path)
            rca_path = rca_manager.get_rca_path(current_app.root_path)
        
        filename = f"RCA_Analiz_{project}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return send_file(
            rca_path,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    
    except Exception as e:
        logger.error(f"Error exporting root cause analysis report: {str(e)}")
        return jsonify({'error': str(e)}), 400


@bp.route('/root-cause', methods=['GET', 'POST'])
@login_required
def root_cause_analysis():
    """Root cause analysis yönetimi"""
    try:
        if request.method == 'POST':
            data = request.get_json()
            
            analysis = RootCauseAnalysis(
                tram_id=data.get('tram_id'),
                sistem=data.get('sistem'),
                alt_sistem=data.get('alt_sistem'),
                failure_description=data.get('failure_description'),
                root_cause=data.get('root_cause'),
                contributing_factors=json.dumps(data.get('contributing_factors', [])),
                preventive_actions=json.dumps(data.get('preventive_actions', [])),
                corrective_actions=json.dumps(data.get('corrective_actions', [])),
                analyzed_by=current_user.id,
                severity_level=data.get('severity_level', 'medium'),
                frequency=int(data.get('frequency', 1))
            )
            
            db.session.add(analysis)
            db.session.commit()
            
            logger.info(f"Root cause analysis created for {data.get('tram_id')}")
            
            return jsonify({
                'success': True,
                'message': 'Root cause analysis başarıyla oluşturuldu',
                'id': analysis.id
            })
        
        # GET - Raporları listele
        tram_ids = [eq.id_tram for eq in Equipment.query.all()]
        analyses = RootCauseAnalysis.query.all()
        
        return render_template('root_cause_analysis.html', 
                             tram_ids=tram_ids,
                             analyses=analyses)
    
    except Exception as e:
        logger.error(f"Error in root cause analysis: {str(e)}")
        return jsonify({'error': str(e)}), 400


@bp.route('/rca/veriler', methods=['GET'])
@login_required
def rca_list_data():
    """RCA Excel verilerini JSON olarak listele"""
    try:
        project = session.get('current_project', 'belgrad')
        period = request.args.get('period', 'toplam')
        
        rca_manager = RCAExcelManager(project)
        
        today = date.today()
        start_date = None
        end_date = None
        if period == 'haftalik':
            start_date = today - timedelta(days=7)
            end_date = today
        elif period == 'aylik':
            start_date = today - timedelta(days=30)
            end_date = today
        elif period == 'ucaylik':
            start_date = today - timedelta(days=90)
            end_date = today
        
        data = rca_manager.get_rca_data(current_app.root_path, start_date, end_date)
        
        # Satır index'ini ekle (güncelleme/silme için)
        for idx, record in enumerate(data):
            record['row_index'] = idx
        
        return jsonify({'success': True, 'data': data, 'total': len(data)})
    except Exception as e:
        logger.error(f"RCA veriler hatası: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 400


@bp.route('/rca/ekle', methods=['POST'])
@login_required
def rca_add():
    """Yeni RCA kaydı ekle (App → Excel)"""
    try:
        data = request.get_json()
        project = session.get('current_project', 'belgrad')
        
        rca_manager = RCAExcelManager(project)
        result = rca_manager.add_rca_record(
            current_app.root_path,
            tram_id=data.get('arac'),
            system=data.get('sistem'),
            subsystem=data.get('alt_sistem', ''),
            category=data.get('kategori', 'servis_disi'),
            description=data.get('aciklama', '')
        )
        
        return jsonify({'success': result, 'message': 'RCA kaydı eklendi'})
    except Exception as e:
        logger.error(f"RCA ekleme hatası: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 400


@bp.route('/rca/guncelle', methods=['POST'])
@login_required
def rca_update():
    """Excel'deki RCA kaydını güncelle"""
    try:
        data = request.get_json()
        project = session.get('current_project', 'belgrad')
        row_index = data.get('row_index')
        
        if row_index is None:
            return jsonify({'success': False, 'error': 'row_index gerekli'}), 400
        
        rca_manager = RCAExcelManager(project)
        update_data = {}
        field_map = {
            'tarih': 'Tarih', 'arac': 'Arac', 'sistem': 'Sistem',
            'alt_sistem': 'Alt Sistem', 'kategori': 'Kategori', 'aciklama': 'Aciklama'
        }
        for key, excel_key in field_map.items():
            if key in data:
                update_data[excel_key] = data[key]
        
        result = rca_manager.update_rca_record(current_app.root_path, int(row_index), update_data)
        return jsonify({'success': result, 'message': 'RCA kaydı güncellendi' if result else 'Güncelleme başarısız'})
    except Exception as e:
        logger.error(f"RCA güncelleme hatası: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 400


@bp.route('/rca/sil', methods=['POST'])
@login_required
def rca_delete():
    """Excel'deki RCA kaydını sil"""
    try:
        data = request.get_json()
        project = session.get('current_project', 'belgrad')
        row_index = data.get('row_index')
        
        if row_index is None:
            return jsonify({'success': False, 'error': 'row_index gerekli'}), 400
        
        rca_manager = RCAExcelManager(project)
        result = rca_manager.delete_rca_record(current_app.root_path, int(row_index))
        return jsonify({'success': result, 'message': 'RCA kaydı silindi' if result else 'Silme başarısız'})
    except Exception as e:
        logger.error(f"RCA silme hatası: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 400


@bp.route('/rca/sync', methods=['POST'])
@login_required
def rca_sync():
    """İki yönlü Excel ↔ DB senkronizasyonu"""
    try:
        project = session.get('current_project', 'belgrad')
        direction = request.get_json().get('direction', 'full') if request.is_json else 'full'
        
        rca_manager = RCAExcelManager(project)
        
        if direction == 'excel_to_db':
            result = rca_manager.sync_excel_to_db(current_app.root_path, db.session, RootCauseAnalysis, project)
        elif direction == 'db_to_excel':
            result = rca_manager.sync_db_to_excel(current_app.root_path, db.session, RootCauseAnalysis, project)
        else:
            result = rca_manager.full_sync(current_app.root_path, db.session, RootCauseAnalysis, project)
        
        return jsonify({'success': True, 'result': result})
    except Exception as e:
        logger.error(f"RCA senkronizasyon hatası: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 400


@bp.route('/rca/indir', methods=['GET'])
@login_required
def rca_download():
    """RCA Excel dosyasını indir"""
    try:
        project = session.get('current_project', 'belgrad')
        rca_manager = RCAExcelManager(project)
        rca_path = rca_manager.get_rca_path(current_app.root_path)
        
        if not os.path.exists(rca_path):
            return jsonify({'success': False, 'error': 'RCA dosyası bulunamadı'}), 404
        
        return send_file(
            rca_path,
            as_attachment=True,
            download_name=f'RCA_Analiz_{project}_{date.today().strftime("%Y%m%d")}.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        logger.error(f"RCA indirme hatası: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 400


@bp.route('/rca/yukle', methods=['POST'])
@login_required
def rca_upload():
    """Harici RCA Excel dosyasını yükle ve mevcut veriye birleştir"""
    try:
        project = session.get('current_project', 'belgrad')
        
        if 'file' not in request.files:
            return jsonify({'success': False, 'error': 'Dosya seçilmedi'}), 400
        
        file = request.files['file']
        if not file.filename.endswith('.xlsx'):
            return jsonify({'success': False, 'error': 'Sadece .xlsx dosyaları kabul edilir'}), 400
        
        # Güvenli geçici dosyaya kaydet
        import tempfile
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            file.save(tmp.name)
            tmp_path = tmp.name
        
        try:
            wb = load_workbook(tmp_path, data_only=True)
            ws = wb.active
            
            # Başlıkları kontrol et
            headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
            required = ['Tarih', 'Arac', 'Sistem']
            # Türkçe/ASCII uyumu: esnek eşleştirme
            header_lower = [str(h).lower().strip() if h else '' for h in headers]
            
            rca_manager = RCAExcelManager(project)
            added = 0
            
            for row_idx in range(2, ws.max_row + 1):
                tarih = ws.cell(row=row_idx, column=1).value
                arac = ws.cell(row=row_idx, column=2).value
                sistem = ws.cell(row=row_idx, column=3).value
                alt_sistem = ws.cell(row=row_idx, column=4).value or ''
                kategori = ws.cell(row=row_idx, column=5).value or 'servis_disi'
                aciklama = ws.cell(row=row_idx, column=6).value or ''
                
                if not arac or not sistem:
                    continue
                
                rca_manager.add_rca_record(
                    current_app.root_path,
                    tram_id=str(arac),
                    system=str(sistem),
                    subsystem=str(alt_sistem),
                    category=str(kategori),
                    description=str(aciklama)
                )
                added += 1
            
            wb.close()
            return jsonify({'success': True, 'message': f'{added} kayıt eklendi', 'added': added})
        finally:
            os.unlink(tmp_path)
    
    except Exception as e:
        logger.error(f"RCA yükleme hatası: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 400


@bp.route('/excel/comprehensive-report', methods=['GET', 'POST'])
@login_required
def export_comprehensive_report():
    """Kapsamlı availability ve root cause raporu"""
    try:
        project = session.get('current_project', 'belgrad')
        # Güvenlik: project adı sadece alfanumerik ve alt çizgi olabilir
        project = ''.join(c for c in project if c.isalnum() or c in ('_', '-')).lower() or 'belgrad'
        # Sadece aktif projenin araçlarını al
        tram_ids = [eq.equipment_code for eq in Equipment.query.filter_by(project_code=project).all()]
        
        if not tram_ids:
            return jsonify({'error': 'Araç bulunamadı'}), 400
        
        # Projeden klasör oluştur
        output_dir = os.path.join('logs', project, 'reports')
        os.makedirs(output_dir, exist_ok=True)
        
        # Dosya adı
        filename = f"Kapsamli_Servis_Durumu_Raporu_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join(output_dir, filename)
        
        # Excel oluştur
        EnhancedExcelGenerator.create_comprehensive_availability_report(tram_ids, filepath)
        
        logger.info(f"Comprehensive report exported to {filepath}")
        
        return send_file(filepath, as_attachment=True, download_name=filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    
    except Exception as e:
        logger.error(f"Error exporting comprehensive report: {str(e)}")
        return jsonify({'error': str(e)}), 400


@bp.route('/excel/root-cause-report', methods=['GET', 'POST'])
@login_required
def export_root_cause_report():
    """Root cause analysis raporu"""
    try:
        project = session.get('current_project', 'belgrad')
        # Güvenlik: project adı sadece alfanumerik ve alt çizgi olabilir
        project = ''.join(c for c in project if c.isalnum() or c in ('_', '-')).lower() or 'belgrad'
        # Sadece aktif projenin araçlarını al
        tram_ids = [eq.equipment_code for eq in Equipment.query.filter_by(project_code=project).all()]
        
        if not tram_ids:
            return jsonify({'error': 'Araç bulunamadı'}), 400
        
        # Klasör oluştur
        output_dir = os.path.join('logs', project, 'reports')
        os.makedirs(output_dir, exist_ok=True)
        
        # Dosya adı
        filename = f"Root_Cause_Analiz_Raporu_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join(output_dir, filename)
        
        # Excel oluştur
        EnhancedExcelGenerator.create_root_cause_analysis_report(tram_ids, filepath)
        
        logger.info(f"Root cause report exported to {filepath}")
        
        return send_file(filepath, as_attachment=True, download_name=filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    
    except Exception as e:
        logger.error(f"Error exporting root cause report: {str(e)}")
        return jsonify({'error': str(e)}), 400


@bp.route('/excel/daily-report/<tram_id>', methods=['GET'])
@login_required
def export_daily_report(tram_id):
    """Günlük servis durumu raporu"""
    try:
        project = session.get('current_project', 'belgrad')
        # Güvenlik: project ve tram_id sanitize
        project = ''.join(c for c in project if c.isalnum() or c in ('_', '-')).lower() or 'belgrad'
        tram_id = ''.join(c for c in str(tram_id) if c.isalnum() or c in ('_', '-'))
        
        if not tram_id:
            return jsonify({'error': 'Geçersiz tram_id'}), 400
        
        # tram_id veritabanında var mı kontrol et
        equipment = Equipment.query.filter(
            (Equipment.equipment_code == tram_id) | (Equipment.id_tram == tram_id)
        ).filter_by(project_code=project).first()
        if not equipment:
            return jsonify({'error': 'Araç bulunamadı'}), 404
        
        # Klasör oluştur
        output_dir = os.path.join('logs', project, 'reports')
        os.makedirs(output_dir, exist_ok=True)
        
        # Dosya adı
        filename = f"Gunluk_Durum_{tram_id}_{date.today().strftime('%Y%m%d')}.xlsx"
        filepath = os.path.join(output_dir, filename)
        
        # Excel oluştur
        EnhancedExcelGenerator.create_detailed_daily_report(tram_id, filepath)
        
        logger.info(f"Daily report exported to {filepath}")
        
        return send_file(filepath, as_attachment=True, download_name=filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    
    except Exception as e:
        logger.error(f"Error exporting daily report: {str(e)}")
        return jsonify({'error': str(e)}), 400


@bp.route('/api/root-cause-summary/<tram_id>', methods=['GET'])
@login_required
def get_root_cause_summary(tram_id):
    """Root cause analizi özeti"""
    try:
        start_date_str = request.args.get('start_date')
        end_date_str = request.args.get('end_date')
        
        start_date = None
        end_date = None
        
        if start_date_str:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        if end_date_str:
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        
        summary = AvailabilityAnalyzer.get_root_cause_summary(tram_id, start_date, end_date)
        
        return jsonify(summary)
    
    except Exception as e:
        logger.error(f"Error getting root cause summary: {str(e)}")
        return jsonify({'error': str(e)}), 400


@bp.route('/durumu/excel/rapor', methods=['POST'])
@login_required
def export_service_status_excel():
    """
    Servis durumu pivot tablosunu Excel'e dışarı aktar (Template'den çağrılır)
    POST: tram_id (opsiyonel - özel bir tramvay için)
    """
    try:
        from io import BytesIO
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.comments import Comment
        
        current_project = session.get('current_project', 'belgrad')
        
        # Veriler.xlsx'ten geçerli araç listesini al
        valid_trams = get_tram_ids_from_veriler(current_project)
        valid_trams_set = set(valid_trams) if valid_trams else None
        
        # İstek verisinden tram_id al (opsiyonel)
        data = request.get_json() or {}
        tram_id = data.get('tram_id')
        
        # Tüm tarihsel verileri al (SADECE aktif proje)
        query = ServiceStatus.query.filter_by(project_code=current_project)
        
        # Eğer spesifik tramvay seçildiyse, sadece onu getir
        if tram_id:
            query = query.filter_by(tram_id=tram_id)
        
        all_status_records = query.order_by(
            ServiceStatus.date.desc(), 
            ServiceStatus.tram_id
        ).all()
        
        if not all_status_records:
            return jsonify({'error': 'Veri bulunamadı'}), 400
        
        # Pivot tablo yapısı oluştur: tramvay -> {tarih -> {status, sistem, alt_sistem, aciklama}}
        pivot_data = {}
        all_trams = set()
        all_dates = set()
        
        for record in all_status_records:
            # Geçerli araç listesinde olmayanları atla
            if valid_trams_set and record.tram_id not in valid_trams_set:
                continue
            if record.tram_id not in pivot_data:
                pivot_data[record.tram_id] = {}
            pivot_data[record.tram_id][record.date] = {
                'status': record.status,
                'sistem': record.sistem or '',
                'alt_sistem': record.alt_sistem or '',
                'aciklama': record.aciklama or ''
            }
            all_trams.add(record.tram_id)
            all_dates.add(record.date)
        
        # Tramvay ve tarihleri sırala
        sorted_trams = sorted(list(all_trams))
        sorted_dates = sorted(list(all_dates), reverse=True)
        
        # En eski ve en yeni tarihi bul
        oldest_date_str = sorted_dates[-1]
        newest_date_str = sorted_dates[0]
        
        oldest_date = datetime.strptime(oldest_date_str, '%Y-%m-%d').date()
        newest_date = datetime.strptime(newest_date_str, '%Y-%m-%d').date()
        
        # En eski ile en yeni arasındaki TÜM günleri generate et
        from datetime import timedelta
        all_dates_filled = []
        current_date = newest_date
        while current_date >= oldest_date:
            all_dates_filled.append(current_date.strftime('%Y-%m-%d'))
            current_date -= timedelta(days=1)
        
        sorted_dates = all_dates_filled
        
        # Pivot tabloyu tüm tarih ve tramvaylar için initialize et
        for tram in sorted_trams:
            if tram not in pivot_data:
                pivot_data[tram] = {}
            for date_str in sorted_dates:
                if date_str not in pivot_data[tram]:
                    pivot_data[tram][date_str] = None
        
        # Excel dosyası oluştur
        wb = Workbook()
        ws = wb.active
        ws.title = 'Servis Durumu'
        
        # Stil tanımları
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Durum renklerini tanımla
        servis_fill = PatternFill(start_color='00B050', end_color='00B050', fill_type='solid')
        isletme_fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
        disi_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
        
        durum_font = Font(bold=True, size=14, color='FFFFFF')
        
        # Header satırı: Tarihler
        ws['A1'] = 'Tramvay'
        ws['A1'].fill = header_fill
        ws['A1'].font = header_font
        ws['A1'].border = border
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        
        for col_idx, date_str in enumerate(sorted_dates, start=2):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = date_str
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Veri satırları: Tramvaylar
        for row_idx, tram_id_row in enumerate(sorted_trams, start=2):
            cell = ws.cell(row=row_idx, column=1)
            cell.value = tram_id_row
            cell.fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
            cell.font = Font(bold=True)
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            for col_idx, date_str in enumerate(sorted_dates, start=2):
                record_info = pivot_data[tram_id_row].get(date_str, None)
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center')
                
                status = record_info.get('status', '') if isinstance(record_info, dict) else (record_info or '')
                sistem = record_info.get('sistem', '') if isinstance(record_info, dict) else ''
                alt_sistem = record_info.get('alt_sistem', '') if isinstance(record_info, dict) else ''
                aciklama = record_info.get('aciklama', '') if isinstance(record_info, dict) else ''
                
                if status and 'Servis' in status:
                    if 'Dışı' in status:
                        if 'İşletme' in status:
                            cell.value = '⚠'
                            cell.fill = isletme_fill
                        else:
                            cell.value = '✗'
                            cell.fill = disi_fill
                        
                        # Servis dışı sebebini Excel yorumu (hover) olarak ekle
                        comment_lines = []
                        if sistem:
                            comment_lines.append(f'Sistem: {sistem}')
                        if alt_sistem:
                            comment_lines.append(f'Alt Sistem: {alt_sistem}')
                        if aciklama:
                            comment_lines.append(f'Açıklama: {aciklama}')
                        if comment_lines:
                            cell.comment = Comment('\n'.join(comment_lines), 'SSH Sistem')
                    else:
                        cell.value = '✓'
                        cell.fill = servis_fill
                else:
                    cell.value = ''
                
                cell.font = durum_font
        
        # Sütun genişliklerini ayarla
        ws.column_dimensions['A'].width = 15
        for col_idx in range(2, 2 + len(sorted_dates)):
            ws.column_dimensions[get_column_letter(col_idx)].width = 15
        
        # Satır yüksekliğini ayarla
        ws.row_dimensions[1].height = 25
        for row_idx in range(2, 2 + len(sorted_trams)):
            ws.row_dimensions[row_idx].height = 30
        
        # Excel dosyasını BytesIO'ya yazla
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"Servis_Durumu_Pivot_{datetime.now().strftime('%d.%m.%Y')}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    
    except Exception as e:
        logger.error(f"Servis durumu Excel export hatası: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'Excel export hatası: {str(e)}'}), 400


@bp.route('/excel/daily-table', methods=['GET'])
@login_required
def export_daily_table_excel():
    """
    Günlük servis durum tablosunu pivot tablo olarak Excel'e dışarı aktar
    Satırlar: Tarihler, Sütunlar: Tramvaylar
    Durum gösterimleri: ✓ (yeşil)=Servis, ⚠ (turuncu)=İşletme Kaynaklı, ✗ (kırmızı)=Servis Dışı
    """
    try:
        from io import BytesIO
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.comments import Comment
        
        current_project = session.get('current_project', 'belgrad')
        
        # Veriler.xlsx'ten geçerli araç listesini al
        valid_trams = get_tram_ids_from_veriler(current_project)
        valid_trams_set = set(valid_trams) if valid_trams else None
        
        # Tüm tarihsel verileri al (sadece aktif proje)
        all_status_records = ServiceStatus.query.filter_by(
            project_code=current_project
        ).order_by(
            ServiceStatus.date.desc(), 
            ServiceStatus.tram_id
        ).all()
        
        if not all_status_records:
            return jsonify({'error': 'Veri bulunamadı'}), 400
        
        # Pivot tablo yapısı oluştur: tramvay -> {tarih -> {status, sistem, alt_sistem, aciklama}}
        pivot_data = {}
        all_trams = set()
        all_dates = set()
        
        for record in all_status_records:
            # Geçerli araç listesinde olmayanları atla
            if valid_trams_set and record.tram_id not in valid_trams_set:
                continue
            if record.tram_id not in pivot_data:
                pivot_data[record.tram_id] = {}
            pivot_data[record.tram_id][record.date] = {
                'status': record.status,
                'sistem': record.sistem or '',
                'alt_sistem': record.alt_sistem or '',
                'aciklama': record.aciklama or ''
            }
            all_trams.add(record.tram_id)
            all_dates.add(record.date)
        
        # Tramvay ve tarihleri sırala
        sorted_trams = sorted(list(all_trams))
        sorted_dates = sorted(list(all_dates), reverse=True)
        
        # En eski ve en yeni tarihi bul
        oldest_date_str = sorted_dates[-1]  # En eski
        newest_date_str = sorted_dates[0]   # En yeni
        
        oldest_date = datetime.strptime(oldest_date_str, '%Y-%m-%d').date()
        newest_date = datetime.strptime(newest_date_str, '%Y-%m-%d').date()
        
        # En eski ile en yeni arasındaki TÜM günleri generate et
        from datetime import timedelta
        all_dates_filled = []
        current_date = newest_date
        while current_date >= oldest_date:
            all_dates_filled.append(current_date.strftime('%Y-%m-%d'))
            current_date -= timedelta(days=1)
        
        sorted_dates = all_dates_filled  # Tüm günleri kullan
        
        # Pivot tabloyu tüm tarih ve tramvaylar için initialize et (boş da olsa göster)
        for tram in sorted_trams:
            if tram not in pivot_data:
                pivot_data[tram] = {}
            for date_str in sorted_dates:
                if date_str not in pivot_data[tram]:
                    pivot_data[tram][date_str] = None  # Boş veri None olarak set
        
        # Excel dosyası oluştur
        wb = Workbook()
        ws = wb.active
        ws.title = 'Servis Durumu'
        
        # Stil tanımları
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Durum renklerini tanımla
        servis_fill = PatternFill(start_color='00B050', end_color='00B050', fill_type='solid')  # Yeşil
        isletme_fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')  # Turuncu
        disi_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')    # Kırmızı
        
        durum_font = Font(bold=True, size=14, color='FFFFFF')
        
        # Header satırı: Tarihler (Sütunlar)
        ws['A1'] = 'Tramvay'
        ws['A1'].fill = header_fill
        ws['A1'].font = header_font
        ws['A1'].border = border
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        
        for col_idx, date_str in enumerate(sorted_dates, start=2):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = date_str
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Veri satırları: Tramvaylar (Satırlar)
        for row_idx, tram_id in enumerate(sorted_trams, start=2):
            # Tramvay hücresi
            cell = ws.cell(row=row_idx, column=1)
            cell.value = tram_id
            cell.fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
            cell.font = Font(bold=True)
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Her tarih için durum
            for col_idx, date_str in enumerate(sorted_dates, start=2):
                record_info = pivot_data[tram_id].get(date_str, None)
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center')
                
                status = record_info.get('status', '') if isinstance(record_info, dict) else (record_info or '')
                sistem = record_info.get('sistem', '') if isinstance(record_info, dict) else ''
                alt_sistem = record_info.get('alt_sistem', '') if isinstance(record_info, dict) else ''
                aciklama = record_info.get('aciklama', '') if isinstance(record_info, dict) else ''
                
                # Durum sembolü ve rengi
                if status and 'Servis' in status:
                    if 'Dışı' in status:
                        if 'İşletme' in status:
                            cell.value = '⚠'
                            cell.fill = isletme_fill
                        else:
                            cell.value = '✗'
                            cell.fill = disi_fill
                        
                        # Servis dışı sebebini Excel yorumu (hover) olarak ekle
                        comment_lines = []
                        if sistem:
                            comment_lines.append(f'Sistem: {sistem}')
                        if alt_sistem:
                            comment_lines.append(f'Alt Sistem: {alt_sistem}')
                        if aciklama:
                            comment_lines.append(f'Açıklama: {aciklama}')
                        if comment_lines:
                            cell.comment = Comment('\n'.join(comment_lines), 'SSH Sistem')
                    else:
                        cell.value = '✓'
                        cell.fill = servis_fill
                else:
                    cell.value = ''
                
                cell.font = durum_font
        
        # Sütun genişliklerini ayarla
        ws.column_dimensions['A'].width = 15
        for col_idx in range(2, 2 + len(sorted_dates)):
            ws.column_dimensions[get_column_letter(col_idx)].width = 15
        
        # Satır yüksekliğini ayarla
        ws.row_dimensions[1].height = 25
        for row_idx in range(2, 2 + len(sorted_trams)):
            ws.row_dimensions[row_idx].height = 30
        
        # Excel dosyasını BytesIO'ya yazla
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"Servis_Durumu_Pivot_{datetime.now().strftime('%d.%m.%Y')}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    
    except Exception as e:
        logger.error(f"Günlük tablo Excel export hatası: {str(e)}")
        import traceback
        traceback.print_exc()
        flash(f'Excel export hatası: {str(e)}', 'danger')
        return redirect(url_for('service_status.service_status_page'))


# TEST ENDPOINT - Debug için
@bp.route('/test/export/<report_type>', methods=['GET'])
def test_export(report_type):
    """Test export endpoint - debug amaçlı (login gerektirmez)"""
    try:
        project = request.args.get('project', 'belgrad')
        output_dir = os.path.join('logs', project, 'reports')
        os.makedirs(output_dir, exist_ok=True)
        
        if report_type == 'comprehensive':
            tram_ids = [eq.equipment_code for eq in Equipment.query.all()]
            filename = f"TEST_Kapsamli_Servis_Durumu_Raporu_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            filepath = os.path.join(output_dir, filename)
            EnhancedExcelGenerator.create_comprehensive_availability_report(tram_ids, filepath)
            
        elif report_type == 'rootcause':
            tram_ids = [eq.equipment_code for eq in Equipment.query.all()]
            filename = f"TEST_Root_Cause_Analiz_Raporu_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            filepath = os.path.join(output_dir, filename)
            EnhancedExcelGenerator.create_root_cause_analysis_report(tram_ids, filepath)
            
        elif report_type == 'daily':
            tram_id = request.args.get('tram_id', Equipment.query.first().equipment_code if Equipment.query.first() else 'TEST')
            filename = f"TEST_Gunluk_Durum_{tram_id}_{date.today().strftime('%Y%m%d')}.xlsx"
            filepath = os.path.join(output_dir, filename)
            EnhancedExcelGenerator.create_detailed_daily_report(tram_id, filepath)
        else:
            return jsonify({'error': 'Bilinmeyen rapor türü'}), 400
        
        logger.info(f"Test report exported to {filepath}")
        
        return send_file(filepath, as_attachment=True, download_name=filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    
    except Exception as e:
        logger.error(f"Error in test export: {str(e)}")
        return jsonify({'error': f'Hata: {str(e)}'}), 400


@bp.route('/analiz/alt-sistem', methods=['GET'])
@login_required
def analiz_alt_sistem():
    """Aylık alt sistem analizi - servis dışı günleri say"""
    try:
        yil = request.args.get('yil', type=int)
        ay = request.args.get('ay', type=int)
        project_code = session.get('current_project', 'belgrad')
        
        if not yil or not ay:
            return jsonify({'success': False, 'error': 'Yıl ve ay parametreleri gerekli'}), 400
        
        # Excel logger'ı kullanarak aylık analizi çek
        from utils.utils_service_status_excel_logger import ServiceStatusExcelLogger
        logger_obj = ServiceStatusExcelLogger(project_code)
        analysis = logger_obj.get_monthly_analysis(yil, ay)
        
        return jsonify({
            'success': True,
            'analysis': analysis
        })
    
    except Exception as e:
        logger.error(f"Error in alt sistem analysis: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 400
