"""
Detaylı RCA Excel Raporu Oluştur
ServiceStatus'ten sistem/alt sistem bilgilerini kulla
"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
from pathlib import Path
import os
from app import create_app, db
from models import RootCauseAnalysis, ServiceStatus, Failure, Equipment

class RCAReportGenerator:
    """RCA Analizini Excel'e aktar"""
    
    def __init__(self, project_code='belgrad'):
        self.project_code = project_code
        self.wb = Workbook()
        self.wb.remove(self.wb.active)
        
        # Stiller
        self.header_font = Font(color="FFFFFF", bold=True, size=12)
        self.header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
        self.title_font = Font(size=14, bold=True, color="1F497D")
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def add_rca_summary(self, rca_records):
        """RCA Özet sayfası ekle"""
        ws = self.wb.create_sheet("RCA Özet", 0)
        
        ws['A1'] = "Root Cause Analysis Özeti"
        ws['A1'].font = self.title_font
        ws.merge_cells('A1:F1')
        
        # Headers
        headers = ['Tramvay', 'Sistem', 'Alt Sistem', 'Temel Neden', 'Şiddet', 'Tarih']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.border = self.border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        ws.row_dimensions[3].height = 25
        
        # Veriler
        row = 4
        for rca in rca_records:
            ws.cell(row=row, column=1).value = rca.tram_id
            ws.cell(row=row, column=2).value = rca.sistem
            ws.cell(row=row, column=3).value = rca.alt_sistem or ''
            ws.cell(row=row, column=4).value = rca.root_cause[:50] + '...' if len(rca.root_cause) > 50 else rca.root_cause
            ws.cell(row=row, column=5).value = rca.severity
            ws.cell(row=row, column=6).value = rca.analysis_date.strftime('%d.%m.%Y') if rca.analysis_date else ''
            
            # Şiddet rengini koy
            severity_color = {
                'critical': 'FF0000',  # Kırmızı
                'high': 'FF6600',      # Orange
                'medium': 'FFC000',    # Sarı
                'low': '92D050'        # Yeşil
            }
            ws.cell(row=row, column=5).fill = PatternFill(
                start_color=severity_color.get(rca.severity, 'FFFFFF'),
                end_color=severity_color.get(rca.severity, 'FFFFFF'),
                fill_type='solid'
            )
            
            for col in range(1, 7):
                ws.cell(row=row, column=col).border = self.border
                ws.cell(row=row, column=col).alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            
            row += 1
        
        # Kolonu genişlik
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 30
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 12
    
    def add_rca_details(self, rca_records):
        """Detaylı RCA sayfası ekle"""
        ws = self.wb.create_sheet("RCA Detaylar")
        
        initial_row = 1
        
        for idx, rca in enumerate(rca_records[:10]):  # Son 10 RCA'yı başsyfaya ekle
            if idx > 0:
                initial_row += 5
            
            # RCA Başlığı
            ws.cell(row=initial_row, column=1).value = f"RCA #{idx+1}: {rca.tram_id} - {rca.sistem}"
            cell = ws.cell(row=initial_row, column=1)
            cell.font = Font(size=11, bold=True, color="1F497D")
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            ws.merge_cells(f'A{initial_row}:F{initial_row}')
            
            # Detaylar
            row = initial_row + 1
            
            details = [
                ('Tramvay', rca.tram_id),
                ('Sistem', rca.sistem),
                ('Alt Sistem', rca.alt_sistem or '-'),
                ('Temel Neden', rca.root_cause),
                ('Acil Neden', rca.immediate_cause or '-'),
                ('Katkıda Bulunan Faktörler', rca.contributing_factors or '-'),
                ('Düzeltici Eylemler', rca.corrective_actions or '-'),
                ('Önleyici Eylemler', rca.preventive_actions or '-'),
                ('Şiddet', rca.severity),
                ('Arıza Modu', rca.failure_mode or '-'),
                ('Durum', rca.status),
                ('Analiz Tarihi', rca.analysis_date.strftime('%d.%m.%Y %H:%M') if rca.analysis_date else '-'),
            ]
            
            for label, value in details:
                ws.cell(row=row, column=1).value = label
                ws.cell(row=row, column=1).font = Font(bold=True, size=10)
                ws.cell(row=row, column=1).fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
                
                ws.cell(row=row, column=2).value = str(value)
                
                for col in range(1, 7):
                    ws.cell(row=row, column=col).border = self.border
                    ws.cell(row=row, column=col).alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                
                row += 1
        
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 60
    
    def add_system_breakdown(self, service_statuses):
        """Sistem bazında breakdown"""
        ws = self.wb.create_sheet("Sistem Analizi")
        
        ws['A1'] = "Sistem & Alt Sistem Bazında Arızalar"
        ws['A1'].font = self.title_font
        ws.merge_cells('A1:D1')
        
        # Sistem sayılarını al
        sistem_counts = {}
        for status in service_statuses:
            sistem = status.sistem or 'Tanımlanmamış'
            alt_sistem = status.alt_sistem or '-'
            
            key = f"{sistem}/{alt_sistem}"
            sistem_counts[key] = sistem_counts.get(key, 0) + 1
        
        # Headers
        headers = ['Sistem', 'Alt Sistem', 'Arıza Sayısı', 'Yüzde']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col)
            cell.value = header
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.border = self.border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Veriler
        total = sum(sistem_counts.values())
        row = 3
        for key, count in sorted(sistem_counts.items(), key=lambda x: x[1], reverse=True):
            sistem, alt_sistem = key.split('/')
            ws.cell(row=row, column=1).value = sistem
            ws.cell(row=row, column=2).value = alt_sistem
            ws.cell(row=row, column=3).value = count
            ws.cell(row=row, column=4).value = f"{(count/total*100):.1f}%"
            
            for col in range(1, 5):
                ws.cell(row=row, column=col).border = self.border
                ws.cell(row=row, column=col).alignment = Alignment(horizontal='center', vertical='center')
            
            row += 1
        
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 12
    
    def save(self, output_path):
        """Dosyayı kaydet"""
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        self.wb.save(output_path)
        print(f"✅ RCA Raporu kaydedildi: {output_path}")

# Helper fonksiyon - Flask app'ı olmadan kullanılabilir
def generate_rca_report(project_code='belgrad', output_path=None):
    """RCA raporunu oluştur"""
    app = create_app()
    
    with app.app_context():
        if output_path is None:
            output_path = f"reports/rca_{project_code}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        # Veri topla
        rca_records = RootCauseAnalysis.query.order_by(RootCauseAnalysis.analysis_date.desc()).limit(50).all()
        service_statuses = ServiceStatus.query.filter_by(project_code=project_code).all()
        
        # Rapor oluştur
        generator = RCAReportGenerator(project_code)
        
        if rca_records:
            generator.add_rca_summary(rca_records)
            generator.add_rca_details(rca_records)
        
        if service_statuses:
            generator.add_system_breakdown(service_statuses)
        
        generator.save(output_path)
    
    return output_path
