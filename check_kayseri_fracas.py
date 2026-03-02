from openpyxl import load_workbook
import os

fracas_file = os.path.join(os.getcwd(), 'logs/kayseri/ariza_listesi/Fracas_KAYSERI.xlsx')
print(f"File: {fracas_file}")
print(f"Exists: {os.path.exists(fracas_file)}")

try:
    wb = load_workbook(fracas_file, data_only=True)
    print(f"Sheets: {wb.sheetnames}")
    
    if 'FRACAS' in wb.sheetnames:
        ws = wb['FRACAS']
        print(f"Max row: {ws.max_row}")
        
        # E sütunun son 10 satırını yakala
        print("\nLast 10 rows, Column E:")
        for row in range(max(5, ws.max_row - 9), ws.max_row + 1):
            val = ws.cell(row=row, column=5).value
            print(f"  Row {row}: {val}")
        
        # Tüm FF'leri bul
        all_ff = []
        for row in range(5, ws.max_row + 1):
            val = ws.cell(row=row, column=5).value
            if val and 'FF-' in str(val):
                all_ff.append((row, val))
        
        if all_ff:
            print(f"\nTotal FF entries: {len(all_ff)}")
            print("Last 3 FF entries:")
            for row, val in all_ff[-3:]:
                print(f"  Row {row}: {val}")
        else:
            print("\nNo FF entries found in Column E")
    else:
        print("FRACAS sheet not found")
    
    wb.close()
except Exception as e:
    print(f"Error: {e}")
    import traceback
    traceback.print_exc()
