from openpyxl import load_workbook
import os
import json

data_dir = 'data/belgrad'
veriler_path = os.path.join(data_dir, 'Veriler.xlsx')

if os.path.exists(veriler_path):
    wb = load_workbook(veriler_path)
    ws = wb['Sayfa2']
    
    sistemler = {}
    
    # Sistem (Sütun F=6) ve Alt Sistem (Sütun G=7) verilerini çek
    for row_idx in range(2, ws.max_row + 1):
        sistem = ws.cell(row=row_idx, column=6).value
        alt_sistem = ws.cell(row=row_idx, column=7).value
        kategori = ws.cell(row=row_idx, column=10).value
        
        # SİSTEM kategorisindeki verileri al
        if sistem and kategori == 'SİSTEM':
            if sistem not in sistemler:
                sistemler[sistem] = set()
        
        # ALT SİSTEM kategorisindeki alt sistemleri sisteme ekle
        if alt_sistem and kategori == 'ALT SİSTEM':
            prev_sistem = ws.cell(row=row_idx-1, column=6).value if row_idx > 2 else None
            if prev_sistem:
                if prev_sistem not in sistemler:
                    sistemler[prev_sistem] = set()
                sistemler[prev_sistem].add(alt_sistem)
    
    # Set'leri list'e çevir
    sistemler = {k: sorted(list(v)) for k, v in sistemler.items()}
    
    print('SISTEMLER VE ALT SİSTEMLER:')
    print('=' * 60)
    for sistem in sorted(sistemler.keys()):
        print(f'\n📌 {sistem}')
        if sistemler[sistem]:
            for alt in sistemler[sistem]:
                print(f'   └─ {alt}')
        else:
            print('   └─ (Alt sistem yok)')
    
    print('\n' + '=' * 60)
    print(f'Toplam Sistem: {len(sistemler)}')
    print(f'Toplam Alt Sistem: {sum(len(v) for v in sistemler.values())}')
else:
    print('Veriler.xlsx dosyası bulunamadı!')
