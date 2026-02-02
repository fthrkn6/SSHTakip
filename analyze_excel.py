from openpyxl import load_workbook

excel_path = 'data/belgrad/Veriler.xlsx'
wb = load_workbook(excel_path)
ws = wb['Sayfa2']

# Tüm header'ları al
header_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]

print("\n" + "="*120)
print("📊 MEVCUT EXCEL YAPISI ANALIZI")
print("="*120)

print("\n1️⃣  SÜTUN BAŞLIKLARI:")
print("-"*120)
for i, h in enumerate(header_row[:20]):
    print(f"  Sütun {i:2d}: {str(h) if h else '(boş)'}")

# Veri satırı sayısı
data_rows = ws.max_row - 1
print(f"\n\n📈 TOPLAM VERİ SATIRI: {data_rows}")
print(f"📊 TOPLAM SÜTUN: {len([h for h in header_row if h])}")

# İlk birkaç satırı analiz et
print("\n" + "="*120)
print("2️⃣  ÖRNEKTİ VERİ SATIRI YAPISI:")
print("="*120)

for i, row in enumerate(ws.iter_rows(min_row=2, max_row=3, values_only=True), 1):
    print(f"\n Satır {i} (Tram ID: {row[0]}):")
    for j, (header, value) in enumerate(zip(header_row[:15], row[:15])):
        if header:
            print(f"   {header:30} => {str(value)[:40] if value else '(boş)'}")

# Verilerin dağılımını analiz et
print("\n" + "="*120)
print("3️⃣  KATEGORİZASYON ÖNERİSİ:")
print("="*120)

categories = {
    'TEMEL BİLGİ': ['tram_id', 'Project', 'Module'],
    'ARIZA SINIFLANDIRMASI': ['Arıza Sınıfı ', 'Arıza Kaynağı'],
    'SİSTEM BİLGİSİ': ['Sistemler'],
    'BILEŞEN BİLGİSİ': [header_row[i] for i in range(6, min(15, len(header_row))) if header_row[i]]
}

for category, cols in categories.items():
    print(f"\n📁 {category}:")
    for col in cols:
        if col in header_row:
            idx = header_row.index(col)
            print(f"   ✓ {col} (Sütun {idx})")
