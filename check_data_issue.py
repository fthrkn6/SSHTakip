import openpyxl
import pandas as pd

# Check with openpyxl
excel_file = r'logs\ariza_listesi\Ariza_Listesi_BELGRAD.xlsx'
wb = openpyxl.load_workbook(excel_file)
ws = wb['Ariza Listesi']

print('=== First 3 Data Rows (openpyxl) ===')
for row_idx in range(5, 8):
    print(f'\nRow {row_idx}:')
    print(f'  Col 20 (Adedi): {ws.cell(row_idx, 20).value}')
    print(f'  Col 30 (Personel Sayısı): {ws.cell(row_idx, 30).value}')

wb.close()

# Check with pandas
print('\n=== Pandas Read ===')
df = pd.read_excel(excel_file, sheet_name='Ariza Listesi', header=3)
print(f'Shape: {df.shape}')
print(f'Total columns: {len(df.columns)}')
if len(df) > 0:
    print(f'Column 19 (Adedi) value: {df.iloc[0, 19]}')
if len(df.columns) > 29:
    print(f'Column 29 (Personnel): {df.iloc[0, 29]}')
else:
    print(f'Only {len(df.columns)} columns found')
