from openpyxl import load_workbook

excel_path = r'c:\Users\ferki\Desktop\bozankaya_ssh_takip\data\belgrad\Veriler.xlsx'
wb = load_workbook(excel_path)
ws = wb['Sayfa2']

print("Tüm sayfalar:", wb.sheetnames)
print(f"\n=== Sayfa2 ilk 5 satır ===")
for row in range(1, min(6, ws.max_row + 1)):
    print(f"Satır {row}:", [ws.cell(row=row, column=col).value for col in range(1, 12)])

# Kırmızı renkli hücreleri bul
print("\n=== KIRMIZI RENKLİ HÜCRELER ===")
KIRMIZI = 'FFFF0000'
for row in range(1, ws.max_row + 1):
    for col in range(1, 50):
        cell = ws.cell(row=row, column=col)
        if cell.fill and cell.fill.start_color:
            color = str(cell.fill.start_color.rgb)
            if color == KIRMIZI:
                print(f"  Hücre {row},{col}: '{cell.value}' (Renk: {color})")

print("\n=== SARI RENKLİ HÜCRELER ===")
SARI = 'FFFFFF00'
for row in range(1, ws.max_row + 1):
    for col in range(1, 50):
        cell = ws.cell(row=row, column=col)
        if cell.fill and cell.fill.start_color:
            color = str(cell.fill.start_color.rgb)
            if color == SARI:
                print(f"  Hücre {row},{col}: '{cell.value}' (Renk: {color})")

print("\n=== MAVI RENKLİ HÜCRELER ===")
MAVI = 'FF0070C0'
for row in range(1, ws.max_row + 1):
    for col in range(1, 50):
        cell = ws.cell(row=row, column=col)
        if cell.fill and cell.fill.start_color:
            color = str(cell.fill.start_color.rgb)
            if color == MAVI:
                print(f"  Hücre {row},{col}: '{cell.value}' (Renk: {color})")
