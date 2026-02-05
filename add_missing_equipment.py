import openpyxl
from models import db, Equipment
from app import create_app

app = create_app()
with app.app_context():
    print("=" * 80)
    print("ADDING MISSING EQUIPMENT FROM EXCEL")
    print("=" * 80)
    
    # Read Excel file
    wb = openpyxl.load_workbook('data/belgrad/Veriler.xlsx')
    ws = wb['Sayfa2']
    
    # Get headers
    headers = [cell.value for cell in ws[1]]
    tram_id_col = 0
    name_col = None
    
    # Find column indices
    for idx, header in enumerate(headers):
        if header and 'tram' in str(header).lower():
            tram_id_col = idx
        elif header and 'name' in str(header).lower():
            name_col = idx
    
    print(f"\nHeaders found: {headers}")
    print(f"Tram ID column: {tram_id_col}")
    print(f"Name column: {name_col}")
    
    # Read data rows
    added_count = 0
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=False), start=2):
        tram_id_cell = row[tram_id_col]
        tram_id = str(tram_id_cell.value).strip() if tram_id_cell.value else None
        
        if not tram_id or tram_id == 'None':
            continue
        
        # Check if already exists
        if Equipment.query.filter_by(equipment_code=tram_id).first():
            print(f"Already exists: {tram_id}")
            continue
        
        # Get name (use second row for name or equipment code)
        name = tram_id
        if name_col is not None and row[name_col].value:
            name = row[name_col].value
        
        # Create new equipment
        eq = Equipment(
            equipment_code=tram_id,
            name=f"Belgrad Tramvay {tram_id}",
            equipment_type='Tramvay',
            status='aktif',
            manufacturer='Belgrad'
        )
        db.session.add(eq)
        print(f"Adding: {tram_id} - Belgrad Tramvay {tram_id}")
        added_count += 1
    
    print("\n" + "=" * 80)
    print(f"Total added: {added_count} records")
    print("=" * 80)
    
    db.session.commit()
    print("\n✓ Veritabanı güncellendi!")
    
    # Verify
    all_eq = Equipment.query.all()
    print(f"\nTotal equipment in database: {len(all_eq)}")
    print("\nTüm Belgrad tramvayları:")
    for eq in sorted(all_eq, key=lambda x: x.equipment_code):
        print(f"  - {eq.equipment_code}: {eq.name}")
