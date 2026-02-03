import pandas as pd
from openpyxl import load_workbook

fracas_file = r'data\belgrad\BEL25_FRACAS.xlsx'

# Workbook'ü aç
wb = load_workbook(fracas_file)
print(f"📊 Sheet isimleri: {wb.sheetnames}")

ws = wb['FRACAS']

# İlk 5 satırı oku (headers arayalım)
print(f"\n📋 İlk 5 satır (headers arayalım):")
for row_idx in range(1, 6):
    values = []
    for col_idx in range(1, 8):
        val = ws.cell(row=row_idx, column=col_idx).value
        values.append(str(val)[:15] if val else "-")
    print(f"   Satır {row_idx}: {values}")

print(f"\n📌 Max row: {ws.max_row}")

# FRACAS ID sütununu bul
print(f"\n🔍 FRACAS ID sütununu ara:")
fracas_col_found = False
for col in df.columns:
    if isinstance(col, str):
        if 'fracas' in col.lower() and 'id' in col.lower():
            print(f"   Bulundu: {col}")
            print(f"   İlk 5 değer: {df[col].dropna().head(5).tolist()}")
            fracas_col_found = True
            break
if not fracas_col_found:
    print(f"   FRACAS ID sütunu bulunamadı!")
    print(f"   Tüm sütunlar: {[col for col in df.columns if isinstance(col, str)]}")
