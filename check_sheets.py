import openpyxl
import os

os.chdir('data/belgrad')

# Dosya adını bul
fracas_file = [f for f in os.listdir('.') if 'fracas' in f.lower() and f.endswith('.xlsx')][0]
print(f"Dosya: {fracas_file}")

# Sheet adlarını oku
wb = openpyxl.load_workbook(fracas_file)
print(f"Sheet adları: {wb.sheetnames}")

# İlk sheet'in ilk birkaç satırını göster
ws = wb.active
print(f"\nAktif sheet: {ws.title}")
print("İlk 5 satır:")
for i, row in enumerate(ws.iter_rows(max_row=5, values_only=True)):
    print(f"Row {i}: {row[:5]}")  # İlk 5 kolonu göster
