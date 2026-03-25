"""
Excel'den dinamik olarak equipment listesini oku - hardcoded değil
"""

from openpyxl import load_workbook
from format_detector import detect_grid_format

def get_equipment_codes_from_file(file_path):
    """
    Excel dosyasından equipment codes'ı dynamically oku.
    Format detection'u kullanarak doğru sütun/satırdan okur.
    
    Returns:
        list: Equipment codes (örn: ['GEBZE-01', 'GEBZE-02', ...])
    """
    
    # Format detection yap
    format_info = detect_grid_format(file_path)
    
    if not format_info.get('is_valid'):
        return []
    
    try:
        wb = load_workbook(file_path, data_only=True)
        ws = wb.active
        
        format_type = format_info['format']
        equipment_codes = []
        
        if format_type == 'standard':
            # STANDART FORMAT: Satır=Tarih, Sütun=Araç
            # B1:?1 sütunlarında araç kodları
            for col_idx in range(2, ws.max_column + 1):
                cell_value = ws.cell(row=1, column=col_idx).value
                if cell_value:
                    equipment_codes.append(cell_value)
        
        elif format_type == 'transposed':
            # TRANSPOSED FORMAT: Satır=Araç, Sütun=Tarih
            # A2:AN satırlarında araç kodları
            for row_idx in range(2, ws.max_row + 1):
                cell_value = ws.cell(row=row_idx, column=1).value
                if cell_value:
                    equipment_codes.append(cell_value)
        
        wb.close()
        return equipment_codes
    
    except Exception as e:
        print(f"Error reading equipment codes: {e}")
        return []


if __name__ == '__main__':
    # Test
    projects = ['belgrad', 'istanbul', 'ankara', 'kayseri', 'kocaeli', 'gebze']
    
    print("Equipment Codes (Dinamik Okuma)")
    print("=" * 60)
    
    for project in projects:
        file_path = f'data/{project}/service_status_grid.xlsx'
        codes = get_equipment_codes_from_file(file_path)
        print(f"\n{project.upper()}: {codes}")
