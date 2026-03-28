"""
KM (Kilometer) Excel Logger
Maintains project-specific Excel files for KM history and auditing.
Logs all KM updates with timestamp, user, tram ID, and change amounts.
"""

import os
from datetime import datetime, timedelta
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import pytz

# Turkish timezone
TR_TZ = pytz.timezone('Europe/Istanbul')


class KMExcelLogger:
    """
    Manages Excel-based KM (kilometer) logging per project.
    Maintains chronological records of all KM updates.
    """
    
    LOG_DIR = 'logs/km'
    
    def __init__(self, project_code):
        """
        Initialize KM Excel logger for a specific project.
        
        Args:
            project_code: Project identifier (e.g., 'belgrad', 'kayseri')
        """
        self.project_code = project_code
        self.excel_path = self._get_excel_file_path()
        self._ensure_excel_file()
    
    def _get_excel_file_path(self):
        """Get the Excel file path for the project."""
        return os.path.join(self.LOG_DIR, f'{self.project_code}_km_log.xlsx')
    
    def _ensure_excel_file(self):
        """
        Create Excel file if it doesn't exist.
        Sets up headers and formatting.
        """
        os.makedirs(self.LOG_DIR, exist_ok=True)
        
        if not os.path.exists(self.excel_path):
            self._create_new_excel()
    
    def _create_new_excel(self):
        """Create a new Excel workbook with proper structure."""
        wb = Workbook()
        ws = wb.active
        ws.title = 'KM Kayıtları'
        
        # Header row
        headers = [
            'Tarih',           # Date
            'Saat',            # Time
            'Tramvay ID',      # Tram ID
            'Önceki KM',       # Previous KM
            'Yeni KM',         # New KM
            'Fark',            # Difference (calculated)
            'Neden',           # Reason
            'Kullanıcı',       # User
            'Sistem'           # System (manual/auto)
        ]
        
        ws.append(headers)
        
        # Format header row
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Set column widths
        column_widths = {
            'A': 12,  # Tarih
            'B': 10,  # Saat
            'C': 12,  # Tramvay ID
            'D': 12,  # Önceki KM
            'E': 12,  # Yeni KM
            'F': 10,  # Fark
            'G': 20,  # Neden
            'H': 15,  # Kullanıcı
            'I': 12   # Sistem
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # Set number format for KM columns
        for row in range(2, 1000):  # Pre-format for future rows
            ws[f'D{row}'].number_format = '#,##0'
            ws[f'E{row}'].number_format = '#,##0'
            ws[f'F{row}'].number_format = '#,##0'
        
        wb.save(self.excel_path)
    
    def log_km_to_excel(self, tram_id, previous_km, new_km, reason='', 
                        user='Sistem', system_type='Manuel'):
        """
        Log a KM update to Excel (append new row).
        
        Args:
            tram_id: Tramway ID (e.g., 'G1001')
            previous_km: Previous KM value
            new_km: New KM value
            reason: Reason for update (optional)
            user: Username who made the change
            system_type: 'Manuel' or 'Otomatik'
        
        Returns:
            bool: True if successful, False if failed
        """
        try:
            wb = load_workbook(self.excel_path)
            ws = wb.active
            
            # Get Turkish date and time
            now = datetime.now(TR_TZ)
            date_str = now.strftime('%d.%m.%Y')
            time_str = now.strftime('%H:%M:%S')
            
            # Calculate difference
            difference = new_km - previous_km
            
            # Append new row
            new_row = [
                date_str,
                time_str,
                tram_id,
                int(previous_km),
                int(new_km),
                int(difference),
                reason,
                user,
                system_type
            ]
            
            ws.append(new_row)
            
            # Format the new row
            row_num = ws.max_row
            self._format_data_row(ws, row_num)
            
            wb.save(self.excel_path)
            return True
            
        except Exception as e:
            print(f'KM Excel logging error (append): {e}')
            return False
    
    def update_km_in_excel(self, tram_id, old_km, new_km, reason='', 
                           user='Sistem', system_type='Manuel'):
        """
        Update an existing KM entry in Excel.
        Finds the row by tram_id and old_km, updates the values.
        
        Args:
            tram_id: Tramway ID
            old_km: Old KM value to find
            new_km: New KM value
            reason: Updated reason
            user: Username
            system_type: 'Manuel' or 'Otomatik'
        
        Returns:
            bool: True if found and updated, False if not found
        """
        try:
            wb = load_workbook(self.excel_path)
            ws = wb.active
            
            # Search for matching row (tram_id and old_km)
            found = False
            for row_idx in range(2, ws.max_row + 1):
                if (ws[f'C{row_idx}'].value == tram_id and 
                    ws[f'E{row_idx}'].value == int(old_km)):
                    
                    # Update the row
                    now = datetime.now(TR_TZ)
                    ws[f'A{row_idx}'].value = now.strftime('%d.%m.%Y')
                    ws[f'B{row_idx}'].value = now.strftime('%H:%M:%S')
                    ws[f'E{row_idx}'].value = int(new_km)
                    ws[f'F{row_idx}'].value = int(new_km) - int(old_km)
                    ws[f'G{row_idx}'].value = reason
                    ws[f'H{row_idx}'].value = user
                    ws[f'I{row_idx}'].value = system_type
                    
                    found = True
                    break
            
            if found:
                wb.save(self.excel_path)
                return True
            else:
                print(f'KM record not found: Tram {tram_id}, KM {old_km}')
                return False
                
        except Exception as e:
            print(f'KM Excel logging error (update): {e}')
            return False
    
    def _format_data_row(self, ws, row_num):
        """Apply formatting to a data row."""
        # Border style
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Alternate row colors for readability
        if row_num % 2 == 0:
            fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        else:
            fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
        
        for col in range(1, 10):
            cell = ws.cell(row_num, col)
            cell.border = thin_border
            cell.fill = fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Reason column (left-aligned for text)
        ws[f'G{row_num}'].alignment = Alignment(horizontal='left', vertical='center')
    
    def get_excel_stats(self):
        """
        Get statistics about the KM log file.
        
        Returns:
            dict: Stats including record count, date range, min/max KM values
        """
        try:
            wb = load_workbook(self.excel_path)
            ws = wb.active
            
            record_count = ws.max_row - 1  # Exclude header
            
            if record_count == 0:
                return {
                    'record_count': 0,
                    'first_date': None,
                    'last_date': None,
                    'min_km': None,
                    'max_km': None
                }
            
            # Get values from columns
            dates = [ws[f'A{i}'].value for i in range(2, ws.max_row + 1)]
            kms = [ws[f'E{i}'].value for i in range(2, ws.max_row + 1)]
            kms = [k for k in kms if k is not None]
            
            return {
                'record_count': record_count,
                'first_date': dates[0] if dates else None,
                'last_date': dates[-1] if dates else None,
                'min_km': min(kms) if kms else None,
                'max_km': max(kms) if kms else None
            }
            
        except Exception as e:
            print(f'Error getting Excel stats: {e}')
            return None


# Convenience function for quick logging
def log_km_excel(project_code, tram_id, previous_km, new_km, reason='', 
                 user='Sistem', system_type='Manuel'):
    """
    Quick function to log KM without creating a logger instance.
    
    Usage:
        from utils_km_excel_logger import log_km_excel
        log_km_excel('belgrad', 'G1001', 125000, 125500, 'Gün sonu sayımı', 'mert.yilmaz')
    """
    logger = KMExcelLogger(project_code)
    return logger.log_km_to_excel(tram_id, previous_km, new_km, reason, user, system_type)
