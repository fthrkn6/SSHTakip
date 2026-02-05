#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel dosyalarına zebra pattern (beyaz-gri alternatif satırlar) uygula
"""
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ariza_listesi_dir = os.path.join(os.path.dirname(__file__), 'logs', 'ariza_listesi')
updated_count = 0

if os.path.exists(ariza_listesi_dir):
    print(f"📁 Klasör: {ariza_listesi_dir}\n")
    
    for file in os.listdir(ariza_listesi_dir):
        if 'Ariza_Listesi' in file and file.endswith('.xlsx'):
            file_path = os.path.join(ariza_listesi_dir, file)
            print(f"📄 Dosya: {file}")
            
            try:
                wb = load_workbook(file_path)
                ws = wb.active
                
                # Renkler
                white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
                gray_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                               top=Side(style='thin'), bottom=Side(style='thin'))
                data_font = Font(size=10)
                
                # Satır sayısını bul (header Row 4, veri Row 5'den başla)
                max_row = ws.max_row
                max_col = 29  # 29 sütun
                
                # Row 5'den sonrası zebra pattern
                for row_idx in range(5, max_row + 1):
                    # Satır numarasına göre rengi belirle
                    # Row 5 = beyaz, Row 6 = gri, Row 7 = beyaz, vb.
                    is_white = (row_idx - 5) % 2 == 0
                    fill = white_fill if is_white else gray_fill
                    
                    for col_idx in range(1, max_col + 1):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        cell.fill = fill
                        cell.border = border
                        cell.font = data_font
                        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                
                # Dosyayı kaydet
                wb.save(file_path)
                print(f"   ✅ Zebra pattern uygulandı ({max_row - 4} satır)\n")
                updated_count += 1
                
            except Exception as e:
                print(f"   ❌ Hata: {e}\n")
else:
    print(f"❌ Klasör bulunamadı: {ariza_listesi_dir}")

print(f"\n{'='*60}")
print(f"✅ İşlem tamamlandı!")
print(f"   Güncellenen dosya: {updated_count}")
print(f"{'='*60}")
