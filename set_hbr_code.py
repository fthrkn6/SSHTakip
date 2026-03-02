#!/usr/bin/env python
# -*- coding: utf-8 -*-

import os
from openpyxl import load_workbook

# Set HBR code in Veriler.xlsx for Kayseri
data_dir = os.path.join(os.path.dirname(__file__), 'data', 'kayseri')

# Find Veriler.xlsx (case-insensitive)
veriler_file = None
if os.path.exists(data_dir):
    for f in os.listdir(data_dir):
        if f.lower() == 'veriler.xlsx':
            veriler_file = os.path.join(data_dir, f)
            break

if veriler_file:
    print(f"Bulunan dosya: {veriler_file}")
    wb = load_workbook(veriler_file)
    
    # Sayfa2'de B2'yi KAY5+6 olarak set et
    if 'Sayfa2' in wb.sheetnames:
        ws = wb['Sayfa2']
        print(f"Sayfa2 B2 eski değeri: {ws['B2'].value}")
        ws['B2'].value = 'KAY5+6'
        wb.save(veriler_file)
        print(f"Sayfa2 B2 yeni değeri: {ws['B2'].value}")
        print("Dosya kaydedildi!")
    else:
        print(f"Sayfa2 bulunamadı. Sheetler: {wb.sheetnames}")
    
    wb.close()
else:
    print("Veriler.xlsx bulunamadı!")
