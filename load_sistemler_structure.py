from openpyxl import load_workbook

data_dir = 'data/belgrad'
veriler_path = f'{data_dir}/Veriler.xlsx'

wb = load_workbook(veriler_path)
ws = wb['Sayfa2']

# Renk tanımları
KIRMIZI = 'FFFF0000'  # SİSTEM
SARI = 'FFFFFF00'     # TEDARİKÇİ
MAVI = 'FF0070C0'     # ALT SİSTEM

# Sütunları tarayarak sistemleri bul (6=F, 7=G, 8=H, 9=I, 11=K)
sütunlar = [6, 7, 8, 9, 11]  # F, G, H, I, K

sistemler_yapı = {}

# Her sütunu kontrol et
for sütun in sütunlar:
    sistem_adi = None
    
    for row in range(1, ws.max_row + 1):
        cell = ws.cell(row=row, column=sütun)
        value = cell.value
        fill = cell.fill
        
        color_hex = None
        if fill and fill.start_color:
            color_hex = str(fill.start_color.rgb) if fill.start_color.rgb else None
        
        # Kırmızı renkli ise sistem bulundu
        if color_hex == KIRMIZI and value:
            sistem_adi = value
            if sistem_adi not in sistemler_yapı:
                sistemler_yapı[sistem_adi] = {
                    'tedarikçiler': set(),
                    'alt_sistemler': set()
                }
        
        # Sarı renkli ise tedarikçi (sistemin altında olmalı)
        elif color_hex == SARI and value and sistem_adi:
            sistemler_yapı[sistem_adi]['tedarikçiler'].add(value)
        
        # Mavi renkli ise alt sistem (sistemin altında olmalı)
        elif color_hex == MAVI and value and sistem_adi:
            sistemler_yapı[sistem_adi]['alt_sistemler'].add(value)

# Set'leri list'e çevir ve sort et
sistemler_final = {
    k: {
        'tedarikçiler': sorted(list(v['tedarikçiler'])),
        'alt_sistemler': sorted(list(v['alt_sistemler']))
    }
    for k, v in sistemler_yapı.items()
}

print('SISTEMLER YAPISI:')
print('=' * 80)

for sistem in sorted(sistemler_final.keys()):
    data = sistemler_final[sistem]
    print(f'\n📌 {sistem}')
    
    if data['tedarikçiler']:
        print(f'\n   TEDARİKÇİLER:')
        for t in data['tedarikçiler']:
            print(f'   🟨 {t}')
    
    if data['alt_sistemler']:
        print(f'\n   ALT SİSTEMLER:')
        for a in data['alt_sistemler']:
            print(f'   🟦 {a}')
    
    print()

print('=' * 80)
print(f'Toplam Sistem: {len(sistemler_final)}')
