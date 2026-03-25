"""
Istanbul, Ankara, Kocaeli dosyalarını bugüne kadar güncelleyen script
"""

from openpyxl import load_workbook
from datetime import date, timedelta

projects = ['istanbul', 'ankara', 'kocaeli']

print("Dosyaları Güncelleme")
print("=" * 60)

for project in projects:
    file_path = f'data/{project}/service_status_grid.xlsx'
    
    try:
        wb = load_workbook(file_path)
        ws = wb.active
        
        # Son satırı bul
        last_row_with_date = None
        for row in range(2, ws.max_row + 1):
            if ws.cell(row, 1).value:
                last_row_with_date = row
        
        if last_row_with_date is None:
            print(f"{project}: Tarih satırı bulunamadı")
            continue
        
        # Son tarih al
        last_date_cell = ws.cell(last_row_with_date, 1).value
        print(f"\n{project.upper()}:")
        print(f"  Son tarih: {last_date_cell}")
        
        # Bugüne kadar veri ekle (her gün bir satır)
        today = date.today()
        from datetime import datetime
        
        if isinstance(last_date_cell, datetime):
            last_date = last_date_cell.date()
        else:
            last_date = last_date_cell
        
        current_date = last_date + timedelta(days=1)
        current_row = last_row_with_date + 1
        
        import random
        random.seed(42)
        
        while current_date <= today:
            ws.cell(current_row, 1, value=current_date)
            
            # Her araç için rastgele status ekle (80% ✓, 15% ⚠, 5% ✗)
            for col in range(2, ws.max_column + 1):
                rand = random.random()
                if rand < 0.80:
                    symbol = '✓'
                elif rand < 0.95:
                    symbol = '⚠'
                else:
                    symbol = '✗'
                ws.cell(current_row, col, value=symbol)
            
            current_date += timedelta(days=1)
            current_row += 1
        
        wb.save(file_path)
        print(f"  ✓ Güncellendi: {(current_row - 1 - last_row_with_date)} satır eklendi")
        print(f"  Yeni son tarih: {today}")
        
    except Exception as e:
        print(f"{project}: Hata - {e}")

print("\n" + "=" * 60)
