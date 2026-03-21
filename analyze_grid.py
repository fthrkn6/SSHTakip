import openpyxl
from pathlib import Path

grid_path = Path('data/belgrad/service_status_grid.xlsx')

if grid_path.exists():
    wb = openpyxl.load_workbook(grid_path, data_only=True)
    ws = wb.active
    
    # Header oku
    headers = []
    for col in range(1, min(12, ws.max_column + 1)):
        headers.append(ws.cell(1, col).value)
    
    print("HEADER:", headers)
    print()
    
    # İlk 5 row oku
    for row in range(2, min(7, ws.max_row + 1)):
        data = []
        for col in range(1, min(7, ws.max_column + 1)):
            val = ws.cell(row, col).value
            data.append(str(val) if val else "-")
        print("ROW", row, ":", " | ".join(data))
    
    print(f"\nGridBoyut: {ws.max_row} rows, {ws.max_column} cols")
    wb.close()
