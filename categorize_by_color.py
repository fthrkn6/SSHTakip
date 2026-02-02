from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment

# Renkler
RED = 'FFFF0000'      # Sistem
YELLOW = 'FFFFFF00'   # Tedarikçi
BLUE = 'FF0070C0'     # Alt Sistem

wb = load_workbook('data/belgrad/Veriler.xlsx')
ws = wb['Sayfa2']

print("🔍 Kategorize ediliyor...\n")

# Kategori mapping
kategoriler = {
    RED: 'SİSTEM',
    YELLOW: 'TEDARİKÇİ',
    BLUE: 'ALT SİSTEM'
}

# Tüm satırları kontrol et ve kategori sütununa yaz
for row_idx in range(2, ws.max_row + 1):
    # Sütun 6'daki hücreyi kontrol et (Sistemler)
    cell = ws.cell(row=row_idx, column=6)
    value = cell.value
    
    # Hücrenin rengini kontrol et
    if cell.fill and cell.fill.start_color:
        color = cell.fill.start_color.rgb if hasattr(cell.fill.start_color, 'rgb') else None
        
        if color in kategoriler:
            category = kategoriler[color]
            
            # Sütun 10'a kategori yaz (I sütunu boş olacak)
            cat_cell = ws.cell(row=row_idx, column=10)
            cat_cell.value = category
            
            # Renk kodu eşleştir - rengini devam ettir
            cat_cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
            cat_cell.alignment = Alignment(horizontal='center', vertical='center')
            
            print(f"Satır {row_idx}: {str(value)[:30]:30} → {category}")

# Header ekle
try:
    header_cell = ws.cell(row=1, column=10)
    header_cell.value = "KATEGORİ"
except:
    # Eğer merged cell ise direkt skipla
    pass

# Sütun genişliğini ayarla
ws.column_dimensions['J'].width = 15

# Dosyayı kaydet
wb.save('data/belgrad/Veriler.xlsx')

print("\n✅ BAŞARILI!")
print("\nVeriler.xlsx dosyası güncellendi:")
print("   • Sütun J: KATEGORİ")
print("   • Renkler korundu")
print("   • FFFF0000 (Kırmızı) → SİSTEM")
print("   • FFFFFF00 (Sarı) → TEDARİKÇİ")
print("   • FF0070C0 (Mavi) → ALT SİSTEM")
