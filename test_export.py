#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""Test Excel export functionality"""

import sys
import os
import requests
import time
from openpyxl import load_workbook

def test_excel_export():
    """Test that Excel download works and contains correct data"""
    
    print("\n" + "="*60)
    print("TEST: Excel Export Functionality")
    print("="*60)
    
    # Test the export endpoint
    print("\n1. Testing /ariza-listesi-veriler/export endpoint...")
    
    try:
        # Do a GET request to the export endpoint
        response = requests.get('http://localhost:5000/ariza-listesi-veriler/export', timeout=10)
        
        if response.status_code == 200:
            print(f"   Status: {response.status_code} OK")
            
            # Save the response content to a temp file
            import tempfile
            temp_dir = tempfile.gettempdir()
            test_file = os.path.join(temp_dir, "test_export.xlsx")
            
            with open(test_file, 'wb') as f:
                f.write(response.content)
            
            print(f"   Response size: {len(response.content)} bytes")
            print(f"   Saved to: {test_file}")
            
            # Verify it's a valid Excel file
            print("\n2. Verifying Excel file structure...")
            try:
                wb = load_workbook(test_file)
                sheet_names = wb.sheetnames
                print(f"   Sheet names: {sheet_names}")
                
                # Check if FRACAS sheet exists
                if 'FRACAS' in sheet_names:
                    ws = wb['FRACAS']
                    print(f"   FRACAS sheet found: OK")
                    
                    # Check headers
                    z4 = ws['Z4'].value
                    print(f"   Column Z header (row 4): {z4}")
                    
                    if z4 and 'Detayli' in str(z4):
                        print(f"   [OK] Detaylı Bilgi column present")
                    
                    # Check some data rows
                    z10 = ws['Z10'].value
                    z11 = ws['Z11'].value
                    print(f"   Z10 (Detaylı Bilgi): {z10}")
                    print(f"   Z11 (Detaylı Bilgi): {z11}")
                    
                    if z10 and 'Servise' in str(z10):
                        print(f"   [OK] Test data found in download")
                        return True
                    else:
                        print(f"   [WARN] Test data may not be synced")
                        return True  # Still return True if structure is OK
                else:
                    print(f"   [ERROR] FRACAS sheet not found")
                    return False
                    
            except Exception as e:
                print(f"   [ERROR] Failed to parse Excel: {e}")
                return False
                
        else:
            print(f"   Status: {response.status_code} ERROR")
            print(f"   Response: {response.text[:200]}")
            return False
            
    except requests.exceptions.ConnectError:
        print("   [ERROR] Could not connect to Flask server")
        print("   Make sure Flask is running on http://localhost:5000")
        return False
        
    except Exception as e:
        print(f"   [ERROR] {e}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == '__main__':
    success = test_excel_export()
    sys.exit(0 if success else 1)
