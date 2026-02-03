#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
FRACAS ID Oluşturma Test Scripti
Yeni FRACAS ID'nin doğru şekilde oluşturulup oluşturulmadığını test eder
"""

import os
import sys
import pandas as pd
from datetime import datetime

# Routes klasörünün yolunu ekle
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

# Test 1: Excel dosyasından son FRACAS ID'yi oku
print("=" * 60)
print("TEST 1: Excel'den Son FRACAS ID'yi Okuma")
print("=" * 60)

excel_path = r"c:\Users\ferki\Desktop\bozankaya_ssh_takip\data\belgrad\BEL25_FRACAS.xlsx"

if os.path.exists(excel_path):
    print(f"✓ Excel dosyası bulundu: {excel_path}")
    
    # Excel'i oku
    df = pd.read_excel(excel_path, sheet_name='FRACAS', header=0, engine='openpyxl')
    df.columns = df.columns.astype(str).str.replace('\n', ' ', regex=False).str.strip()
    
    # FRACAS ID sütununu bul
    fracas_col = None
    for col in df.columns:
        if 'fracas' in col.lower() and 'id' in col.lower():
            fracas_col = col
            break
    
    if fracas_col:
        print(f"✓ FRACAS ID sütunu bulundu: {fracas_col}")
        
        # Boş olmayan FRACAS ID'leri al
        fracas_ids = df[df[fracas_col].notna()][fracas_col].astype(str).unique()
        print(f"✓ Toplam {len(fracas_ids)} FRACAS ID bulundu")
        
        # Son 10 ID'yi göster
        print("\nSon 10 FRACAS ID:")
        for i, fid in enumerate(sorted(fracas_ids)[-10:], 1):
            print(f"  {i}. {fid}")
        
        # Son ID'yi bul (sayısal olarak en yüksek olanı)
        max_id = None
        max_number = 0
        
        for fid in fracas_ids:
            parts = str(fid).split('-')
            if len(parts) >= 1:
                try:
                    number = int(parts[-1])
                    if number > max_number:
                        max_number = number
                        max_id = fid
                except:
                    pass
        
        print(f"\n✓ Son FRACAS ID: {max_id}")
        print(f"✓ Son ID'deki sayı: {max_number}")
        
        # Yeni ID oluştur
        if max_id:
            next_number = max_number + 1
            parts = str(max_id).split('-')
            prefix = '-'.join(parts[:-1])
            new_fracas_id = f"{prefix}-{next_number:03d}"
            
            print(f"\n✓ Yeni FRACAS ID: {new_fracas_id}")
            print(f"✓ Format korundu: prefix='{prefix}', yeni numara={next_number:03d}")
    else:
        print("✗ FRACAS ID sütunu bulunamadı")
else:
    print(f"✗ Excel dosyası bulunamadı: {excel_path}")

print("\n" + "=" * 60)
print("TEST 2: Fonksiyonları Import Etme")
print("=" * 60)

try:
    # Flask uygulamasını başlat
    from app import create_app
    app = create_app()
    
    with app.app_context():
        from routes.fracas import get_next_fracas_id, save_fracas_to_excel
        
        print("✓ Fonksiyonlar başarıyla import edildi")
        
        # get_next_fracas_id fonksiyonunu test et (project parametresiyle)
        print("\nTEST: get_next_fracas_id(project='belgrad') fonksiyonu:")
        next_id = get_next_fracas_id(project='belgrad')
        if next_id:
            print(f"✓ Oluşturulan FRACAS ID: {next_id}")
        else:
            print("✗ FRACAS ID oluşturulamadı")
        
except Exception as e:
    print(f"✗ Hata: {str(e)}")
    import traceback
    traceback.print_exc()

print("\n" + "=" * 60)
print("TEST 3: Excel'e Veri Yazma Simülasyonu")
print("=" * 60)

try:
    from openpyxl import load_workbook
    
    wb = load_workbook(excel_path)
    ws = wb['FRACAS']
    
    # Header satırını oku
    headers = {}
    for col_num, cell in enumerate(ws[1], 1):
        if cell.value:
            headers[cell.value] = col_num
    
    print(f"✓ Header satırından {len(headers)} sütun okundu")
    print("\nÖnemli sütunlar:")
    for key in ['FRACAS ID', 'Araç Numarası  Vehicle Number', 'Arıza Tanımı Failure Description', 'Hata Tarih /Date']:
        if key in headers:
            print(f"  ✓ {key}: Sütun {headers[key]}")
        else:
            print(f"  ✗ {key}: Bulunamadı")
    
    wb.close()
    
except Exception as e:
    print(f"✗ Hata: {str(e)}")

print("\n" + "=" * 60)
print("✓ Test Tamamlandı")
print("=" * 60)
