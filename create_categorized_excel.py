from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Renkler
COLORS = {
    'L1_SYSTEM': 'D9E8F5',      # Açık mavi
    'L2_SUBSYSTEM': 'E8F5E9',   # Açık yeşil
    'L3_COMPONENT': 'FFF9E6',   # Açık sarı
    'L4_PART': 'FFE6CC',        # Açık turuncu
    'HEADER': '4472C4'           # Koyu mavi
}

# Excel dosyasını aç
wb = load_workbook('data/belgrad/Veriler.xlsx')
ws = wb['Sayfa2']

# Header satırı ekle (Satır 0 - Şu anda satır 1'in üstüne)
ws.insert_rows(1)

# Yeni header'ları yazı
headers = [
    'Tram_ID', 'Proje', 'Modül', 'Arıza_Sınıfı', 'Arıza_Kaynağı',
    'L1_ANA_SİSTEM', 'L2_ALT_SİSTEM', 'L3_BİLEŞEN', 'L4_SPESIFIK_PARÇA'
]

for col_idx, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_idx)
    cell.value = header
    cell.fill = PatternFill(start_color=COLORS['HEADER'], end_color=COLORS['HEADER'], fill_type='solid')
    cell.font = Font(bold=True, color='FFFFFF')
    cell.alignment = Alignment(horizontal='center', vertical='center')

# Veriyi işle
print("📊 Excel verisi kategorize ediliyor...\n")

for row_idx in range(3, ws.max_row + 1):  # 3'ten başla (header + 1)
    # Mevcut verileri oku
    tram_id = ws.cell(row=row_idx, column=1).value
    proje = ws.cell(row=row_idx, column=2).value
    modul = ws.cell(row=row_idx, column=3).value
    ariza_sinifi = ws.cell(row=row_idx, column=4).value
    ariza_kaynagi = ws.cell(row=row_idx, column=5).value
    sistem = ws.cell(row=row_idx, column=6).value
    
    # L1, L2, L3, L4'ü dolumla
    l1 = sistem if sistem else None
    l2 = ws.cell(row=row_idx, column=7).value
    l3 = ws.cell(row=row_idx, column=8).value
    l4 = ws.cell(row=row_idx, column=9).value
    
    # Yeni sütunlara yaz
    ws.cell(row=row_idx, column=6).value = l1 if l1 else None
    ws.cell(row=row_idx, column=7).value = l2 if l2 else None
    ws.cell(row=row_idx, column=8).value = l3 if l3 else None
    ws.cell(row=row_idx, column=9).value = l4 if l4 else None
    
    # Reng uygula
    if l1:
        ws.cell(row=row_idx, column=6).fill = PatternFill(start_color=COLORS['L1_SYSTEM'], end_color=COLORS['L1_SYSTEM'], fill_type='solid')
    
    if l2:
        ws.cell(row=row_idx, column=7).fill = PatternFill(start_color=COLORS['L2_SUBSYSTEM'], end_color=COLORS['L2_SUBSYSTEM'], fill_type='solid')
    
    if l3:
        ws.cell(row=row_idx, column=8).fill = PatternFill(start_color=COLORS['L3_COMPONENT'], end_color=COLORS['L3_COMPONENT'], fill_type='solid')
    
    if l4:
        ws.cell(row=row_idx, column=9).fill = PatternFill(start_color=COLORS['L4_PART'], end_color=COLORS['L4_PART'], fill_type='solid')
    
    # Alignment
    for col in [6, 7, 8, 9]:
        ws.cell(row=row_idx, column=col).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

# Sütun genişlikleri ayarla
ws.column_dimensions['A'].width = 12
ws.column_dimensions['B'].width = 15
ws.column_dimensions['C'].width = 12
ws.column_dimensions['D'].width = 20
ws.column_dimensions['E'].width = 20
ws.column_dimensions['F'].width = 25
ws.column_dimensions['G'].width = 25
ws.column_dimensions['H'].width = 25
ws.column_dimensions['I'].width = 25

# Freeze Panes (İlk satırı sabitle)
ws.freeze_panes = 'A2'

# AutoFilter ekle
ws.auto_filter.ref = f'A1:I{ws.max_row}'

# Dosyayı kaydet
wb.save('data/belgrad/Veriler_KATEGORIZE.xlsx')

print("✅ BAŞARILI!")
print("\n📄 Yeni dosya oluşturuldu: Veriler_KATEGORIZE.xlsx")
print("\n📋 Yapı:")
print("   • Sütun F: ANA SİSTEM (Mavi)")
print("   • Sütun G: ALT SİSTEM (Yeşil)")
print("   • Sütun H: BİLEŞEN (Sarı)")
print("   • Sütun I: SPESIFIK PARÇA (Turuncu)")
print("\n💡 Özellikler:")
print("   ✓ Renk kodlama ile kategorisasyon")
print("   ✓ AutoFilter etkin")
print("   ✓ Satır 1 sabitlendi (Freeze)")
print("   ✓ Optimal sütun genişlikleri")
print("\n🎯 Kullanım:")
print("   • Her sütunun üstündeki filtreyi tıklayın")
print("   • İstediğiniz sistemi/parçayı seçin")
print("   • Veriler otomatik filtrelenir")
