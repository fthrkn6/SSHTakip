#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Excel dosyasının içeriğini kontrol et"""
import sys
from pathlib import Path
sys.path.insert(0, '.')

from openpyxl import load_workbook

# En yeni Excel dosyasını bul
rca_dir = Path('logs/root_cause_analysis')
excel_files = sorted(rca_dir.glob('*.xlsx'), key=lambda x: x.stat().st_mtime, reverse=True)

if not excel_files:
    print("Excel dosyasi bulunamadi!")
    sys.exit(1)

latest = excel_files[0]
print(f"Dosya: {latest}")
print(f"Boyut: {latest.stat().st_size} bytes")

# Excel'i aç
wb = load_workbook(latest)
print(f"\nSayfalar: {wb.sheetnames}")

# İlk sayfayı kontrol et
ws = wb.active
print(f"\nAktif Sayfa: {ws.title}")
print(f"Boyut: {ws.dimensions}")

# Verileri göster
print("\nIlk 40 satir (Tablolar):")
for i, row in enumerate(ws.iter_rows(min_row=1, max_row=40, values_only=False), 1):
    # Değer topla (encoding hatasını önle)
    values = []
    for cell in row[:8]:
        if cell.value is None:
            values.append('.')
        else:
            try:
                val = str(cell.value)[:20]
                values.append(val)
            except:
                values.append('[?]')
    
    if any(v != '.' for v in values):
        print(f"  R{i}: {' | '.join(values)}")

print("\nSayfanin Tum Satir Sayisi:", ws.max_row)
print("Tum Sutun Sayisi:", ws.max_column)

# Tablolar var mı?
print("\nTablelar:")
for table in ws.tables.values():
    print(f"  - {table.name}: {table.ref}")
