import pandas as pd
import openpyxl

excel_file = r'logs\ariza_listesi\Ariza_Listesi_BELGRAD.xlsx'

# Excel'de personel sayısı sütunundaki verileri kontrol et
wb = openpyxl.load_workbook(excel_file)
ws = wb['Ariza Listesi']

print('Personel Sayısı Sütunu (Column 30) Verileri:')
print('=' * 50)

for row_idx in range(5, 13):  # Row 5-12
    cell_value = ws.cell(row_idx, 30).value
    araç = ws.cell(row_idx, 2).value
    print(f'Row {row_idx}: Araç={araç}, Personel Sayısı={cell_value}')

wb.close()

# Pandas ile oku
print('\n' + '=' * 50)
print('Pandas ile Okunan Veriler:')
print('=' * 50)

df = pd.read_excel(excel_file, sheet_name='Ariza Listesi', header=3)
if len(df.columns) > 29:
    print(f'Column 29 (index 29) adı: {df.columns[29]}')
    print('İlk 5 satır:')
    for i, val in enumerate(df.iloc[:5, 29]):
        print(f'  Row {i}: {val} (type: {type(val).__name__})')
