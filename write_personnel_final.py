import openpyxl
import shutil
import os

excel_dir = r'logs\ariza_listesi'
original_file = os.path.join(excel_dir, 'Ariza_Listesi_BELGRAD.xlsx')
temp_file = os.path.join(excel_dir, 'temp_working.xlsx')

# Temp dosyayı aç
wb = openpyxl.load_workbook(temp_file)
ws = wb['Ariza Listesi']

# Personel sayılarını yaz (Row 5-12, Column 30)
personel_sayilari = [2, 1, 3, 2, 1, 2, 1, 3]

print('Personel Sayılarını Excel\'e Yazılıyor:')
print('=' * 50)

for idx, personel_sayı in enumerate(personel_sayilari):
    row_idx = 5 + idx
    araç = ws.cell(row_idx, 2).value
    ws.cell(row_idx, 30).value = personel_sayı
    print(f'Row {row_idx}: Araç={araç}, Personel Sayısı={personel_sayı}')

# Temp dosyayı kaydet
wb.save(temp_file)
wb.close()

print('\n✅ Temp dosya kaydedildi')

# Orijinal dosyayı sil ve temp'yi kopyala
try:
    os.remove(original_file)
    print(f'✅ Orijinal silinidi: {original_file}')
except Exception as e:
    print(f'❌ Silinemedi: {e}')
    exit(1)

shutil.copy(temp_file, original_file)
print(f'✅ Temp dosya orijinal konuma kopyalandı')

# Temp dosyayı sil
os.remove(temp_file)
print(f'✅ Temp dosya silindi')

print('\n✅ Tüm işlem başarıyla tamamlandı!')
