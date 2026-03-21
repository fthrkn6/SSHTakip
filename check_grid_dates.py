from openpyxl import load_workbook
import os
from datetime import datetime

projects = ['belgrad', 'istanbul', 'ankara', 'iasi', 'timisoara', 'kayseri', 'kocaeli', 'gebze', 'samsun']
base_path = r'c:\Users\fatiherkin\Desktop\bozankaya_ssh_takip\data'

print("\n" + "="*90)
print("SERVICE STATUS GRID EXCEL - DATE RANGES CHECK")
print("="*90 + "\n")

for project in projects:
    grid_path = os.path.join(base_path, project, 'service_status_grid.xlsx')
    
    if not os.path.exists(grid_path):
        print(f"❌ {project:12} | File not found")
        continue
    
    try:
        wb = load_workbook(grid_path, data_only=True)
        ws = wb.active
        
        # Get first and last dates
        first_date = ws['A2'].value
        last_date = None
        
        # Find actual last row with data
        for row in range(2, ws.max_row + 1):
            cell_val = ws[f'A{row}'].value
            if cell_val:
                last_date = cell_val
        
        # Format dates if they are datetime objects
        if isinstance(first_date, datetime):
            first_str = first_date.strftime('%Y-%m-%d')
        else:
            first_str = str(first_date)
            
        if isinstance(last_date, datetime):
            last_str = last_date.strftime('%Y-%m-%d')
        else:
            last_str = str(last_date)
        
        row_count = ws.max_row - 1
        col_count = ws.max_column - 1  # Excluding date column
        
        print(f"✓ {project:12} | START: {first_str:12} → END: {last_str:12} | Rows: {row_count:3} | Cols: {col_count}")
        
        wb.close()
    except Exception as e:
        print(f"⚠️ {project:12} | Error: {str(e)[:60]}")

print("\n" + "="*90)
print(f"Today's date: {datetime.now().strftime('%Y-%m-%d')}")
print("="*90 + "\n")
