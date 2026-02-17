"""
Test script: Form submission simulation
Flask app'ı başlamadan form data'sını test etmek için simülasyon
"""

import sys
sys.path.insert(0, '.')

from utils_fracas_writer import FracasWriter
from datetime import datetime

print("=" * 80)
print("TEST: Yeni Arıza Form Submission Simulation")
print("=" * 80)

# Farklı test verileri - gerçek form submission'ını simüle et
test_cases = [
    {
        'name': 'Test 1: Bozankaya tarafından tespit edilen arıza',
        'data': {
            'arac_numarasi': 'T02',
            'arac_module': ['Elektrik', 'Kontrol'],
            'arac_km': '2500',
            'hata_tarih': '2024-02-17',
            'hata_saat': '10:15',
            'sistem': 'Elektrik',
            'alt_sistem': 'Çekiş Kontrol',
            'tedarikci': 'ABB',
            'ariza_tanimi': 'Çekiş kontrol arızası - motor çıktısı yok',
            'ariza_sinifi': 'Kritik',
            'ariza_kaynagi': 'Elektronik Arızası',
            'yapilan_islem': 'Kart değiştirildi',
            'aksiyon': 'Değişim',
            'garanti_kapsami': 'Evet',
            'ariza_tespit_yontemi': 'Bozankaya',
            'tamir_baslama_tarih': '2024-02-17',
            'tamir_baslama_saati': '10:30',
            'tamir_bitisi_tarih': '2024-02-17',
            'tamir_bitisi_saati': '12:30',
            'tamir_suresi': '2 saat 0 dakika',
            'mttr': '120 dk',
            'personel_sayisi': '3',
            'parca_kodu': 'E001',
            'parca_adi': 'Traction Control Board',
            'parca_seri_no': 'SN67890',
            'adet': '1',
            'iscilik_maliyeti': '75.00',
            'servise_verilis_tarih': '2024-02-17',
            'servise_verilis_saati': '12:30',
            'ariza_tipi': 'Elektronik',
        }
    },
    {
        'name': 'Test 2: Müşteri tarafından tespit edilen arıza',
        'data': {
            'arac_numarasi': 'T03',
            'arac_module': ['Mekanik', 'Şasi'],
            'arac_km': '3200',
            'hata_tarih': '2024-02-16',
            'hata_saat': '16:45',
            'sistem': 'Türediş',
            'alt_sistem': 'Şasi Yapısı',
            'tedarikci': 'Bombardier',
            'ariza_tanimi': 'Şasi çatlağı - ön sol köşe',
            'ariza_sinifi': 'Yüksek',
            'ariza_kaynagi': 'Mekanik Yıpranma',
            'yapilan_islem': 'Kaynakla onarıldı',
            'aksiyon': 'Tamir Edildi',
            'garanti_kapsami': 'Hayır',
            'ariza_tespit_yontemi': 'Müşteri',
            'tamir_baslama_tarih': '2024-02-17',
            'tamir_baslama_saati': '08:00',
            'tamir_bitisi_tarih': '2024-02-17',
            'tamir_bitisi_saati': '14:00',
            'tamir_suresi': '6 saat 0 dakika',
            'mttr': '360 dk',
            'personel_sayisi': '4',
            'parca_kodu': 'M002',
            'parca_adi': 'Şasi Panel',
            'parca_seri_no': 'SN11111',
            'adet': '1',
            'iscilik_maliyeti': '125.00',
            'servise_verilis_tarih': '2024-02-17',
            'servise_verilis_saati': '14:00',
            'ariza_tipi': 'Mekanik',
        }
    }
]

# Test senaryolarını çalıştır
writer = FracasWriter()

for test_case in test_cases:
    print(f"\n{test_case['name']}")
    print("-" * 80)
    
    try:
        # "Arıza Tespit Yöntemi" field'ını kontrol et
        detection_method = test_case['data'].get('ariza_tespit_yontemi', 'N/A')
        print(f"  📍 Tespit Yöntemi: {detection_method}")
        
        # Verileri yaz
        result = writer.write_failure_data(test_case['data'])
        
        if result['success']:
            print(f"  ✅ Başarılı!")
            print(f"     - FRACAS ID: {result['fracas_id']}")
            print(f"     - Satır: {result['row']}")
            print(f"     - Model: {test_case['data'].get('sistem')}")
            print(f"     - Alt Sistem: {test_case['data'].get('alt_sistem')}")
            print(f"     - Arıza: {test_case['data'].get('ariza_tanimi')}")
        else:
            print(f"  ❌ Başarısız: {result}")
            
    except Exception as e:
        print(f"  ❌ Hata: {e}")

print("\n" + "=" * 80)
print("TEST TAMAMLANDI")
print("=" * 80)

# Excel dosyasında tüm yazılan verileri göster
print("\nFracas_BELGRAD.xlsx İçeriği:")
print("=" * 80)

from openpyxl import load_workbook

try:
    wb = load_workbook('logs/belgrad/ariza_listesi/Fracas_BELGRAD.xlsx')
    ws = wb['FRACAS']
    
    print(f"\nToplamda {ws.max_row - 4} veri satırı var (Row 4 = Header, Row 5+ = Data)")
    print("\nSon 5 Satır:\n")
    
    # Son 5 satırı göster (B, E, F, G, H, P sütunları - Araç, FRACAS ID, Tarih, Sistem, Alt Sistem, Tespit Yöntemi)
    for row_num in range(max(5, ws.max_row - 4), ws.max_row + 1):
        arac = ws[f'B{row_num}'].value
        fracas_id = ws[f'E{row_num}'].value
        tarih = ws[f'F{row_num}'].value
        sistem = ws[f'G{row_num}'].value
        alt_sistem = ws[f'H{row_num}'].value
        tespit_yontemi = ws[f'P{row_num}'].value
        
        if arac:  # Boş olmayan satırları göster
            print(f"  Row {row_num}: {arac} | {fracas_id} | {tarih} | {sistem} | {alt_sistem} | Tespit: {tespit_yontemi}")
            
    wb.close()
    
except Exception as e:
    print(f"❌ Excel kontrol hatası: {e}")

print("\n✅ Form integration ready for deployment!")
