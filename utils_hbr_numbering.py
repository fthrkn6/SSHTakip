#!/usr/bin/env python
# coding: utf-8
"""
HBR Belge Adlandırma Sistemi
Projekt koduna göre HBR belge numaralandırmasını otomatik yönetir
Format: {PROJECT_CODE}-NCR-{SEQUENCE_NUMBER}
Örnek: BEL25-NCR-001, BEL25-NCR-002, vb.
"""

import os
import re
import openpyxl
from pathlib import Path
from flask import current_app, session
import logging

logger = logging.getLogger(__name__)


def get_project_code_from_veriler(project_name='belgrad'):
    """
    Veriler.xlsx'ten proje kodunu oku
    B2 hücresindeki değeri döndür
    """
    try:
        # Verilen proje klasöründen Veriler.xlsx'i aç
        veriler_path = os.path.join(
            current_app.root_path, 
            'data', 
            project_name, 
            'Veriler.xlsx'
        )
        
        if not os.path.exists(veriler_path):
            logger.warning(f"[HBR] Veriler.xlsx bulunamadı: {veriler_path}")
            return project_name.upper()[:3] + "25"  # Fallback: BEL25, etc.
        
        wb = openpyxl.load_workbook(veriler_path, data_only=True)
        # Sayfa2 veya Sheet1'den oku
        ws = wb.active if 'Sayfa2' not in wb.sheetnames else wb['Sayfa2']
        
        # B2 hücresinden proje kodunu oku
        project_code = ws['B2'].value
        
        if not project_code:
            logger.warning(f"[HBR] B2 hücresi boş, Sayfa1'den oku")
            ws = wb['Sayfa1'] if 'Sayfa1' in wb.sheetnames else wb.active
            project_code = ws['B2'].value
        
        project_code = str(project_code).strip().upper()
        logger.info(f"[HBR] Proje kodu: {project_code}")
        
        return project_code
        
    except Exception as e:
        logger.error(f"[HBR] Proje kodu okunamadı: {e}")
        return project_name.upper()[:3] + "25"  # Fallback


def get_next_ncr_number(project_code, hbr_folder='HBR'):
    """
    HBR klasöründe {PROJECT_CODE}-NCR-XXX formatındaki dosyaları tara
    En yüksek numarayı bul, +1 döndür
    """
    try:
        hbr_path = Path(hbr_folder)
        
        if not hbr_path.exists():
            logger.info(f"[HBR] {hbr_folder} klasörü oluşturuluyor")
            hbr_path.mkdir(parents=True, exist_ok=True)
            return 1  # İlk belge
        
        # Klasördeki tüm dosyaları tara ve pattern eşleştir
        pattern = re.compile(rf'^{re.escape(project_code)}-NCR-(\d+)', re.IGNORECASE)
        
        max_number = 0
        for item in hbr_path.rglob('*'):
            match = pattern.match(item.name)
            if match:
                number = int(match.group(1))
                if number > max_number:
                    max_number = number
                    logger.debug(f"[HBR] Bulunan: {item.name} (No: {number})")
        
        next_number = max_number + 1
        logger.info(f"[HBR] Son kaydı numarası: {max_number}, Sonraki: {next_number}")
        
        return next_number
        
    except Exception as e:
        logger.error(f"[HBR] Numuara araması hatası: {e}")
        return 1


def generate_ncr_filename(project_code, ncr_number, extension='.docx'):
    """
    HBR belge adını oluştur
    Format: {PROJECT_CODE}-NCR-{SEQUENCE_NUMBER:03d}{extension}
    Örnek: BEL25-NCR-001.docx
    """
    filename = f"{project_code}-NCR-{ncr_number:03d}{extension}"
    return filename


def get_next_hbr_filename(project_name='belgrad', hbr_folder='HBR', extension='.docx'):
    """
    Sonraki HBR belge adını oluştur
    İşlem sırası:
    1. Proje kodunu Veriler.xlsx'ten oku
    2. HBR klasöründe en yüksek numarayı bul
    3. Bir sonraki numarayı döndür
    """
    try:
        # 1. Proje kodunu oku
        project_code = get_project_code_from_veriler(project_name)
        
        # 2. Sonraki numarayı bul
        next_number = get_next_ncr_number(project_code, hbr_folder)
        
        # 3. Dosya adını oluştur
        filename = generate_ncr_filename(project_code, next_number, extension)
        
        logger.info(f"[HBR] Yeni belge adı: {filename}")
        return filename
        
    except Exception as e:
        logger.error(f"[HBR] Belge adı oluşturulamadı: {e}")
        return None


def get_hbr_folder_path(hbr_folder='HBR'):
    """HBR klasöründün tam yolunu döndür"""
    return Path(hbr_folder).resolve()


def create_hbr_document_placeholder(filename, hbr_folder='HBR'):
    """
    HBR klasöründe placeholder belge oluştur (test için)
    """
    try:
        hbr_path = Path(hbr_folder)
        hbr_path.mkdir(parents=True, exist_ok=True)
        
        file_path = hbr_path / filename
        file_path.touch()
        
        logger.info(f"[HBR] Placeholder oluşturuldu: {file_path}")
        return file_path
        
    except Exception as e:
        logger.error(f"[HBR] Placeholder oluşturulamadı: {e}")
        return None


if __name__ == '__main__':
    # Test
    import sys
    from app import create_app
    
    app = create_app()
    
    with app.app_context():
        print("[TEST] HBR Numaralandırma Sistemi")
        print("=" * 50)
        
        # Proje kodunu oku
        project_code = get_project_code_from_veriler('belgrad')
        print(f"Proje Kodu: {project_code}")
        
        # Sonraki dosya adını al
        next_filename = get_next_hbr_filename('belgrad')
        print(f"Sonraki Dosya Adı: {next_filename}")
        
        # Örnek: 3 dosya oluştur
        print("\nTest dosyaları oluşturuluyor...")
        for i in range(3):
            filename = get_next_hbr_filename('belgrad')
            if filename:
                create_hbr_document_placeholder(filename)
                print(f"  ✓ {filename}")
        
        # HBR klasöründeki dosyaları listele
        print(f"\nHBR Klasöründeki dosyalar:")
        hbr_path = get_hbr_folder_path()
        if hbr_path.exists():
            for file in sorted(hbr_path.glob('*')):
                print(f"  - {file.name}")
