"""
Servis Durumu Excel Logger
logs/{project}/service_status/ klasörüne tüm servis durum değişikliklerini kaydet
"""
import os
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
import logging

logger = logging.getLogger(__name__)


class ServiceStatusExcelLogger:
    """Servis durumu değişikliklerini Excel'e logla - Sadece ekleme (append)"""
    
    def __init__(self, project_code='belgrad'):
        self.project_code = project_code
        self.log_dir = Path(__file__).parent / 'logs' / project_code / 'service_status'
        self.log_dir.mkdir(parents=True, exist_ok=True)
        self.log_file = self.log_dir / 'service_status_log.xlsx'
    
    def _ensure_workbook(self):
        """Excel dosyasını oluştur veya yükle"""
        headers = [
            'Saat',
            'Tramvay ID',
            'Tramvay Adı',
            'Durum',
            'Sistem',
            'Alt Sistem',
            'Açıklama',
            'Kullanıcı',
            'İşlem'
        ]
        
        if not self.log_file.exists():
            wb = Workbook()
            ws = wb.active
            ws.title = 'Günlük'
            
            # Başlık satırını ekle
            ws.append(headers)
            
            # Başlık stillemesi
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF', size=11)
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # Sütun genişlikleri
            ws.column_dimensions['A'].width = 20  # Saat
            ws.column_dimensions['B'].width = 15  # Tramvay ID
            ws.column_dimensions['C'].width = 20  # Tramvay Adı
            ws.column_dimensions['D'].width = 20  # Durum
            ws.column_dimensions['E'].width = 18  # Sistem
            ws.column_dimensions['F'].width = 18  # Alt Sistem
            ws.column_dimensions['G'].width = 25  # Açıklama
            ws.column_dimensions['H'].width = 15  # Kullanıcı
            ws.column_dimensions['I'].width = 12  # İşlem
            
            wb.save(self.log_file)
            logger.info(f"Service status log dosyası oluşturuldu: {self.log_file}")
            return wb
        
        return load_workbook(self.log_file)
    
    def log_status_change(self, tram_id, tram_name, status, sistem='', alt_sistem='', 
                         aciklama='', user='system', action='OLUŞTUR'):
        """Servis durumu değişikliğini Excel'e ekle"""
        try:
            wb = self._ensure_workbook()
            ws = wb.active
            
            timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            
            # Yeni satır ekle
            new_row = [
                timestamp,
                str(tram_id),
                str(tram_name),
                str(status),
                str(sistem) if sistem else '-',
                str(alt_sistem) if alt_sistem else '-',
                str(aciklama) if aciklama else '-',
                str(user),
                action
            ]
            
            ws.append(new_row)
            
            # Son satırın stillemesi (durum rengine göre)
            last_row = ws.max_row
            status_lower = status.lower() if status else ''
            
            # Durum rengine göre arka plan rengi
            if 'aktif' in status_lower or 'servis' in status_lower:
                fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')  # Yeşil
                font = Font(color='006100')
            elif 'dışı' in status_lower or 'arıza' in status_lower:
                fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')  # Kırmızı
                font = Font(color='B10000')
            elif 'işletme' in status_lower:
                fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')  # Sarı
                font = Font(color='856404')
            else:
                fill = None
                font = Font()
            
            # Satır stillemesi
            for col in range(1, len(new_row) + 1):
                cell = ws.cell(row=last_row, column=col)
                if fill:
                    cell.fill = fill
                cell.font = font
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            
            # Excel dosyasını kaydet
            wb.save(self.log_file)
            logger.info(f"Service status logged: {tram_id} - {status} - {user}")
            return True
            
        except Exception as e:
            logger.error(f"Service status Excel logging error: {str(e)}")
            return False
    
    def get_monthly_analysis(self, year: int, month: int):
        """Aylık alt sistem analizini getir"""
        try:
            if not self.log_file.exists():
                return {}
            
            wb = load_workbook(self.log_file)
            ws = wb.active
            
            analysis = {}
            
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not row[0]:  # Saat boşsa atla
                    continue
                
                try:
                    date_str = str(row[0]).split()[0]  # YYYY-MM-DD
                    date_parts = date_str.split('-')
                    row_year, row_month = int(date_parts[0]), int(date_parts[1])
                    
                    if row_year != year or row_month != month:
                        continue
                    
                    tram_id = str(row[1]).strip()
                    tram_name = str(row[2]).strip() if row[2] else f'Tramvay {tram_id}'
                    status = str(row[3]).strip() if row[3] else ''
                    sistem = str(row[4]).strip() if row[4] else '-'
                    alt_sistem = str(row[5]).strip() if row[5] else '-'
                    
                    # Sadece "Servis Dışı" olanları say
                    if 'dışı' not in status.lower() and 'arıza' not in status.lower():
                        continue
                    
                    # Key: tram_id + alt_sistem
                    key = f"{tram_id}|{tram_name}|{alt_sistem}"
                    
                    if key not in analysis:
                        analysis[key] = {
                            'tram_id': tram_id,
                            'tram_name': tram_name,
                            'alt_sistem': alt_sistem,
                            'days_count': 0,
                            'dates': []
                        }
                    
                    # Tarihi ekle (eğer yoksa)
                    if date_str not in analysis[key]['dates']:
                        analysis[key]['dates'].append(date_str)
                        analysis[key]['days_count'] += 1
                
                except Exception as e:
                    logger.warning(f"Row parsing error: {e}")
                    continue
            
            return analysis
        
        except Exception as e:
            logger.error(f"Monthly analysis error: {str(e)}")
            return {}


def log_service_status_to_excel(tram_id, tram_name, status, sistem='', alt_sistem='', 
                               aciklama='', user='system', action='OLUŞTUR', project_code='belgrad'):
    """Servis durumunu Excel'e logla - Wrapper fonksiyonu"""
    logger_obj = ServiceStatusExcelLogger(project_code)
    return logger_obj.log_status_change(tram_id, tram_name, status, sistem, alt_sistem, aciklama, user, action)
