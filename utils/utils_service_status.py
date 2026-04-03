"""
Servis Durumu Utility'leri
Availability analizi, Excel raporlama ve log kaydı
"""

import os
import json
import logging
from datetime import datetime, timedelta, date
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from models import db, ServiceLog, AvailabilityMetrics, RootCauseAnalysis, Equipment, Failure

logger = logging.getLogger(__name__)

# Availability log klasörü (proje-agnostic)
AVAILABILITY_LOG_DIR = os.path.join(os.path.dirname(__file__), 'logs', 'belgrad', 'availability')
os.makedirs(AVAILABILITY_LOG_DIR, exist_ok=True)


class AvailabilityAnalyzer:
    """Servis durumu ve availability analizi"""
    
    @staticmethod
    def calculate_daily_availability(tram_id, target_date=None):
        """Günlük availability hesapla"""
        if target_date is None:
            target_date = date.today()
        
        # O günün tüm log'larını getir
        logs = ServiceLog.query.filter(
            ServiceLog.tram_id == tram_id,
            db.func.date(ServiceLog.log_date) == target_date
        ).order_by(ServiceLog.log_date.asc()).all()
        
        total_hours = 24
        downtime = 0
        operational = 24
        
        for log in logs:
            if log.new_status != 'operasyonel':
                downtime += log.duration_hours or 0
        
        operational = max(0, total_hours - downtime)
        availability = (operational / total_hours * 100) if total_hours > 0 else 0
        
        # Veritabanına kaydet veya güncelle
        metric = AvailabilityMetrics.query.filter_by(
            tram_id=tram_id,
            metric_date=target_date,
            report_period='daily'
        ).first()
        
        if not metric:
            metric = AvailabilityMetrics(
                tram_id=tram_id,
                metric_date=target_date,
                report_period='daily'
            )
            db.session.add(metric)
        
        metric.total_hours = total_hours
        metric.operational_hours = operational
        metric.downtime_hours = downtime
        metric.availability_percentage = round(availability, 2)
        metric.failure_count = len([l for l in logs if l.new_status != 'operasyonel'])
        
        db.session.commit()
        return metric
    
    @staticmethod
    def calculate_period_availability(tram_id, start_date, end_date, period_type):
        """Dönem bazlı availability hesapla"""
        
        # Tüm log'ları getir
        logs = ServiceLog.query.filter(
            ServiceLog.tram_id == tram_id,
            ServiceLog.log_date >= datetime.combine(start_date, datetime.min.time()),
            ServiceLog.log_date <= datetime.combine(end_date, datetime.max.time())
        ).all()
        
        # Gün sayısını hesapla
        days = (end_date - start_date).days + 1
        total_hours = days * 24
        
        downtime = sum(log.duration_hours or 0 for log in logs)
        operational = max(0, total_hours - downtime)
        availability = (operational / total_hours * 100) if total_hours > 0 else 0
        
        # Sistem bazında downtime analizi
        system_downtime = {}
        for log in logs:
            if log.sistem:
                if log.sistem not in system_downtime:
                    system_downtime[log.sistem] = {'downtime': 0, 'count': 0, 'subsystems': {}}
                system_downtime[log.sistem]['downtime'] += log.duration_hours or 0
                system_downtime[log.sistem]['count'] += 1
                
                if log.alt_sistem:
                    if log.alt_sistem not in system_downtime[log.sistem]['subsystems']:
                        system_downtime[log.sistem]['subsystems'][log.alt_sistem] = {'downtime': 0, 'count': 0}
                    system_downtime[log.sistem]['subsystems'][log.alt_sistem]['downtime'] += log.duration_hours or 0
                    system_downtime[log.sistem]['subsystems'][log.alt_sistem]['count'] += 1
        
        # Metrik kaydet
        metric = AvailabilityMetrics(
            tram_id=tram_id,
            metric_date=start_date,
            report_period=period_type,
            total_hours=total_hours,
            operational_hours=operational,
            downtime_hours=downtime,
            availability_percentage=round(availability, 2),
            failure_count=len([l for l in logs if l.new_status != 'operasyonel']),
            sistem=json.dumps(system_downtime) if system_downtime else None
        )
        
        db.session.add(metric)
        db.session.commit()
        
        metric.system_breakdown = system_downtime
        return metric
    
    @staticmethod
    def get_root_cause_summary(tram_id, start_date=None, end_date=None):
        """Root cause analizi özeti - sistem ve alt sistem bazında"""
        
        if start_date is None:
            start_date = date.today() - timedelta(days=30)
        if end_date is None:
            end_date = date.today()
        
        analyses = RootCauseAnalysis.query.filter(
            RootCauseAnalysis.tram_id == tram_id,
            RootCauseAnalysis.analysis_date >= datetime.combine(start_date, datetime.min.time()),
            RootCauseAnalysis.analysis_date <= datetime.combine(end_date, datetime.max.time())
        ).all()
        
        summary = {
            'total_analyses': len(analyses),
            'by_system': {},
            'by_severity': {'kritik': 0, 'yüksek': 0, 'orta': 0, 'düşük': 0},
            'by_status': {'open': 0, 'closed': 0, 'pending': 0},
            'most_frequent_causes': {},
            'details': []
        }
        
        for analysis in analyses:
            # Sistem bazında
            if analysis.sistem not in summary['by_system']:
                summary['by_system'][analysis.sistem] = {
                    'count': 0,
                    'subsystems': {},
                    'causes': []
                }
            summary['by_system'][analysis.sistem]['count'] += 1
            
            if analysis.alt_sistem:
                if analysis.alt_sistem not in summary['by_system'][analysis.sistem]['subsystems']:
                    summary['by_system'][analysis.sistem]['subsystems'][analysis.alt_sistem] = 0
                summary['by_system'][analysis.sistem]['subsystems'][analysis.alt_sistem] += 1
            
            summary['by_system'][analysis.sistem]['causes'].append(analysis.root_cause)
            
            # Severity bazında
            if analysis.severity_level:
                summary['by_severity'][analysis.severity_level] += 1
            
            # Status bazında
            summary['by_status'][analysis.status] += 1
            
            # En sık neden
            if analysis.root_cause:
                summary['most_frequent_causes'][analysis.root_cause] = \
                    summary['most_frequent_causes'].get(analysis.root_cause, 0) + 1
            
            summary['details'].append({
                'sistem': analysis.sistem,
                'alt_sistem': analysis.alt_sistem,
                'root_cause': analysis.root_cause,
                'severity': analysis.severity_level,
                'frequency': analysis.frequency,
                'status': analysis.status,
                'date': analysis.analysis_date.strftime('%d.%m.%Y %H:%M')
            })
        
        return summary


class ExcelReportGenerator:
    """Excel rapor üreteci"""
    
    @staticmethod
    def create_comprehensive_availability_report(tram_ids, output_path):
        """Kapsamlı availability raporu oluştur - Tüm dönemler"""
        
        try:
            wb = Workbook()
            ws_summary = wb.active
            ws_summary.title = "Özet"
            
            # Özet sayfası başlığı
            ws_summary['A1'] = "SERVİS DURUMU VE AVAILABILITY ANALİZİ"
            ws_summary['A1'].font = Font(bold=True, size=14, color="FFFFFF")
            ws_summary['A1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            ws_summary.merge_cells('A1:H1')
            ws_summary['A2'] = f"Rapor Tarihi: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
            
            # Özet tablosu
            row = 4
            headers = ['Araç', 'Günlük %', 'Haftalık %', 'Aylık %', '3 Aylık %', '6 Aylık %', 'Yıllık %', 'Total %']
            
            for col, header in enumerate(headers, 1):
                cell = ws_summary.cell(row=row, column=col)
                cell.value = header
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            row += 1
            
            for tram_id in tram_ids:
                equipment = Equipment.query.filter_by(equipment_code=tram_id).first()
                equipment_name = equipment.name if equipment else tram_id
                
                ws_summary.cell(row=row, column=1).value = equipment_name
                
                # Dönemler
                periods = [
                    ('daily', date.today()),
                    ('weekly', date.today() - timedelta(days=7)),
                    ('monthly', date.today() - timedelta(days=30)),
                    ('quarterly', date.today() - timedelta(days=90)),
                    ('biannual', date.today() - timedelta(days=180)),
                    ('yearly', date.today() - timedelta(days=365)),
                ]
                
                col = 2
                for period, period_date in periods:
                    metric = AvailabilityMetrics.query.filter_by(
                        tram_id=tram_id,
                        report_period=period
                    ).order_by(AvailabilityMetrics.created_at.desc()).first()
                    
                    if metric:
                        cell = ws_summary.cell(row=row, column=col)
                        cell.value = metric.availability_percentage
                        cell.number_format = '0.00"%"'
                        # Renk kodlama
                        if metric.availability_percentage >= 95:
                            cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                        elif metric.availability_percentage >= 80:
                            cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                        else:
                            cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    
                    col += 1
                
                # Total
                total_metric = AvailabilityMetrics.query.filter_by(
                    tram_id=tram_id,
                    report_period='total'
                ).order_by(AvailabilityMetrics.created_at.desc()).first()
                
                if total_metric:
                    cell = ws_summary.cell(row=row, column=8)
                    cell.value = total_metric.availability_percentage
                    cell.number_format = '0.00"%"'
                
                row += 1
            
            # Detaylı sheet'ler
            for tram_id in tram_ids:
                equipment = Equipment.query.filter_by(equipment_code=tram_id).first()
                equipment_name = (equipment.name if equipment else tram_id)[:30]
                
                ws = wb.create_sheet(title=equipment_name)
                
                ws['A1'] = f"Araç: {equipment_name} ({tram_id})"
                ws['A1'].font = Font(bold=True, size=12)
                ws.merge_cells('A1:F1')
                
                # Detaylı veriler
                row = 3
                headers = ['Dönem', 'Başlang. Tarihi', 'Bitiş Tarihi', 'Availability %', 'Operational Saatleri', 'Downtime Saatleri', 'Arıza Sayısı']
                
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=row, column=col)
                    cell.value = header
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                
                row += 1
                
                # Tüm metrikler
                metrics = AvailabilityMetrics.query.filter_by(tram_id=tram_id).order_by(
                    AvailabilityMetrics.created_at.desc()
                ).limit(100).all()
                
                for metric in metrics:
                    ws.cell(row=row, column=1).value = metric.report_period or '-'
                    ws.cell(row=row, column=2).value = metric.metric_date.strftime('%d.%m.%Y') if metric.metric_date else '-'
                    ws.cell(row=row, column=3).value = (metric.metric_date + timedelta(days=1)).strftime('%d.%m.%Y') if metric.metric_date else '-'
                    ws.cell(row=row, column=4).value = metric.availability_percentage
                    ws.cell(row=row, column=4).number_format = '0.00"%"'
                    ws.cell(row=row, column=5).value = round(metric.operational_hours, 2)
                    ws.cell(row=row, column=6).value = round(metric.downtime_hours, 2)
                    ws.cell(row=row, column=7).value = metric.failure_count
                    
                    row += 1
            
            wb.save(output_path)
            logger.info(f"Comprehensive availability report created: {output_path}")
            
        except Exception as e:
            logger.error(f"Error creating comprehensive report: {str(e)}")
            raise

    
    
    @staticmethod
    def create_root_cause_analysis_report(tram_ids, output_path):
        """Root cause analysis raporu"""
        
        try:
            wb = Workbook()
            ws_summary = wb.active
            ws_summary.title = "Root Cause Özet"
            
            ws_summary['A1'] = "ROOT CAUSE ANALİZİ"
            ws_summary['A1'].font = Font(bold=True, size=14, color="FFFFFF")
            ws_summary['A1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            ws_summary.merge_cells('A1:H1')
            
            # Sistem bazında özet
            ws_sys = wb.create_sheet(title="Sistem Bazında")
            ws_sys['A1'] = "SİSTEM BAZINDA ROOT CAUSE ANALİZİ"
            ws_sys['A1'].font = Font(bold=True, size=12)
            ws_sys.merge_cells('A1:G1')
            
            row = 3
            headers = ['Araç', 'Sistem', 'Alt Sistem', 'Toplam Analiz', 'Açık', 'Kapalı', 'Beklemede']
            
            for col, header in enumerate(headers, 1):
                cell = ws_sys.cell(row=row, column=col)
                cell.value = header
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            
            row += 1
            
            for tram_id in tram_ids:
                equipment = Equipment.query.filter_by(equipment_code=tram_id).first()
                equipment_name = equipment.name if equipment else tram_id
                
                analyses = RootCauseAnalysis.query.filter_by(tram_id=tram_id).all()
                
                for analysis in analyses:
                    ws_sys.cell(row=row, column=1).value = equipment_name
                    ws_sys.cell(row=row, column=2).value = analysis.sistem
                    ws_sys.cell(row=row, column=3).value = analysis.alt_sistem or '-'
                    ws_sys.cell(row=row, column=4).value = 1
                    ws_sys.cell(row=row, column=5).value = 1 if analysis.status == 'submitted' else 0
                    ws_sys.cell(row=row, column=6).value = 1 if analysis.status == 'closed' else 0
                    ws_sys.cell(row=row, column=7).value = 1 if analysis.status == 'draft' else 0
                    
                    row += 1
            
            # Detaylı analiz
            ws_detail = wb.create_sheet(title="Detaylı Analiz")
            ws_detail['A1'] = "DETAYLI ROOT CAUSE ANALİZİ"
            ws_detail['A1'].font = Font(bold=True, size=12)
            ws_detail.merge_cells('A1:H1')
            
            row = 3
            headers = ['Araç', 'Sistem', 'Alt Sistem', 'Arıza Modu', 'Kök Neden', 'Severity', 'Status']
            
            for col, header in enumerate(headers, 1):
                cell = ws_detail.cell(row=row, column=col)
                cell.value = header
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            
            row += 1
            
            for tram_id in tram_ids:
                equipment = Equipment.query.filter_by(equipment_code=tram_id).first()
                equipment_name = equipment.name if equipment else tram_id
                
                analyses = RootCauseAnalysis.query.filter_by(tram_id=tram_id).all()
                
                for analysis in analyses:
                    ws_detail.cell(row=row, column=1).value = equipment_name
                    ws_detail.cell(row=row, column=2).value = analysis.sistem
                    ws_detail.cell(row=row, column=3).value = analysis.alt_sistem or '-'
                    ws_detail.cell(row=row, column=4).value = analysis.failure_mode or '-'
                    ws_detail.cell(row=row, column=5).value = analysis.root_cause
                    ws_detail.cell(row=row, column=6).value = analysis.severity
                    ws_detail.cell(row=row, column=7).value = analysis.status
                    
                    # Renk kodlama
                    severity_colors = {
                        'critical': 'FFC7CE',
                        'high': 'FFEB9C',
                        'medium': 'BDD7EE',
                        'low': 'C6EFCE'
                    }
                    color = severity_colors.get(analysis.severity, 'FFFFFF')
                    for col_idx in range(1, 8):
                        ws_detail.cell(row=row, column=col_idx).fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                    
                    row += 1
            
            wb.save(output_path)
            logger.info(f"Root cause analysis report created: {output_path}")
            
        except Exception as e:
            logger.error(f"Error creating root cause report: {str(e)}")
            raise

    
    @staticmethod
    def create_detailed_daily_report(tram_id, output_path):
        """Detaylı günlük rapor"""
        
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Günlük Durum"
            
            ws['A1'] = f"GÜNLÜK SERVİS DURUMU RAPORU - {tram_id}"
            ws['A1'].font = Font(bold=True, size=14, color="FFFFFF")
            ws['A1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            ws.merge_cells('A1:G1')
            
            # Bugünün logları
            today = date.today()
            logs = ServiceLog.query.filter(
                ServiceLog.tram_id == tram_id,
                db.func.date(ServiceLog.log_date) == today
            ).order_by(ServiceLog.log_date.asc()).all()
            
            row = 3
            headers = ['Saat', 'Status', 'Sistem', 'Alt Sistem', 'Süre (Saat)', 'Neden', 'Kaydeden']
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col)
                cell.value = header
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            
            row += 1
            
            for log in logs:
                ws.cell(row=row, column=1).value = log.log_date.strftime('%H:%M')
                ws.cell(row=row, column=2).value = log.new_status
                ws.cell(row=row, column=3).value = log.sistem or '-'
                ws.cell(row=row, column=4).value = log.alt_sistem or '-'
                ws.cell(row=row, column=5).value = log.duration_hours or 0
                ws.cell(row=row, column=6).value = log.reason or '-'
                ws.cell(row=row, column=7).value = log.user.full_name if log.user else '-'
                
                row += 1
            
            wb.save(output_path)
            logger.info(f"Daily report created: {output_path}")
            
        except Exception as e:
            logger.error(f"Error creating daily report: {str(e)}")
            raise



def log_service_status_change(tram_id, new_status, sistem=None, alt_sistem=None, 
                              reason=None, duration_hours=0, user_id=None):
    """Servis durumu değişikliğini kaydet"""
    
    try:
        # Son status'u getir
        last_log = ServiceLog.query.filter_by(tram_id=tram_id).order_by(
            ServiceLog.log_date.desc()
        ).first()
        
        previous_status = last_log.new_status if last_log else 'bilinmiyor'
        
        # Yeni log oluştur
        log_entry = ServiceLog(
            tram_id=tram_id,
            previous_status=previous_status,
            new_status=new_status,
            sistem=sistem,
            alt_sistem=alt_sistem,
            reason=reason,
            duration_hours=duration_hours,
            created_by=user_id
        )
        
        db.session.add(log_entry)
        db.session.commit()
        
        # Günlük availability'i güncelle
        AvailabilityAnalyzer.calculate_daily_availability(tram_id)
        
        # Log dosyasına kaydet
        log_file = os.path.join(AVAILABILITY_LOG_DIR, f"{tram_id}.log")
        with open(log_file, 'a', encoding='utf-8') as f:
            f.write(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] "
                   f"Status: {previous_status} -> {new_status} | "
                   f"System: {sistem} | SubSystem: {alt_sistem} | "
                   f"Reason: {reason} | Duration: {duration_hours}h\n")
        
        logger.info(f"Service status logged for {tram_id}: {new_status}")
        return log_entry
    
    except Exception as e:
        logger.error(f"Error logging service status: {str(e)}")
        raise
