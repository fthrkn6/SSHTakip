import os
from datetime import datetime, date, timedelta
from openpyxl import Workbook, load_workbook


def normalize_status(status: str) -> str:
    """Durum string'ini standart forma normalize et"""
    if not status:
        return ''
    s = status.strip()
    s_upper = s.upper()
    # İşletme Kaynaklı variants
    if 'İŞLETME' in s_upper or 'ISLETME' in s_upper or 'IŞLETME' in s_upper:
        return 'İşletme Kaynaklı Servis Dışı'
    # Servis Dışı variants
    if 'DIŞI' in s_upper or 'DISI' in s_upper or s_upper == 'ARIZA':
        return 'Servis Dışı'
    # Servis/Serviste/Aktif variants
    if s_upper in ('SERVİSTE', 'SERVISTE', 'SERVİS', 'SERVIS', 'AKTİF', 'AKTIF'):
        return 'Servis'
    return s


def _project_dir(project_code: str):
    return os.path.join(os.path.dirname(__file__), '..', 'data', project_code)


def _ensure_workbook(path: str, headers):
    os.makedirs(os.path.dirname(path), exist_ok=True)
    if not os.path.exists(path):
        wb = Workbook()
        ws = wb.active
        ws.append(headers)
        wb.save(path)
        return wb
    return load_workbook(path)


def get_km_excel_path(project_code: str):
    return os.path.join(_project_dir(project_code), 'km_data.xlsx')


def get_service_excel_path(project_code: str):
    """
    Varsayılan dosya yolu (eski): service_status.xlsx
    """
    return os.path.join(_project_dir(project_code), 'service_status.xlsx')


def get_servis_durumu_excel_path(project_code: str):
    """
    Yeni dosya yolu: Servis_Durumu.xlsx
    """
    return os.path.join(_project_dir(project_code), 'Servis_Durumu.xlsx')


def read_all_km(project_code: str):
    path = get_km_excel_path(project_code)
    wb = _ensure_workbook(path, ['tram_id', 'current_km', 'notes', 'updated_at', 'updated_by'])
    ws = wb.active

    results = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        tram_id = str(row[0]).strip() if row and row[0] is not None else None
        if not tram_id:
            continue
        try:
            current_km = int(float(row[1])) if row[1] is not None and str(row[1]).strip() != '' else 0
        except Exception:
            current_km = 0
        results[tram_id] = {
            'current_km': current_km,
            'notes': str(row[2] or ''),
            'updated_at': str(row[3] or ''),
            'updated_by': str(row[4] or '')
        }

    return results


def upsert_km(project_code: str, tram_id: str, current_km: int, notes: str = '', updated_by: str = 'system'):
    path = get_km_excel_path(project_code)
    wb = _ensure_workbook(path, ['tram_id', 'current_km', 'notes', 'updated_at', 'updated_by'])
    ws = wb.active

    tram_id = str(tram_id)
    updated = False
    now_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    for row_idx in range(2, ws.max_row + 1):
        val = ws.cell(row=row_idx, column=1).value
        if val is not None and str(val).strip() == tram_id:
            ws.cell(row=row_idx, column=2).value = int(current_km)
            ws.cell(row=row_idx, column=3).value = notes or ''
            ws.cell(row=row_idx, column=4).value = now_str
            ws.cell(row=row_idx, column=5).value = updated_by
            updated = True
            break

    if not updated:
        ws.append([tram_id, int(current_km), notes or '', now_str, updated_by])

    wb.save(path)


def sync_km_excel_to_equipment(project_code: str):
    from models import Equipment, db

    km_map = read_all_km(project_code)
    if not km_map:
        return 0

    updated = 0
    for tram_id, info in km_map.items():
        equipment = Equipment.query.filter_by(equipment_code=str(tram_id), project_code=project_code).first()
        if not equipment:
            equipment = Equipment(
                equipment_code=str(tram_id),
                name=f'Tramvay {tram_id}',
                equipment_type='Tramvay',
                current_km=info['current_km'],
                monthly_km=0,
                notes=info.get('notes', ''),
                project_code=project_code
            )
            db.session.add(equipment)
            updated += 1
            continue

        new_km = int(info['current_km']) if info.get('current_km') is not None else 0
        if (equipment.current_km or 0) != new_km:
            equipment.current_km = new_km
            updated += 1

    db.session.commit()
    return updated


def bootstrap_km_excel_from_equipment(project_code: str):
    from models import Equipment

    existing = read_all_km(project_code)
    if existing:
        return 0

    seeded = 0
    equipments = Equipment.query.filter_by(parent_id=None, project_code=project_code).all()
    for eq in equipments:
        upsert_km(
            project_code=project_code,
            tram_id=str(eq.equipment_code),
            current_km=int(eq.current_km or 0),
            notes=str(getattr(eq, 'notes', '') or ''),
            updated_by='bootstrap'
        )
        seeded += 1

    return seeded


def _get_status_symbol_and_fill(status: str):
    """Duruma göre sembol, hücre dolgu rengi ve font döndür"""
    from openpyxl.styles import PatternFill, Font, Alignment
    s = status.strip() if status else ''
    s_upper = s.upper()
    # Beyaz, kalın font - tüm semboller için
    symbol_font = Font(color='FFFFFF', bold=True, size=12)
    # Önce "İşletme Kaynaklı" kontrol et (çünkü "Servis Dışı" da içerir)
    if 'İŞLETME' in s_upper or 'ISLETME' in s_upper or 'IŞLETME' in s_upper:
        return '⚠', PatternFill(start_color="00FFC000", end_color="00FFC000", fill_type="solid"), symbol_font
    elif 'SERVİS DIŞI' in s_upper or 'SERVIS DIŞI' in s_upper or 'SERVIS DISI' in s_upper or 'ARIZA' in s_upper or s_upper == 'PASİF' or s_upper == 'PASIF':
        return '✗', PatternFill(start_color="00FF0000", end_color="00FF0000", fill_type="solid"), symbol_font
    elif s_upper in ('SERVİSTE', 'SERVISTE', 'SERVİS', 'SERVIS', 'AKTİF', 'AKTIF'):
        return '✓', PatternFill(start_color="0000B050", end_color="0000B050", fill_type="solid"), symbol_font
    else:
        return status, None, None


def _status_from_symbol(symbol: str, fill_rgb: str = None):
    """Excel sembolü ve/veya dolgu renginden durum string'i döndür"""
    sym = str(symbol or '').strip()
    if sym == '✓':
        return 'Servis'
    elif sym == '✗':
        return 'Servis Dışı'
    elif sym == '⚠':
        return 'İşletme Kaynaklı Servis Dışı'
    # Renk bazlı fallback
    if fill_rgb:
        rgb = fill_rgb.upper().replace('00', '', 1) if fill_rgb.startswith('00') and len(fill_rgb) == 8 else fill_rgb.upper()
        if '00B050' in rgb or '92D050' in rgb:
            return 'Servis'
        elif 'FFC000' in rgb:
            return 'İşletme Kaynaklı Servis Dışı'
        elif 'FF0000' in rgb:
            return 'Servis Dışı'
    if sym:
        return sym
    return ''


def _normalize_date_str(cell_val):
    """Hücre değerini YYYY-MM-DD string'ine normalize et (datetime veya string olabilir)"""
    if cell_val is None:
        return ''
    from datetime import datetime as dt_cls, date as date_cls
    if isinstance(cell_val, dt_cls):
        return cell_val.strftime('%Y-%m-%d')
    if isinstance(cell_val, date_cls):
        return cell_val.strftime('%Y-%m-%d')
    return str(cell_val).strip()[:10]  # '2026-04-09 00:00:00' -> '2026-04-09'


def _write_servis_durumu_excel(project_code: str, status_date: str, tram_id: str, status: str,
                               sistem: str = '', alt_sistem: str = '', aciklama: str = ''):
    """Servis_Durumu.xlsx dosyasına grid formatında yaz (Satır=Tramvay, Sütun=Tarih)"""
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.comments import Comment
    path = get_servis_durumu_excel_path(project_code)
    if not os.path.exists(path):
        return  # Dosya yoksa oluşturma, kullanıcının kendi dosyası olmalı

    wb = load_workbook(path)
    ws = wb.active

    # Tramvay satırını bul
    tram_row = None
    for r in range(2, ws.max_row + 1):
        cell_val = ws.cell(row=r, column=1).value
        if cell_val is not None and str(cell_val).strip() == str(tram_id).strip():
            tram_row = r
            break

    if tram_row is None:
        # Tramvay yoksa yeni satır ekle
        tram_row = ws.max_row + 1
        ws.cell(row=tram_row, column=1, value=str(tram_id))
        ws.cell(row=tram_row, column=1).fill = PatternFill(start_color="00E7E6E6", end_color="00E7E6E6", fill_type="solid")

    # Tarih sütununu bul (datetime ve string uyumlu)
    tarih_str = str(status_date).strip()[:10]
    tarih_col = None
    for c in range(2, ws.max_column + 1):
        cell_val = ws.cell(row=1, column=c).value
        if _normalize_date_str(cell_val) == tarih_str:
            tarih_col = c
            break

    if tarih_col is None:
        # Tarih yoksa → doğru sıraya ekle (tarihler soldan sağa azalan: en yeni sol, en eski sağ)
        # Col 2'den itibaren ilk kendi'nden küçük tarihi bul, onun önüne ekle
        insert_col = ws.max_column + 1  # default: en sona
        for c in range(2, ws.max_column + 1):
            col_date = _normalize_date_str(ws.cell(row=1, column=c).value)
            if col_date and col_date < tarih_str:
                insert_col = c
                break

        # Sütun ekle (insert_col'dan itibaren mevcut sütunları sağa kaydır)
        ws.insert_cols(insert_col)
        tarih_col = insert_col

        header_cell = ws.cell(row=1, column=tarih_col, value=tarih_str)
        header_cell.fill = PatternFill(start_color="004472C4", end_color="004472C4", fill_type="solid")
        header_cell.font = Font(color="FFFFFF", bold=True)
        header_cell.alignment = Alignment(horizontal="center")

    # Duruma göre sembol, renk ve font
    from openpyxl.styles import Border, Side
    symbol, fill, symbol_font = _get_status_symbol_and_fill(status)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    cell = ws.cell(row=tram_row, column=tarih_col, value=symbol)
    if fill:
        cell.fill = fill
    if symbol_font:
        cell.font = symbol_font
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border

    # Servis Dışı veya İşletme Kaynaklı durumlarında not (comment) ekle
    s_upper = (status.strip().upper() if status else '')
    is_disi = ('DIŞI' in s_upper or 'DISI' in s_upper or 'İŞLETME' in s_upper or 
               'IŞLETME' in s_upper or 'ISLETME' in s_upper or 'ARIZA' in s_upper)
    if is_disi and (sistem or alt_sistem or aciklama):
        comment_lines = []
        if sistem:
            comment_lines.append(f'Sistem: {sistem}')
        if alt_sistem:
            comment_lines.append(f'Alt Sistem: {alt_sistem}')
        if aciklama:
            comment_lines.append(f'Açıklama: {aciklama}')
        cell.comment = Comment('\n'.join(comment_lines), 'SSH Sistem')
    elif not is_disi:
        # Serviste/Aktif durumda not varsa kaldır
        cell.comment = None

    # Tüm mevcut hücrelere border uygula (eğer yoksa)
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            existing_cell = ws.cell(row=r, column=c)
            if not existing_cell.border or not existing_cell.border.left or existing_cell.border.left.style is None:
                existing_cell.border = thin_border

    wb.save(path)


def _read_servis_durumu_excel(project_code: str, status_date: str):
    """Servis_Durumu.xlsx dosyasından grid formatında oku (Satır=Tramvay, Sütun=Tarih)"""
    path = get_servis_durumu_excel_path(project_code)
    if not os.path.exists(path):
        return None  # Dosya yoksa None döndür (fallback'a geç)

    wb = load_workbook(path)
    ws = wb.active

    # Tarih sütununu bul (datetime ve string uyumlu)
    tarih_str = str(status_date).strip()[:10]
    tarih_col = None
    for c in range(2, ws.max_column + 1):
        cell_val = ws.cell(row=1, column=c).value
        if _normalize_date_str(cell_val) == tarih_str:
            tarih_col = c
            break

    if tarih_col is None:
        return None  # Bu tarih sütunu bulunamadı, fallback'a geç

    out = {}
    for r in range(2, ws.max_row + 1):
        t_id = ws.cell(row=r, column=1).value
        if t_id is None:
            continue
        t_id = str(t_id).strip()
        cell = ws.cell(row=r, column=tarih_col)
        symbol = cell.value
        if symbol is None:
            continue

        # Dolgu rengini al
        fill_rgb = None
        try:
            if cell.fill and cell.fill.start_color and cell.fill.patternType:
                fill_rgb = cell.fill.start_color.rgb
        except Exception:
            pass

        status_str = _status_from_symbol(symbol, fill_rgb)

        # Hücre notundan (comment) sistem/alt_sistem/açıklama bilgisini oku
        c_sistem = ''
        c_alt_sistem = ''
        c_aciklama = ''
        if cell.comment and cell.comment.text:
            for line in cell.comment.text.split('\n'):
                line = line.strip()
                if line.startswith('Sistem:'):
                    c_sistem = line[len('Sistem:'):].strip()
                elif line.startswith('Alt Sistem:'):
                    c_alt_sistem = line[len('Alt Sistem:'):].strip()
                elif line.startswith('Açıklama:'):
                    c_aciklama = line[len('Açıklama:'):].strip()

        out[t_id] = {
            'status': status_str,
            'sistem': c_sistem,
            'alt_sistem': c_alt_sistem,
            'aciklama': c_aciklama,
            'updated_at': '',
            'updated_by': '',
        }

    return out


def upsert_service_status(project_code: str, status_date: str, tram_id: str, status: str,
                          sistem: str = '', alt_sistem: str = '', aciklama: str = '', updated_by: str = 'system'):
    # 1) service_status.xlsx'e log formatında yaz
    path = get_service_excel_path(project_code)
    wb = _ensure_workbook(path, [
        'date', 'tram_id', 'status', 'sistem', 'alt_sistem', 'aciklama', 'updated_at', 'updated_by'
    ])
    ws = wb.active

    key_date = str(status_date)
    key_tram = str(tram_id)
    now_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    updated = False

    for row_idx in range(2, ws.max_row + 1):
        row_date = ws.cell(row=row_idx, column=1).value
        row_tram = ws.cell(row=row_idx, column=2).value
        if str(row_date or '').strip() == key_date and str(row_tram or '').strip() == key_tram:
            ws.cell(row=row_idx, column=3).value = status
            ws.cell(row=row_idx, column=4).value = sistem or ''
            ws.cell(row=row_idx, column=5).value = alt_sistem or ''
            ws.cell(row=row_idx, column=6).value = aciklama or ''
            ws.cell(row=row_idx, column=7).value = now_str
            ws.cell(row=row_idx, column=8).value = updated_by
            updated = True
            break

    if not updated:
        ws.append([key_date, key_tram, status, sistem or '', alt_sistem or '', aciklama or '', now_str, updated_by])

    wb.save(path)

    # 2) Servis_Durumu.xlsx'e grid formatında yaz (renkli, sembolik + not)
    try:
        _write_servis_durumu_excel(project_code, status_date, tram_id, status,
                                   sistem=sistem or '', alt_sistem=alt_sistem or '', aciklama=aciklama or '')
    except Exception as e:
        import logging
        logging.getLogger(__name__).warning(f'Servis_Durumu.xlsx yazma hatası: {e}')


def read_service_status_by_date(project_code: str, status_date: str):
    # Öncelik: Servis_Durumu.xlsx grid formatından oku
    servis_data = _read_servis_durumu_excel(project_code, status_date)
    if servis_data is not None:
        return servis_data

    # Fallback: service_status.xlsx log formatından oku
    path = get_service_excel_path(project_code)
    if not os.path.exists(path):
        return {}

    wb = _ensure_workbook(path, [
        'date', 'tram_id', 'status', 'sistem', 'alt_sistem', 'aciklama', 'updated_at', 'updated_by'
    ])
    ws = wb.active

    out = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_date = str(row[0]).strip() if row and row[0] is not None else ''
        if row_date != str(status_date):
            continue
        tram_id = str(row[1]).strip() if row[1] is not None else ''
        if not tram_id:
            continue
        out[tram_id] = {
            'status': str(row[2] or ''),
            'sistem': str(row[3] or ''),
            'alt_sistem': str(row[4] or ''),
            'aciklama': str(row[5] or ''),
            'updated_at': str(row[6] or ''),
            'updated_by': str(row[7] or ''),
        }

    return out


def sync_grid_excel_to_db_recent(project_code: str, days: int = 7):
    """Servis_Durumu.xlsx grid formatından son N günün verilerini DB'ye sync et (sayfa yüklenirken çağrılır)"""
    import logging
    logger = logging.getLogger(__name__)
    from models import ServiceStatus, db

    path = get_servis_durumu_excel_path(project_code)
    if not os.path.exists(path):
        return 0

    try:
        wb = load_workbook(path)
        ws = wb.active

        # Son N gün tarihlerini hesapla
        recent_dates = set()
        for i in range(days):
            d = (datetime.now() - timedelta(days=i)).strftime('%Y-%m-%d')
            recent_dates.add(d)

        # Tarih sütunlarını bul (sadece son N gün)
        date_cols = {}
        for c in range(2, ws.max_column + 1):
            d = _normalize_date_str(ws.cell(row=1, column=c).value)
            if d and d in recent_dates:
                date_cols[c] = d

        if not date_cols:
            wb.close()
            return 0

        # Tramvay satırlarını oku
        tram_rows = {}
        for r in range(2, ws.max_row + 1):
            t = ws.cell(row=r, column=1).value
            if t is not None:
                tram_rows[r] = str(t).strip()

        changed = 0
        for r, tram_id in tram_rows.items():
            for c, date_str in date_cols.items():
                cell = ws.cell(row=r, column=c)
                symbol = cell.value
                if symbol is None:
                    continue

                fill_rgb = None
                try:
                    if cell.fill and cell.fill.start_color and cell.fill.patternType:
                        fill_rgb = cell.fill.start_color.rgb
                except Exception:
                    pass

                status_str = _status_from_symbol(symbol, fill_rgb)
                if not status_str:
                    continue

                # Normalize status
                status_str = normalize_status(status_str)

                # Hücre notlarından sistem/alt_sistem/açıklama
                c_sistem = ''
                c_alt_sistem = ''
                c_aciklama = ''
                if cell.comment and cell.comment.text:
                    for line in cell.comment.text.split('\n'):
                        line = line.strip()
                        if line.startswith('Sistem:'):
                            c_sistem = line[len('Sistem:'):].strip()
                        elif line.startswith('Alt Sistem:'):
                            c_alt_sistem = line[len('Alt Sistem:'):].strip()
                        elif line.startswith('Açıklama:'):
                            c_aciklama = line[len('Açıklama:'):].strip()

                existing = ServiceStatus.query.filter_by(
                    tram_id=tram_id, date=date_str, project_code=project_code
                ).first()

                if existing:
                    if existing.status != status_str:
                        existing.status = status_str
                        changed += 1
                else:
                    db.session.add(ServiceStatus(
                        tram_id=tram_id, date=date_str, status=status_str,
                        sistem=c_sistem, alt_sistem=c_alt_sistem, aciklama=c_aciklama,
                        project_code=project_code
                    ))
                    changed += 1

        if changed:
            db.session.commit()
        wb.close()
        return changed
    except Exception as e:
        logger.warning(f'Grid sync hatası ({project_code}): {e}')
        return 0


def sync_service_excel_to_db(project_code: str, only_date: str = None):
    from models import ServiceStatus, db

    path = get_service_excel_path(project_code)
    wb = _ensure_workbook(path, [
        'date', 'tram_id', 'status', 'sistem', 'alt_sistem', 'aciklama', 'updated_at', 'updated_by'
    ])
    ws = wb.active

    changed = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_date = str(row[0]).strip() if row and row[0] is not None else ''
        tram_id = str(row[1]).strip() if row and row[1] is not None else ''
        if not row_date or not tram_id:
            continue
        if only_date and row_date != str(only_date):
            continue

        status = str(row[2] or '')
        sistem = str(row[3] or '')
        alt_sistem = str(row[4] or '')
        aciklama = str(row[5] or '')

        existing = ServiceStatus.query.filter_by(
            tram_id=tram_id,
            date=row_date,
            project_code=project_code
        ).first()

        if existing:
            if existing.status != status or (existing.sistem or '') != sistem or (existing.alt_sistem or '') != alt_sistem or (existing.aciklama or '') != aciklama:
                existing.status = status
                existing.sistem = sistem
                existing.alt_sistem = alt_sistem
                existing.aciklama = aciklama
                changed += 1
        else:
            db.session.add(ServiceStatus(
                tram_id=tram_id,
                date=row_date,
                status=status,
                sistem=sistem,
                alt_sistem=alt_sistem,
                aciklama=aciklama,
                project_code=project_code
            ))
            changed += 1

    db.session.commit()
    return changed

def get_tramvay_list_with_km(project_code: str):
    """
    Get all tramvay/equipment for a project with their KM values.
    Single source of truth: Database (Equipment table)
    
    Returns: List of Equipment objects with current_km populated
    Strategy:
      1. Get tram ID list from Excel (Veriler.xlsx Sayfa2)
      2. Pull KM from Database (Equipment table) 
      3. Sync Excel->DB first to ensure consistency
      
    This replaces complex logic in /tramvay-km route
    """
    import pandas as pd
    from models import Equipment, db
    
    # NOTE: Equipment tablosu tek kaynak - Excel sync devre dışı
    # KM verileri sadece /tramvay-km sayfasından girilmelidir
    
    # Get tram IDs from Excel
    tram_ids = []
    project_folder = os.path.join(os.path.dirname(__file__), '..', 'data', project_code)
    excel_path = os.path.join(project_folder, 'Veriler.xlsx')
    
    if os.path.exists(excel_path):
        try:
            xls = pd.ExcelFile(excel_path)
            sheet_name = 'Sayfa2' if 'Sayfa2' in xls.sheet_names else (xls.sheet_names[0] if xls.sheet_names else None)
            
            if sheet_name:
                df = pd.read_excel(excel_path, sheet_name=sheet_name, header=0, engine='openpyxl')
                
                # Find tram_id column
                tram_col = None
                for col in df.columns:
                    if col.lower() in ['tram_id', 'araç', 'equipment_code', 'a']:
                        tram_col = col
                        break
                
                if tram_col:
                    for idx, row in df.iterrows():
                        tram_id = str(row[tram_col]).strip() if pd.notna(row[tram_col]) else None
                        if tram_id and tram_id.lower() not in ['project', 'proje', 'nan', '']:
                            tram_ids.append(tram_id)
        except Exception:
            pass
    
    # Query Database for equipments
    equipments = []
    if tram_ids:
        # Get from DB by equipment_code (which is the tramvay number: 1531, 1532, etc)
        equipments = Equipment.query.filter(
            Equipment.equipment_code.in_(tram_ids),
            Equipment.project_code == project_code
        ).all()
    
    # If no results, fallback to all equipment in project
    if not equipments:
        equipments = Equipment.query.filter_by(
            equipment_type='Tramvay',
            project_code=project_code
        ).all()
    
    return equipments


def import_servis_durumu_grid_to_db(project_code: str):
    """
    Servis_Durumu.xlsx grid dosyasından TÜM verileri DB ServiceStatus tablosuna aktar.
    Excel'deki semboller (✓/✗/⚠) + hücre notlarından (sistem/alt_sistem/açıklama) DB'ye yazar.
    Mevcut kayıtları günceller, yeni kayıtları ekler.
    
    Returns:
        dict: {'inserted': int, 'updated': int, 'skipped': int, 'total_cells': int}
    """
    from models import ServiceStatus, db
    
    path = get_servis_durumu_excel_path(project_code)
    if not os.path.exists(path):
        return {'inserted': 0, 'updated': 0, 'skipped': 0, 'total_cells': 0, 'error': 'Dosya bulunamadı'}
    
    wb = load_workbook(path)
    ws = wb.active
    
    # Tarih sütunlarını oku
    date_cols = {}  # col_index -> date_str
    for c in range(2, ws.max_column + 1):
        d = _normalize_date_str(ws.cell(row=1, column=c).value)
        if d:
            date_cols[c] = d
    
    # Tramvay satırlarını oku
    tram_rows = {}  # row_index -> tram_id
    for r in range(2, ws.max_row + 1):
        t = ws.cell(row=r, column=1).value
        if t is not None:
            tram_rows[r] = str(t).strip()
    
    inserted = 0
    updated = 0
    skipped = 0
    total_cells = 0
    
    for r, tram_id in tram_rows.items():
        for c, date_str in date_cols.items():
            cell = ws.cell(row=r, column=c)
            symbol = cell.value
            if symbol is None:
                skipped += 1
                continue
            
            total_cells += 1
            
            # Dolgu rengini al
            fill_rgb = None
            try:
                if cell.fill and cell.fill.start_color and cell.fill.patternType:
                    fill_rgb = cell.fill.start_color.rgb
            except Exception:
                pass
            
            status_str = _status_from_symbol(symbol, fill_rgb)
            if not status_str:
                skipped += 1
                continue
            
            # Hücre notundan sistem/alt_sistem/açıklama oku
            c_sistem = ''
            c_alt_sistem = ''
            c_aciklama = ''
            if cell.comment and cell.comment.text:
                for line in cell.comment.text.split('\n'):
                    line = line.strip()
                    if line.startswith('Sistem:'):
                        c_sistem = line[len('Sistem:'):].strip()
                    elif line.startswith('Alt Sistem:'):
                        c_alt_sistem = line[len('Alt Sistem:'):].strip()
                    elif line.startswith('Açıklama:'):
                        c_aciklama = line[len('Açıklama:'):].strip()
            
            # DB'de mevcut mu?
            existing = ServiceStatus.query.filter_by(
                tram_id=tram_id,
                date=date_str,
                project_code=project_code
            ).first()
            
            # Status normalizasyonu: 'Serviste' -> 'Servis' (DB tutarlılığı)
            if status_str == 'Serviste':
                status_str = 'Servis'
            
            if existing:
                changed = False
                if existing.status != status_str:
                    existing.status = status_str
                    changed = True
                if (existing.sistem or '') != c_sistem:
                    existing.sistem = c_sistem
                    changed = True
                if (existing.alt_sistem or '') != c_alt_sistem:
                    existing.alt_sistem = c_alt_sistem
                    changed = True
                if (existing.aciklama or '') != c_aciklama:
                    existing.aciklama = c_aciklama
                    changed = True
                if changed:
                    updated += 1
                else:
                    skipped += 1
            else:
                db.session.add(ServiceStatus(
                    tram_id=tram_id,
                    date=date_str,
                    status=status_str,
                    sistem=c_sistem,
                    alt_sistem=c_alt_sistem,
                    aciklama=c_aciklama,
                    project_code=project_code
                ))
                inserted += 1
    
    db.session.commit()
    wb.close()
    
    return {
        'inserted': inserted,
        'updated': updated,
        'skipped': skipped,
        'total_cells': total_cells,
        'date_range': f"{min(date_cols.values())} - {max(date_cols.values())}" if date_cols else 'N/A',
        'tram_count': len(tram_rows)
    }


def import_historical_excel_to_system(project_code: str, excel_path: str):
    """
    Harici bir Excel dosyasından geçmiş veriyi hem DB'ye hem Servis_Durumu.xlsx'e aktar.
    
    Excel formatı (grid): 
        Satır 1 = Tarihler (YYYY-MM-DD)
        Sütun A = Tramvay IDleri
        Hücre = ✓ / ✗ / ⚠ veya Servis / Servis Dışı / İşletme Kaynaklı
        Hücre Notu = Sistem: X / Alt Sistem: Y / Açıklama: Z
    
    Returns:
        dict: {'db_inserted': int, 'db_updated': int, 'excel_written': int, 'errors': list}
    """
    from models import ServiceStatus, db
    
    if not os.path.exists(excel_path):
        return {'error': 'Dosya bulunamadı', 'db_inserted': 0, 'db_updated': 0, 'excel_written': 0}
    
    wb = load_workbook(excel_path)
    ws = wb.active
    
    # Tarih sütunlarını oku
    date_cols = {}
    for c in range(2, ws.max_column + 1):
        d = _normalize_date_str(ws.cell(row=1, column=c).value)
        if d:
            date_cols[c] = d
    
    # Tramvay satırlarını oku
    tram_rows = {}
    for r in range(2, ws.max_row + 1):
        t = ws.cell(row=r, column=1).value
        if t is not None:
            tram_rows[r] = str(t).strip()
    
    db_inserted = 0
    db_updated = 0
    excel_written = 0
    errors = []
    
    for r, tram_id in tram_rows.items():
        for c, date_str in date_cols.items():
            cell = ws.cell(row=r, column=c)
            raw_value = cell.value
            if raw_value is None:
                continue
            
            raw_str = str(raw_value).strip()
            
            # Sembol veya metin → status
            fill_rgb = None
            try:
                if cell.fill and cell.fill.start_color and cell.fill.patternType:
                    fill_rgb = cell.fill.start_color.rgb
            except Exception:
                pass
            
            status_str = _status_from_symbol(raw_str, fill_rgb)
            if not status_str:
                # Metin bazlı kontrol
                upper = raw_str.upper()
                if 'İŞLETME' in upper or 'ISLETME' in upper:
                    status_str = 'İşletme Kaynaklı Servis Dışı'
                elif 'DIŞI' in upper or 'DISI' in upper or 'ARIZA' in upper:
                    status_str = 'Servis Dışı'
                elif 'SERVİS' in upper or 'SERVIS' in upper or 'AKTİF' in upper or 'AKTIF' in upper:
                    status_str = 'Servis'
                else:
                    errors.append(f'{tram_id}/{date_str}: Bilinmeyen değer "{raw_str}"')
                    continue
            
            if status_str == 'Serviste':
                status_str = 'Servis'
            
            # Notlardan bilgi oku
            c_sistem = ''
            c_alt_sistem = ''
            c_aciklama = ''
            if cell.comment and cell.comment.text:
                for line in cell.comment.text.split('\n'):
                    line = line.strip()
                    if line.startswith('Sistem:'):
                        c_sistem = line[len('Sistem:'):].strip()
                    elif line.startswith('Alt Sistem:'):
                        c_alt_sistem = line[len('Alt Sistem:'):].strip()
                    elif line.startswith('Açıklama:'):
                        c_aciklama = line[len('Açıklama:'):].strip()
            
            # 1) DB'ye yaz
            try:
                existing = ServiceStatus.query.filter_by(
                    tram_id=tram_id,
                    date=date_str,
                    project_code=project_code
                ).first()
                
                if existing:
                    existing.status = status_str
                    existing.sistem = c_sistem
                    existing.alt_sistem = c_alt_sistem
                    existing.aciklama = c_aciklama
                    db_updated += 1
                else:
                    db.session.add(ServiceStatus(
                        tram_id=tram_id,
                        date=date_str,
                        status=status_str,
                        sistem=c_sistem,
                        alt_sistem=c_alt_sistem,
                        aciklama=c_aciklama,
                        project_code=project_code
                    ))
                    db_inserted += 1
            except Exception as e:
                errors.append(f'{tram_id}/{date_str}: DB hatası - {str(e)}')
                continue
            
            # 2) Servis_Durumu.xlsx'e yaz
            try:
                _write_servis_durumu_excel(project_code, date_str, tram_id, status_str,
                                           sistem=c_sistem, alt_sistem=c_alt_sistem, aciklama=c_aciklama)
                excel_written += 1
            except Exception as e:
                errors.append(f'{tram_id}/{date_str}: Excel yazma hatası - {str(e)}')
    
    db.session.commit()
    wb.close()
    
    return {
        'db_inserted': db_inserted,
        'db_updated': db_updated,
        'excel_written': excel_written,
        'date_range': f"{min(date_cols.values())} - {max(date_cols.values())}" if date_cols else 'N/A',
        'tram_count': len(tram_rows),
        'errors': errors[:50]  # İlk 50 hata
    }