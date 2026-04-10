"""
Excel Grid Manager - Servis Durumu ve RCA Analiz Excel Dosya Yönetimi
Satırlar: Tarihler, Sütunlar: Araçlar, İçerik: Durum (Yeşil/Kırmızı/Turuncu)
"""

import pandas as pd
import os
from datetime import datetime, timedelta
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import logging

logger = logging.getLogger(__name__)

# Durum kodları ve renkleri
STATUS_COLORS = {
    'aktif': 'FF00B050',  # Yeşil
    'servis_disi': 'FFFF0000',  # Kırmızı
    'isletme_kaynakli': 'FFFFC000'  # Turuncu
}

STATUS_SYMBOLS = {
    'aktif': '✓',
    'servis_disi': '✗',
    'isletme_kaynakli': '⚠'
}

class ExcelGridManager:
    """Excel grid dosyalarını yönet (service_status_grid.xlsx)"""
    
    def __init__(self, project_code='belgrad'):
        self.project_code = project_code
        self.grid_file = None
    
    def get_grid_path(self, base_path):
        """Grid Excel dosyasının yolunu al"""
        data_dir = os.path.join(base_path, 'data', self.project_code)
        os.makedirs(data_dir, exist_ok=True)
        return os.path.join(data_dir, 'service_status_grid.xlsx')
    
    def init_grid(self, base_path, equipment_codes):
        """Grid Excel dosyasını oluştur veya sütunlar uyuşmuyorsa yeniden oluştur
        
        Yapı:
        - Satırlar: Tarihler (son 90 gün)
        - Sütunlar: Araçlar (equipment_code)
        - İçerik: Durum sembolü (✓/✗/⚠)
        """
        grid_path = self.get_grid_path(base_path)
        
        if os.path.exists(grid_path):
            # Mevcut sütunları kontrol et - uyuşmuyorsa yeniden oluştur
            try:
                existing_wb = load_workbook(grid_path, data_only=True)
                existing_ws = existing_wb.active
                existing_codes = []
                for col_idx in range(2, existing_ws.max_column + 1):
                    val = existing_ws.cell(row=1, column=col_idx).value
                    if val:
                        existing_codes.append(str(val))
                
                expected_codes = [str(c) for c in equipment_codes]
                
                if set(existing_codes) == set(expected_codes):
                    existing_wb.close()
                    return grid_path  # Kodlar doğru, değişiklik yok
                
                # Kodlar uyuşmuyor - mevcut veriyi koru ve yeniden oluştur
                logger.warning(f'Grid sütunları uyuşmuyor ({self.project_code}): '
                             f'mevcut={existing_codes[:5]}... beklenen={expected_codes[:5]}...')
                
                # Mevcut verileri oku (tarih -> {equipment_code -> value})
                old_data = {}
                for row_idx in range(2, existing_ws.max_row + 1):
                    date_cell = existing_ws.cell(row=row_idx, column=1).value
                    if date_cell:
                        date_key = str(date_cell)[:10]
                        old_data[date_key] = {}
                        for col_idx, code in enumerate(existing_codes, start=2):
                            val = existing_ws.cell(row=row_idx, column=col_idx).value
                            if val:
                                old_data[date_key][code] = val
                
                existing_wb.close()
                os.remove(grid_path)
                logger.info(f'Eski grid silindi, yeniden oluşturuluyor: {self.project_code}')
            except Exception as e:
                logger.error(f'Grid kontrol hatası: {e}')
                return grid_path
        else:
            old_data = {}
        
        # Workbook oluştur
        wb = Workbook()
        ws = wb.active
        ws.title = 'Durum Tablosu'
        
        # Başlık satırı: Araçlar (Sütun A'dan başla)
        ws['A1'] = 'Tarih'
        for idx, equipment_code in enumerate(equipment_codes, start=2):
            col_letter = get_column_letter(idx)
            ws[f'{col_letter}1'] = equipment_code
        
        # Son 90 günün tarihlerini ekle
        base_date = datetime.now().date()
        for day_offset in range(89, -1, -1):  # 90 gün geriye
            current_date = base_date - timedelta(days=day_offset)
            row = 91 - day_offset  # En eski tarih en alt
            ws[f'A{row}'] = current_date
        
        # Stil uygula
        self._apply_grid_styles(ws)
        
        # Eski verilerden eşleşen kodları geri yükle
        if old_data:
            code_to_col = {}
            for idx, code in enumerate(equipment_codes, start=2):
                code_to_col[str(code)] = idx
            
            date_to_row = {}
            for row_idx in range(2, ws.max_row + 1):
                date_cell = ws.cell(row=row_idx, column=1).value
                if date_cell:
                    date_to_row[str(date_cell)[:10]] = row_idx
            
            migrated = 0
            for date_key, code_vals in old_data.items():
                row_idx = date_to_row.get(date_key)
                if not row_idx:
                    continue
                for code, val in code_vals.items():
                    col_idx = code_to_col.get(str(code))
                    if col_idx:
                        cell = ws.cell(row=row_idx, column=col_idx)
                        cell.value = val
                        migrated += 1
            
            if migrated > 0:
                logger.info(f'Grid rebuild: {migrated} hücre eski veriden aktarıldı ({self.project_code})')
        
        wb.save(grid_path)
        logger.info(f'Grid Excel oluşturuldu: {grid_path}')
        return grid_path
    
    def batch_update_status(self, base_path, updates):
        """Toplu durum güncelleme (performans için)
        
        Args:
            updates: list of {'tram_id', 'date', 'status'}
        """
        grid_path = self.get_grid_path(base_path)
        
        if not os.path.exists(grid_path):
            logger.warning(f'Grid dosyası bulunamadı: {grid_path}')
            return False
        
        wb = load_workbook(grid_path)
        ws = wb.active
        
        # Tüm sütun/satır mappings'i bul
        col_map = {}  # tram_id -> col_idx
        row_map = {}  # date -> row_idx
        
        for idx, cell in enumerate(ws[1], start=1):
            if cell.value:
                col_map[cell.value] = idx
        
        for idx, cell in enumerate(ws['A'], start=1):
            if cell.value:
                row_map[str(cell.value)[:10]] = idx  # İlk 10 char (tarih)
        
        # Batch güncelleme
        for update in updates:
            tram_id = update['tram_id']
            date_str = update['date']
            status = update['status']
            
            col_idx = col_map.get(tram_id)
            row_idx = row_map.get(date_str)
            
            if col_idx and row_idx:
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = STATUS_SYMBOLS.get(status, '?')
                
                if status in STATUS_COLORS:
                    cell.fill = PatternFill(start_color=STATUS_COLORS[status], 
                                           end_color=STATUS_COLORS[status], 
                                           fill_type='solid')
                
                cell.font = Font(bold=True, size=14)
                cell.alignment = Alignment(horizontal='center', vertical='center')
        
        wb.save(grid_path)
        logger.info(f'Batch güncelleme tamamlandı: {len(updates)} hücre')
        return True
    
    def update_status(self, base_path, tram_id, date_str, status):
        """Tramvay durumunu güncelle
        
        Args:
            tram_id: Araç ID (sütun)
            date_str: Tarih (satır) - '2026-03-21'
            status: 'aktif' | 'servis_disi' | 'isletme_kaynakli'
        """
        grid_path = self.get_grid_path(base_path)
        
        if not os.path.exists(grid_path):
            raise FileNotFoundError(f'Grid dosyası bulunamadı: {grid_path}')
        
        wb = load_workbook(grid_path)
        ws = wb.active
        
        # Sütun bul (tram_id)
        col_idx = None
        for idx, cell in enumerate(ws[1], start=1):
            if cell.value == tram_id:
                col_idx = idx
                break
        
        if col_idx is None:
            logger.warning(f'Araç bulunamadı: {tram_id}')
            return False
        
        # Satır bul (tarih)
        row_idx = None
        for idx, cell in enumerate(ws['A'], start=1):
            if cell.value and str(cell.value).startswith(date_str):
                row_idx = idx
                break
        
        if row_idx is None:
            logger.warning(f'Tarih bulunamadı: {date_str}')
            return False
        
        # Hücreyi güncelle
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.value = STATUS_SYMBOLS.get(status, '?')
        
        # Renk uygula
        if status in STATUS_COLORS:
            cell.fill = PatternFill(start_color=STATUS_COLORS[status], 
                                   end_color=STATUS_COLORS[status], 
                                   fill_type='solid')
        
        cell.font = Font(bold=True, size=14)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
        wb.save(grid_path)
        logger.info(f'Durum güncellendi: {tram_id} - {date_str} - {status}')
        return True
    
    def get_availability_data(self, base_path, start_date=None, end_date=None):
        """Excel'den availability verilerini oku (Hızlı - Openpyxl kullanır)
        
        Returns:
            dict: {'Araç-101': 95%, 'Araç-102': 87%, ...}
        """
        grid_path = self.get_grid_path(base_path)
        
        if not os.path.exists(grid_path):
            return {}
        
        try:
            wb = load_workbook(grid_path, data_only=True)
            ws = wb.active
            
            # Başlık satırını al (araç adları)
            equipment_codes = []
            for col_idx in range(2, ws.max_column + 1):  # A'yı atla (Tarih sütunu)
                cell_value = ws.cell(row=1, column=col_idx).value
                if cell_value:
                    equipment_codes.append(cell_value)
            
            # Tarih aralığına göre satırları filter et
            date_rows = []
            from datetime import datetime as dt
            
            for row_idx in range(2, ws.max_row + 1):
                date_cell = ws.cell(row=row_idx, column=1).value
                
                if date_cell:
                    try:
                        # Tarih parse et
                        if isinstance(date_cell, str):
                            cell_date = dt.strptime(date_cell, '%Y-%m-%d').date()
                        else:
                            cell_date = date_cell.date() if hasattr(date_cell, 'date') else date_cell
                        
                        # Filtre uygula
                        if start_date and cell_date < start_date:
                            continue
                        if end_date and cell_date > end_date:
                            continue
                        
                        date_rows.append(row_idx)
                    except:
                        continue
            
            # Her araç için availability hesapla
            availability = {}
            for col_idx, equipment_code in enumerate(equipment_codes, start=2):
                active_count = 0
                total_count = 0
                
                for row_idx in date_rows:
                    cell_value = ws.cell(row=row_idx, column=col_idx).value
                    if cell_value:
                        total_count += 1
                        # Sembol normalizasyonu: √ → ✓
                        normalized_value = str(cell_value).strip()
                        if normalized_value in ('✓', '√'):  # Both check mark variants
                            active_count += 1
                
                if total_count > 0:
                    availability[equipment_code] = round((active_count / total_count) * 100)
            
            wb.close()
            return availability
        
        except Exception as e:
            logger.error(f'Excel okuma hatası: {e}')
            return {}
    
    def get_availability_trend(self, base_path, start_date=None, end_date=None, granularity='daily'):
        """Availability trend verilerini oku - Zamansal trend analizi
        
        granularity: 'daily' | 'weekly' | 'monthly'
        
        Returns:
            dict: {
                'dates': ['2026-03-21', '2026-03-20', ...],
                'averages': [85.5, 88.2, ...],
                'equipments': {
                    'eq1': [85, 88, ...],
                    'eq2': [90, 92, ...],
                    ...
                }
            }
        """
        grid_path = self.get_grid_path(base_path)
        
        if not os.path.exists(grid_path):
            return {'dates': [], 'averages': [], 'equipments': {}}
        
        try:
            from datetime import datetime as dt
            from collections import defaultdict
            
            wb = load_workbook(grid_path, data_only=True)
            ws = wb.active
            
            # Başlık satırını al (araç adları)
            equipment_codes = []
            for col_idx in range(2, ws.max_column + 1):
                cell_value = ws.cell(row=1, column=col_idx).value
                if cell_value:
                    equipment_codes.append(cell_value)
            
            # Tüm verileri oku
            all_data = {}  # {date: {'eq1': symbol, 'eq2': symbol, ...}}
            
            for row_idx in range(2, ws.max_row + 1):
                date_cell = ws.cell(row=row_idx, column=1).value
                
                if date_cell:
                    try:
                        # Tarih parse et
                        if isinstance(date_cell, str):
                            cell_date = dt.strptime(date_cell, '%Y-%m-%d').date()
                        else:
                            cell_date = date_cell.date() if hasattr(date_cell, 'date') else date_cell
                        
                        # Filtre uygula
                        if start_date and cell_date < start_date:
                            continue
                        if end_date and cell_date > end_date:
                            continue
                        
                        # Bu tarih için verileri topla
                        row_data = {}
                        for col_idx, equipment_code in enumerate(equipment_codes, start=2):
                            cell_value = ws.cell(row=row_idx, column=col_idx).value
                            row_data[equipment_code] = cell_value
                        
                        all_data[cell_date] = row_data
                    except:
                        continue
            
            # Granülarite'ye göre bucket'la
            buckets = defaultdict(lambda: defaultdict(list))  # {bucket_key: {eq: [symbols]}}
            
            for date_obj, row_data in all_data.items():
                # Bucket key'i belirle
                if granularity == 'daily':
                    bucket_key = date_obj
                elif granularity == 'weekly':
                    # ISO week
                    year, week, _ = date_obj.isocalendar()
                    bucket_key = f"{year}-W{week:02d}"
                elif granularity == 'monthly':
                    bucket_key = date_obj.strftime('%Y-%m')
                else:
                    bucket_key = date_obj
                
                # Bu bucket'a veri ekle
                for equipment_code, symbol in row_data.items():
                    if symbol:
                        normalized = str(symbol).strip()
                        buckets[bucket_key][equipment_code].append(normalized)
            
            # Availability hesapla
            dates = sorted(buckets.keys())
            averages = []
            equipments_dict = defaultdict(list)
            
            for bucket_key in dates:
                bucket_data = buckets[bucket_key]
                
                # Bu bucket'taki tüm araçlar ve semboller
                total_availability = 0
                total_count = 0
                
                for equipment_code in equipment_codes:
                    symbols = bucket_data.get(equipment_code, [])
                    
                    if symbols:
                        # Bu araç için availability
                        active = sum(1 for s in symbols if s in ('✓', '√'))
                        eq_availability = round((active / len(symbols)) * 100)
                    else:
                        eq_availability = 0
                    
                    equipments_dict[equipment_code].append(eq_availability)
                    total_availability += eq_availability
                    total_count += 1
                
                # Ortalama
                avg = round(total_availability / total_count) if total_count > 0 else 0
                averages.append(avg)
            
            wb.close()
            
            return {
                'dates': [str(d) for d in dates],
                'averages': averages,
                'equipments': dict(equipments_dict)
            }
        
        except Exception as e:
            logger.error(f'Trend okuma hatası: {e}')
            return {'dates': [], 'averages': [], 'equipments': {}}
    
    def _apply_grid_styles(self, ws):
        """Grid'e stil uygula"""
        # Başlık satırı stilini uygula
        header_fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
        header_font = Font(color='FFFFFFFF', bold=True, size=11)
        
        for cell in ws[1]:
            if cell.value:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Tarih sütunu stilini uygula
        date_fill = PatternFill(start_color='FFF2F2F2', end_color='FFF2F2F2', fill_type='solid')
        for cell in ws['A'][1:]:
            cell.fill = date_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Sütun genişliğini ayarla
        ws.column_dimensions['A'].width = 15
        for col_idx in range(2, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 12
        
        # Satır yüksekliğini ayarla
        ws.row_dimensions[1].height = 25


class RCAExcelManager:
    """Root Cause Analysis Excel Dosya Yönetimi (rca_analysis.xlsx)"""
    
    def __init__(self, project_code='belgrad'):
        self.project_code = project_code
    
    def get_rca_path(self, base_path):
        """RCA Excel dosyasının yolunu al"""
        rca_dir = os.path.join(base_path, 'logs', self.project_code, 'rca')
        os.makedirs(rca_dir, exist_ok=True)
        return os.path.join(rca_dir, 'rca_analysis.xlsx')
    
    def init_rca(self, base_path):
        """RCA Excel dosyasını oluştur (ilk kez)
        
        Yapı:
        - Sütunlar: Tarih, Araç, Sistem, Alt Sistem, Kategori
        - Satırlar: Her arıza kaydı
        """
        rca_path = self.get_rca_path(base_path)
        
        if os.path.exists(rca_path):
            return rca_path
        
        wb = Workbook()
        ws = wb.active
        ws.title = 'RCA Verileri'
        
        # Başlıklar
        headers = ['Tarih', 'Araç', 'Sistem', 'Alt Sistem', 'Kategori', 'Açıklama']
        for idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=idx)
            cell.value = header
            cell.fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
            cell.font = Font(color='FFFFFFFF', bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Sütun genişliğini ayarla
        column_widths = [15, 12, 15, 20, 20, 30]
        for idx, width in enumerate(column_widths, start=1):
            ws.column_dimensions[get_column_letter(idx)].width = width
        
        wb.save(rca_path)
        logger.info(f'RCA Excel oluşturuldu: {rca_path}')
        return rca_path
    
    def add_rca_record(self, base_path, tram_id, system, subsystem, category, description=''):
        """RCA kaydı ekle
        
        Args:
            tram_id: Araç ID
            system: Ana sistem (BRAKE, ENGINE, vb.)
            subsystem: Alt sistem (Kaliper, Piston, vb.)
            category: Kategori ('servis_disi' | 'isletme_kaynakli')
            description: Açıklama metni
        """
        rca_path = self.get_rca_path(base_path)
        
        if not os.path.exists(rca_path):
            self.init_rca(base_path)
        
        wb = load_workbook(rca_path)
        ws = wb.active
        
        # Son satırı bul
        next_row = ws.max_row + 1
        
        # Kaydı ekle
        ws.cell(row=next_row, column=1).value = datetime.now().strftime('%Y-%m-%d %H:%M')
        ws.cell(row=next_row, column=2).value = tram_id
        ws.cell(row=next_row, column=3).value = system
        ws.cell(row=next_row, column=4).value = subsystem
        ws.cell(row=next_row, column=5).value = category
        ws.cell(row=next_row, column=6).value = description
        
        # Satırı stillendir
        for col_idx in range(1, 7):
            cell = ws.cell(row=next_row, column=col_idx)
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            if col_idx == 5:  # Kategori sütunu renk kodlu
                if category == 'servis_disi':
                    cell.fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
                elif category == 'isletme_kaynakli':
                    cell.fill = PatternFill(start_color='FFFFC000', end_color='FFFFC000', fill_type='solid')
        
        wb.save(rca_path)
        logger.info(f'RCA kaydı eklendi: {tram_id} - {system} - {subsystem}')
        return True
    
    def get_rca_data(self, base_path, start_date=None, end_date=None):
        """RCA verilerini oku (Hızlı - Openpyxl kullanır)
        
        Returns:
            list: RCA kayıtları
        """
        rca_path = self.get_rca_path(base_path)
        
        if not os.path.exists(rca_path):
            return []
        
        try:
            wb = load_workbook(rca_path, data_only=True)
            ws = wb.active
            
            # Başlık satırını al
            headers = []
            for col_idx in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=1, column=col_idx).value
                headers.append(cell_value)
            
            # Veri satırlarını oku
            records = []
            from datetime import datetime as dt
            
            for row_idx in range(2, ws.max_row + 1):
                # Tarih kontrolü
                date_cell = ws.cell(row=row_idx, column=1).value
                if date_cell:
                    try:
                        if isinstance(date_cell, str):
                            cell_date = dt.strptime(date_cell[:10], '%Y-%m-%d').date()
                        else:
                            cell_date = date_cell.date() if hasattr(date_cell, 'date') else date_cell
                        
                        if start_date and cell_date < start_date:
                            continue
                        if end_date and cell_date > end_date:
                            continue
                    except:
                        continue
                
                # Satır verisini dic'ye çevir
                record = {}
                for col_idx, header in enumerate(headers, start=1):
                    record[header] = ws.cell(row=row_idx, column=col_idx).value
                
                records.append(record)
            
            wb.close()
            return records
        
        except Exception as e:
            logger.error(f'RCA okuma hatası: {e}')
            return []
    
    def get_system_stats(self, base_path, start_date=None, end_date=None):
        """Sistem bazında arıza istatistikleri
        
        Returns:
            dict: {'BRAKE': 10, 'ENGINE': 5, ...}
        """
        rca_data = self.get_rca_data(base_path, start_date, end_date)
        stats = {}
        
        for record in rca_data:
            system = record.get('Sistem') or 'Bilinmiyor'
            stats[system] = stats.get(system, 0) + 1
        
        return dict(sorted(stats.items(), key=lambda x: (-x[1], x[0] or ''), reverse=False))
    
    def get_subsystem_stats(self, base_path, start_date=None, end_date=None):
        """Alt sistem bazında arıza istatistikleri
        
        Returns:
            dict: {'Kaliper': 8, 'Piston': 4, ...}
        """
        rca_data = self.get_rca_data(base_path, start_date, end_date)
        stats = {}
        
        for record in rca_data:
            subsystem = record.get('Alt Sistem') or 'Bilinmiyor'
            stats[subsystem] = stats.get(subsystem, 0) + 1
        
        return dict(sorted(stats.items(), key=lambda x: (-x[1], x[0] or ''), reverse=False))
    
    def update_rca_record(self, base_path, row_index, data):
        """Excel'deki belirli bir RCA kaydını güncelle
        
        Args:
            row_index: Güncellenecek satır (0-tabanlı, başlık hariç)
            data: dict {'Tarih', 'Arac', 'Sistem', 'Alt Sistem', 'Kategori', 'Aciklama'}
        """
        rca_path = self.get_rca_path(base_path)
        if not os.path.exists(rca_path):
            return False
        
        wb = load_workbook(rca_path)
        ws = wb.active
        
        excel_row = row_index + 2  # +1 başlık, +1 zero-index
        if excel_row < 2 or excel_row > ws.max_row:
            wb.close()
            return False
        
        col_map = {'Tarih': 1, 'Arac': 2, 'Sistem': 3, 'Alt Sistem': 4, 'Kategori': 5, 'Aciklama': 6}
        for key, col_idx in col_map.items():
            if key in data and data[key] is not None:
                ws.cell(row=excel_row, column=col_idx).value = data[key]
                # Kategori renk kodlama
                if key == 'Kategori':
                    cell = ws.cell(row=excel_row, column=col_idx)
                    if data[key] == 'servis_disi':
                        cell.fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
                        cell.font = Font(color='FFFFFFFF', bold=True)
                    elif data[key] == 'isletme_kaynakli':
                        cell.fill = PatternFill(start_color='FFFFC000', end_color='FFFFC000', fill_type='solid')
                        cell.font = Font(bold=False)
        
        wb.save(rca_path)
        logger.info(f'RCA kaydı güncellendi: satır {row_index}')
        return True
    
    def delete_rca_record(self, base_path, row_index):
        """Excel'deki belirli bir RCA kaydını sil
        
        Args:
            row_index: Silinecek satır (0-tabanlı, başlık hariç)
        """
        rca_path = self.get_rca_path(base_path)
        if not os.path.exists(rca_path):
            return False
        
        wb = load_workbook(rca_path)
        ws = wb.active
        
        excel_row = row_index + 2
        if excel_row < 2 or excel_row > ws.max_row:
            wb.close()
            return False
        
        ws.delete_rows(excel_row, 1)
        wb.save(rca_path)
        logger.info(f'RCA kaydı silindi: satır {row_index}')
        return True
    
    def sync_excel_to_db(self, base_path, db_session, RootCauseAnalysis, project_code=None):
        """Excel'deki RCA verilerini veritabanına senkronize et (Excel → DB)
        
        Excel'deki kayıtları okur, DB'de olmayanları ekler.
        """
        from datetime import datetime as dt
        
        if not project_code:
            project_code = self.project_code
        
        rca_data = self.get_rca_data(base_path)
        if not rca_data:
            return {'added': 0, 'skipped': 0}
        
        added = 0
        skipped = 0
        
        for record in rca_data:
            tarih_str = record.get('Tarih', '')
            arac = str(record.get('Arac', ''))
            sistem = record.get('Sistem', '')
            alt_sistem = record.get('Alt Sistem', '')
            kategori = record.get('Kategori', '')
            aciklama = record.get('Aciklama', '') or ''
            
            if not arac or not sistem:
                skipped += 1
                continue
            
            # Tarih parse
            analysis_date = None
            if tarih_str:
                try:
                    if isinstance(tarih_str, str):
                        analysis_date = dt.strptime(tarih_str[:16], '%Y-%m-%d %H:%M')
                    elif hasattr(tarih_str, 'strftime'):
                        analysis_date = tarih_str
                except Exception:
                    analysis_date = dt.now()
            
            # Duplicate kontrolü: aynı araç + sistem + tarih
            existing = RootCauseAnalysis.query.filter_by(
                tram_id=arac,
                sistem=sistem,
                alt_sistem=alt_sistem or None
            ).filter(
                db_session.func.date(RootCauseAnalysis.analysis_date) == (analysis_date.date() if analysis_date else dt.now().date())
            ).first()
            
            if existing:
                skipped += 1
                continue
            
            # Severity belirle
            severity = 'high' if kategori == 'servis_disi' else 'medium'
            
            rca = RootCauseAnalysis(
                tram_id=arac,
                sistem=sistem,
                alt_sistem=alt_sistem or None,
                root_cause=aciklama or f'{sistem} arızası',
                severity=severity,
                status='submitted',
                analysis_date=analysis_date or dt.now(),
                notes=f'Excel senkronizasyonu - Kategori: {kategori}'
            )
            db_session.add(rca)
            added += 1
        
        if added > 0:
            db_session.commit()
        
        logger.info(f'Excel→DB senkronizasyonu: {added} eklendi, {skipped} atlandı')
        return {'added': added, 'skipped': skipped}
    
    def sync_db_to_excel(self, base_path, db_session, RootCauseAnalysis, project_code=None):
        """Veritabanındaki RCA kayıtlarını Excel'e senkronize et (DB → Excel)
        
        DB'deki kayıtları okur, Excel'de olmayanları ekler.
        """
        if not project_code:
            project_code = self.project_code
        
        # DB'den kayıtları çek
        db_records = RootCauseAnalysis.query.filter(
            RootCauseAnalysis.tram_id.isnot(None),
            RootCauseAnalysis.sistem.isnot(None)
        ).order_by(RootCauseAnalysis.analysis_date).all()
        
        if not db_records:
            return {'added': 0, 'skipped': 0}
        
        # Mevcut Excel verilerini oku (duplicate kontrolü için)
        existing_data = self.get_rca_data(base_path)
        existing_keys = set()
        for rec in existing_data:
            key = (
                str(rec.get('Arac', '')),
                str(rec.get('Sistem', '')),
                str(rec.get('Alt Sistem', '')),
                str(rec.get('Tarih', ''))[:10]
            )
            existing_keys.add(key)
        
        added = 0
        skipped = 0
        
        for record in db_records:
            tarih_str = record.analysis_date.strftime('%Y-%m-%d') if record.analysis_date else ''
            key = (
                str(record.tram_id),
                str(record.sistem),
                str(record.alt_sistem or ''),
                tarih_str[:10]
            )
            
            if key in existing_keys:
                skipped += 1
                continue
            
            # Kategoriyi severity'den türet
            kategori = 'servis_disi' if record.severity in ('high', 'critical') else 'isletme_kaynakli'
            
            self.add_rca_record(
                base_path,
                tram_id=record.tram_id,
                system=record.sistem,
                subsystem=record.alt_sistem or '',
                category=kategori,
                description=record.root_cause or ''
            )
            added += 1
        
        logger.info(f'DB→Excel senkronizasyonu: {added} eklendi, {skipped} atlandı')
        return {'added': added, 'skipped': skipped}
    
    def full_sync(self, base_path, db_session, RootCauseAnalysis, project_code=None):
        """Tam iki yönlü senkronizasyon (Excel ↔ DB)
        
        1. Excel → DB (önce Excel'deki yeni kayıtları DB'ye aktar)
        2. DB → Excel (sonra DB'deki yeni kayıtları Excel'e aktar)
        """
        excel_to_db = self.sync_excel_to_db(base_path, db_session, RootCauseAnalysis, project_code)
        db_to_excel = self.sync_db_to_excel(base_path, db_session, RootCauseAnalysis, project_code)
        
        result = {
            'excel_to_db': excel_to_db,
            'db_to_excel': db_to_excel,
            'total_added': excel_to_db['added'] + db_to_excel['added'],
            'total_skipped': excel_to_db['skipped'] + db_to_excel['skipped']
        }
        logger.info(f'Tam senkronizasyon tamamlandı: {result}')
        return result


def init_excel_files(app, project_code, equipment_codes):
    """İlk kullanım için Excel dosyalarını oluştur"""
    try:
        grid_manager = ExcelGridManager(project_code)
        rca_manager = RCAExcelManager(project_code)
        
        grid_manager.init_grid(app.root_path, equipment_codes)
        rca_manager.init_rca(app.root_path)
        
        logger.info(f'Excel dosyaları hazır: {project_code}')
        return True
    except Exception as e:
        logger.error(f'Excel dosyası oluşturma hatası: {e}')
        return False
