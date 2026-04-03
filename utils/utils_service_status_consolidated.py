"""
Unified Service Status Logger
Servis durumu değişikliklerini JSON ve Excel formatında tek bir sınıf aracılığıyla logla
Consolidates: utils_service_status_logger, utils_service_status_data_logger, utils_service_status_excel_logger
"""

import json
import os
import logging
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

logger = logging.getLogger(__name__)


class ServiceStatusLogger:
    """
    Unified Service Status Logger - Servis durumu değişikliklerini yönet
    Supports: JSON logging, Excel logging, Data history management
    """
    
    BASE_DIR = 'logs/service_status'
    
    def __init__(self, project_code: str = 'belgrad'):
        """Initialize logger for specific project"""
        self.project_code = project_code
        self.project_dir: Path = Path(self.BASE_DIR) / project_code
        self.project_dir.mkdir(parents=True, exist_ok=True)
        self.json_file: Path = self.project_dir / 'service_status_history.json'
        self.excel_file: Path = self.project_dir / 'service_status_log.xlsx'
    
    # ==================== JSON LOGGING ====================
    
    def log_json(self, tram_id: str, date: str, status: str, sistem: str = '', 
                 alt_sistem: str = '', aciklama: str = '', user_id: Optional[int] = None) -> bool:
        """Log service status to JSON file"""
        try:
            # Load existing logs
            if self.json_file.exists():
                with open(self.json_file, 'r', encoding='utf-8') as f:
                    logs = json.load(f)
            else:
                logs = []
            
            # Create new log entry
            new_log = {
                'timestamp': datetime.now().isoformat(),
                'tram_id': str(tram_id),
                'date': str(date),
                'status': status,
                'sistem': sistem,
                'alt_sistem': alt_sistem,
                'aciklama': aciklama,
                'user_id': user_id
            }
            logs.append(new_log)
            
            # Write to file
            with open(self.json_file, 'w', encoding='utf-8') as f:
                json.dump(logs, f, indent=2, ensure_ascii=False, default=str)
            
            logger.info(f"JSON log saved: {tram_id} -> {status}")
            return True
            
        except Exception as e:
            logger.error(f"JSON logging error: {e}")
            return False
    
    # ==================== EXCEL LOGGING ====================
    
    def log_excel(self, tram_id: str, date: str, status: str, sistem: str = '', 
                  alt_sistem: str = '', aciklama: str = '', user_id: Optional[int] = None) -> bool:
        """Log service status to Excel file"""
        try:
            timestamp = datetime.now().strftime('%d.%m.%Y %H:%M:%S')
            
            # Load or create workbook
            if self.excel_file.exists():
                wb = load_workbook(self.excel_file)
                ws = wb.active
                next_row = ws.max_row + 1
            else:
                wb = Workbook()
                ws = wb.active
                ws.title = 'Service Status History'
                next_row = 1
                self._create_excel_headers(ws)
                next_row = 2
            
            # Determine status color
            status_fill = self._get_status_fill(status)
            status_font = Font(bold=True, color='FFFFFF') if status_fill else Font()
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Add data row
            data = [timestamp, str(tram_id), str(date), status, sistem, alt_sistem, aciklama, user_id or '']
            
            for col_idx, value in enumerate(data, 1):
                cell = ws.cell(row=next_row, column=col_idx)
                cell.value = value
                cell.border = border
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                
                if col_idx == 4:  # Status column
                    if status_fill:
                        cell.fill = status_fill
                        cell.font = status_font
            
            # Save workbook
            wb.save(self.excel_file)
            logger.info(f"Excel log saved: {tram_id} -> {status}")
            return True
            
        except Exception as e:
            logger.error(f"Excel logging error: {e}")
            return False
    
    # ==================== UNIFIED LOGGING ====================
    
    def log_status_change(self, tram_id: str, date: str, status: str, sistem: str = '', 
                         alt_sistem: str = '', aciklama: str = '', user_id: Optional[int] = None,
                         format: str = 'both') -> bool:
        """
        Log status change to both JSON and Excel (or specified format)
        format: 'json', 'excel', 'both'
        """
        if format in ('json', 'both'):
            json_ok = self.log_json(tram_id, date, status, sistem, alt_sistem, aciklama, user_id)
        else:
            json_ok = True
        
        if format in ('excel', 'both'):
            excel_ok = self.log_excel(tram_id, date, status, sistem, alt_sistem, aciklama, user_id)
        else:
            excel_ok = True
        
        return json_ok and excel_ok
    
    # ==================== DATA RETRIEVAL ====================
    
    def get_json_logs(self, limit: Optional[int] = None) -> List[Dict]:
        """Get logs from JSON file"""
        try:
            if not self.json_file.exists():
                return []
            
            with open(self.json_file, 'r', encoding='utf-8') as f:
                logs = json.load(f)
            
            if limit:
                return logs[-limit:]
            return logs
            
        except Exception as e:
            logger.error(f"Error reading JSON logs: {e}")
            return []
    
    def get_excel_logs(self) -> pd.DataFrame:
        """Get logs from Excel file as DataFrame"""
        try:
            if not self.excel_file.exists():
                return pd.DataFrame()
            
            return pd.read_excel(self.excel_file)
            
        except Exception as e:
            logger.error(f"Error reading Excel logs: {e}")
            return pd.DataFrame()
    
    def get_status_history(self, tram_id: str, limit: Optional[int] = None) -> List[Dict]:
        """Get status history for specific equipment"""
        logs = self.get_json_logs()
        tram_logs = [log for log in logs if log.get('tram_id') == str(tram_id)]
        
        if limit:
            return tram_logs[-limit:]
        return tram_logs
    
    def get_latest_status(self, tram_id: str) -> Optional[Dict]:
        """Get latest status for equipment"""
        history = self.get_status_history(tram_id, limit=1)
        return history[0] if history else None
    
    # ==================== STATISTICS ====================
    
    def get_downtime_duration(self, tram_id: str, start_date: str, end_date: str) -> float:
        """Calculate downtime duration in hours for equipment during period"""
        try:
            history = self.get_status_history(tram_id)
            
            # Filter by date range
            period_logs = [
                log for log in history
                if start_date <= log.get('date', '') <= end_date
            ]
            
            # Calculate total downtime
            downtime_hours = 0
            for log in period_logs:
                if 'Dışı' in log.get('status', ''):
                    # Approximate as 1 hour per log entry for now
                    downtime_hours += 1
            
            return downtime_hours
            
        except Exception as e:
            logger.error(f"Error calculating downtime: {e}")
            return 0
    
    # ==================== HELPER METHODS ====================
    
    @staticmethod
    def _create_excel_headers(ws) -> None:
        """Create Excel file headers"""
        headers = ['Timestamp', 'Equipment ID', 'Date', 'Status', 'System', 'Subsystem', 'Description', 'User ID']
        
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Set column widths
        widths = [20, 14, 12, 20, 20, 20, 30, 12]
        for col_idx, width in enumerate(widths, 1):
            ws.column_dimensions[chr(64 + col_idx)].width = width
    
    @staticmethod
    def _get_status_fill(status: str) -> Optional[PatternFill]:
        """Get fill color for status"""
        if 'In Service' in status or 'Servis' in status:
            if 'Out' in status or 'Dışı' in status:
                if 'Maintenance' in status or 'İşletme' in status:
                    return PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')  # Orange
                else:
                    return PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')  # Red
            else:
                return PatternFill(start_color='00B050', end_color='00B050', fill_type='solid')  # Green
        return None


# ==================== BACKWARD COMPATIBILITY ====================

# Export for backward compatibility
def log_service_status(project_code: str, tram_id: str, date: str, status: str,
                      sistem: str = '', alt_sistem: str = '', user_id: Optional[int] = None) -> bool:
    """Backward compatible function for existing code"""
    logger = ServiceStatusLogger(project_code)
    return logger.log_status_change(tram_id, date, status, sistem, alt_sistem, '', user_id)
