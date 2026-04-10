#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Root Cause Analysis Utility - Sistem ve Alt Sistem analizi
Tramvayların servis dışı kaldığı nedenleri derinlemesine analiz eder
"""
from collections import defaultdict, Counter
from datetime import datetime, timedelta
from pathlib import Path
import json

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, 
    numbers as xl_numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from models import ServiceStatus


class RootCauseAnalyzer:
    """Servis durumu verilerini analiz ederek root cause raporları üretir"""
    
    @staticmethod
    def analyze_service_disruptions(start_date=None, end_date=None, tram_id=None, project_code=None):
        """
        Servis dışı kalışları sistem/alt sistem bazında analiz eder
        
        Args:
            start_date: YYYY-MM-DD formatında başlangıç tarihi
            end_date: YYYY-MM-DD formatında bitiş tarihi
            tram_id: Belirli bir araç için analiz (None = tüm araçlar)
            project_code: Proje kodu (None = tüm projeler)
            
        Returns:
            dict: Analiz sonuçları
        """
        
        # Tarih aralığı varsayılanları
        if not end_date:
            end_date = datetime.now().strftime('%Y-%m-%d')
        if not start_date:
            start_date = (datetime.now() - timedelta(days=90)).strftime('%Y-%m-%d')
        
        # Servis dışı kayıtları sor
        query = ServiceStatus.query.filter(
            ServiceStatus.date >= start_date,
            ServiceStatus.date <= end_date,
            ServiceStatus.status.in_([
                'Servis Dışı', 
                'İşletme Kaynaklı Servis Dışı',
                'Servis Disi',  # ASCII version
                'Isletme Kaynaklı Servis Disi'  # ASCII version
            ])
        )
        
        if project_code:
            query = query.filter_by(project_code=project_code)
        
        if tram_id:
            query = query.filter_by(tram_id=tram_id)
        
        records = query.all()
        
        # Analiz yapıları
        sistem_analysis = defaultdict(lambda: {
            'count': 0,
            'days': 0,
            'tramvays': set(),
            'alt_sistemler': defaultdict(lambda: {'count': 0, 'days': 0})
        })
        
        tram_analysis = defaultdict(lambda: {
            'total_days_out': 0,
            'issues': [],
            'sistemler': set()
        })
        
        # Her kayıt için bir gün olarak sayıyoruz
        for record in records:
            sistem = record.sistem or 'Belirtilmedi'
            alt_sistem = record.alt_sistem or 'İçeri'
            
            # Sistem analizi
            sistem_analysis[sistem]['count'] += 1
            sistem_analysis[sistem]['days'] += 1
            sistem_analysis[sistem]['tramvays'].add(record.tram_id)
            sistem_analysis[sistem]['alt_sistemler'][alt_sistem]['count'] += 1
            sistem_analysis[sistem]['alt_sistemler'][alt_sistem]['days'] += 1
            
            # Araç analizi
            tram_analysis[record.tram_id]['total_days_out'] += 1
            tram_analysis[record.tram_id]['issues'].append({
                'date': record.date,
                'sistem': sistem,
                'alt_sistem': alt_sistem,
                'status': record.status,
                'aciklama': record.aciklama
            })
            tram_analysis[record.tram_id]['sistemler'].add(sistem)
        
        # Set'leri listeye çevir
        for sistem in sistem_analysis:
            sistem_analysis[sistem]['tramvays'] = list(sistem_analysis[sistem]['tramvays'])
            sistem_analysis[sistem]['affected_tram_count'] = len(sistem_analysis[sistem]['tramvays'])
        
        return {
            'start_date': start_date,
            'end_date': end_date,
            'total_records': len(records),
            'sistem_analysis': dict(sistem_analysis),
            'tram_analysis': dict(tram_analysis),
            'top_systems': sorted(
                sistem_analysis.items(),
                key=lambda x: x[1]['days'],
                reverse=True
            )[:10]
        }
    
    @staticmethod
    def generate_rca_excel(analysis_data, filename=None):
        """
        Root Cause Analysis raporunu Excel formatında oluşturur
        
        Args:
            analysis_data: analyze_service_disruptions() sonucu
            filename: Çıkış dosyası adı (None = logs/rca_{tarih}.xlsx)
            
        Returns:
            str: Oluşturulan dosyanın yolu
        """
        
        if not filename:
            now = datetime.now().strftime('%Y%m%d_%H%M%S')
            log_dir = Path('logs/root_cause_analysis')
            log_dir.mkdir(parents=True, exist_ok=True)
            filename = log_dir / f'rca_report_{now}.xlsx'
        else:
            filename = Path(filename)
        
        wb = Workbook()
        ws = wb.active
        ws.title = 'Root Cause Analysis'
        
        # Stiller
        header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=12)
        
        title_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        title_font = Font(bold=True, size=11)
        
        critical_fill = PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid')
        critical_font = Font(bold=True, color='FFFFFF')
        
        warning_fill = PatternFill(start_color='FFC107', end_color='FFC107', fill_type='solid')
        warning_font = Font(bold=True, color='000000')
        
        info_fill = PatternFill(start_color='E8F4F8', end_color='E8F4F8', fill_type='solid')
        info_font = Font(size=10, bold=True)
        
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left_align = Alignment(horizontal='left', vertical='top', wrap_text=True)
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Başlık
        ws['A1'] = 'ROOT CAUSE ANALYSIS - SERVIS DURUMU RAPORU'
        ws['A1'].font = Font(bold=True, size=14, color='FFFFFF')
        ws['A1'].fill = PatternFill(start_color='203864', end_color='203864', fill_type='solid')
        ws.merge_cells('A1:H1')
        ws['A1'].alignment = center_align
        ws.row_dimensions[1].height = 25
        
        # Analiz tarihi
        current_row = 3
        ws[f'A{current_row}'] = 'Analiz Tarihi:'
        ws[f'B{current_row}'] = datetime.now().strftime('%d.%m.%Y %H:%M:%S')
        ws[f'A{current_row}'].font = info_font
        
        current_row += 1
        ws[f'A{current_row}'] = 'Dönem:'
        ws[f'B{current_row}'] = f"{analysis_data['start_date']} → {analysis_data['end_date']}"
        ws[f'A{current_row}'].font = info_font
        
        current_row += 1
        ws[f'A{current_row}'] = 'Toplam Servis Dışı Gün:'
        ws[f'B{current_row}'] = analysis_data['total_records']
        ws[f'B{current_row}'].font = critical_font
        ws[f'B{current_row}'].fill = critical_fill
        ws[f'A{current_row}'].font = info_font
        
        # ===== SİSTEM BAZLI ANALİZ TABLOSU =====
        current_row += 3
        ws[f'A{current_row}'] = 'SİSTEM BAZLI ANALİZ'
        ws[f'A{current_row}'].font = title_font
        ws[f'A{current_row}'].fill = title_fill
        ws.merge_cells(f'A{current_row}:H{current_row}')
        
        current_row += 1
        headers = ['Sistem', 'Servis Dışı Gün', 'Olay Sayısı', 'Etkilenen Araç', 'Başlıca Alt Sistem', 'Yüzde', 'Risk Düzeyi', 'Öneri']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
        
        ws.row_dimensions[current_row].height = 20
        
        # Sistem verilerini sıralı şekilde ekle
        total_days = sum(v['days'] for v in analysis_data['sistem_analysis'].values()) or 1
        
        current_row += 1
        start_data_row = current_row
        
        for sistem, data in sorted(
            analysis_data['sistem_analysis'].items(),
            key=lambda x: x[1]['days'],
            reverse=True
        ):
            percentage = (data['days'] / total_days * 100) if total_days > 0 else 0
            
            # Risk seviyesi belirleme
            if percentage > 30:
                risk_level = 'KRITIK'
                risk_fill = critical_fill
                risk_font = critical_font
            elif percentage > 15:
                risk_level = 'YÜKSEK'
                risk_fill = warning_fill
                risk_font = warning_font
            else:
                risk_level = 'DÜŞÜK'
                risk_fill = PatternFill(start_color='D4EDDA', end_color='D4EDDA', fill_type='solid')
                risk_font = Font(bold=True, color='155724')
            
            # Alt sistemleri sırala
            top_alt_sistem = max(
                data['alt_sistemler'].items(),
                key=lambda x: x[1]['days'],
                default=('', {'days': 0})
            )[0]
            
            # Öneriler
            if risk_level == 'KRITIK':
                recommendation = 'Acil önlem alınmalı. Sistem gözden geçirilmesi gerekli.'
            elif risk_level == 'YÜKSEK':
                recommendation = 'Planlı bakım yapılmalı. Root cause analizi derinleştirilmesi gerekli.'
            else:
                recommendation = 'Düzenli izleme yapılmalı.'
            
            ws.cell(row=current_row, column=1).value = sistem
            ws.cell(row=current_row, column=2).value = data['days']
            ws.cell(row=current_row, column=3).value = data['count']
            ws.cell(row=current_row, column=4).value = data['affected_tram_count']
            ws.cell(row=current_row, column=5).value = top_alt_sistem
            ws.cell(row=current_row, column=6).value = percentage
            ws.cell(row=current_row, column=7).value = risk_level
            ws.cell(row=current_row, column=8).value = recommendation
            
            # Formatting
            for col in range(1, 9):
                cell = ws.cell(row=current_row, column=col)
                cell.border = thin_border
                if col == 7:  # Risk Düzeyi sütunu
                    cell.fill = risk_fill
                    cell.font = risk_font
                    cell.alignment = center_align
                elif col == 6:  # Yüzde
                    cell.number_format = '0.0"%"'
                    cell.alignment = center_align
                elif col in [2, 3, 4]:
                    cell.alignment = center_align
                else:
                    cell.alignment = left_align
            
            current_row += 1
        
        # ===== ALT SİSTEM DETAY TABLOSU =====
        current_row += 2
        ws[f'A{current_row}'] = 'ALT SİSTEM DETAY'
        ws[f'A{current_row}'].font = title_font
        ws[f'A{current_row}'].fill = title_fill
        ws.merge_cells(f'A{current_row}:D{current_row}')
        
        current_row += 1
        headers = ['Sistem', 'Alt Sistem', 'Olay Sayısı', 'Gün']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
        
        current_row += 1
        for sistem, data in sorted(
            analysis_data['sistem_analysis'].items(),
            key=lambda x: x[1]['days'],
            reverse=True
        ):
            for alt_sistem in sorted(
                data['alt_sistemler'].items(),
                key=lambda x: x[1]['days'],
                reverse=True
            ):
                ws.cell(row=current_row, column=1).value = sistem
                ws.cell(row=current_row, column=2).value = alt_sistem[0]
                ws.cell(row=current_row, column=3).value = alt_sistem[1]['count']
                ws.cell(row=current_row, column=4).value = alt_sistem[1]['days']
                
                for col in range(1, 5):
                    cell = ws.cell(row=current_row, column=col)
                    cell.border = thin_border
                    cell.alignment = left_align if col <= 2 else center_align
                
                current_row += 1
        
        # ===== EN ÇOK ETKILENEN ARAÇLAR =====
        current_row += 2
        ws[f'A{current_row}'] = 'EN ÇOK ETKILENEN ARAÇLAR'
        ws[f'A{current_row}'].font = title_font
        ws[f'A{current_row}'].fill = title_fill
        ws.merge_cells(f'A{current_row}:C{current_row}')
        
        current_row += 1
        headers = ['Araç ID', 'Servis Dışı Gün', 'Başlıca Sistemler']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
        
        current_row += 1
        for tram_id, data in sorted(
            analysis_data['tram_analysis'].items(),
            key=lambda x: x[1]['total_days_out'],
            reverse=True
        )[:15]:
            ws.cell(row=current_row, column=1).value = tram_id
            ws.cell(row=current_row, column=2).value = data['total_days_out']
            ws.cell(row=current_row, column=3).value = ', '.join(sorted(data['sistemler']))
            
            for col in range(1, 4):
                cell = ws.cell(row=current_row, column=col)
                cell.border = thin_border
                cell.alignment = center_align if col == 2 else left_align
            
            current_row += 1
        
        # Sütun genişlikleri
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 40
        
        wb.save(filename)
        return str(filename)


# Test
if __name__ == '__main__':
    print("RootCauseAnalyzer utility loaded successfully")
    print("Available methods:")
    print("  - analyze_service_disruptions(start_date, end_date, tram_id)")
    print("  - generate_rca_excel(analysis_data, filename)")
