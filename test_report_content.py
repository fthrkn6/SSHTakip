from routes.service_status import bp
from app import create_app
from models import db
from openpyxl import load_workbook

app = create_app()

# Rapor oluştur
app.test_client().get('/servis/excel/comprehensive-report')

# Son oluşturulan Excel'i bul
import os
from pathlib import Path

report_dir = Path('logs/rapor_cikti')
if report_dir.exists():
    files = list(report_dir.glob('*.xlsx'))
    if files:
        latest_file = max(files, key=lambda f: f.stat().st_mtime)
        print(f"Latest report: {latest_file}")
        print(f"Created: {latest_file.stat().st_mtime}")
        
        # Excel'i oku
        wb = load_workbook(latest_file)
        ws = wb.active
        
        print(f"\nSheet name: {ws.title}")
        print(f"Max row: {ws.max_row}")
        print("\nFirst 5 rows:")
        for row in ws.iter_rows(min_row=1, max_row=6, values_only=True):
            print(row)
