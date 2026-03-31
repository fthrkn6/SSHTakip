"""
KM Takip Excel - Matris formatında KM veri takip dosyası
Satırlar: Araçlar (tramvay ID)
Sütunlar: Tarihler (veri girilen günler)
Hücreler: O günkü KM değeri
Renklendirme: Artış yeşil, değişim yok sarı, azalış kırmızı, yeni giriş mavi
"""

import os
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
import pytz

TR_TZ = pytz.timezone('Europe/Istanbul')

# Renk paleti
HEADER_FILL = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
HEADER_FONT = Font(bold=True, color='FFFFFF', size=11, name='Calibri')
TRAM_FILL = PatternFill(start_color='D6DCE4', end_color='D6DCE4', fill_type='solid')
TRAM_FONT = Font(bold=True, size=11, name='Calibri')

# Hücre renkleri
GREEN_FILL = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')     # KM arttı
YELLOW_FILL = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')    # KM aynı
RED_FILL = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')       # KM azaldı
BLUE_FILL = PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid')      # İlk giriş
WHITE_FILL = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

GREEN_FONT = Font(color='006100', size=10, name='Calibri')
YELLOW_FONT = Font(color='9C6500', size=10, name='Calibri')
RED_FONT = Font(color='9C0006', bold=True, size=10, name='Calibri')
BLUE_FONT = Font(color='1F4E79', size=10, name='Calibri')
NORMAL_FONT = Font(size=10, name='Calibri')

THIN_BORDER = Border(
    left=Side(style='thin', color='B4C6E7'),
    right=Side(style='thin', color='B4C6E7'),
    top=Side(style='thin', color='B4C6E7'),
    bottom=Side(style='thin', color='B4C6E7')
)

LOG_DIR = 'logs/km'


def _get_takip_path(project_code):
    """Proje için takip Excel dosya yolunu döndür."""
    return os.path.join(LOG_DIR, f'{project_code}_km_takip.xlsx')


def _create_new_workbook(path):
    """Yeni takip dosyası oluştur."""
    os.makedirs(os.path.dirname(path), exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = 'KM Takip'

    # A1 hücresi
    cell = ws.cell(row=1, column=1, value='Araç No')
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = THIN_BORDER
    ws.column_dimensions['A'].width = 14

    # Lejant sayfası
    ws_legend = wb.create_sheet('Lejant')
    legend_data = [
        ('Renk', 'Anlam'),
        ('Yeşil', 'KM arttı (normal)'),
        ('Sarı', 'KM değişmedi'),
        ('Kırmızı', 'KM azaldı (düzeltme)'),
        ('Mavi', 'İlk giriş'),
    ]
    fills = [HEADER_FILL, GREEN_FILL, YELLOW_FILL, RED_FILL, BLUE_FILL]
    fonts = [HEADER_FONT, GREEN_FONT, YELLOW_FONT, RED_FONT, BLUE_FONT]
    for i, (renk, anlam) in enumerate(legend_data, 1):
        c1 = ws_legend.cell(row=i, column=1, value=renk)
        c2 = ws_legend.cell(row=i, column=2, value=anlam)
        c1.fill = fills[i - 1]
        c1.font = fonts[i - 1]
        c2.font = fonts[i - 1] if i > 0 else NORMAL_FONT
        c1.border = THIN_BORDER
        c2.border = THIN_BORDER
    ws_legend.column_dimensions['A'].width = 14
    ws_legend.column_dimensions['B'].width = 30

    wb.save(path)
    return wb


def _find_tram_row(ws, tram_id):
    """Araç satırını bul, yoksa None döndür."""
    tram_str = str(tram_id).strip()
    for row_idx in range(2, ws.max_row + 1):
        val = ws.cell(row=row_idx, column=1).value
        if val is not None and str(val).strip() == tram_str:
            return row_idx
    return None


def _find_date_col(ws, date_str):
    """Tarih sütununu bul, yoksa None döndür."""
    for col_idx in range(2, ws.max_column + 1):
        val = ws.cell(row=1, column=col_idx).value
        if val is not None and str(val).strip() == date_str:
            return col_idx
    return None


def _add_tram_row(ws, tram_id):
    """Yeni araç satırı ekle (sıralı)."""
    tram_str = str(tram_id).strip()
    new_row = ws.max_row + 1

    # Sıralı ekleme: mevcut araçları tara, doğru yere yerleştir
    insert_at = None
    for row_idx in range(2, ws.max_row + 1):
        existing = ws.cell(row=row_idx, column=1).value
        if existing is not None and str(existing).strip() > tram_str:
            insert_at = row_idx
            break

    if insert_at:
        ws.insert_rows(insert_at)
        target_row = insert_at
    else:
        target_row = new_row

    cell = ws.cell(row=target_row, column=1, value=tram_str)
    cell.fill = TRAM_FILL
    cell.font = TRAM_FONT
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = THIN_BORDER
    return target_row


def _add_date_col(ws, date_str):
    """Yeni tarih sütunu ekle (sona)."""
    new_col = ws.max_column + 1
    cell = ws.cell(row=1, column=new_col, value=date_str)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = THIN_BORDER
    ws.column_dimensions[get_column_letter(new_col)].width = 13
    return new_col


def _get_previous_km(ws, row_idx, col_idx):
    """Bir önceki tarih sütunundaki KM değerini bul."""
    for c in range(col_idx - 1, 1, -1):
        val = ws.cell(row=row_idx, column=c).value
        if val is not None:
            try:
                return int(float(val))
            except (ValueError, TypeError):
                continue
    return None


def _apply_cell_style(cell, km_value, previous_km):
    """Hücreye KM değişimine göre renk uygula."""
    cell.number_format = '#,##0'
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = THIN_BORDER

    if previous_km is None:
        # İlk giriş
        cell.fill = BLUE_FILL
        cell.font = BLUE_FONT
    elif km_value > previous_km:
        # KM arttı - normal
        cell.fill = GREEN_FILL
        cell.font = GREEN_FONT
    elif km_value == previous_km:
        # KM aynı
        cell.fill = YELLOW_FILL
        cell.font = YELLOW_FONT
    else:
        # KM azaldı - dikkat
        cell.fill = RED_FILL
        cell.font = RED_FONT


def log_km_takip(project_code, tram_id, km_value):
    """
    KM takip matrisine yeni veri ekle.

    Args:
        project_code: Proje kodu (belgrad, kayseri, vb.)
        tram_id: Araç kodu
        km_value: Güncel KM değeri

    Returns:
        bool: Başarılı ise True
    """
    try:
        path = _get_takip_path(project_code)

        # Dosyayı aç veya oluştur
        if os.path.exists(path):
            wb = load_workbook(path)
            ws = wb['KM Takip']
        else:
            wb = _create_new_workbook(path)
            ws = wb['KM Takip']

        # Bugünün tarihi
        now = datetime.now(TR_TZ)
        date_str = now.strftime('%d.%m.%Y')

        # Tarih sütununu bul veya oluştur
        col_idx = _find_date_col(ws, date_str)
        if col_idx is None:
            col_idx = _add_date_col(ws, date_str)

        # Araç satırını bul veya oluştur
        row_idx = _find_tram_row(ws, tram_id)
        if row_idx is None:
            row_idx = _add_tram_row(ws, tram_id)
            # Satır eklendiyse, tarih sütununu tekrar bul (insert_rows kayabilir)
            col_idx = _find_date_col(ws, date_str)

        # KM değerini yaz
        km_int = int(float(km_value))
        cell = ws.cell(row=row_idx, column=col_idx, value=km_int)

        # Önceki değeri bul ve renk uygula
        previous_km = _get_previous_km(ws, row_idx, col_idx)
        _apply_cell_style(cell, km_int, previous_km)

        # Freeze panes: ilk satır ve ilk sütun sabit
        ws.freeze_panes = 'B2'

        wb.save(path)
        return True

    except Exception as e:
        import logging
        logging.getLogger(__name__).warning(f'KM takip Excel hatasi: {e}')
        return False


def get_km_takip_path(project_code):
    """Dışarıdan erişim için dosya yolu döndür."""
    return _get_takip_path(project_code)
