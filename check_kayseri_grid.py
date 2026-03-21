from openpyxl import load_workbook
import os
from datetime import datetime, date

project = 'kayseri'
base_path = r'c:\Users\fatiherkin\Desktop\bozankaya_ssh_takip\data'
grid_path = os.path.join(base_path, project, 'service_status_grid.xlsx')

wb = load_workbook(grid_path, data_only=True)
ws = wb.active

print("\n" + "="*90)
print(f"KAYSERI GRID - İÇERİK ANALİZİ")
print("="*90 + "\n")

# Başlık
print("BAŞLIK SATIRI (Sütunlar):")
header = []
for col_idx in range(1, ws.max_column + 1):
    cell_val = ws.cell(row=1, column=col_idx).value
    header.append(cell_val)
    print(f"  Col {col_idx}: {cell_val}")

# Tarih aralığını kontrol et
print(f"\nTARİH ARALIGI:")
print(f"  Toplam Satır Sayısı: {ws.max_row}")

# İlk 5 tarih
print(f"\n  İlk 5 tarih (satır 2-6):")
for row_idx in range(2, 7):
    date_val = ws.cell(row=row_idx, column=1).value
    print(f"    Satır {row_idx}: {date_val}")

# Son 5 tarih
print(f"\n  Son 5 tarih (satır {ws.max_row-4} - {ws.max_row}):")
for row_idx in range(ws.max_row - 4, ws.max_row + 1):
    date_val = ws.cell(row=row_idx, column=1).value
    print(f"    Satır {row_idx}: {date_val}")

# Bugün (2026-03-21) için veri kontrol et
print(f"\nBUGÜNÜN (2026-03-21) VERİSİ:")
today = date.today()
found_today = False

for row_idx in range(2, ws.max_row + 1):
    date_cell = ws.cell(row=row_idx, column=1).value
    
    if date_cell:
        if isinstance(date_cell, str):
            try:
                cell_date = datetime.strptime(date_cell, '%Y-%m-%d').date()
            except:
                continue
        else:
            cell_date = date_cell.date() if hasattr(date_cell, 'date') else date_cell
        
        if cell_date == today:
            found_today = True
            print(f"  ✓ BULUNDU! Satır {row_idx}")
            print(f"    Tarih: {date_cell}")
            print("    Araçlar:")
            for col_idx in range(2, ws.max_column + 1):
                status = ws.cell(row=row_idx, column=col_idx).value
                equipment = header[col_idx - 1]
                print(f"      {equipment}: {status}")
            break

if not found_today:
    print(f"  ❌ BULUNAMADI - 2026-03-21 tarihi grid'de yok!")
    print(f"\n  Disponibilitede var olan tarihler (yakın zamanda):")
    count = 0
    for row_idx in range(ws.max_row, max(1, ws.max_row - 30), -1):
        date_cell = ws.cell(row=row_idx, column=1).value
        if date_cell:
            print(f"    Satır {row_idx}: {date_cell}")
            count += 1
            if count >= 5:
                break

# İçeriği kontrol et (boş hücreler)
print(f"\nVERİ KONTROLÜ:")
non_empty = 0
empty = 0
for row_idx in range(2, ws.max_row + 1):
    for col_idx in range(2, ws.max_column + 1):
        if ws.cell(row=row_idx, column=col_idx).value:
            non_empty += 1
        else:
            empty += 1

total = non_empty + empty
fill_rate = (non_empty / total * 100) if total > 0 else 0
print(f"  Toplam Hücre: {total}")
print(f"  Dolu Hücre: {non_empty} ({fill_rate:.1f}%)")
print(f"  Boş Hücre: {empty} ({100-fill_rate:.1f}%)")

wb.close()
print("\n" + "="*90 + "\n")
