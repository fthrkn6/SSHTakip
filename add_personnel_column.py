#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""Add Personnel Count column to Excel file using openpyxl"""

from openpyxl import load_workbook
import os

excel_file = r'c:\Users\ferki\Desktop\bozankaya_ssh_takip\logs\ariza_listesi\Ariza_Listesi_BELGRAD.xlsx'

if os.path.exists(excel_file):
    try:
        # Load workbook
        wb = load_workbook(excel_file)
        ws = wb['Ariza Listesi']
        
        # Find the last column
        max_col = ws.max_column
        new_col = max_col + 1
        
        # Add header in row 4 (header row)
        ws.cell(row=4, column=new_col, value='Personel Sayısı')
        
        print(f"✅ 'Personel Sayısı' sütunu eklendi (Sütun {new_col})")
        
        # Save the file
        wb.save(excel_file)
        
        print(f"✅ Excel dosyası başarıyla güncellendi!")
        print(f"Dosya: {excel_file}")
        
        # Verify
        wb_check = load_workbook(excel_file)
        ws_check = wb_check['Ariza Listesi']
        print(f"\nMax sütun: {ws_check.max_column}")
        print(f"Son sütun adı: {ws_check.cell(row=4, column=ws_check.max_column).value}")
        
    except Exception as e:
        print(f"❌ Hata: {str(e)}")
else:
    print(f"❌ Dosya bulunamadı: {excel_file}")
