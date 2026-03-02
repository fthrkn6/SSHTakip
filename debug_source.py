import os
from openpyxl import load_workbook

log = open('debug.log', 'w')

try:
    log.write("Starting debug...\n")
    from app import app
    log.write("App loaded\n")
    
    project = 'belgrad'
    fracas_file = os.path.join(app.root_path, 'logs', project, 'ariza_listesi', f'Fracas_{project.upper()}.xlsx')
    
    log.write(f'Dosya: {fracas_file}\n')
    log.write(f'Var mı: {os.path.exists(fracas_file)}\n')
    
    if os.path.exists(fracas_file):
        wb = load_workbook(fracas_file, data_only=True)
        ws = wb['FRACAS']
        
        # Last 3 IDs
        log.write(f'\nSon 3 ID:\n')
        for row in range(max(5, ws.max_row - 2), ws.max_row + 1):
            val = ws.cell(row=row, column=5).value
            log.write(f"  Row {row}: {val}\n")
    else:
        log.write("Dosya BULUNAMADI!\n")
        
except Exception as e:
    log.write(f"Hata: {e}\n")
    import traceback
    traceback.print_exc(file=log)
finally:
    log.close()
