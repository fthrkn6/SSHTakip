"""
Improved Excel Report Generator - Profesyonel raporlar
Güzel formatlandırılmış, veri-dolu Excel çıktıları
"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from pathlib import Path
import os

class ProfessionalExcelReporter:
    """Profesyonel Excel raporları oluştur"""
    
    def __init__(self, filename, project_name='Belgrad'):
        self.filename = filename
        self.project_name = project_name
        self.wb = Workbook()
        self.wb.remove(self.wb.active)  # Varsayılan sheet'i kaldır
        
        # Stiller
        self.header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
        self.header_font = Font(color="FFFFFF", bold=True, size=12)
        self.title_font = Font(size=14, bold=True, color="1F497D")
        self.subheader_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        self.subheader_font = Font(bold=True, color="1F497D", size=11)
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Renk kodları
        self.color_aktif = "92D050"      # Yeşil - Aktif
        self.color_isletme = "FFC000"    # Sarı - İşletme
        self.color_ariza = "FF0000"      # Kırmızı - Arız
    
    def add_title_sheet(self):
        """Başlık sayfası ekle"""
        ws = self.wb.create_sheet("Özet", 0)
        
        ws['A1'] = f"Bozankaya Hafif Raylı Sistem - {self.project_name}"
        ws['A1'].font = Font(size=16, bold=True, color="1F497D")
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells('A1:E1')
        ws.row_dimensions[1].height = 25
        
        ws['A3'] = "Günlük Servis Durumu & Availability Raporu"
        ws['A3'].font = self.title_font
        ws.merge_cells('A3:E3')
        
        ws['A4'] = f"Rapor Tarihi: {datetime.now().strftime('%d.%m.%Y %H:%M:%S')}"
        ws['A4'].font = Font(italic=True)
        
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 20
    
    def add_service_status_sheet(self, data):
        """Servis durumu özeti ekle"""
        ws = self.wb.create_sheet("Servis Durumu")
        
        # Başlık
        ws['A1'] = "Güne İtibarla Servis Durumu"
        ws['A1'].font = self.title_font
        ws.merge_cells('A1:E1')
        
        # Headers
        headers = ['Tramvay', 'Durum', 'Sistem', 'Alt Sistem', 'Son Güncelleme']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.border = self.border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Veriler
        row = 4
        for item in data:
            ws.cell(row=row, column=1).value = item['tram_id']
            ws.cell(row=row, column=2).value = item['status']
            
            # Durum rengini ekle
            status = item['status'].lower()
            if 'aktif' in status or 'işletme' in status:
                color = self.color_aktif if 'aktif' in status else self.color_isletme
            else:
                color = self.color_ariza
            
            ws.cell(row=row, column=2).fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            
            ws.cell(row=row, column=3).value = item.get('sistem', '')
            ws.cell(row=row, column=4).value = item.get('alt_sistem', '')
            ws.cell(row=row, column=5).value = item.get('updated_at', '')
            
            for col in range(1, 6):
                ws.cell(row=row, column=col).border = self.border
                ws.cell(row=row, column=col).alignment = Alignment(horizontal='left', vertical='center')
            
            row += 1
        
        # Kolonu genişlik
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 18
    
    def add_availability_sheet(self, equipment_list):
        """Availability tablosu ekle"""
        ws = self.wb.create_sheet("Availability")
        
        # Başlık
        ws['A1'] = "Günlük Availability Metrikleri"
        ws['A1'].font = self.title_font
        ws.merge_cells('A1:H1')
        
        # Headers
        headers = ['Tramvay', 'Durum', 'Operatif Saatler', 'Kapalı Saatler', 'Availability %', 'Arıza Sayısı', 'Notlar']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.border = self.border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        ws.row_dimensions[3].height = 25
        
        # Veriler
        row = 4
        for item in equipment_list:
            ws.cell(row=row, column=1).value = item['equipment_code']
            ws.cell(row=row, column=2).value = item['status_badge']
            ws.cell(row=row, column=3).value = item['operational']
            ws.cell(row=row, column=4).value = item['downtime']
            ws.cell(row=row, column=5).value = item['availability']
            ws.cell(row=row, column=6).value = 0  # Failure count
            ws.cell(row=row, column=7).value = item.get('status_record').aciklama if item.get('status_record') else ''
            
            # Availability'ye göre renk kodu
            availability = item['availability']
            if availability >= 80:
                color = self.color_aktif
            elif availability >= 50:
                color = self.color_isletme
            else:
                color = self.color_ariza
            
            ws.cell(row=row, column=5).fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            
            for col in range(1, 8):
                ws.cell(row=row, column=col).border = self.border
                ws.cell(row=row, column=col).alignment = Alignment(horizontal='center', vertical='center')
            
            row += 1
        
        # Kolonu genişlik
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 16
        ws.column_dimensions['D'].width = 16
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 25
    
    def add_summary_stats(self, stats):
        """Özet istatistikleri ekle"""
        ws = self.wb.create_sheet("İstatistikler", 1)
        
        # Başlık
        ws['A1'] = "Günlük İstatistikler"
        ws['A1'].font = self.title_font
        ws.merge_cells('A1:C1')
        
        # Veriler
        row = 3
        stats_items = [
            ('Toplam Araç', stats.get('toplam', 0)),
            ('Aktif Araç', stats.get('aktif', 0)),
            ('İşletme Kaynaklı', stats.get('isletme', 0)),
            ('Servis Dışı', stats.get('servis_disi', 0)),
            ('Ortalama Availability', f"{stats.get('availability', 0):.1f}%")
        ]
        
        for label, value in stats_items:
            ws.cell(row=row, column=1).value = label
            ws.cell(row=row, column=1).font = self.subheader_font
            ws.cell(row=row, column=1).fill = self.subheader_fill
            
            ws.cell(row=row, column=2).value = value
            ws.cell(row=row, column=2).font = Font(size=11, bold=True)
            
            ws.cell(row=row, column=1).border = self.border
            ws.cell(row=row, column=2).border = self.border
            
            row += 1
        
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
    
    def save(self):
        """Dosyayı kaydet"""
        self.wb.save(self.filename)
        print(f"✅ Excel raporu kaydedildi: {self.filename}")

# Helper fonksiyonlar
def generate_service_status_report(equipment_list, stats, project_name='Belgrad', output_path=None):
    """Servis durumu raporu oluştur"""
    if output_path is None:
        output_path = f"reports/servis_durumu_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    
    reporter = ProfessionalExcelReporter(output_path, project_name)
    reporter.add_title_sheet()
    reporter.add_summary_stats(stats)
    reporter.add_availability_sheet(equipment_list)
    reporter.save()
    
    return output_path
