from flask import Flask
from app import app
import sys

# Test: sistemler_data'nın doğru render edilip edilmediğini kontrol et
with app.app_context():
    from models import db
    from flask_sqlalchemy import SQLAlchemy
    import os
    from openpyxl import load_workbook
    from datetime import datetime
    
    # Sistemleri yükle (app.py'daki logic'le aynı)
    sistemler = {}
    data_dir = os.path.join(os.path.dirname(__file__), 'data', 'belgrad')
    
    veriler_path = None
    if os.path.exists(data_dir):
        for file in os.listdir(data_dir):
            if 'veriler' in file.lower() and file.endswith('.xlsx'):
                veriler_path = os.path.join(data_dir, file)
                break
    
    if veriler_path and os.path.exists(veriler_path):
        try:
            wb = load_workbook(veriler_path)
            ws = wb['Sayfa2']
            
            # Renk tanımları
            KIRMIZI = 'FFFF0000'
            SARI = 'FFFFFF00'
            MAVI = 'FF0070C0'
            
            # Sütunları tarayarak sistemleri bul
            sütunlar = [6, 7, 8, 9, 11]
            
            for sütun in sütunlar:
                sistem_adi = None
                
                for row in range(1, ws.max_row + 1):
                    cell = ws.cell(row=row, column=sütun)
                    value = cell.value
                    fill = cell.fill
                    
                    color_hex = None
                    if fill and fill.start_color:
                        color_hex = str(fill.start_color.rgb) if fill.start_color.rgb else None
                    
                    if color_hex == KIRMIZI and value:
                        sistem_adi = str(value).strip()
                        if sistem_adi not in sistemler:
                            sistemler[sistem_adi] = {
                                'tedarikçiler': set(),
                                'alt_sistemler': set()
                            }
                    
                    elif color_hex == SARI and value and sistem_adi:
                        sistemler[sistem_adi]['tedarikçiler'].add(str(value).strip())
                    
                    elif color_hex == MAVI and value and sistem_adi:
                        sistemler[sistem_adi]['alt_sistemler'].add(str(value).strip())
            
            sistemler = {
                k: {
                    'tedarikçiler': sorted(list(v['tedarikçiler'])),
                    'alt_sistemler': sorted(list(v['alt_sistemler']))
                }
                for k, v in sistemler.items()
            }
        except Exception as e:
            print(f"Hata: {e}")
            sistemler = {}
    
    print("Python'da sistemler_data:")
    print("=" * 60)
    import json
    print(json.dumps(sistemler, ensure_ascii=False, indent=2))
    
    print("\n" + "=" * 60)
    print("JavaScript'e aktarılacak JSON:")
    print("=" * 60)
    
    # JavaScript'e aktarılacak JSON'u oluştur
    js_json = "{"
    first_sistem = True
    for sistem_adi, sistem_data in sorted(sistemler.items()):
        if not first_sistem:
            js_json += ", "
        first_sistem = False
        
        js_json += f'"{sistem_adi}": {{'
        
        js_json += f'"tedarikçiler": ['
        first_t = True
        for t in sistem_data['tedarikçiler']:
            if not first_t:
                js_json += ', '
            first_t = False
            js_json += f'"{t}"'
        js_json += '], '
        
        js_json += f'"alt_sistemler": ['
        first_a = True
        for a in sistem_data['alt_sistemler']:
            if not first_a:
                js_json += ', '
            first_a = False
            js_json += f'"{a}"'
        js_json += ']'
        
        js_json += '}'
    js_json += "}"
    
    print(js_json)
