from openpyxl import load_workbook

data_dir = 'data/belgrad'
veriler_path = f'{data_dir}/Veriler.xlsx'

wb = load_workbook(veriler_path)
ws = wb['Sayfa2']

print('EXCEL YAPISINI KONTROL EDIYORUM:')
print('=' * 100)

# İlk 25 satırda renkleri kontrol et
for row in range(1, min(25, ws.max_row + 1)):
    for col in range(6, 12):  # F, G, H, I, J, K sütunları
        cell = ws.cell(row=row, column=col)
        value = cell.value
        fill = cell.fill
        
        if fill and fill.start_color:
            color_hex = str(fill.start_color.rgb) if fill.start_color.rgb else 'None'
        else:
            color_hex = 'None'
        
        col_letter = chr(64 + col)
        val_str = str(value)[:20] if value else '-'
        if value:
            print(f'Satır {row:2} {col_letter}: {val_str:<20} | Renk: {color_hex}')
