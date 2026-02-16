"""
Servis Durumu Log Sistemi
Her servis durumu değişikliğini JSON ve Excel format'ında kayıt tutar
"""

import json
import os
from datetime import datetime
from pathlib import Path
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


class ServiceStatusLogger:
    """Servis durumu değişikliklerini log'la"""
    
    BASE_LOG_DIR = 'logs/service_status_history'
    
    @staticmethod
    def ensure_log_directory():
        """Log klasörünü oluştur"""
        os.makedirs(ServiceStatusLogger.BASE_LOG_DIR, exist_ok=True)
    
    @staticmethod
    def get_log_file_path(file_format='json'):
        """Log dosyası yolunu döndür"""
        ServiceStatusLogger.ensure_log_directory()
        
        if file_format == 'json':
            filename = f"service_status_log_{datetime.now().strftime('%Y-%m-%d')}.json"
        elif file_format == 'excel':
            filename = f"service_status_log_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
        else:
            raise ValueError(f"Unknown format: {file_format}")
        
        return os.path.join(ServiceStatusLogger.BASE_LOG_DIR, filename)
    
    @staticmethod
    def log_json(tram_id, date, status, sistem='', alt_sistem='', aciklama='', user_id=None):
        """
        Servis durumunu JSON'a log'la
        Mevcut dosyaya satır ekle veya yeni dosya oluştur
        """
        try:
            log_file = ServiceStatusLogger.get_log_file_path('json')
            
            # Mevcut logu oku
            if os.path.exists(log_file):
                with open(log_file, 'r', encoding='utf-8') as f:
                    logs = json.load(f)
            else:
                logs = []
            
            # Yeni kayıt ekle
            new_log = {
                'timestamp': datetime.now().isoformat(),
                'tram_id': str(tram_id),
                'date': str(date),
                'status': status,
                'sistem': sistem,
                'alt_sistem': alt_sistem,
                'aciklama': aciklama,
                'user_id': user_id
            }
            logs.append(new_log)
            
            # Fichier'a yaz
            with open(log_file, 'w', encoding='utf-8') as f:
                json.dump(logs, f, indent=2, ensure_ascii=False, default=str)
            
            return True
        except Exception as e:
            print(f"❌ JSON log error: {str(e)}")
            return False
    
    @staticmethod
    def log_excel(tram_id, date, status, sistem='', alt_sistem='', aciklama='', user_id=None):
        """
        Servis durumunu Excel'e log'la
        Mevcut dosyaya satır ekle veya yeni dosya oluştur
        """
        try:
            log_file = ServiceStatusLogger.get_log_file_path('excel')
            
            timestamp = datetime.now().strftime('%d.%m.%Y %H:%M:%S')
            
            # Mevcut dosyayı kontrol et
            if os.path.exists(log_file):
                # Yükle ve satır ekle
                wb = load_workbook(log_file)
                ws = wb.active
                
                # Sonraki boş satırı bul
                next_row = ws.max_row + 1
            else:
                # Yeni dosya oluştur
                wb = Workbook()
                ws = wb.active
                ws.title = 'Servis Durumu Log'
                next_row = 1
                
                # Header ekle
                headers = ['Timestamp', 'Tramvay', 'Tarih', 'Durum', 'Sistem', 'Alt Sistem', 'Açıklama', 'User ID']
                
                header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                header_font = Font(bold=True, color='FFFFFF')
                border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                for col_idx, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_idx)
                    cell.value = header
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.border = border
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                
                ws.column_dimensions['A'].width = 20  # Timestamp
                ws.column_dimensions['B'].width = 12  # Tramvay
                ws.column_dimensions['C'].width = 12  # Tarih
                ws.column_dimensions['D'].width = 20  # Durum
                ws.column_dimensions['E'].width = 20  # Sistem
                ws.column_dimensions['F'].width = 20  # Alt Sistem
                ws.column_dimensions['G'].width = 30  # Açıklama
                ws.column_dimensions['H'].width = 10  # User ID
                
                next_row = 2
            
            # Durum rengi belirle
            status_fill = None
            if 'Servis' in str(status):
                if 'Dışı' in str(status):
                    if 'İşletme' in str(status):
                        status_fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')  # Turuncu
                    else:
                        status_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')  # Kırmızı
                else:
                    status_fill = PatternFill(start_color='00B050', end_color='00B050', fill_type='solid')  # Yeşil
            
            status_font = Font(bold=True, color='FFFFFF') if status_fill else Font()
            
            # Border
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Satır ekle
            row_data = [
                timestamp,
                str(tram_id),
                str(date),
                status,
                sistem,
                alt_sistem,
                aciklama,
                user_id or ''
            ]
            
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=next_row, column=col_idx)
                cell.value = value
                cell.border = border
                cell.alignment = Alignment(horizontal='left', vertical='center')
                
                # Durum sütununa renk ekle
                if col_idx == 4 and status_fill:
                    cell.fill = status_fill
                    cell.font = status_font
            
            # Dosyayı kaydet
            wb.save(log_file)
            return True
        except Exception as e:
            print(f"❌ Excel log error: {str(e)}")
            return False
    
    @staticmethod
    def log_status_change(tram_id, date, status, sistem='', alt_sistem='', aciklama='', user_id=None):
        """
        Servis durumu değişikliğini hem JSON hem Excel'e log'la
        """
        json_ok = ServiceStatusLogger.log_json(tram_id, date, status, sistem, alt_sistem, aciklama, user_id)
        excel_ok = ServiceStatusLogger.log_excel(tram_id, date, status, sistem, alt_sistem, aciklama, user_id)
        
        if json_ok and excel_ok:
            print(f"✅ Log saved: {tram_id} - {date} - {status}")
            return True
        else:
            print(f"⚠️  Partial log save: JSON={json_ok}, Excel={excel_ok}")
            return False
    
    @staticmethod
    def get_all_logs(start_date=None, end_date=None):
        """
        Tüm logları al (isteğe bağlı olarak tarih filtrelemesi)
        """
        try:
            ServiceStatusLogger.ensure_log_directory()
            
            all_logs = []
            
            # Tüm log dosyalarını tara
            for filename in os.listdir(ServiceStatusLogger.BASE_LOG_DIR):
                if filename.endswith('.json'):
                    file_path = os.path.join(ServiceStatusLogger.BASE_LOG_DIR, filename)
                    with open(file_path, 'r', encoding='utf-8') as f:
                        logs = json.load(f)
                        all_logs.extend(logs)
            
            # Tarih filtrelemesi yap
            if start_date or end_date:
                from datetime import datetime as dt
                filtered_logs = []
                for log in all_logs:
                    log_date = log['date']
                    if start_date and log_date < start_date:
                        continue
                    if end_date and log_date > end_date:
                        continue
                    filtered_logs.append(log)
                all_logs = filtered_logs
            
            return sorted(all_logs, key=lambda x: x['timestamp'], reverse=True)
        except Exception as e:
            print(f"❌ Error getting logs: {str(e)}")
            return []


if __name__ == '__main__':
    # Test
    ServiceStatusLogger.log_status_change('1531', '2026-02-16', 'Servis', 'Elektrik', 'Motor', 'Test log', 1)
    ServiceStatusLogger.log_status_change('1532', '2026-02-16', 'Servis Dışı', 'Mekanik', '', 'Arıza', 2)
    ServiceStatusLogger.log_status_change('1533', '2026-02-16', 'İşletme Kaynaklı Servis Dışı', '', '', 'Operatör hatası', 1)
    
    print("\n✅ Log system test completed")
