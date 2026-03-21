from openpyxl import load_workbook
import os
import time

project = 'kayseri'
base_path = r'c:\Users\fatiherkin\Desktop\bozankaya_ssh_takip\data'
grid_path = os.path.join(base_path, project, 'service_status_grid.xlsx')

print(f"\n{'='*80}")
print(f"KAYSERI GRID - SEMBOL NORMALİZASYONU")
print(f"{'='*80}\n")

# Dosya kilidi olabilir, 2 saniye bekle
time.sleep(2)

try:
    wb = load_workbook(grid_path)
    ws = wb.active
    
    # Symbol replacement mapping
    symbol_map = {
        '√': '✓',      # SQUARE ROOT → CHECK MARK
    }
    
    # Tüm veri hücrelerini kontrol et
    replaced = 0
    for row_idx in range(2, ws.max_row + 1):
        for col_idx in range(2, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell_value = cell.value
            
            if cell_value:
                str_value = str(cell_value).strip()
                
                # Sembol değiştir
                if str_value in symbol_map:
                    new_symbol = symbol_map[str_value]
                    if str_value != new_symbol:
                        cell.value = new_symbol
                        replaced += 1
                        
                        if replaced % 100 == 0:
                            print(f"  {replaced} hücre değiştirildi...")
    
    wb.save(grid_path)
    print(f"\n✓ Normalizasyon tamamlandı!")
    print(f"  Toplam değiştirilen: √ → ✓ ({replaced} hücre)")
    print(f"  Dosya kaydedildi: {grid_path}")
    
except PermissionError as e:
    print(f"❌ Dosya kilitli! Hata: {e}")
    print(f"  Flask uygulaması kapatılmamış olabilir.")
    print(f"  Ya da Excel dosyası başka yerde açık olabilir.")
    exit(1)

print(f"\n{'='*80}\n")
