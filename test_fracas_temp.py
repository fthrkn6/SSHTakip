#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""Test FracasWriter with temp copy of template"""

import os
import sys
import shutil
import tempfile

# Workspace root'u al
workspace_root = os.path.dirname(os.path.abspath(__file__))
print(f"Workspace: {workspace_root}\n")

# Original file
original_file = os.path.join(workspace_root, "logs/belgrad/ariza_listesi/Fracas_BELGRAD.xlsx")

if not os.path.exists(original_file):
    print(f"ERROR: File not found: {original_file}")
    sys.exit(1)

# Create a temp copy to test with
temp_dir = tempfile.gettempdir()
test_file = os.path.join(temp_dir, "Fracas_BELGRAD_TEST.xlsx")

# Copy to temp
shutil.copy(original_file, test_file)
print(f"Test file created: {test_file}\n")

from utils_fracas_writer import FracasWriter

# Test: Override file_path before calling write
print("TEST: Writing to temp copy")
writer = FracasWriter(base_path=workspace_root)

# Override the file path to test file
original_path = writer.file_path
writer.file_path = test_file
print(f"  Original path: {original_path}")
print(f"  Test path: {test_file}")
print(f"  Test file exists: {os.path.exists(test_file)}\n")

sample_data = {
    'arac_numarasi': 'T99_TEST',
    'sistem': 'OKS',
    'alt_sistem': 'Sistem',
    'tedarikci': 'Test',
    'ariza_tanimi': 'Test ariza',
    'ariza_sinifi': 'Test',
    'hata_tarih': '2026-02-17',
    'hata_saat': '16:30',
    'ariza_tespit_yontemi': 'Bozankaya',
}

print("Writing data...")
try:
    result = writer.write_failure_data(sample_data)
    if result.get('success'):
        print(f"  SUCCESS!")
        print(f"    Row: {result['row']}")
        print(f"    ID: {result['fracas_id']}")
        print(f"    Path: {result['file_path']}\n")
        
        # Verify data in temp file
        from openpyxl import load_workbook
        wb = load_workbook(test_file)
        ws = wb['FRACAS']
        row_num = result['row']
        print("Verification:")
        print(f"    B{row_num}: {ws[f'B{row_num}'].value}")  # Arac
        print(f"    E{row_num}: {ws[f'E{row_num}'].value}")  # ID
        print(f"    P{row_num}: {ws[f'P{row_num}'].value}")  # Tespit Yontemi
        
    else:
        print(f"  FAILED: {result}")
except Exception as e:
    print(f"  ERROR: {e}")

# Cleanup
if os.path.exists(test_file):
    os.remove(test_file)
    print("\nTest file cleaned up")

print("\nDone!")
