"""
FracasWriter - Yeni arıza verilerini Fracas_BELGRAD.xlsx template'ine yazmak için utility
Özellikler:
- Form verilerini Fracas columns'larıyla eşleştirme
- Template'i yükleme ve yeni satırlar ekleme
- Otomatik next FRACAS ID belirleme
- Verileri validate etme
- Excel formatting'i koruma
"""

import os
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import logging

logger = logging.getLogger(__name__)

class FracasWriter:
    """Fracas_BELGRAD.xlsx template'ine yeni arıza verisi yazma"""
    
    # Form alanları -> Fracas Excel Sütunları eşlemesi 
    FORM_TO_FRACAS_MAPPING = {
        # Araç Bilgileri
        'fracas_id': 'E',              # FRACAS ID
        'arac_numarasi': 'B',          # Araç Numarası
        'arac_module': 'C',            # Araç Module (multiple -> virgülle birleş)
        'arac_km': 'D',                # Araç Kilometresi
        
        # Arıza Zamanı
        'hata_tarih': 'F',             # Hata Tarih (datetime format oluşturulacak)
        'hata_saat': 'F',              # Hata Saat (combine with hata_tarih)
        
        # Sistem Bilgileri
        'sistem': 'G',                 # Sistem
        'alt_sistem': 'H',             # Alt Sistem
        'tedarikci': 'I',              # İlgili Tedarikçi
        
        # Arıza Detayları
        'ariza_tanimi': 'J',           # Arıza Tanımı
        'ariza_sinifi': 'K',           # Arıza Sınıfı
        'ariza_kaynagi': 'L',          # Arıza Kaynağı
        'yapilan_islem': 'M',          # Arıza Tespitini Takiben Yapılan İşlem
        'aksiyon': 'N',                # Aksiyon
        'garanti_kapsami': 'O',        # Garanti Kapsamı
        'ariza_tespit_yontemi': 'P',   # Arıza Tespit Yöntemi
        'personel_sayisi': 'Q',        # Tamir için Gerekli Personel Sayısı
        
        # Tamir Zamanları
        'tamir_baslama_tarih': 'R',    # Tamir Başlama Tarihi
        'tamir_baslama_saati': 'S',    # Tamir Başlama Saati
        'tamir_bitisi_tarih': 'T',     # Tamir Bitiş Tarihi
        'tamir_bitisi_saati': 'U',     # Tamir Bitiş Saati
        'tamir_suresi': 'V',           # Tamir Süresi (dakika)
        
        # Servise Veriliş
        'servise_verilis_tarih': 'W',  # Servise Veriliş Tarih
        'servise_verilis_saati': 'X',  # Servise Veriliş Saati
        
        # Diğer Arıza Bilgileri
        'ariza_tipi': 'Y',             # Arıza Tipi
        'detayli_bilgi': 'Z',          # Detaylı Bilgi (Servise Engel / Değil) (YENİ)
        
        # MTTR/MDT Hesaplamaları (otomatik hesaplanacak)
        # 'mttr_araç': 'AA',           # Araç MTTR / MDT
        # 'mttr_komponent': 'AB',      # Kompanent MTTR / MDT
        
        # Parça Bilgileri
        'parca_kodu': 'AC',            # Parça Kodu
        'parca_seri_no': 'AD',         # Seri Numarası
        'parca_adi': 'AE',             # Parça Adı
        'adet': 'AF',                  # Adeti
        'iscilik_maliyeti': 'AG',      # İşçilik Maliyeti
    }
    
    # Header satır numarası
    HEADER_ROW = 4
    # İlk veri satırı
    FIRST_DATA_ROW = 5
    
    # Üretim / Test ortamı
    FRACAS_FILE_PATH = 'logs/belgrad/ariza_listesi/Fracas_BELGRAD.xlsx'
    PROJECT_NAME = 'Bozankaya'  # Fracas template'de bu sabit oluyor
    
    def __init__(self, base_path=None):
        """FracasWriter'ı başlat
        
        Args:
            base_path: Root workspace path (eğer None ise otomatik tespit et)
        """
        self.base_path = base_path
        self.workbook = None
        self.worksheet = None
        self.file_path = self.get_fracas_file_path()
        
    def get_fracas_file_path(self):
        """Fracas template dosya yolunu al"""
        if self.base_path:
            # Eğer base_path verilmişse onu kullan
            file_path = os.path.join(self.base_path, self.FRACAS_FILE_PATH)
            if os.path.exists(file_path):
                return os.path.abspath(file_path)
        
        # Auto-detect: workspace root'u bul (app.py veya utils_fracas_writer.py'in bulunduğu yer)
        current_dir = os.path.dirname(os.path.abspath(__file__))
        file_path = os.path.join(current_dir, self.FRACAS_FILE_PATH)
        
        # Eğer dosya bulunamadıysa, bir üst klasöre bak
        if not os.path.exists(file_path):
            parent_dir = os.path.dirname(current_dir)
            alt_path = os.path.join(parent_dir, self.FRACAS_FILE_PATH)
            if os.path.exists(alt_path):
                file_path = alt_path
        
        return os.path.abspath(file_path)
    
    def load_template(self):
        """Fracas_BELGRAD.xlsx template'i yükle"""
        try:
            if not os.path.exists(self.file_path):
                raise FileNotFoundError(f"Fracas template bulunamadı: {self.file_path}")
            
            self.workbook = load_workbook(self.file_path)
            self.worksheet = self.workbook['FRACAS']
            logger.info(f"Fracas template yüklendi: {self.file_path}")
            return True
            
        except Exception as e:
            logger.error(f"Template yükleme hatası: {e}")
            raise
    
    def get_next_data_row(self):
        """Bir sonraki boş satırı bul (veri yazılacak satır)"""
        row = self.FIRST_DATA_ROW
        # FRACAS ID sütununu kontrol ederek boş satırı bul
        while self.worksheet[f'E{row}'].value is not None:
            row += 1
        return row
    
    def generate_next_fracas_id(self):
        """BOZ-BEL25-FF-NNN formatında FRACAS ID oluştur (Arıza Listesi ile senkronize)"""
        from openpyxl import load_workbook
        import os
        
        # Arıza Listesi dosyasından son numarayı al (aynı sistem)
        ariza_listesi_file = os.path.join(
            os.path.dirname(__file__), 
            'logs', 'belgrad', 'ariza_listesi', 
            'Ariza_Listesi_BELGRAD.xlsx'
        )
        
        next_num = 1
        if os.path.exists(ariza_listesi_file):
            try:
                wb = load_workbook(ariza_listesi_file, data_only=True)
                ws = wb.active
                ids = []
                # A sütununda FF numaralarını ara
                for row in range(5, ws.max_row + 1):
                    cell_val = ws.cell(row=row, column=1).value
                    if cell_val and 'FF-' in str(cell_val):
                        num = int(str(cell_val).split('FF-')[-1])
                        ids.append(num)
                wb.close()
                if ids:
                    next_num = max(ids) + 1
            except Exception as e:
                print(f"⚠️ Arıza Listesi'nden FRACAS ID hesaplanamadı: {e}")
                pass
        
        fracas_id = f'BOZ-BEL25-FF-{next_num:03d}'
        return fracas_id
    
    def validate_form_data(self, form_data):
        """Form verilerini valide et"""
        errors = []
        
        # Zorunlu alanlar
        required_fields = ['arac_numarasi', 'sistem', 'hata_tarih', 'hata_saat', 'ariza_tanimi', 'ariza_sinifi']
        for field in required_fields:
            if field not in form_data or not form_data[field]:
                errors.append(f"Zorunlu alan boş: {field}")
        
        return errors
    
    def prepare_datetime(self, date_str, time_str):
        """Tarih ve saati datetime formatında birleştir"""
        try:
            if not date_str or not time_str:
                return None
            datetime_str = f"{date_str} {time_str}"
            dt = datetime.strptime(datetime_str, '%Y-%m-%d %H:%M')
            return dt.strftime('%Y-%m-%d %H:%M')
        except Exception as e:
            logger.warning(f"DateTime birleştirme hatası: {e}")
            return None
    
    def prepare_form_data(self, form_data):
        """Form verilerini Excel yazılımına hazırla"""
        prepared = {}
        
        # Hata tarihini ve saatini birleştir
        if 'hata_tarih' in form_data and 'hata_saat' in form_data:
            prepared['F'] = self.prepare_datetime(form_data['hata_tarih'], form_data['hata_saat'])
        
        # Tamir başlama zamanını birleştir
        if 'tamir_baslama_tarih' in form_data and 'tamir_baslama_saati' in form_data:
            prepared['R'] = form_data.get('tamir_baslama_tarih', '')
            prepared['S'] = form_data.get('tamir_baslama_saati', '')
        
        # Tamir bitiş zamanını birleştir
        if 'tamir_bitisi_tarih' in form_data and 'tamir_bitisi_saati' in form_data:
            prepared['T'] = form_data.get('tamir_bitisi_tarih', '')
            prepared['U'] = form_data.get('tamir_bitisi_saati', '')
        
        # Servise veriliş zamanını birleştir
        if 'servise_verilis_tarih' in form_data and 'servise_verilis_saati' in form_data:
            prepared['W'] = form_data.get('servise_verilis_tarih', '')
            prepared['X'] = form_data.get('servise_verilis_saati', '')
        
        # Araç module'ü virgülle birleştir (multiple select)
        if 'arac_module' in form_data and form_data['arac_module']:
            modules = form_data['arac_module']
            if isinstance(modules, list):
                prepared['C'] = ', '.join(modules)
            else:
                prepared['C'] = str(modules)
        
        # Diğer alanları harita kullanarak at
        for form_field, fracas_col in self.FORM_TO_FRACAS_MAPPING.items():
            # Yukarıda özel işlem yapılanları atla
            if form_field in ['hata_tarih', 'hata_saat', 'tamir_baslama_tarih', 'tamir_baslama_saati',
                             'tamir_bitisi_tarih', 'tamir_bitisi_saati', 'servise_verilis_tarih',
                             'servise_verilis_saati', 'arac_module']:
                continue
            
            # Eğer sütun zaten atandıysa atla
            if fracas_col in prepared:
                continue
            
            value = form_data.get(form_field, '')
            if value:
                prepared[fracas_col] = value
        
        # Project alanını Always set
        prepared['A'] = self.PROJECT_NAME
        
        # FRACAS ID'yi otomatik oluştur (eğer verilmediyse)
        if 'E' not in prepared or not prepared['E']:
            prepared['E'] = self.generate_next_fracas_id()
        
        return prepared
    
    def write_failure_data(self, form_data):
        """
        Form verilerini Fracas template'ine yaz
        
        Args:
            form_data: dict - Form'tan gelen veriler
        
        Returns:
            dict - Yazılan veri satırı numarası ve FRACAS ID
        
        Raises:
            ValueError: Validasyon hatası
            IOError: File write hatası
        """
        import tempfile
        import shutil
        import time
        
        try:
            # Verileri valide et
            errors = self.validate_form_data(form_data)
            if errors:
                raise ValueError(f"Validasyon hatası: {', '.join(errors)}")
            
            # Temp dosya konumu
            temp_dir = tempfile.gettempdir()
            temp_file = os.path.join(temp_dir, f"Fracas_write_{int(time.time() * 1000)}.xlsx")
            
            # Ana dosyayı temp'e kopyala (lock'u çözmek için)
            if not os.path.exists(self.file_path):
                raise FileNotFoundError(f"Fracas template bulunamadı: {self.file_path}")
            
            shutil.copy(self.file_path, temp_file)
            time.sleep(0.2)
            logger.debug(f"Template temp'e kopyalandı: {temp_file}")
            
            # Temp dosyayı aç ve modify et
            wb = load_workbook(temp_file)
            ws = wb['FRACAS']
            
            # Template'i self.'e ataprogramlanewsheet olarak ayarla (generate_next_fracas_id için)
            self.worksheet = ws
            self.workbook = wb
            
            # Bir sonraki satırı bul
            next_row = self.FIRST_DATA_ROW
            while ws[f'E{next_row}'].value is not None:
                next_row += 1
            
            # Form verilerini hazırla
            prepared_data = self.prepare_form_data(form_data)
            
            # Hazırlanan verileri Excel'e yaz
            for col_letter, value in prepared_data.items():
                if value:  # Sadece boş olmayan değerleri yaz
                    cell = ws[f'{col_letter}{next_row}']
                    cell.value = value
                    logger.debug(f"Yazıldı: {col_letter}{next_row} = {value}")
            
            # Temp dosyayı kaydet
            wb.save(temp_file)
            wb.close()
            time.sleep(0.3)
            logger.debug(f"Temp dosya kaydedildi: {temp_file}")
            
            # Eski dosyayı sil, temp'i ana konuma taşı (atomic)
            # Windows dosya lock problemi için retry ekle
            max_retries = 3
            for attempt in range(max_retries):
                try:
                    if os.path.exists(self.file_path):
                        os.remove(self.file_path)
                        logger.debug(f"Eski dosya silindi: {self.file_path}")
                    
                    shutil.move(temp_file, self.file_path)
                    time.sleep(0.2)
                    break
                    
                except (PermissionError, OSError) as e:
                    if attempt < max_retries - 1:
                        logger.warning(f"Dosya işlemi retry {attempt + 1}/{max_retries}: {e}")
                        time.sleep(1 + attempt)  # Progressively increase wait
                    else:
                        raise
            
            logger.info(f"Fracas verileri yazıldı - Satır: {next_row}, FRACAS ID: {prepared_data.get('E', 'N/A')}")
            
            return {
                'success': True,
                'row': next_row,
                'fracas_id': prepared_data.get('E', 'N/A'),
                'file_path': self.file_path
            }
            
        except Exception as e:
            logger.error(f"Fracas yazma hatası: {e}")
            # Temp dosyayı temizle eğer varsa
            try:
                if 'temp_file' in locals() and os.path.exists(temp_file):
                    os.remove(temp_file)
            except:
                pass
            raise
    
    def get_current_excel_file(self):
        """Güncel Excel dosyasının yolunu döndür (download için)"""
        return self.file_path


# Test
if __name__ == '__main__':
    # Test verileri
    test_data = {
        'arac_numarasi': 'T01',
        'arac_module': ['Wheel', 'Brake'],
        'arac_km': '1500',
        'hata_tarih': '2024-02-17',
        'hata_saat': '14:30',
        'sistem': 'Traksiyon',
        'alt_sistem': 'Motor',
        'tedarikci': 'Siemens',
        'ariza_tanimi': 'Motor arızası',
        'ariza_sinifi': 'Kritik',
        'ariza_kaynagi': 'Mekanik',
        'yapilan_islem': 'Motor tamir edildi',
        'aksiyon': 'Tamir Edildi',
        'garanti_kapsami': 'Evet',
        'ariza_tespit_yontemi': 'Bozankaya',
        'tamir_baslama_tarih': '2024-02-17',
        'tamir_baslama_saati': '15:00',
        'tamir_bitisi_tarih': '2024-02-17',
        'tamir_bitisi_saati': '18:00',
        'tamir_suresi': '3 saat 0 dakika',
        'mttr': '180 dk',
        'personel_sayisi': '2',
        'parca_kodu': 'M001',
        'parca_adi': 'Motor Assembly',
        'parca_seri_no': 'SN12345',
        'adet': '1',
        'iscilik_maliyeti': '50.00',
        'servise_verilis_tarih': '2024-02-17',
        'servise_verilis_saati': '18:00',
        'ariza_tipi': 'Mekanik',
    }
    
    writer = FracasWriter()
    try:
        result = writer.write_failure_data(test_data)
        print(f"✅ Başarılı: {result}")
    except Exception as e:
        print(f"❌ Hata: {e}")
