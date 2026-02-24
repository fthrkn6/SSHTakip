#!/usr/bin/env python
# -*- coding: utf-8 -*-
from models import db, Equipment
from app import app

with app.app_context():
    eq1 = Equipment.query.filter_by(id=1, project_code='belgrad').first()
    eq1531 = Equipment.query.filter_by(equipment_code=1531, project_code='belgrad').first()
    
    print("\n====== VERİTABANI KONTROL ======")
    print("\nEquipment (id=1):")
    if eq1:
        print(f"  code={eq1.equipment_code}, km={eq1.current_km}")
    else:
        print("  NOT FOUND")
    
    print("\nEquipment (code=1531):")
    if eq1531:
        print(f"  id={eq1531.id}, km={eq1531.current_km}")
    else:
        print("  NOT FOUND")
    
    print("\n====== EXCEL KONTROL ======")
    from openpyxl import load_workbook
    wb = load_workbook('data/belgrad/km_data.xlsx')
    ws = wb.active
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if row[0]:
            print(f"  tram={row[0]}, km={row[1]}")
    
    print("\n====== MESAJ HATASININ NEDENİ ======")
    print("Eğer 'id=1' kaydının km=100 ise, yanlış yere yazılmıştır!")
    print("1531'e yazılması gerekiyordu.")
