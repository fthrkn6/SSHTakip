import openpyxl
from models import db, Equipment
from app import create_app

app = create_app()
with app.app_context():
    print("=" * 80)
    print("TRAMWAY DATA SOURCE ANALYSIS")
    print("=" * 80)
    
    # Read Excel file
    print("\n1. READING VERILER.XLSX - SAYFA2 (tram_id column):")
    print("-" * 80)
    
    try:
        wb = openpyxl.load_workbook('data/belgrad/Veriler.xlsx')
        ws = wb['Sayfa2']
        
        # Find tram_id column
        tram_id_col = None
        for col_idx, cell in enumerate(ws[1], 1):
            if cell.value and 'tram' in str(cell.value).lower():
                tram_id_col = col_idx
                print(f"Found 'tram_id' column at: Column {col_idx} ({openpyxl.utils.get_column_letter(col_idx)})")
                break
        
        if not tram_id_col:
            print("ERROR: 'tram_id' column not found in Sayfa2")
            print("Available columns:", [cell.value for cell in ws[1]])
        else:
            # Extract all tram_ids from Excel
            excel_tram_ids = set()
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=False), start=2):
                cell = row[tram_id_col - 1]
                if cell.value:
                    excel_tram_ids.add(str(cell.value).strip())
            
            print(f"\nTram IDs in Veriler.xlsx (Sayfa2): {len(excel_tram_ids)} records")
            excel_tram_ids_sorted = sorted(excel_tram_ids)
            for tram_id in excel_tram_ids_sorted:
                print(f"  - {tram_id}")
    
    except Exception as e:
        print(f"ERROR reading Excel: {e}")
        excel_tram_ids = set()
    
    # Read Database Equipment
    print("\n2. READING DATABASE - EQUIPMENT TABLE (equipment_code column):")
    print("-" * 80)
    
    db_equipment = Equipment.query.all()
    db_tram_ids = set([eq.equipment_code for eq in db_equipment])
    
    print(f"Tram IDs in Database: {len(db_tram_ids)} records")
    db_tram_ids_sorted = sorted(db_tram_ids)
    for tram_id in db_tram_ids_sorted:
        eq = Equipment.query.filter_by(equipment_code=tram_id).first()
        print(f"  - {tram_id}: {eq.name}")
    
    # Compare
    print("\n3. COMPARISON:")
    print("-" * 80)
    
    in_excel_not_db = excel_tram_ids - db_tram_ids
    in_db_not_excel = db_tram_ids - excel_tram_ids
    in_both = excel_tram_ids & db_tram_ids
    
    print(f"\nIn EXCEL and DATABASE (matched): {len(in_both)}")
    for tram_id in sorted(in_both):
        print(f"  ✓ {tram_id}")
    
    if in_excel_not_db:
        print(f"\nIn EXCEL but NOT in DATABASE: {len(in_excel_not_db)}")
        for tram_id in sorted(in_excel_not_db):
            print(f"  ⚠ {tram_id} (exists in Excel but NOT in DB)")
    
    if in_db_not_excel:
        print(f"\nIn DATABASE but NOT in EXCEL: {len(in_db_not_excel)}")
        print("These are TEST/EXTRA equipment that could be deleted:")
        for tram_id in sorted(in_db_not_excel):
            eq = Equipment.query.filter_by(equipment_code=tram_id).first()
            print(f"  ✗ {tram_id}: {eq.name}")
    
    print("\n" + "=" * 80)
    print("RECOMMENDATION:")
    print("=" * 80)
    if in_db_not_excel:
        print(f"\nSilmek için önerilen {len(in_db_not_excel)} kayıt:")
        for tram_id in sorted(in_db_not_excel):
            eq = Equipment.query.filter_by(equipment_code=tram_id).first()
            print(f"  - {tram_id}: {eq.name}")
        print("\nEğer 'sil' dersen, bu kayıtlar veritabanından silinecek.")
    else:
        print("\nSilinecek kayıt yok. Database ve Excel'de aynı veriler var.")
    
