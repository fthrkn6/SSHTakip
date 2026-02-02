from openpyxl import load_workbook
import os
from jinja2 import Environment

# Sistemleri yükle
sistemler = {}
data_dir = os.path.join(os.path.dirname(__file__), 'data', 'belgrad')

veriler_path = None
if os.path.exists(data_dir):
    for file in os.listdir(data_dir):
        if 'veriler' in file.lower() and file.endswith('.xlsx'):
            veriler_path = os.path.join(data_dir, file)
            break

if veriler_path and os.path.exists(veriler_path):
    try:
        wb = load_workbook(veriler_path)
        ws = wb['Sayfa2']
        
        # Renk tanımları
        KIRMIZI = 'FFFF0000'
        SARI = 'FFFFFF00'
        MAVI = 'FF0070C0'
        
        sütunlar = [6, 7, 8, 9, 11]
        
        for sütun in sütunlar:
            sistem_adi = None
            
            for row in range(1, ws.max_row + 1):
                cell = ws.cell(row=row, column=sütun)
                value = cell.value
                fill = cell.fill
                
                color_hex = None
                if fill and fill.start_color:
                    color_hex = str(fill.start_color.rgb) if fill.start_color.rgb else None
                
                if color_hex == KIRMIZI and value:
                    sistem_adi = str(value).strip()
                    if sistem_adi not in sistemler:
                        sistemler[sistem_adi] = {
                            'tedarikçiler': set(),
                            'alt_sistemler': set()
                        }
                
                elif color_hex == SARI and value and sistem_adi:
                    sistemler[sistem_adi]['tedarikçiler'].add(str(value).strip())
                
                elif color_hex == MAVI and value and sistem_adi:
                    sistemler[sistem_adi]['alt_sistemler'].add(str(value).strip())
        
        sistemler = {
            k: {
                'tedarikçiler': sorted(list(v['tedarikçiler'])),
                'alt_sistemler': sorted(list(v['alt_sistemler']))
            }
            for k, v in sistemler.items()
        }
    except Exception as e:
        print(f"Hata: {e}")

# Test template snippet
test_template = """
<script>
    const sistemler_data = {
    {% for sistem_adi, sistem_data in sistemler.items() %}
    "{{ sistem_adi }}": {
        "tedarikçiler": [{% for t in sistem_data.tedarikçiler %}"{{ t }}"{{ ", " if not loop.last else "" }}{% endfor %}],
        "alt_sistemler": [{% for a in sistem_data.alt_sistemler %}"{{ a }}"{{ ", " if not loop.last else "" }}{% endfor %}]
    }{{ ", " if not loop.last else "" }}
    {% endfor %}
    };
    
    console.log(sistemler_data);
</script>
"""

env = Environment()
template = env.from_string(test_template)
rendered = template.render(sistemler=sistemler)

print("RENDER EDILEN ÇIKTI:")
print("=" * 80)
print(rendered)
print("=" * 80)
