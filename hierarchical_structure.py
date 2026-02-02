from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
import re

wb = load_workbook('data/belgrad/Veriler.xlsx')
ws = wb['Sayfa2']

print("=" * 120)
print("📊 HIERARCHICAL KATEGORIZASYON YAPISI")
print("=" * 120)

print("\n🔍 MEVCUT VERİ YAPISI ANALIZI:\n")

# Veriye bak
data_structure = {}
for i, row in enumerate(ws.iter_rows(min_row=2, max_row=15, values_only=True), 1):
    tram_id = row[0]
    sistem = row[5]
    
    if sistem and sistem not in data_structure:
        # Sistemi parse et
        if 'Sis.' in str(sistem):
            # Şeklini kontrol et: "Tahrik Konvertörü Sis. Alt Sistem"
            parts = str(sistem).split('Sis.')
            main_sys = parts[0].strip() + ' (ANA SİSTEM)'
            sub_sys = parts[1].strip() if len(parts) > 1 else 'Bilinmiyor'
        else:
            main_sys = str(sistem)
            sub_sys = 'Direkt Sistem'
        
        data_structure[sistem] = {
            'main': main_sys,
            'sub': sub_sys,
            'components': []
        }

print("\n📋 ÖNERILEN KATEGORIZASYON YAPISI:\n")

categories = {
    'LEVEL 1 - ANA SİSTEM': 'Traction_Converter, Medcom, ABB, vb.',
    'LEVEL 2 - ALT SİSTEM': 'Auxiliary_Power_Unit, Battery, Control, vb.',
    'LEVEL 3 - BİLEŞEN TİPİ': 'Hoppecke, ABB, VEM, vb. (Üretici)',
    'LEVEL 4 - SPESIFIK PARÇA': 'Motor, Pantograf, ESS Battery, vb.'
}

for level, description in categories.items():
    print(f"  {level}: {description}")

print("\n\n✅ YENİ EXCEL YAPISI - KATEGORIZE İLE:\n")
print("-" * 120)
print(f"{'Tram':<8} | {'L1: ANA SİS':<25} | {'L2: ALT SİS':<25} | {'L3: BİLEŞEN':<25} | {'L4: PARÇA':<25}")
print("-" * 120)

for i, row in enumerate(ws.iter_rows(min_row=2, max_row=15, values_only=True), 1):
    tram_id = row[0]
    col6 = row[5]  # Sistemler
    col7 = row[6]  # Alt sistem 1
    col8 = row[7]  # Alt sistem 2 / Bileşen
    col9 = row[8]  # Spesifik parça
    
    # Kırılım göstermek için indentation
    l1 = str(col6) if col6 else "---"
    l2 = f"  └─ {col7}" if col7 else "---"
    l3 = f"      └─ {col8}" if col8 else "---"
    l4 = f"          └─ {col9}" if col9 else "---"
    
    print(f"{int(tram_id):<8} | {l1:<25} | {l2:<25} | {l3:<25} | {l4:<25}")

print("\n\n" + "=" * 120)
print("💡 ÇÖZÜM SEÇENEKLERI:")
print("=" * 120)

print("""
1️⃣  SEPARATİ SÜTUNLAR (En basit):
   ✓ Sütun F: Sistem (ANA)
   ✓ Sütun G: Alt Sistem Seviyesi 1
   ✓ Sütun H: Alt Sistem Seviyesi 2
   ✓ Sütun I: Spesifik Parça
   
2️⃣  RENKLE KATEGORİZASYON (Görsel):
   ✓ ANA SİSTEM: Mavi arka plan
   ✓ ALT SİSTEM: Yeşil arka plan
   ✓ BİLEŞEN: Sarı arka plan
   ✓ PARÇA: Turuncu arka plan
   
3️⃣  İNDENTATION İLE AYNASÜTUNDA (Gelişmiş):
   ✓ "SISTEM > ALT_SISTEM > BİLEŞEN > PARÇA"
   ✓ Tree view gösterimi
   ✓ Hiyerarşik görünüm
   
4️⃣  VERİ TABANI YAPISI (Optimal):
   ✓ Sistem (ID, Adı, Tür)
   ✓ Alt_Sistem (ID, Sistem_ID, Adı)
   ✓ Bileşen (ID, Alt_Sistem_ID, Üretici, Model)
   ✓ Parça (ID, Bileşen_ID, Adı, Seri_No)
""")

print("\n🎯 En Praktik Çözüm: 1 + 2 Kombinasyonu")
print("   • 4 ayrı sütun (F, G, H, I)")
print("   • Her seviye farklı renk")
print("   • Excel Filtresi ile kolay seçim")
print("   • Raportlamada düzenli görünüm")
