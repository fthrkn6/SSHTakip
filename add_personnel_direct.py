#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""Add Personnel Count column directly to original Excel file"""

from openpyxl import load_workbook
import os

excel_file = r'c:\Users\ferki\Desktop\bozankaya_ssh_takip\logs\ariza_listesi\Ariza_Listesi_BELGRAD.xlsx'

try:
    print("Dosya açılıyor...")
    # Load workbook with data_only=False to preserve formulas
    wb = load_workbook(excel_file)
    ws = wb['Ariza Listesi']
    
    # Find the last column
    max_col = ws.max_column
    new_col = max_col + 1
    
    print(f"Son sütun: {max_col}")
    print(f"Yeni sütun: {new_col}")
    
    # Add header in row 4 (header row for data)
    ws.cell(row=4, column=new_col, value='Personel Sayısı')
    print(f"✅ 'Personel Sayısı' başlığı eklendi (Sütun {new_col}, Satır 4)")
    
    # Save the file
    print("Dosya kaydediliyor...")
    wb.save(excel_file)
    
    print(f"\n✅ BAŞARILI!")
    print(f"Dosya: {excel_file}")
    
    # Verify
    wb_check = load_workbook(excel_file)
    ws_check = wb_check['Ariza Listesi']
    header_value = ws_check.cell(row=4, column=ws_check.max_column).value
    print(f"En son sütun başlığı: {header_value}")
    print(f"Toplam sütun: {ws_check.max_column}")
    
except PermissionError as e:
    print(f"❌ Dosya kilitli! Hata: {e}")
    print("\nÇözüm: Dosyayı açan programı kapatın (VSCode, Excel, vs.)")
except Exception as e:
    print(f"❌ Hata: {str(e)}")
    import traceback
    traceback.print_exc()
