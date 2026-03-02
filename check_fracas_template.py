#!/usr/bin/env python
# -*- coding: utf-8 -*-

import os
from openpyxl import load_workbook

# Template dosya yolu
template_path = r"logs\kayseri\ariza_listesi\Fracas_KAYSERI.xlsx"

if os.path.exists(template_path):
    print(f"✓ Template found: {template_path}")
    wb = load_workbook(template_path)
    
    if 'FRACAS' in wb.sheetnames:
        ws = wb['FRACAS']
        print(f"✓ FRACAS sheet found")
        print(f"\nRow 4 (Header):")
        print(f"  A4: {ws['A4'].value}")
        print(f"  B4: {ws['B4'].value}")
        print(f"  E4: {ws['E4'].value}")
        
        print(f"\nRow 5 (First data row):")
        print(f"  A5: {ws['A5'].value}")
        print(f"  B5: {ws['B5'].value}")
        print(f"  E5: {ws['E5'].value}")
        
    else:
        print(f"✗ FRACAS sheet not found. Sheets: {wb.sheetnames}")
    
    wb.close()
else:
    print(f"✗ Template not found: {template_path}")
