from openpyxl import load_workbook
import os
import shutil

project = 'kayseri'
base_path = r'c:\Users\fatiherkin\Desktop\bozankaya_ssh_takip\data'
grid_path = os.path.join(base_path, project, 'service_status_grid.xlsx')
temp_path = grid_path + '.temp'
old_path = grid_path + '.old'

print(f"\n{'='*80}")
print(f"KAYSERI GRID - SEMBOL NORMALİZASYONU (Geçici Dosya Yöntemi)")
print(f"{'='*80}\n")

try:
    # Dosyayı aç (geçici olarak de olsa)
    print(f"1️⃣  Excel dosyası okunuyor...")
    wb = load_workbook(grid_path)
    ws = wb.active
    
    # Symbol replacement mapping
    symbol_map = {'√': '✓'}
    
    # Tüm veri hücrelerini kontrol et
    replaced = 0
    for row_idx in range(2, ws.max_row + 1):
        for col_idx in range(2, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell_value = cell.value
            
            if cell_value:
                str_value = str(cell_value).strip()
                if str_value in symbol_map:
                    new_symbol = symbol_map[str_value]
                    cell.value = new_symbol
                    replaced += 1
                    
                    if replaced % 500 == 0:
                        print(f"  {replaced} hücre hazırlandı...")
    
    # Geçici dosyaya kaydet
    print(f"2️⃣  Geçici dosyaya kaydediliyor ({replaced} değişim)...")
    wb.save(temp_path)
    print(f"  ✓ {temp_path}")
    
    # Orijinal dosyayı yeniden adlandır
    print(f"3️⃣  Orijinal dosya arşivleniyor...")
    if os.path.exists(old_path):
        os.remove(old_path)
    try:
        os.replace(grid_path, old_path)  # Windows'da atomik rename
    except:
        import time
        time.sleep(2)
        os.replace(grid_path, old_path)
    print(f"  ✓ {old_path}")
    
    # Geçici dosyayı asıl yere koy
    print(f"4️⃣  Güncellenmiş dosya uygulanıyor...")
    os.rename(temp_path, grid_path)
    print(f"  ✓ {grid_path}")
    
    print(f"\n✅ Başarılı! Kayseri sembolü güncellendi.")
    print(f"  Değiştirilen: {replaced} hücre (√ → ✓)")
    print(f"  Yedek: {old_path}")
    
except Exception as e:
    print(f"\n❌ Hata: {e}")
    # Rollback
    if os.path.exists(temp_path):
        os.remove(temp_path)
    if os.path.exists(old_path) and not os.path.exists(grid_path):
        os.rename(old_path, grid_path)
    exit(1)

print(f"\n{'='*80}\n")
