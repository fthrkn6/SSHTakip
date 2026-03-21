#!/usr/bin/env python3
"""
Scenarios Analysis - Excel Data Debug
Analiz endpointinin Excel'den veri çekip çekmediğini kontrol et
"""

from app import app, db
from models import Equipment
from utils_excel_grid_manager import ExcelGridManager
from datetime import date, timedelta
import os

project = 'belgrad'

with app.app_context():
    print("=" * 70)
    print("📊 SENARYO ANALİZ VERİ DEBUG")
    print("=" * 70)
    
    # 1. Excel dosyasını kontrol et
    print("\n1️⃣ EXCEL DOSYASI KONTROLÜ:")
    print("-" * 70)
    
    grid_manager = ExcelGridManager(project)
    grid_path = grid_manager.get_grid_path(app.root_path)
    
    print(f"Grid Dosyası: {grid_path}")
    print(f"Dosya Var mı: {'✅ VAR' if os.path.exists(grid_path) else '❌ YOK'}")
    
    if os.path.exists(grid_path):
        file_size = os.path.getsize(grid_path)
        print(f"Dosya Boyutu: {file_size} bytes")
    
    # 2. Excel içeriğini analiz et
    print("\n2️⃣ EXCEL İÇERİĞİ:")
    print("-" * 70)
    
    if os.path.exists(grid_path):
        import openpyxl
        wb = openpyxl.load_workbook(grid_path, data_only=True)
        ws = wb.active
        
        print(f"Grid Boyut: {ws.max_row} rows × {ws.max_column} cols")
        
        # Header oku
        headers = []
        for col in range(1, min(8, ws.max_column + 1)):
            val = ws.cell(1, col).value
            headers.append(val)
        print(f"Header (İlk 7 col): {headers}")
        
        # İlk 5 veri satırı
        print(f"\nİlk 5 Veri Satırı:")
        for row in range(2, min(7, ws.max_row + 1)):
            tarih = ws.cell(row, 1).value
            veri = [ws.cell(row, col).value for col in range(2, min(8, ws.max_column + 1))]
            print(f"  Row {row}: {tarih} | {veri}")
        
        wb.close()
    
    # 3. Availability hesabını test et
    print("\n3️⃣ AVAILABILITY HESABI (GUNLUK):")
    print("-" * 70)
    
    today = date.today()
    availability = grid_manager.get_availability_data(app.root_path, today, today)
    
    print(f"Bugün: {today}")
    print(f"Availability Data: {availability}")
    
    if availability:
        for tram, av in list(availability.items())[:5]:
            print(f"  {tram}: {av}%")
    else:
        print("  ⚠️ HIÇBIR VERİ BULUNAMADI!")
    
    # 4. Availability hesabını test et (TOPLAM)
    print("\n4️⃣ AVAILABILITY HESABI (TOPLAM):")
    print("-" * 70)
    
    availability_total = grid_manager.get_availability_data(app.root_path, None, None)
    print(f"Availability Data (TOPLAM): {availability_total}")
    
    if availability_total:
        for tram, av in list(availability_total.items())[:5]:
            print(f"  {tram}: {av}%")
    else:
        print("  ⚠️ TOPLAM HESABDA VERİ YOK!")
    
    # 5. Equipment kontrol et
    print("\n5️⃣ EQUIPMENT KONTROL:")
    print("-" * 70)
    
    equips = Equipment.query.filter_by(project_code=project).all()
    print(f"Toplam Equipment: {len(equips)}")
    print("İlk 5 Equipment:")
    for eq in equips[:5]:
        print(f"  {eq.equipment_code}: {eq.name}")
    
    print("\n" + "=" * 70)
    print("✅ DEBUG TAMAMLANDI")
    print("=" * 70)
