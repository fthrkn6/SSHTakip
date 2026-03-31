"""
SSH Takip - Bilgisayarlı Bakım Yönetim Sistemi
Bozankaya Hafif Raylı Sistem için Kapsamlı Bakım Yönetimi
EN 13306, ISO 55000, EN 15341, ISO 27001 Standartlarına Uygun
"""

from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, session, after_this_request
from flask_login import LoginManager, login_user, logout_user, login_required, current_user
from datetime import datetime
from models import db, User, Equipment, Failure, WorkOrder, MaintenancePlan, SparePartInventory, ServiceStatus
from werkzeug.utils import secure_filename
from routes.fracas import bp as fracas_bp, get_excel_path, get_column
from routes.kpi import bp as kpi_bp
from routes.service_status import bp as service_status_bp
from routes.dashboard import bp as dashboard_bp
from routes.reports import reports_bp
from routes.admin import bp as admin_bp
from routes.role_management import bp_roles
from routes.hbr import bp as hbr_bp
from routes.equipment import bp as equipment_bp
from routes.maintenance import bp as maintenance_bp
from routes.api import bp as api_bp
from utils_service_status_logger import ServiceStatusLogger
from utils_root_cause_analysis import RootCauseAnalyzer
from utils.project_manager import ProjectManager
from utils.backup_manager import BackupManager
from utils.auth_decorators import require_admin, require_project_access, check_project_in_session
from utils_km_logger import log_km_change
from utils_km_excel_logger import KMExcelLogger
from utils_km_manager import KMDataManager
from utils_daily_service_logger import log_service_status
from utils_equipment_sync import sync_equipment_with_excel
from utils_project_excel_store import (
    read_all_km,
    upsert_km,
    sync_km_excel_to_equipment,
    bootstrap_km_excel_from_equipment,
    upsert_service_status,
    sync_service_excel_to_db,
)
from utils_excel_grid_manager import ExcelGridManager, RCAExcelManager, init_excel_files
import os
import shutil
import tempfile
import logging
from datetime import timedelta

# Infrastructure: Caching & Performance
from utils_performance import init_cache, CacheManager, CacheConfig, cache_result
from utils_report_manager import template_manager, report_builder, ReportTemplateManager
from utils_ui_config import UIConfig, DARK_MODE_SCRIPT, CUSTOM_CSS

# Celery for async tasks
try:
    from celery import Celery
    from celery_config import make_celery, CeleryConfig
    CELERY_AVAILABLE = True
except ImportError:
    CELERY_AVAILABLE = False
    logger_module = logging.getLogger('ssh_takip')
    logger_module.warning("Celery not installed - async tasks disabled")

# Initialize reporting system
from utils_reporting import init_reporting_system

# Initialize logger at module level
logger = logging.getLogger('ssh_takip')
logger.setLevel(logging.INFO)
if not logger.handlers:
    handler = logging.StreamHandler()
    formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
    handler.setFormatter(formatter)
    logger.addHandler(handler)


ALLOWED_EXTENSIONS = {'pdf', 'doc', 'docx', 'xls', 'xlsx', 'dwg', 'jpg', 'png', 'jpeg'}

# Projeler listesi
PROJECTS = [
    {'code': 'belgrad', 'name': 'Belgrad', 'country': 'Sırbistan', 'flag': '🇷🇸'},
    {'code': 'iasi', 'name': 'Iași', 'country': 'Romanya', 'flag': '🇷🇴'},
    {'code': 'timisoara', 'name': 'Timișoara', 'country': 'Romanya', 'flag': '�🇴'},
    {'code': 'kayseri', 'name': 'Kayseri', 'country': 'Türkiye', 'flag': '🇹🇷'},
    {'code': 'kocaeli', 'name': 'Kocaeli', 'country': 'Türkiye', 'flag': '🇹🇷'},
    {'code': 'gebze', 'name': 'Gebze', 'country': 'Türkiye', 'flag': '🇹🇷'},
    {'code': 'samsun', 'name': 'Samsun', 'country': 'Türkiye', 'flag': '🇹🇷'},
    {'code': 'istanbul', 'name': 'İstanbul', 'country': 'Türkiye', 'flag': '🇹🇷'},
]

# Sistem üzerindeki tüm sayfalar - Yetkilendirme için
PAGES = [
    # Ana Menü
    {'id': 1, 'code': 'dashboard', 'name': 'Dashboard', 'section': 'Ana Menü'},
    
    # İş Yönetimi
    {'id': 2, 'code': 'ariza_bildir', 'name': 'Yeni Arıza Bildir', 'section': 'İş Yönetimi'},
    {'id': 3, 'code': 'ariza_listesi', 'name': 'Arıza Listesi', 'section': 'İş Yönetimi'},
    {'id': 4, 'code': 'hbr_listesi', 'name': 'HBR Listesi', 'section': 'İş Yönetimi'},
    {'id': 5, 'code': 'bakim_planlari', 'name': 'Bakım Planları', 'section': 'İş Yönetimi'},
    {'id': 6, 'code': 'tramvay_km', 'name': 'Tramvay KM', 'section': 'İş Yönetimi'},
    {'id': 7, 'code': 'servis_durumu', 'name': 'Servis Durumu', 'section': 'İş Yönetimi'},
    
    # Envanter
    {'id': 8, 'code': 'yedek_parca', 'name': 'Yedek Parça', 'section': 'Envanter'},
    
    # Raporlar & Analiz
    {'id': 9, 'code': 'fracas', 'name': 'FRACAS Analiz', 'section': 'Raporlar & Analiz'},
    {'id': 10, 'code': 'kpi', 'name': 'KPI Dashboard', 'section': 'Raporlar & Analiz'},
    {'id': 11, 'code': 'scenarios', 'name': 'Senaryo Analizi', 'section': 'Raporlar & Analiz'},
    {'id': 12, 'code': 'raporlar', 'name': 'Son Raporlar', 'section': 'Raporlar & Analiz'},
    {'id': 13, 'code': 'syslog', 'name': 'Sistem Logları', 'section': 'Raporlar & Analiz'},
    
    # Dokümantasyon
    {'id': 14, 'code': 'dokumantasyon', 'name': 'Teknik Dokümanlar', 'section': 'Dokümantasyon'},
    
    # Yönetim
    {'id': 15, 'code': 'admin_dashboard', 'name': 'Admin Paneli', 'section': 'Yönetim'},
    {'id': 16, 'code': 'kullanici_yonetimi', 'name': 'Kullanıcı Yönetimi', 'section': 'Yönetim'},
    {'id': 17, 'code': 'proje_yonetimi', 'name': 'Proje Yönetimi', 'section': 'Yönetim'},
    {'id': 18, 'code': 'yedekleme', 'name': 'Yedekleme', 'section': 'Yönetim'},
    {'id': 19, 'code': 'denetim', 'name': 'Denetim Günlüğü', 'section': 'Yönetim'},
    {'id': 20, 'code': 'yetkilendirme', 'name': 'Yetki Yönetimi', 'section': 'Yönetim'},
]


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


# Global parts cache - initialized per project
_parts_cache = {}
_parts_cache_time = {}

def load_parts_cache(project=None):
    """Excel dosyasını proje bazında yükle ve cache'e al"""
    global _parts_cache, _parts_cache_time
    import pandas as pd
    from datetime import datetime
    
    if project is None:
        project = 'belgrad'
    
    project = project.lower()
    
    # Cache'te varsa döndür
    if project in _parts_cache and _parts_cache[project]:
        return _parts_cache[project]
    
    data_dir = os.path.join(os.path.dirname(__file__), 'data', project)
    
    # data/{project}/ içinde GÜNCEL ile başlayan Excel'i bul
    part_file = None
    if os.path.exists(data_dir):
        files = os.listdir(data_dir)
        for file in files:
            # Case-insensitive extensions check
            file_upper = file.upper()
            if file_upper.startswith('GÜNCEL') and file_upper.endswith('.XLSX') and not file.startswith('~$'):
                part_file = os.path.join(data_dir, file)
                break
    else:
        logger.warning(f"Parts directory not found: {data_dir}")
    
    if not part_file or not os.path.exists(part_file):
        _parts_cache[project] = []
        return []
    
    try:
        df = pd.read_excel(part_file)
        parts = []
        
        # Excel'den parçaları oku ve cache'e koy
        for idx, row in df.iterrows():
            bilesen_no = str(row['Bileşen numarası']).strip().upper() if pd.notna(row['Bileşen numarası']) else ''
            nesne_metni = str(row['Nesne kısa metni']).strip().upper() if pd.notna(row['Nesne kısa metni']) else ''
            
            if bilesen_no or nesne_metni:
                parts.append({
                    'bilesen_no': bilesen_no,
                    'nesne_metni': nesne_metni,
                    'bilesen_no_lower': bilesen_no.lower(),
                    'nesne_metni_lower': nesne_metni.lower()
                })
        
        _parts_cache[project] = parts
        _parts_cache_time[project] = datetime.now()
        logger.info(f"Parts cache loaded ({project}): {len(parts)} parts - {part_file}")
        return parts
    except Exception as e:
        logger.error(f"Parts cache error ({project}): {e}")
        _parts_cache[project] = []
        return []


def create_app():
    try:
        app = Flask(__name__, static_folder='static', static_url_path='/static')
        
        # Configuration
        app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'bozankaya-ssh_takip-2024-gizli')
        app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///ssh_takip_bozankaya.db'
        app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
        app.config['SQLALCHEMY_ENGINE_OPTIONS'] = {
            'pool_size': 10,
            'pool_recycle': 3600,
            'pool_pre_ping': True,
        }
        app.config['UPLOAD_FOLDER'] = os.path.join(os.path.dirname(__file__), 'uploads')
        app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024
        
        # Performance Optimization
        app.config['JSON_SORT_KEYS'] = False  # Smaller JSON responses
        app.config['SEND_FILE_MAX_AGE_DEFAULT'] = 0  # No cache for HTML/templates
        app.config['JINJA_AUTO_RELOAD'] = True  # Enable auto reload for templates
        app.config['TEMPLATES_AUTO_RELOAD'] = True  # Enable auto reload for templates
        app.config['SESSION_COOKIE_HTTPONLY'] = True  # Security: prevent JavaScript access
        app.config['SESSION_COOKIE_SECURE'] = False  # Set to True if using HTTPS
        app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'  # CSRF protection
        app.config['PERMANENT_SESSION_LIFETIME'] = 3600  # 1 hour session timeout
        app.config['COMPRESS_LEVEL'] = 6  # Gzip compression level
        
        # Infrastructure Configuration
        app.config['REDIS_URL'] = os.environ.get('REDIS_URL', 'redis://localhost:6379/0')
        app.config['CACHE_DEFAULT_TIMEOUT'] = 300  # 5 minutes
        
        # Celery Configuration
        if CELERY_AVAILABLE:
            app.config.from_object(CeleryConfig)
            celery = make_celery(app)
            app.celery = celery
        
        # Enable compression
        try:
            from flask_compress import Compress
            Compress(app)
        except ImportError:
            pass  # Flask-Compress not installed, skipping
        
        # Initialize Cache Manager
        try:
            import redis
            redis_client = redis.from_url(app.config['REDIS_URL'])
            redis_client.ping()  # Test connection
            cache_mgr = init_cache(redis_client)
            logger.info("Redis cache initialized successfully")
        except Exception as e:
            logger.warning(f"Redis not available, using local cache: {e}")
            cache_mgr = init_cache(None)  # Fallback to local cache
        
        app.cache_manager = cache_mgr

        # Initialize database
        db.init_app(app)
        # Flask-Migrate entegrasyonu (optional)
        try:
            from flask_migrate import Migrate
            migrate = Migrate(app, db)
        except ImportError:
            pass

        # Initialize LoginManager
        login_manager = LoginManager()
        login_manager.init_app(app)
        login_manager.login_view = 'login'
        login_manager.login_message = 'Please log in to view this page.'
        login_manager.login_message_category = 'warning'

        @login_manager.user_loader
        def load_user(user_id):
            try:
                return db.session.get(User, int(user_id))
            except (TypeError, ValueError):
                return None
            except Exception as e:
                logger.error(f"user_loader error for user_id={user_id}: {e}")
                return None

        # Register blueprints
        app.register_blueprint(fracas_bp)
        app.register_blueprint(kpi_bp)
        app.register_blueprint(service_status_bp)
        app.register_blueprint(dashboard_bp)
        app.register_blueprint(reports_bp)
        app.register_blueprint(admin_bp)  # Admin paneli
        app.register_blueprint(bp_roles)  # Rol yönetimi
        app.register_blueprint(hbr_bp)  # HBR Yönetimi
        app.register_blueprint(equipment_bp)  # Ekipman Yönetimi
        app.register_blueprint(maintenance_bp)  # Bakım Yönetimi
        app.register_blueprint(api_bp)  # API Endpoints
        
        # Register infrastructure blueprints
        from routes.performance import bp as performance_bp
        app.register_blueprint(performance_bp)  # Performance monitoring
        
        from routes.reporting import bp as reporting_bp
        app.register_blueprint(reporting_bp)  # Advanced reporting
        
        # Project session handler
        @app.before_request
        def set_project_session():
            """
            Navbar proje seçicisinden gelen project query parametresini session'a kaydet
            """
            if current_user.is_authenticated:
                # Query parameter'dan proje kodunu al
                project_code = request.args.get('project')
                
                if project_code:
                    # Kullanıcının bu projeye erişim yetkisi var mı kontrol et
                    if current_user.can_access_project(project_code):
                        session['current_project'] = project_code
                        # Proje adını da sessionda sakla
                        projects = ProjectManager.get_all_projects()
                        for p in projects:
                            if p['code'] == project_code:
                                session['project_name'] = p['name']
                                break
                
                # Session'da proje kodu yoksa varsayılan ayarla
                if 'current_project' not in session:
                    # Admin ise belgrad, saha ise ilk atanan projeyi al
                    if current_user.is_admin():
                        session['current_project'] = 'belgrad'
                        session['project_name'] = 'Belgrad'
                    else:
                        assigned = current_user.get_assigned_projects()
                        if assigned:
                            session['current_project'] = assigned[0]
                            projects = ProjectManager.get_all_projects()
                            for p in projects:
                                if p['code'] == assigned[0]:
                                    session['project_name'] = p['name']
                                    break
        
        # ========== GLOBAL SYNC MIDDLEWARE (OPTIMIZED) ==========
        # Sync time tracking (her project için ayrı)
        _sync_cache = {}  # {project: last_sync_time}
        SYNC_INTERVAL = 3600  # 1 saat (3600 saniye)
        
        def should_sync_excel(project_code):
            """Sync'in yapılması gerekip gerekmediğini kontrol et (1 saatte 1 kere)"""
            import time
            current_time = time.time()
            last_sync = _sync_cache.get(project_code, 0)
            
            # İlk kez veya 1 saat geçmişse sync yap
            if (current_time - last_sync) >= SYNC_INTERVAL:
                _sync_cache[project_code] = current_time
                return True
            return False
        
        @app.before_request
        def global_excel_sync():
            """Excel senkronizasyonu (OPTIMIZED: 1 saatte 1 kere)"""
            if current_user.is_authenticated:
                current_project = session.get('current_project', 'belgrad')
                
                # ========== YENİ: Excel Files Init (İlk Kullanım) ==========
                try:
                    from routes.service_status import get_tram_ids_from_veriler
                    
                    # Excel dosyaları varsa init et (idempotent - zaten varsa re-create etmez)
                    equipment_codes = get_tram_ids_from_veriler(current_project)
                    if equipment_codes:
                        init_excel_files(app, current_project, equipment_codes)
                except Exception as excel_init_error:
                    logger.warning(f'Excel init hatası (devam et): {excel_init_error}')
                
                # ========== Excel Sync (1 saatte 1 kere) ==========
                
                # Sync'in gerekli olup olmadığını kontrol et
                if should_sync_excel(current_project):
                    try:
                        from routes.dashboard import sync_excel_to_equipment
                        sync_excel_to_equipment(current_project)
                        print(f'[SYNC OK] {current_project} senkronize edildi - {datetime.now().strftime("%H:%M:%S")}')
                    except Exception as e:
                        logger.error(f'Global sync hatası: {e}')
                # Sync gerekli değilse hiçbir şey yapma (çok hızlı!)
        
        # Cache control headers - disable browser cache for HTML/templates
        @app.after_request
        def add_cache_headers(response):
            """Add cache control headers to prevent browser caching of HTML/templates"""
            if response.content_type and ('text/html' in response.content_type or 'application/json' in response.content_type):
                response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate, public, max-age=0'
                response.headers['Pragma'] = 'no-cache'
                response.headers['Expires'] = '0'
            return response
        
        # Initialize reporting system
        with app.app_context():
            init_reporting_system()

        # ==================== ROUTES ====================

        @app.route('/')
        def index():
            if current_user.is_authenticated:
                return redirect(url_for('dashboard.index'))
            return redirect(url_for('login'))

        @app.route('/favicon.ico')
        def favicon():
            favicon_path = os.path.join(app.root_path, 'static', 'favicon.ico')
            if os.path.exists(favicon_path):
                return app.send_static_file('favicon.ico')
            return ('', 204)

        @app.route('/login', methods=['GET', 'POST'])
        def login():
            if current_user.is_authenticated:
                return redirect(url_for('dashboard.index'))
            
            if request.method == 'POST':
                try:
                    username = request.form.get('username')
                    password = request.form.get('password')
                    logger.info(f"\1")
                    
                    user = User.query.filter_by(username=username).first()
                    logger.info(f"\1")
                    
                    if user and user.check_password(password):
                        logger.info(f"\1")
                        login_user(user, remember=request.form.get('remember'))
                        user.last_login = datetime.now()
                        db.session.commit()
                        logger.info(f"\1")
                        
                        # Set project from form selection or default to belgrad
                        selected_project = request.form.get('project', '')
                        logger.info(f"\1")
                        logger.info(f"\1")
                        
                        if selected_project and selected_project in [p['code'] for p in PROJECTS]:
                            project = next((p for p in PROJECTS if p['code'] == selected_project), None)
                            if project:
                                session['current_project'] = selected_project
                                session['project_code'] = selected_project
                                session['project_name'] = f"{project['flag']} {project['name']}"
                                logger.info(f"\1")
                        else:
                            logger.info(f"\1")
                            session['current_project'] = 'belgrad'
                            session['project_code'] = 'belgrad'
                            session['project_name'] = '🇷🇸 Belgrad'
                        
                        next_page = request.args.get('next')
                        if next_page and next_page.startswith('/'):
                            return redirect(next_page)
                        return redirect(url_for('dashboard.index'))
                    else:
                        logger.info(f"\1")
                        flash('Invalid username or password', 'danger')
                except Exception as e:
                    logger.info(f"\1")
                    import traceback
                    traceback.print_exc()
                    flash(f'Login error: {str(e)}', 'danger')
            
            return render_template('login.html')

        @app.route('/logout')
        @login_required
        def logout():
            logout_user()
            flash('You have been logged out.', 'success')
            return redirect(url_for('login'))

        @app.route('/profile')
        @login_required
        def profile():
            """User profile page"""
            return render_template('profil.html', user=current_user)

        @app.route('/profile/update', methods=['POST'])
        @login_required
        def update_profile():
            """Update user profile"""
            current_user.full_name = request.form.get('full_name', current_user.full_name)
            current_user.email = request.form.get('email', current_user.email)
            db.session.commit()
            flash('Profile updated successfully', 'success')
            return redirect(url_for('profile'))

        # Stub routes for template compatibility
        @app.route('/yeni-ariza-bildir', methods=['GET', 'POST'])
        @login_required
        def yeni_ariza_bildir():
            import pandas as pd
            from openpyxl import load_workbook
            from openpyxl.utils import get_column_letter
            from datetime import datetime
            import os
            
            project = session.get('current_project', 'belgrad')
            # APP.ROOT_PATH kullanarak doğru lokasyonu bul
            data_dir = os.path.join(app.root_path, 'data', project)
            
            # FRACAS Excel dosyasını bul
            excel_path = None
            if os.path.exists(data_dir):
                for file in os.listdir(data_dir):
                    if 'FRACAS' in file.upper() and file.endswith('.xlsx'):
                        excel_path = os.path.join(data_dir, file)
                        break
            
            if request.method == 'GET':
                # Son FRACAS ID'yi FRACAS TEMPLATE dosyasından bul (Ariza Listesi'nden değil!)
                import tempfile
                import time
                next_fracas_id = 1
                
                # APP.ROOT_PATH kullanarak doğru lokasyonu bul
                fracas_file = os.path.join(app.root_path, 'logs', project, 'ariza_listesi', f'Fracas_{project.upper()}.xlsx')
                
                logger.info(f"\1")
                logger.info(f"\1")
                
                if os.path.exists(fracas_file):
                    try:
                        # FRACAS template dosyasını oku
                        from openpyxl import load_workbook
                        wb = load_workbook(fracas_file, data_only=True)
                        ws = wb['FRACAS']
                        
                        # E sütununda FRACAS ID'leri ara (Row 5'ten başla, Row 4 header)
                        # FRACAS template'de ID sütunu E'dir
                        ids = []
                        for row in range(5, ws.max_row + 1):
                            cell_val = ws.cell(row=row, column=5).value  # E = 5th column
                            if cell_val and isinstance(cell_val, str) and 'FF-' in str(cell_val):
                                try:
                                    num = int(str(cell_val).split('FF-')[-1])
                                    ids.append(num)
                                    logger.info(f"\1")
                                except:
                                    pass
                        
                        wb.close()
                        
                        if ids:
                            next_fracas_id = max(ids) + 1
                        else:
                            next_fracas_id = 1
                    except Exception as e:
                        logger.info(f"\1")
                        next_fracas_id = 1
                
                # Tramvaylar ve sistemler
                tramvaylar = []
                sistemler = {}
                
                # Veriler.xlsx'den sistemleri yükle
                veriler_path = None
                if os.path.exists(data_dir):
                    for file in os.listdir(data_dir):
                        if 'veriler' in file.lower() and file.endswith('.xlsx'):
                            veriler_path = os.path.join(data_dir, file)
                            break
                
                if veriler_path and os.path.exists(veriler_path):
                    try:
                        wb = load_workbook(veriler_path)
                        ws = wb['Sayfa1']
                        
                        # Sistem renk tanımları
                        KIRMIZI = 'FFFF0000'
                        SARI = 'FFFFFF00'
                        MAVI = 'FF0070C0'
                        
                        # Sütun sütun tarama
                        for col in range(1, ws.max_column + 1):
                            sistem_adi = None
                            
                            for row in range(1, ws.max_row + 1):
                                cell = ws.cell(row=row, column=col)
                                value = cell.value
                                fill = cell.fill
                                
                                color_hex = None
                                if fill and fill.start_color:
                                    color_hex = str(fill.start_color.rgb) if fill.start_color.rgb else None
                                
                                # Kırmızı = Sistem
                                if color_hex == KIRMIZI and value:
                                    sistem_adi = str(value).strip()
                                    if sistem_adi not in sistemler:
                                        sistemler[sistem_adi] = {
                                            'tedarikciler': [],
                                            'alt_sistemler': []
                                        }
                                
                                # Sarı = Tedarikçi
                                elif color_hex == SARI and value and sistem_adi:
                                    sistemler[sistem_adi]['tedarikciler'].append(str(value).strip())
                                
                                # Mavi = Alt Sistem
                                elif color_hex == MAVI and value and sistem_adi:
                                    sistemler[sistem_adi]['alt_sistemler'].append(str(value).strip())
                    except Exception as e:
                        pass
                
                # Tramvaylar, Modüller, Arıza Sınıfları ve Arıza Kaynakları - Sayfa2'den
                modules = []  # default
                ariza_siniflari = ['Kritik', 'Yüksek', 'Orta', 'Düşük']  # default
                ariza_kaynaklari = ['Fabrika Hatası', 'Kullanıcı Hatası', 'Yıpranma', 'Bilinmiyor']  # default
                ariza_tipleri = []  # default
                
                if os.path.exists(os.path.join(data_dir, 'Veriler.xlsx')):
                    try:
                        import unicodedata
                        
                        df_trams = pd.read_excel(os.path.join(data_dir, 'Veriler.xlsx'), sheet_name='Sayfa2', header=0)

                        
                        # Sütun adlarını normalize et (Türkçe karakterleri ASCII'ye çevir)
                        def normalize_col(col_name):
                            """Türkçe karakterleri normalize et"""
                            # Türkçe karakterleri değiştir
                            replacements = {
                                'ı': 'i', 'ş': 's', 'ç': 'c', 'ğ': 'g', 'ü': 'u', 'ö': 'o',
                                'I': 'I', 'Ş': 'S', 'Ç': 'C', 'Ğ': 'G', 'Ü': 'U', 'Ö': 'O'
                            }
                            result = col_name.strip().lower()
                            for tr, en in replacements.items():
                                result = result.replace(tr, en)
                            return result
                        
                        # tram_id sütununu bul
                        for col in df_trams.columns:
                            col_norm = normalize_col(col)
                            if 'tram' in col_norm and 'id' in col_norm:
                                tramvaylar = df_trams[col].dropna().unique().tolist()
                                tramvaylar = [str(int(t)) if isinstance(t, (int, float)) else str(t) for t in tramvaylar]

                                break
                        
                        # Modül sütununu bul
                        for col in df_trams.columns:
                            col_norm = normalize_col(col)
                            if col_norm == 'module':
                                modules = [str(m).strip() for m in df_trams[col].dropna().unique().tolist() if str(m).strip()]

                                break
                        
                        # Arıza Sınıfı sütununu bul
                        for col in df_trams.columns:
                            col_norm = normalize_col(col)
                            if 'ariza' in col_norm and 'sinif' in col_norm:
                                ariza_siniflari = [str(s).strip() for s in df_trams[col].dropna().unique().tolist() if str(s).strip()]

                                break
                        
                        # Arıza Kaynağı sütununu bul
                        for col in df_trams.columns:
                            col_norm = normalize_col(col)
                            if 'ariza' in col_norm and 'kaynag' in col_norm:
                                ariza_kaynaklari = [str(k).strip() for k in df_trams[col].dropna().unique().tolist() if str(k).strip()]
                                break
                        
                        # Arıza Tipi sütununu bul
                        ariza_tipleri = []
                        for col in df_trams.columns:
                            col_norm = normalize_col(col)
                            if 'ariza' in col_norm and 'tip' in col_norm:
                                ariza_tipleri = [str(t).strip() for t in df_trams[col].dropna().unique().tolist() if str(t).strip()]
                                ariza_tipleri = sorted(list(set(ariza_tipleri)))
                                break
                    except Exception as e:
                        logger.info(f"\1")
                
                sistem_detay = {k: {'tedarikciler': list(set(v['tedarikciler'])), 'alt_sistemler': list(set(v['alt_sistemler']))} for k, v in sistemler.items()}
                
                # Veriler.xlsx'in Sayfa2 B2'sinden FRACAS kodunu oku
                fracas_code = None
                if veriler_path and os.path.exists(veriler_path):
                    try:
                        wb_veriler = load_workbook(veriler_path)
                        if 'Sayfa2' in wb_veriler.sheetnames:
                            ws_sayfa2 = wb_veriler['Sayfa2']
                            fracas_code = ws_sayfa2['B2'].value
                            logger.info(f"\1")
                        wb_veriler.close()
                    except Exception as e:
                        logger.info(f"\1")
                
                # Fallback: Proje kodu kısaltması
                if not fracas_code:
                    project_code_map = {
                        'belgrad': 'BEL25',
                        'iasi': 'IAS25',
                        'timisoara': 'TIM25',
                        'kayseri': 'KAY25',
                        'kocaeli': 'KOC25',
                        'gebze': 'GEB25'
                    }
                    fracas_code = project_code_map.get(project, 'BOZ')
                
                return render_template('yeni_ariza_bildir.html', 
                                     sistem_detay=sistem_detay, 
                                     modules=modules,
                                     next_fracas_id=f"BOZ-{fracas_code}-FF-{next_fracas_id:03d}",
                                     tramvaylar=tramvaylar,
                                     sistemler=list(sistemler.keys()),
                                     ariza_siniflari=ariza_siniflari,
                                     ariza_kaynaklari=ariza_kaynaklari,
                                     ariza_tipleri=ariza_tipleri)
            else:
                # POST - Excel'e kayıt et
                try:
                    form_data = request.form.to_dict()
                    
                    # Araç Modülü - birden fazla seçim desteği
                    arac_modules = request.form.getlist('arac_module')
                    arac_module_str = ', '.join(arac_modules) if arac_modules else ''
                    
                    logger.info(f"\1")
                    logger.info(f"\1")
                    logger.info(f"\1")
                    logger.info(f"\1")
                    logger.info(f"\1")
                    
                    # FRACAS ID'yi form'dan al veya hesapla
                    fracas_id = form_data.get('fracas_id', '').strip()
                    logger.info(f"\1")
                    
                    # FRACAS TEMPLATE dosyasından FRACAS ID'yi hesapla (Ariza Listesi değil!)
                    import tempfile
                    import time
                    
                    # APP.ROOT_PATH kullanarak doğru lokasyonu bul
                    fracas_file = os.path.join(app.root_path, 'logs', project, 'ariza_listesi', f'Fracas_{project.upper()}.xlsx')
                    os.makedirs(os.path.dirname(fracas_file), exist_ok=True)
                    
                    temp_dir = tempfile.gettempdir()
                    
                    # FRACAS ID hesapla (Fracas_PROJE.xlsx'ten - FracasWriter'la senkron)
                    next_fracas_num = 1
                    if os.path.exists(fracas_file):
                        try:
                            # Main dosyayı temp'e kopyala (lock'u çözmek için)
                            temp_read_file = os.path.join(temp_dir, f"Fracas_check_{int(time.time())}.xlsx")
                            shutil.copy(fracas_file, temp_read_file)
                            time.sleep(0.2)
                            
                            # Temp'ten FRACAS sheet'ini oku
                            wb_check = load_workbook(temp_read_file, data_only=True)
                            ws_check = wb_check['FRACAS']
                            
                            # E sütununda (FRACAS ID) max numarayı bul
                            logger.info(f"\1")
                            for row in range(5, ws_check.max_row + 1):
                                cell_val = ws_check.cell(row=row, column=5).value  # E sütunu = column 5
                                if cell_val:
                                    try:
                                        if isinstance(cell_val, str) and 'FF-' in cell_val:
                                            num = int(cell_val.split('FF-')[-1])
                                            next_fracas_num = max(next_fracas_num, num + 1)
                                            logger.info(f"\1")
                                    except Exception as e:
                                        pass
                            wb_check.close()
                            os.remove(temp_read_file)
                            logger.info(f"\1")
                        except Exception as e:
                            logger.info(f"\1")
                            next_fracas_num = 1
                    
                    if not fracas_id:
                        # Veriler.xlsx'in Sayfa2 B2'sinden FRACAS kodunu oku
                        fracas_code = None
                        veriler_path = None
                        if os.path.exists(data_dir):
                            for file in os.listdir(data_dir):
                                if 'veriler' in file.lower() and file.endswith('.xlsx'):
                                    veriler_path = os.path.join(data_dir, file)
                                    break
                        
                        if veriler_path and os.path.exists(veriler_path):
                            try:
                                wb_veriler = load_workbook(veriler_path)
                                if 'Sayfa2' in wb_veriler.sheetnames:
                                    ws_sayfa2 = wb_veriler['Sayfa2']
                                    fracas_code = ws_sayfa2['B2'].value
                                wb_veriler.close()
                            except Exception as e:
                                logger.info(f"\1")
                        
                        # Fallback
                        if not fracas_code:
                            project_code_map = {
                                'belgrad': 'BEL25',
                                'iasi': 'IAS25',
                                'timisoara': 'TIM25',
                                'kayseri': 'KAY25',
                                'kocaeli': 'KOC25',
                                'gebze': 'GEB25'
                            }
                            fracas_code = project_code_map.get(project, 'BOZ')
                        
                        fracas_id = f"BOZ-{fracas_code}-FF-{next_fracas_num:03d}"
                        logger.info(f"\1")
                        form_data['fracas_id'] = fracas_id
                    
                    # YENI: Arıza Listesi Excel dosyasına yaz
                    from openpyxl.styles import Border, Side, Font, PatternFill, Alignment
                    import shutil
                    
                    ariza_listesi_dir = os.path.join(app.root_path, 'logs', project, 'ariza_listesi')
                    os.makedirs(ariza_listesi_dir, exist_ok=True)
                    
                    # Güncellik Arıza Listesi dosyasını bul
                    today_date = datetime.now().strftime('%Y%m%d')
                    ariza_listesi_file = os.path.join(ariza_listesi_dir, f"Ariza_Listesi_{project.upper()}.xlsx")
                    
                    # Temp klasörü tanımla (tüm işlemler için)
                    import tempfile
                    import time
                    temp_dir = tempfile.gettempdir()
                    
                    # Yoksa yeni dosya oluştur (temp'te, sonra taşı)
                    if not os.path.exists(ariza_listesi_file):
                        temp_file = os.path.join(temp_dir, f"Ariza_Listesi_{project.upper()}_temp_{int(time.time())}.xlsx")
                        
                        from openpyxl import Workbook
                        wb_new = Workbook()
                        ws_new = wb_new.active
                        ws_new.title = "Ariza Listesi"
                        
                        title_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
                        title_font = Font(bold=True, size=12, color="FFFFFF")
                        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                        header_font = Font(bold=True, color="FFFFFF", size=11)
                        
                        project_display = [p['name'] for p in PROJECTS if p['code'] == project][0] if project in [p['code'] for p in PROJECTS] else project.upper()
                        ws_new['A1'] = f"ARIZA LİSTESİ - {project_display.upper()} PROJESİ"
                        ws_new.merge_cells('A1:U1')
                        ws_new['A1'].font = title_font
                        ws_new['A1'].fill = title_fill
                        ws_new['A1'].alignment = Alignment(horizontal="center", vertical="center")
                        ws_new.row_dimensions[1].height = 25
                        
                        ws_new['A2'] = f"Oluşturma Tarihi: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
                        ws_new.merge_cells('A2:U2')
                        ws_new['A2'].font = Font(italic=True, size=10)
                        ws_new['A2'].alignment = Alignment(horizontal="right")
                        
                        headers = ['FRACAS ID', 'Araç No', 'Araç Modül', 'Kilometre', 'Arıza Tarihi', 'Arıza Saati', 
                                  'Sistem', 'Alt Sistem', 'Tedarikçi', 'Arıza Sınıfı', 'Arıza Kaynağı', 'Arıza Tipi',
                                  'Garanti Kapsamı', 'Arıza Tanımı', 'Yapılan İşlem', 'Aksiyon', 'Parça Kodu', 'Parça Adı', 'Parça Seri No', 'Adet',
                                  'Tamir Başlama Tarihi', 'Tamir Başlama Saati', 'Tamir Bitişi Tarihi', 'Tamir Bitişi Saati', 'Tamir Süresi', 'MTTR (dk)',
                                  'Servise Veriliş Tarihi', 'Servise Veriliş Saati', 'Kayıt Durumu', 'Personel Sayısı']
                        
                        # Sütun genişlikleri (30 sütun)
                        column_widths = [13, 10, 12, 10, 12, 10, 12, 12, 12, 14, 14, 18, 12, 20, 14, 10, 12, 12, 12, 10, 15, 14, 14, 14, 14, 12, 14, 14, 10, 12]
                        for idx, width in enumerate(column_widths, 1):
                            ws_new.column_dimensions[get_column_letter(idx)].width = width
                        border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                                       top=Side(style='thin'), bottom=Side(style='thin'))
                        
                        for col_idx, header in enumerate(headers, 1):
                            cell = ws_new.cell(row=4, column=col_idx)
                            cell.value = header
                            cell.font = header_font
                            cell.fill = header_fill
                            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                            cell.border = border
                        
                        ws_new.row_dimensions[4].height = 30
                        ws_new.freeze_panes = "A5"
                        
                        # Temp'e kaydet
                        wb_new.save(temp_file)
                        wb_new.close()
                        
                        # Final konuma taşı
                        time.sleep(0.5)
                        try:
                            shutil.move(temp_file, ariza_listesi_file)
                            logger.info(f"\1")
                        except:
                            # Zaten var olabilir, kontrol et
                            if not os.path.exists(ariza_listesi_file) and os.path.exists(temp_file):
                                shutil.copy(temp_file, ariza_listesi_file)
                                os.remove(temp_file)
                    
                    # Arıza Listesi dosyasına veri ekle (Temp dosya ile atomic işlem)
                    try:
                        # Ana dosyayı temp'e kopyala
                        temp_write_file = os.path.join(temp_dir, f"Ariza_write_{today_date}_{int(time.time())}.xlsx")
                        shutil.copy(ariza_listesi_file, temp_write_file)
                        time.sleep(0.3)
                        
                        # Temp dosyada aç, veri yaz
                        wb = load_workbook(temp_write_file)
                        ws = wb.active
                        
                        # Son satırı bul (Header 4. satırdan sonra)
                        next_row = 5
                        for row in range(5, ws.max_row + 100):
                            if not ws.cell(row=row, column=1).value:
                                next_row = row
                                break
                        else:
                            next_row = ws.max_row + 1
                        
                        logger.info(f"\1")
                        
                        # Form verilerini Excel'e yaz
                        data = [
                            form_data.get('fracas_id', ''),
                            form_data.get('arac_numarasi', ''),
                            arac_module_str,  # Birden fazla modül desteği
                            form_data.get('arac_km', ''),
                            form_data.get('hata_tarih', ''),
                            form_data.get('hata_saat', ''),
                            form_data.get('sistem', ''),
                            form_data.get('alt_sistem', ''),
                            form_data.get('tedarikci', ''),
                            form_data.get('ariza_sinifi', ''),
                            form_data.get('ariza_kaynagi', ''),
                            form_data.get('ariza_tipi', ''),
                            form_data.get('garanti_kapsami', ''),
                            form_data.get('ariza_tanimi', ''),
                            form_data.get('yapilan_islem', ''),
                            form_data.get('aksiyon', ''),
                            form_data.get('parca_kodu', ''),
                            form_data.get('parca_adi', ''),
                            form_data.get('parca_seri_no', ''),
                            form_data.get('adet', ''),
                            form_data.get('tamir_baslama_tarih', ''),
                            form_data.get('tamir_baslama_saati', ''),
                            form_data.get('tamir_bitisi_tarih', ''),
                            form_data.get('tamir_bitisi_saati', ''),
                            form_data.get('tamir_suresi', ''),
                            form_data.get('mttr', ''),
                            form_data.get('servise_verilis_tarih', ''),
                            form_data.get('servise_verilis_saati', ''),
                            'Kaydedildi',
                            form_data.get('personel_sayisi', '')
                        ]
                        
                        border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                                       top=Side(style='thin'), bottom=Side(style='thin'))
                        
                        # Zebra pattern: Row 5 = beyaz, Row 6 = gri, vb.
                        is_white = (next_row - 5) % 2 == 0
                        white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
                        gray_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                        fill = white_fill if is_white else gray_fill
                        
                        for col_idx, value in enumerate(data, 1):
                            cell = ws.cell(row=next_row, column=col_idx)
                            cell.value = value
                            cell.border = border
                            cell.font = Font(size=10)
                            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                            cell.fill = fill
                        
                        # Temp dosyaya kaydet
                        wb.save(temp_write_file)
                        wb.close()
                        time.sleep(0.5)
                        
                        # Eski dosyayı sil, temp'i ana konuma taşı (atomic)
                        if os.path.exists(ariza_listesi_file):
                            os.remove(ariza_listesi_file)
                        
                        shutil.move(temp_write_file, ariza_listesi_file)
                        time.sleep(0.3)
                        
                        # YENI: Dosya değişikliğini otomatik yedekle
                        try:
                            project = session.get('current_project', 'belgrad')
                            backup_success, backup_msg = BackupManager.backup_file(ariza_listesi_file, project)
                            if backup_success:
                                logger.info(f"\1")
                            else:
                                logger.info(f"\1")
                        except Exception as backup_err:
                            logger.info(f"\1")
                        
                        logger.info(f"\1")
                        
                        # YENI: Fracas_BELGRAD.xlsx template'ine de yaz
                        try:
                            from utils_fracas_writer import FracasWriter
                            
                            # Form verilerini FracasWriter'a hazırla
                            fracas_data = form_data.copy()
                            if arac_modules:
                                fracas_data['arac_module'] = arac_modules
                            
                            logger.info(f"\1")
                            logger.info(f"\1")
                            logger.info(f"\1")
                            
                            # Aktif proje'yi al
                            current_project = session.get('current_project', 'belgrad')
                            
                            # FracasWriter ile Fracas template'ine yaz (Flask root_path ve project geçerek)
                            writer = FracasWriter(base_path=app.root_path, project=current_project)
                            logger.info(f"\1")
                            logger.info(f"\1")
                            
                            result = writer.write_failure_data(fracas_data)
                            
                            if result.get('success'):
                                project_upper = current_project.upper()
                                logger.info(f"\1")
                                
                                # FRACAS dosyasını da yedekle
                                try:
                                    project = session.get('current_project', 'belgrad')
                                    fracas_file = ProjectManager.get_fracas_file(project)
                                    backup_success, backup_msg = BackupManager.backup_file(fracas_file, project)
                                    if backup_success:
                                        logger.info(f"\1")
                                except Exception as backup_err:
                                    logger.info(f"\1")
                                
                                flash(f'[OK] Arıza başarıyla kaydedildi: {form_data.get("fracas_id")} (Fracas + Arıza Listesi)', 'success')
                            else:
                                logger.info(f"\1")
                                flash(f'[OK] Arıza kaydedildi (Arıza Listesi), Fracas: {result.get("error", "Bilinmeyen hata")}', 'warning')
                        except Exception as fracas_error:
                            import traceback
                            logger.info(f"\1")
                            logger.info(f"\1")
                            logger.info(f"\1")
                            flash(f'[ERROR] FRACAS yazma hatası (Arıza Listesi OK): {str(fracas_error)[:150]}', 'danger')
                    except Exception as e:
                        flash(f'[ERROR] Arıza Listesi yazma hatası: {str(e)}', 'danger')
                    
                    # ===== HBR (HATA BİLDİRİM RAPORU) OLUŞTURMA =====
                    logger.info(f"\1")
                    logger.info(f"\1")
                    logger.info(f"\1")
                    
                    if request.form.get('create_hbr') == 'true':
                        try:
                            logger.info(f"\1")
                            import time
                            from openpyxl import load_workbook
                            from openpyxl.drawing.image import Image as XLImage
                            from io import BytesIO
                            from PIL import Image as PILImage
                            
                            project = session.get('current_project', 'belgrad')
                            hbr_dir = os.path.join(app.root_path, 'logs', project, 'HBR')
                            os.makedirs(hbr_dir, exist_ok=True)
                            
                            # Template yükleme
                            template_path = os.path.join(app.root_path, 'data', project, 'FR_010_R06_SSH HBR.xlsx')
                            if not os.path.exists(template_path):
                                logger.info(f"\1")
                            else:
                                logger.info(f"\1")
                                # HBR PREFIX'ini oku (Veriler.xlsx'ten veya mapping'ten)
                                hbr_code = None
                                # Case-insensitive dosya bulma
                                data_dir = os.path.join(app.root_path, 'data', project)
                                veriler_path = None
                                if os.path.exists(data_dir):
                                    for file in os.listdir(data_dir):
                                        if file.lower() == 'veriler.xlsx':
                                            veriler_path = os.path.join(data_dir, file)
                                            break
                                
                                if veriler_path and os.path.exists(veriler_path):
                                    try:
                                        wb_veriler = load_workbook(veriler_path)
                                        if 'Sayfa2' in wb_veriler.sheetnames:
                                            ws_sayfa2 = wb_veriler['Sayfa2']
                                            hbr_code = ws_sayfa2['B2'].value
                                            logger.info(f"\1")
                                        wb_veriler.close()
                                    except Exception as e:
                                        logger.info(f"\1")
                                else:
                                    logger.info(f"\1")
                                
                                # Fallback: Proje mapping'i
                                if not hbr_code:
                                    hbr_code_map = {
                                        'belgrad': 'BEL25',
                                        'iasi': 'IAS25',
                                        'timisoara': 'TIM25',
                                        'kayseri': 'KAY5+6',
                                        'kocaeli': 'KOC25',
                                        'gebze': 'GEB25'
                                    }
                                    hbr_code = hbr_code_map.get(project, 'BOZ')
                                    logger.info(f"\1")
                                
                                logger.info(f"\1")
                                logger.info(f"\1")
                                # Klasördeki mevcut HBR dosyalarını tara ve en yüksek numarayı bul
                                ncr_numbers = []
                                if os.path.exists(hbr_dir):
                                    existing_files = os.listdir(hbr_dir)
                                    logger.info(f"\1")
                                    for f in existing_files:
                                        logger.info(f"\1")
                                        if f.endswith('.xlsx') and '-NCR-' in f:
                                            try:
                                                # {code}-NCR-{number}.xlsx formatından number'ı çıkar
                                                num_str = f.split('-NCR-')[-1].replace('.xlsx', '')
                                                num = int(num_str)
                                                ncr_numbers.append(num)
                                                logger.info(f"\1")
                                            except (ValueError, IndexError) as e:
                                                logger.info(f"\1")
                                else:
                                    logger.info(f"\1")
                                
                                # Sonraki numarayı hesapla
                                ncr_counter = (max(ncr_numbers) if ncr_numbers else 0) + 1
                                ncr_number = f"{hbr_code}-NCR-{ncr_counter:03d}"
                                hbr_filename = f"{ncr_number}.xlsx"
                                hbr_filepath = os.path.join(hbr_dir, hbr_filename)
                                
                                # Template'i temp dosyaya kopyala
                                temp_hbr_file = hbr_filepath.replace('.xlsx', '_temp.xlsx')
                                shutil.copy2(template_path, temp_hbr_file)
                                
                                # Excel dosyasını aç
                                wb = load_workbook(temp_hbr_file)
                                ws = wb.active
                                
                                # Form verilerinden gerekli alanları çıkar (Gerçek form field names'lerine göre)
                                parca_kodu = form_data.get('parca_kodu', '')  # A6 - Bileşen Numarası
                                parca_adi = form_data.get('parca_adi', '')    # D6 - Nesne Kısa Metni
                                arac_numarasi = form_data.get('arac_numarasi', '')
                                arac_modules = request.form.getlist('arac_module')
                                arac_modulu = ', '.join(arac_modules) if arac_modules else ''
                                arac_km = form_data.get('arac_km', '')
                                hata_tarih = form_data.get('hata_tarih', '')
                                tedarikci = form_data.get('tedarikci', '')
                                ariza_tanimi = form_data.get('ariza_tanimi', '')
                                parca_seri_no = form_data.get('parca_seri_no', '')
                                ariza_sinifi = form_data.get('ariza_sinifi', '').strip()  # A, B, C - STRIP() EKLEDIK
                                ariza_tipi = form_data.get('ariza_tipi', '').strip()  # Çeşitli türler - STRIP() EKLEDIK
                                
                                # DEBUG: Form'dan gelen değerleri loglayalım
                                logger.info(f"\1")
                                logger.info(f"\1")
                                logger.info(f"\1")
                                
                                # Hücre yazma helper fonksiyonu - merged cells için unmerge + write + remerge
                                def write_cell(worksheet, cell_ref, value, append=False):
                                    """
                                    Hücreye değer yaz. Merged hücreler için önce unmerge et, yaz, sonra remerge et.
                                    append=True ise mevcut değerin yanına ekle.
                                    """
                                    try:
                                        # Hücreyi oku
                                        cell = worksheet[cell_ref]
                                        
                                        # Append modunu yaz
                                        if append and cell.value:
                                            value = f"{cell.value} {value}"
                                        
                                        # Merged range kontrolü
                                        merged_ranges = list(worksheet.merged_cells.ranges)
                                        merged_range = None
                                        
                                        for mr in merged_ranges:
                                            if cell_ref in mr:
                                                merged_range = mr
                                                break
                                        
                                        if merged_range:
                                            # MERGED HÜCRE YAZMA PROSEDÜRÜ
                                            merged_str = str(merged_range)
                                            logger.info(f"\1")
                                            
                                            # Adım 1: Unmerge
                                            try:
                                                worksheet.unmerge_cells(merged_str)
                                                logger.info(f"\1")
                                            except Exception as umr_err:
                                                logger.info(f"\1")
                                            
                                            # Adım 2: Değeri yaz
                                            try:
                                                worksheet[cell_ref].value = value
                                                logger.info(f"\1")
                                            except Exception as write_err:
                                                logger.info(f"\1")
                                            
                                            # Adım 3: Remerge
                                            try:
                                                worksheet.merge_cells(merged_str)
                                                logger.info(f"\1")
                                            except Exception as mrg_err:
                                                logger.info(f"\1")
                                                logger.info(f"\1")
                                        else:
                                            # NORMAL HÜCRE - basit yazma
                                            worksheet[cell_ref].value = value
                                            logger.info(f"\1")
                                    
                                    except Exception as e:
                                        logger.info(f"\1")
                                        # Son çare - zorla yaz
                                        try:
                                            worksheet[cell_ref].value = value
                                            logger.info(f"\1")
                                        except:
                                            logger.info(f"\1")

                                
                                # Hücreleri doldur (Excel satır/sütun 1-bazlı)
                                write_cell(ws, 'A6', parca_kodu)          # Bileşen Numarası
                                write_cell(ws, 'D6', parca_adi)           # Nesne Kısa Metni
                                write_cell(ws, 'E6', datetime.now().strftime("%d.%m.%Y"))  # Rapor Tarihi
                                write_cell(ws, 'G6', hata_tarih)          # Arıza Tarihi
                                write_cell(ws, 'I6', ncr_number)          # NCR Numarası
                                
                                # A8: Araç No (template'de boş olacak, A8'in eski hali korunacak)
                                write_cell(ws, 'A8', arac_numarasi, append=True)
                                
                                write_cell(ws, 'G7', arac_km)
                                write_cell(ws, 'I7', tedarikci)
                                
                                # E8: Müşteri bilgisi (veriler.xlsx'den B3'ten çek)
                                musteri_code = ''
                                veriler_path = os.path.join(app.root_path, 'data', project, 'veriler.xlsx')
                                if os.path.exists(veriler_path):
                                    try:
                                        veriler_wb = load_workbook(veriler_path)
                                        veriler_ws = veriler_wb.active
                                        musteri_code = veriler_ws['B3'].value or ''
                                        logger.info(f"\1")
                                    except Exception as e:
                                        logger.info(f"\1")
                                write_cell(ws, 'E8', musteri_code, append=True)
                                
                                # Tespit Yöntemi (Bozankaya ise F8, Müşteri ise H8)
                                ariza_tespit_yontemi = form_data.get('ariza_tespit_yontemi', '')
                                if 'bozankaya' in current_user.username.lower() or 'Bozankaya' in ariza_tespit_yontemi:
                                    write_cell(ws, 'F8', '[X]', append=True)
                                    logger.info(f"\1")
                                elif 'müşteri' in ariza_tespit_yontemi.lower():
                                    write_cell(ws, 'H8', '[X]', append=True)
                                    logger.info(f"\1")
                                
                                # NOT: muslteri_bildirimi form'da olmadığı için bu alan yazılmıyor
                                
                                # Arıza Sınıfı (Kritik, Yüksek, Orta, Düşük -> tümü için [X] yaz)
                                # Form'dan gelen değerler: başında "A-", "B-", "C-" gibi prefixler var
                                sinif_mapping = {}
                                if 'A' in ariza_sinifi or 'Kritik' in ariza_sinifi:
                                    sinif_mapping = {'cell': 'G9', 'type': 'Kritik'}
                                elif 'B' in ariza_sinifi or 'Yüksek' in ariza_sinifi or 'Fonksiyonel' in ariza_sinifi:
                                    sinif_mapping = {'cell': 'G10', 'type': 'Yüksek'}
                                elif 'C' in ariza_sinifi or 'Orta' in ariza_sinifi:
                                    sinif_mapping = {'cell': 'G11', 'type': 'Orta'}
                                elif 'D' in ariza_sinifi or 'Düşük' in ariza_sinifi:
                                    sinif_mapping = {'cell': 'G11', 'type': 'Düşük'}
                                
                                if sinif_mapping and 'cell' in sinif_mapping:
                                    write_cell(ws, sinif_mapping['cell'], '[X]', append=True)
                                    logger.info(f"\1")
                                else:
                                    logger.info(f"\1")
                                
                                # Arıza Tipi yerleştirme - Form'dan gelen değerler incelenecek
                                # Olası değerler: 'ilk_defa', 'tekrarlayan_ayni_arac', 'tekrarlayan_farkli_arac'
                                logger.info(f"\1")
                                if ariza_tipi:
                                    ariza_tipi_lower = str(ariza_tipi).lower().strip()
                                    
                                    # İlk Defa
                                    if 'ilk' in ariza_tipi_lower or 'first' in ariza_tipi_lower or 'ilk_defa' in ariza_tipi_lower:
                                        write_cell(ws, 'H9', '[X]', append=True)
                                        logger.info(f"\1")
                                    
                                    # Aynı araçta tekrarlayan
                                    if ('tekrarlayan' in ariza_tipi_lower or 'repeat' in ariza_tipi_lower) and ('aynı' in ariza_tipi_lower or 'same' in ariza_tipi_lower or 'ayni_arac' in ariza_tipi_lower):
                                        write_cell(ws, 'A12', '[X]', append=True)
                                        logger.info(f"\1")
                                    
                                    # Farklı araçta tekrarlayan
                                    if ('tekrarlayan' in ariza_tipi_lower or 'repeat' in ariza_tipi_lower) and ('farklı' in ariza_tipi_lower or 'different' in ariza_tipi_lower or 'farkli_arac' in ariza_tipi_lower):
                                        write_cell(ws, 'E12', '[X]', append=True)
                                        logger.info(f"\1")
                                else:
                                    logger.info(f"\1")
                                
                                # Arıza Tanımı
                                write_cell(ws, 'B17', ariza_tanimi)
                                
                                # Araç Modülü
                                write_cell(ws, 'D19', arac_modulu)
                                
                                # Parça Seri No
                                write_cell(ws, 'G19', parca_seri_no)
                                
                                # Fotoğraf (B20)
                                if 'hbr_fotograf' in request.files:
                                    hbr_foto = request.files['hbr_fotograf']
                                    if hbr_foto and hbr_foto.filename != '':
                                        try:
                                            # Resmi oku ve boyutlandır
                                            img = PILImage.open(hbr_foto.stream)
                                            # Boyut: 300x300 pixel (daha yüksek kalite)
                                            img.thumbnail((300, 300), PILImage.Resampling.LANCZOS)
                                            
                                            # BytesIO'ya kaydet
                                            img_buffer = BytesIO()
                                            img.save(img_buffer, format='PNG')
                                            img_buffer.seek(0)
                                            
                                            # Excel'e ekle
                                            xl_img = XLImage(img_buffer)
                                            ws.add_image(xl_img, 'B20')
                                        except Exception as img_err:
                                            logger.info(f"\1")
                                
                                # SSH Sorumlusu (B22) - GÜNCELLEME YAPMA!
                                # write_cell(ws, 'B22', current_user.username)  # SSH Sorumlusu
                                kullanici_adi_soyadi = current_user.full_name if current_user.full_name else current_user.username
                                write_cell(ws, 'B24', kullanici_adi_soyadi)  # Hata Bildiren (B24)
                                write_cell(ws, 'C24', kullanici_adi_soyadi)  # Hata Bildiren Kullanıcı
                                write_cell(ws, 'D24', kullanici_adi_soyadi)  # Hata Bildiren Kullanıcı (Tekrar)
                                
                                # Dosyayı kaydet
                                wb.save(temp_hbr_file)
                                wb.close()
                                time.sleep(0.3)
                                
                                # Temp dosyasını ana konuma taşı
                                if os.path.exists(hbr_filepath):
                                    os.remove(hbr_filepath)
                                shutil.move(temp_hbr_file, hbr_filepath)
                                
                                logger.info(f"\1")
                                flash(f'✅ HBR başarıyla oluşturuldu: {ncr_number}', 'success')
                        
                        except Exception as hbr_error:
                            import traceback
                            logger.info(f"\1")
                            logger.info(f"\1")
                            flash(f'⚠️  HBR oluşturulamadı: {str(hbr_error)[:100]}', 'warning')
                    
                    return redirect(url_for('yeni_ariza_bildir'))
                except Exception as e:
                    flash(f'❌ Kayıt hatası: {str(e)}', 'danger')
                    return redirect(url_for('yeni_ariza_bildir'))

        # HBR LİSTESİ ROUTE'U
        @app.route('/hbr-listesi')
        @login_required
        def hbr_listesi():
            """HBR (Hata Bildirim Raporu) Listesi"""
            import os
            from datetime import datetime
            
            project = session.get('current_project', 'belgrad')
            hbr_dir = os.path.join(app.root_path, 'logs', project, 'HBR')
            
            logger.info(f"\1")
            logger.info(f"\1")
            logger.info(f"\1")
            logger.info(f"\1")
            
            hbr_files = []
            total_size = 0
            latest_date = None
            
            if os.path.exists(hbr_dir):
                try:
                    files_found = os.listdir(hbr_dir)
                    logger.info(f"\1")
                    logger.info(f"\1")
                    
                    for filename in files_found:
                        # Tüm -NCR- formatındaki .xlsx dosyalarını al
                        if filename.endswith('.xlsx') and '-NCR-' in filename:
                            filepath = os.path.join(hbr_dir, filename)
                            
                            try:
                                file_stat = os.stat(filepath)
                                file_size = file_stat.st_size
                                file_time = datetime.fromtimestamp(file_stat.st_mtime)
                                
                                # Boyutu MB'ye dönüştür
                                if file_size > 1024 * 1024:
                                    size_display = f"{file_size / (1024 * 1024):.2f} MB"
                                else:
                                    size_display = f"{file_size / 1024:.2f} KB"
                                
                                hbr_files.append({
                                    'name': filename,
                                    'date': file_time.strftime('%d.%m.%Y %H:%M:%S'),
                                    'size': file_size,
                                    'size_display': size_display,
                                    'download_url': f'/hbr-download/{filename}',
                                    'open_url': f'/hbr-download/{filename}'
                                })
                                
                                total_size += file_size
                                if not latest_date or file_time > latest_date:
                                    latest_date = file_time.strftime('%d.%m.%Y %H:%M')
                                
                                logger.info(f"\1")
                            except Exception as e:
                                logger.info(f"\1")
                    
                    # Tarihe göre sırala (yeni dosyalar önce)
                    hbr_files.sort(key=lambda x: x['date'], reverse=True)
                    logger.info(f"\1")
                    
                except Exception as e:
                    logger.info(f"\1")
            else:
                logger.info(f"\1")
                try:
                    os.makedirs(hbr_dir, exist_ok=True)
                    logger.info(f"\1")
                except Exception as e:
                    logger.info(f"\1")
            
            total_size_mb = total_size / (1024 * 1024) if total_size > 0 else 0
            
            return render_template('hbr_listesi.html',
                                 hbr_files=hbr_files,
                                 total_size_mb=total_size_mb,
                                 latest_date=latest_date,
                                 project=project)

        @app.route('/hbr-listesi/sil/<filename>', methods=['POST'])
        @login_required
        def hbr_delete(filename):
            """HBR dosyasını sil"""
            import os
            from werkzeug.utils import secure_filename
            
            # Güvenlik: sadece {CODE}-NCR-{NUMBER}.xlsx formatındaki dosyaları sil
            if not (filename.endswith('.xlsx') and '-NCR-' in filename):
                return {'error': 'Geçersiz dosya adı'}, 400
            
            project = session.get('current_project', 'belgrad')
            hbr_dir = os.path.join(app.root_path, 'logs', project, 'HBR')
            filepath = os.path.join(hbr_dir, secure_filename(filename))
            
            try:
                if os.path.exists(filepath) and filepath.startswith(hbr_dir):
                    os.remove(filepath)
                    logger.info(f"\1")
                    return {'success': True, 'message': 'HBR dosyası silindi'}, 200
                else:
                    return {'error': 'Dosya bulunamadı'}, 404
            except Exception as e:
                logger.info(f"\1")
                return {'error': f'Silme hatası: {str(e)}'}, 500

        @app.route('/hbr-download')
        @login_required
        def hbr_download():
            """HBR dosyasını indir"""
            from flask import send_file
            from werkzeug.utils import secure_filename
            import os
            
            filename = request.args.get('filename', '')
            
            # Güvenlik: sadece {CODE}-NCR-{NUMBER}.xlsx formatındaki dosyaları izin ver
            if not (filename.endswith('.xlsx') and '-NCR-' in filename):
                flash('❌ Geçersiz dosya adı', 'danger')
                return redirect(url_for('hbr_listesi'))
            
            project = session.get('current_project', 'belgrad')
            hbr_dir = os.path.join(app.root_path, 'logs', project, 'HBR')
            filepath = os.path.join(hbr_dir, secure_filename(filename))
            
            # Güvenlik: path traversal kontrol et
            if not filepath.startswith(hbr_dir) or not os.path.exists(filepath):
                flash('❌ Dosya bulunamadı', 'danger')
                return redirect(url_for('hbr_listesi'))
            
            try:
                return send_file(
                    filepath, 
                    as_attachment=True, 
                    download_name=filename,
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
            except Exception as e:
                logger.info(f"\1")
                flash(f'❌ İndirme başarısız: {str(e)}', 'danger')
                return redirect(url_for('hbr_listesi'))

        @app.route('/ariza-listesi-veriler')
        @login_required
        def ariza_listesi_veriler():
            """Arıza Listesi sayfası - logs/{project}/ariza_listesi/Fracas_*.xlsx'den FRACAS verilerini oku ve göster"""
            import pandas as pd
            import numpy as np
            from openpyxl import load_workbook
            
            project = session.get('current_project', 'belgrad')
            logger.info(f"\1")
            
            # Dropdown data (yeni_ariza_bildir ile aynı)
            data_dir = os.path.join(app.root_path, 'data', project)
            sistemler = {}
            modules = []
            ariza_siniflari = ['Kritik', 'Yüksek', 'Orta', 'Düşük']
            ariza_kaynaklari = ['Fabrika Hatası', 'Kullanıcı Hatası', 'Yıpranma', 'Bilinmiyor']
            ariza_tipleri = []
            
            # Veriler.xlsx'den sistemleri yükle (Sayfa1)
            veriler_path = None
            if os.path.exists(data_dir):
                for file in os.listdir(data_dir):
                    if 'veriler' in file.lower() and file.endswith('.xlsx'):
                        veriler_path = os.path.join(data_dir, file)
                        break
            
            if veriler_path and os.path.exists(veriler_path):
                try:
                    wb = load_workbook(veriler_path)
                    ws = wb['Sayfa1']
                    
                    # Sistem renk tanımları
                    KIRMIZI = 'FFFF0000'
                    SARI = 'FFFFFF00'
                    MAVI = 'FF0070C0'
                    
                    # Sütun sütun tarama
                    for col in range(1, ws.max_column + 1):
                        sistem_adi = None
                        
                        for row in range(1, ws.max_row + 1):
                            cell = ws.cell(row=row, column=col)
                            value = cell.value
                            fill = cell.fill
                            
                            color_hex = None
                            if fill and fill.start_color:
                                color_hex = str(fill.start_color.rgb) if fill.start_color.rgb else None
                            
                            # Kırmızı = Sistem
                            if color_hex == KIRMIZI and value:
                                sistem_adi = str(value).strip()
                                if sistem_adi not in sistemler:
                                    sistemler[sistem_adi] = {
                                        'tedarikciler': [],
                                        'alt_sistemler': []
                                    }
                            
                            # Sarı = Tedarikçi
                            elif color_hex == SARI and value and sistem_adi:
                                sistemler[sistem_adi]['tedarikciler'].append(str(value).strip())
                            
                            # Mavi = Alt Sistem
                            elif color_hex == MAVI and value and sistem_adi:
                                sistemler[sistem_adi]['alt_sistemler'].append(str(value).strip())
                    
                    wb.close()
                except Exception as e:
                    logger.error(f"\2")
            
            # Sayfa2'den modules, ariza_siniflari, ariza_kaynaklari, ariza_tipleri oku
            if veriler_path and os.path.exists(veriler_path):
                try:
                    def normalize_col(col_name):
                        """Türkçe karakterleri normalize et"""
                        replacements = {
                            'ı': 'i', 'ş': 's', 'ç': 'c', 'ğ': 'g', 'ü': 'u', 'ö': 'o',
                            'I': 'I', 'Ş': 'S', 'Ç': 'C', 'Ğ': 'G', 'Ü': 'U', 'Ö': 'O'
                        }
                        result = col_name.strip().lower()
                        for tr, en in replacements.items():
                            result = result.replace(tr, en)
                        return result
                    
                    df_trams = pd.read_excel(veriler_path, sheet_name='Sayfa2', header=0)
                    
                    # Modül sütununu bul
                    for col in df_trams.columns:
                        col_norm = normalize_col(col)
                        if col_norm == 'module':
                            modules = [str(m).strip() for m in df_trams[col].dropna().unique().tolist() if str(m).strip()]
                            break
                    
                    # Arıza Sınıfı sütununu bul
                    for col in df_trams.columns:
                        col_norm = normalize_col(col)
                        if 'ariza' in col_norm and 'sinif' in col_norm:
                            ariza_siniflari = [str(s).strip() for s in df_trams[col].dropna().unique().tolist() if str(s).strip()]
                            break
                    
                    # Arıza Kaynağı sütununu bul
                    for col in df_trams.columns:
                        col_norm = normalize_col(col)
                        if 'ariza' in col_norm and 'kaynag' in col_norm:
                            ariza_kaynaklari = [str(k).strip() for k in df_trams[col].dropna().unique().tolist() if str(k).strip()]
                            break
                    
                    # Arıza Tipi sütununu bul
                    for col in df_trams.columns:
                        col_norm = normalize_col(col)
                        if 'ariza' in col_norm and 'tip' in col_norm:
                            ariza_tipleri = [str(t).strip() for t in df_trams[col].dropna().unique().tolist() if str(t).strip()]
                            ariza_tipleri = sorted(list(set(ariza_tipleri)))
                            break
                except Exception as e:
                    logger.error(f"\2")
            
            sistem_detay = {k: {'tedarikciler': list(set(v['tedarikciler'])), 'alt_sistemler': list(set(v['alt_sistemler']))} for k, v in sistemler.items()}
            
            # Birincil konum: logs/{project}/ariza_listesi/Fracas_{PROJECT}.xlsx
            ariza_listesi_dir = os.path.join(app.root_path, 'logs', project, 'ariza_listesi')
            
            ariza_listesi_file = None
            use_sheet = None
            header_row = 0
            
            if os.path.exists(ariza_listesi_dir):
                # FRACAS template dosyasını ara (Fracas_BELGRAD.xlsx, Fracas_ISTANBUL.xlsx, vb.)
                for file in os.listdir(ariza_listesi_dir):
                    if file.upper().startswith('FRACAS_') and file.endswith('.xlsx') and not file.startswith('~$'):
                        ariza_listesi_file = os.path.join(ariza_listesi_dir, file)
                        use_sheet = 'FRACAS'
                        header_row = 3  # FRACAS headers row 4 (index 3)
                        break
            
            # Fallback: Ariza_Listesi dosyası
            if not ariza_listesi_file and os.path.exists(ariza_listesi_dir):
                for file in os.listdir(ariza_listesi_dir):
                    if 'Ariza_Listesi' in file and file.endswith('.xlsx') and not file.startswith('~$'):
                        ariza_listesi_file = os.path.join(ariza_listesi_dir, file)
                        use_sheet = 'Ariza Listesi'
                        header_row = 3
                        break
            
            # Fallback: data/{project}/Veriler.xlsx
            if not ariza_listesi_file:
                veriler_file = os.path.join(os.path.dirname(__file__), 'data', project, 'Veriler.xlsx')
                if os.path.exists(veriler_file):
                    ariza_listesi_file = veriler_file
                    use_sheet = 'Veriler'
                    header_row = 0
            
            rows = []
            row_count = 0
            file_date = 'Bilinmiyor'
            column_names = []
            
            if ariza_listesi_file and os.path.exists(ariza_listesi_file):
                try:
                    # Excel'i oku
                    df = pd.read_excel(ariza_listesi_file, sheet_name=use_sheet, header=header_row)
                    column_names = list(df.columns)  # DataFrame column names
                    
                    # Kolon isimlerinden indeks haritası oluştur (hardcoded index yerine)
                    def find_col_idx(keywords, col_names=column_names):
                        """Kolon isimlerinde anahtar kelimeleri ara, indeks döndür"""
                        for idx_c, name in enumerate(col_names):
                            name_lower = str(name).lower()
                            if all(kw in name_lower for kw in keywords):
                                return idx_c
                        return None
                    
                    # Kritik kolon indekslerini otomatik tespit et
                    idx_tamir_suresi = find_col_idx(['tamir', 'süresi']) or find_col_idx(['tamir']) or 21
                    idx_arac_mttr = find_col_idx(['araç', 'mttr']) or find_col_idx(['araç', 'mdt']) or 26
                    idx_komp_mttr = find_col_idx(['komponent', 'mttr']) or find_col_idx(['komponent', 'mdt']) or 27
                    idx_parca_kodu = find_col_idx(['parça', 'kod']) or 28
                    idx_parca_adi = find_col_idx(['parça', 'ad']) or find_col_idx(['parça', 'ismi']) or 30
                    idx_parca_adedi = find_col_idx(['parça', 'adet']) or find_col_idx(['adet']) or 31
                    idx_ariza_sinifi = find_col_idx(['arıza', 'sınıf']) or 10
                    idx_detayli_bilgi = find_col_idx(['detaylı']) or find_col_idx(['detay']) or 25
                    
                    # Verileri hazırla
                    for idx, row in df.iterrows():
                        row_data = list(row)
                        # Boş satırları atla
                        if any(row_data):  # Eğer satırda herhangi bir veri varsa
                            # NaN değerlerini 'Yok' ile değiştir, sayıları int'e çevir
                            processed_row = []
                            for i, val in enumerate(row_data):
                                # Parça Adedi NaN ise 0 yap
                                if i == idx_parca_adedi and (pd.isna(val) or (isinstance(val, float) and np.isnan(val))):
                                    processed_row.append(0)
                                # Parça Kodu ve Parça Adı NaN ise 'Yok' yap
                                elif (i == idx_parca_kodu or i == idx_parca_adi) and (pd.isna(val) or (isinstance(val, float) and np.isnan(val))):
                                    processed_row.append('Yok')
                                # MTTR değerleri NaN ise 0 yap
                                elif (i == idx_arac_mttr or i == idx_komp_mttr) and (pd.isna(val) or (isinstance(val, float) and np.isnan(val))):
                                    processed_row.append(0)
                                # Float değerini int'e çevir (1.0 → 1) - NaN değilse
                                elif isinstance(val, float) and not (pd.isna(val) or np.isnan(val)) and val == int(val):
                                    processed_row.append(int(val))
                                else:
                                    processed_row.append(val)
                            
                            # Eğer Parça Kodu ve Parça Adı her ikisi de 'Yok' ise, Adedi 0 yap
                            if len(processed_row) > idx_parca_adedi and len(processed_row) > idx_parca_kodu and len(processed_row) > idx_parca_adi:
                                if processed_row[idx_parca_kodu] == 'Yok' and processed_row[idx_parca_adi] == 'Yok':
                                    processed_row[idx_parca_adedi] = 0
                            
                            # ===== MTTR HESAPLAMA =====
                            # Tamir Süresi sütunu - "3 saat 0 dakika" formatını dakikaya çevir
                            tamir_suresi_text = processed_row[idx_tamir_suresi] if len(processed_row) > idx_tamir_suresi else None
                            mttr_minutes = 0
                            
                            if tamir_suresi_text and isinstance(tamir_suresi_text, str):
                                try:
                                    # "3 saat 0 dakika" → dakika sayısına çevir
                                    parts = tamir_suresi_text.lower().split()
                                    saat = 0
                                    dakika = 0
                                    
                                    for i, part in enumerate(parts):
                                        if 'saat' in part and i > 0:
                                            try:
                                                saat = int(parts[i-1])
                                            except:
                                                pass
                                        if 'dakika' in part and i > 0:
                                            try:
                                                dakika = int(parts[i-1])
                                            except:
                                                pass
                                    
                                    mttr_minutes = saat * 60 + dakika
                                except:
                                    mttr_minutes = 0
                            
                            # Eğer processed_row'un uzunluğu yeterli değilse doldur
                            max_idx = max(idx_parca_adedi, idx_komp_mttr, idx_arac_mttr) + 1
                            while len(processed_row) < max_idx:
                                processed_row.append(0)
                            
                            # Araç MTTR / MDT ve Komponent MTTR / MDT
                            
                            ariza_sinifi = processed_row[idx_ariza_sinifi] if len(processed_row) > idx_ariza_sinifi else None
                            detayli_bilgi = processed_row[idx_detayli_bilgi] if len(processed_row) > idx_detayli_bilgi else None
                            
                            # Komponent MTTR = MTTR (dk)
                            processed_row[idx_komp_mttr] = mttr_minutes
                            
                            # Araç MTTR: IF (A or B) OR (Servise Engel) THEN Komponent MTTR ELSE 0
                            if (ariza_sinifi in ['A', 'B']) or (detayli_bilgi and 'Servise Engel' in str(detayli_bilgi)):
                                processed_row[idx_arac_mttr] = mttr_minutes
                            else:
                                processed_row[idx_arac_mttr] = 0
                            
                            rows.append(processed_row)
                    
                    row_count = len(rows)
                    
                    # Dosya tarihi
                    file_mtime = os.path.getmtime(ariza_listesi_file)
                    file_date = datetime.fromtimestamp(file_mtime).strftime('%d.%m.%Y %H:%M')
                    
                    logger.info(f"\2")
                    
                except Exception as e:
                    logger.error(f"\2")
                    flash(f'[WARNING] Veri okuma hatası: {str(e)}', 'warning')
            else:
                flash(f'⚠️ Bugünün Arıza Listesi dosyası bulunamadı', 'warning')
            
            return render_template('ariza_listesi.html', 
                                 rows=rows, 
                                 column_names=column_names,
                                 row_count=row_count,
                                 file_date=file_date,
                                 enumerate=enumerate,
                                 sistem_detay=sistem_detay,
                                 modules=modules,
                                 sistemler=list(sistemler.keys()),
                                 ariza_siniflari=ariza_siniflari,
                                 ariza_kaynaklari=ariza_kaynaklari,
                                 ariza_tipleri=ariza_tipleri)

        @app.route('/ariza-listesi-veriler/process', methods=['POST'])
        @login_required
        def ariza_listesi_veriler_process():
            """Arıza Listesi verilerini işle - FRACAS dosyasından veri okuma"""
            
            logger.info("\1")
            
            try:
                project = session.get('current_project', 'belgrad')
                ariza_listesi_dir = os.path.join(app.root_path, 'logs', project, 'ariza_listesi')
                
                # FRACAS dosyasını bul
                fracas_file = None
                if os.path.exists(ariza_listesi_dir):
                    for file in os.listdir(ariza_listesi_dir):
                        if file.upper().startswith('FRACAS_') and file.endswith('.xlsx') and not file.startswith('~$'):
                            fracas_file = os.path.join(ariza_listesi_dir, file)
                            break
                
                if not fracas_file or not os.path.exists(fracas_file):
                    flash('❌ FRACAS dosyası bulunamadı', 'danger')
                    return redirect(url_for('ariza_listesi_veriler'))
                
                flash(f'✅ FRACAS verileri başarıyla işlendi!', 'success')
                
            except Exception as e:
                flash(f'❌ İşlem hatası: {str(e)}', 'danger')
            
            return redirect(url_for('ariza_listesi_veriler'))

        @app.route('/ariza-listesi-veriler/export')
        @login_required
        def ariza_listesi_veriler_export():
            """FRACAS Excel dosyasını indir (project'e göre Fracas_BELGRAD.xlsx, Fracas_ISTANBUL.xlsx vb.)"""
            import pandas as pd
            import tempfile
            import shutil
            import time
            
            project = session.get('current_project', 'belgrad')
            ariza_listesi_dir = os.path.join(app.root_path, 'logs', project, 'ariza_listesi')
            
            # FRACAS dosyasını bul
            fracas_file = None
            if os.path.exists(ariza_listesi_dir):
                for file in os.listdir(ariza_listesi_dir):
                    if file.upper().startswith('FRACAS_') and file.endswith('.xlsx') and not file.startswith('~$'):
                        fracas_file = os.path.join(ariza_listesi_dir, file)
                        break
            
            if not fracas_file or not os.path.exists(fracas_file):
                flash('❌ FRACAS dosyası bulunamadı', 'danger')
                return redirect(url_for('ariza_listesi_veriler'))
            
            try:
                from flask import send_file
                # Temp dosya oluştur (Windows locking sorunu için)
                temp_dir = tempfile.gettempdir()
                temp_file = os.path.join(temp_dir, f"{int(time.time() * 1000)}_{os.path.basename(fracas_file)}")
                shutil.copy(fracas_file, temp_file)
                
                # Dosya adını belirle (Fracas_BELGRAD.xlsx, Fracas_ISTANBUL.xlsx, vb.)
                filename = os.path.basename(fracas_file)
                
                @after_this_request
                def cleanup(response):
                    try:
                        if os.path.exists(temp_file):
                            os.remove(temp_file)
                    except OSError:
                        pass
                    return response
                
                return send_file(
                    temp_file,
                    as_attachment=True,
                    download_name=filename,
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
            except Exception as e:
                flash(f'[ERROR] İndirme hatas: {str(e)}', 'danger')
                return redirect(url_for('ariza_listesi_veriler'))

        @app.route('/api/projects')
        def get_projects():
            """Systemde mevcut tüm projeleri döndür (data/ + projects_config.json birleştir)"""
            import os
            try:
                projects_map = {}
                
                # 1) projects_config.json'dan aktif projeleri oku
                try:
                    config_path = os.path.join(os.path.dirname(__file__), 'projects_config.json')
                    if os.path.exists(config_path):
                        with open(config_path, 'r', encoding='utf-8') as f:
                            config = json.load(f)
                        for p in config.get('projects', []):
                            if p.get('status') == 'aktif':
                                projects_map[p['code']] = {
                                    'code': p['code'],
                                    'name': p.get('name', p['code'].capitalize()),
                                    'has_veriler': False
                                }
                except Exception:
                    pass
                
                # 2) data/ klasöründen Veriler.xlsx olan projeleri ekle
                projects_dir = os.path.join(os.path.dirname(__file__), 'data')
                if os.path.exists(projects_dir):
                    for item in os.listdir(projects_dir):
                        item_path = os.path.join(projects_dir, item)
                        if os.path.isdir(item_path):
                            excel_path = os.path.join(item_path, 'Veriler.xlsx')
                            has_veriler = os.path.exists(excel_path)
                            if item in projects_map:
                                projects_map[item]['has_veriler'] = has_veriler
                            elif has_veriler:
                                projects_map[item] = {
                                    'code': item,
                                    'name': item.capitalize(),
                                    'has_veriler': True
                                }
                
                projects = sorted(projects_map.values(), key=lambda x: x['code'])
                return jsonify(projects)
            except Exception as e:
                print(f'[ERROR] /api/projects: {e}')
                return jsonify([]), 200

        @app.route('/api/failure-by-fracas-id', methods=['GET'])
        @login_required
        def get_failure_by_fracas_id():
            """FRACAS ID'den Failure record'unu bul"""
            try:
                from models import Failure
                
                fracas_id = request.args.get('fracas_id', '')
                project = session.get('current_project', 'belgrad')
                
                if not fracas_id:
                    return jsonify({'error': 'FRACAS ID required'}), 400
                
                # Database'de failure kaydını ara (fracas_id field'ından)
                failure = Failure.query.filter_by(
                    fracas_id=fracas_id,
                    project_code=project
                ).first()
                
                if failure:
                    return jsonify({
                        'id': failure.id,
                        'failure_code': failure.failure_code,
                        'found': True
                    })
                else:
                    return jsonify({
                        'found': False,
                        'fracas_id': fracas_id
                    })
            except Exception as e:
                logger.error(f"[API] failure-by-fracas-id hatası: {e}")
                return jsonify({
                    'error': str(e),
                    'found': False
                }), 500

        @app.route('/api/create-hbr-from-fracas', methods=['POST'])
        @login_required
        def create_hbr_from_fracas():
            """FRACAS verilerinden HBR oluştur"""
            try:
                import shutil
                from openpyxl import load_workbook
                
                data = request.get_json()
                fracas_id = data.get('fracas_id', '')
                arac_no = data.get('arac_no', '')
                ariza_tanimi = data.get('ariza_tanimi', '')
                ariza_sinifi = data.get('ariza_sinifi', '')
                
                if not fracas_id:
                    return jsonify({'success': False, 'error': 'FRACAS ID required'}), 400
                
                project = session.get('current_project', 'belgrad')
                
                # HBR klasörü oluştur
                hbr_dir = os.path.join(app.root_path, 'logs', project, 'HBR')
                os.makedirs(hbr_dir, exist_ok=True)
                
                # HBR template dosyasını bul
                template_paths = [
                    os.path.join(app.root_path, 'data', project, 'FR_010_R06_SSH HBR.xlsx'),
                    os.path.join(app.root_path, 'logs', project, 'FR_010_R06_SSH HBR.xlsx'),
                    os.path.join(app.root_path, 'data', 'belgrad', 'FR_010_R06_SSH HBR.xlsx'),
                ]
                
                template_file = None
                for path in template_paths:
                    if os.path.exists(path):
                        template_file = path
                        break
                
                if not template_file:
                    return jsonify({'success': False, 'error': 'HBR template dosyası bulunamadı'}), 400
                
                # HBR dosya adı oluştur
                hbr_filename = f"{arac_no}-{fracas_id}.xlsx"
                hbr_filepath = os.path.join(hbr_dir, hbr_filename)
                
                # Template'i kopyala
                shutil.copy(template_file, hbr_filepath)
                
                try:
                    # Excel dosyasını aç var data'ları doldur
                    wb = load_workbook(hbr_filepath)
                    ws = wb.active
                    
                    # Merged cell'leri skip et - sadece normal cell'lere yaz
                    # B3: Araç No, B4: FRACAS ID, B5: Arıza Tanımı, B6: Arıza Sınıfı
                    cells_to_fill = [
                        ('B3', arac_no),
                        ('B4', fracas_id),
                        ('B5', str(ariza_tanimi)[:100]),
                        ('B6', ariza_sinifi),
                    ]
                    
                    for cell_addr, value in cells_to_fill:
                        try:
                            cell = ws[cell_addr]
                            # Merged cell'ler skip et
                            if not isinstance(cell, type(ws['A1'])) or cell.data_type != 'f':
                                cell.value = value
                        except:
                            pass  # Merged cell veya write-protected ise skip
                    
                    wb.save(hbr_filepath)
                    wb.close()
                except Exception as excel_error:
                    logger.warning(f"[HBR] Excel header doldurma hatası (template OK): {excel_error}")
                    # Template copy edildi, header doldurma başarısız olsa da dosya OK
                
                return jsonify({
                    'success': True,
                    'file_name': hbr_filename,
                    'message': 'HBR başarıyla oluşturuldu'
                })
            
            except Exception as e:
                logger.error(f"[API] HBR oluşturma hatası: {e}")
                return jsonify({
                    'success': False,
                    'error': str(e)
                }), 500

        @app.route('/api/edit-fracas-inline', methods=['POST'])
        @login_required
        def edit_fracas_inline():
            """FRACAS Excel dosyasını doğrudan düzenle (Inline)"""
            try:
                from openpyxl import load_workbook
                
                data = request.get_json()
                fracas_id = data.get('fracas_id', '')
                arac_no = data.get('arac_no', '')
                sistem = data.get('sistem', '')
                hata_tarihi = data.get('hata_tarihi', '')
                ariza_tanimi = data.get('ariza_tanimi', '')
                ariza_sinifi = data.get('ariza_sinifi', '')
                yapilan_islem = data.get('yapilan_islem', '')
                ariza_kaynagi = data.get('ariza_kaynagi', '')
                ariza_tipi = data.get('ariza_tipi', '')
                garanti_kapsami = data.get('garanti_kapsami', '')
                tedarikci = data.get('tedarikci', '')
                alt_sistem = data.get('alt_sistem', '')
                arac_modul = data.get('arac_modul', '')
                kilometre = data.get('kilometre', '')
                ariza_tespit_yontemi = data.get('ariza_tespit_yontemi', '')
                aksiyon = data.get('aksiyon', '')
                detayli_bilgi = data.get('detayli_bilgi', '')
                tamir_baslama_tarih = data.get('tamir_baslama_tarih', '')
                tamir_baslama_saati = data.get('tamir_baslama_saati', '')
                tamir_bitisi_tarih = data.get('tamir_bitisi_tarih', '')
                tamir_bitisi_saati = data.get('tamir_bitisi_saati', '')
                
                if not fracas_id:
                    return jsonify({'success': False, 'error': 'FRACAS ID gereklidir'}), 400
                
                project = session.get('current_project', 'belgrad')
                
                # FRACAS dosya yolunu al
                fracas_file = os.path.join(app.root_path, 'logs', project, 'ariza_listesi', f'Fracas_{project.upper()}.xlsx')
                
                if not os.path.exists(fracas_file):
                    return jsonify({'success': False, 'error': f'FRACAS dosyası bulunamadı: {fracas_file}'}), 404
                
                logger.info(f"[EDIT-FRACAS] Dosya: {fracas_file}, FRACAS ID: {fracas_id}")
                
                # Excel dosyasını aç
                wb = load_workbook(fracas_file)
                ws = wb.active
                
                # Header satırından kolon indekslerini otomatik tespit et
                header_row_data = list(ws.iter_rows(min_row=1, max_row=1))[0] if ws.max_row > 0 else []
                col_map = {}
                for ci, cell in enumerate(header_row_data):
                    if cell.value:
                        col_map[str(cell.value).lower().strip()] = ci
                
                def find_edit_col(keywords, default_idx):
                    """Header'dan kolon indeksini bul, bulamazsa varsayılanı kullan"""
                    for key, idx_v in col_map.items():
                        if all(kw in key for kw in keywords):
                            return idx_v
                    return default_idx
                
                # Kolon indeksleri (header'dan otomatik, fallback hardcoded)
                ci_arac_no = find_edit_col(['araç', 'no'], 1)
                ci_arac_modul = find_edit_col(['araç', 'modül'], 2)
                ci_km = find_edit_col(['kilometre'], 3) if find_edit_col(['kilometre'], None) is not None else find_edit_col(['km'], 3)
                ci_fracas_id = find_edit_col(['fracas'], 4)
                ci_hata_tarih = find_edit_col(['hata', 'tarih'], 5)
                ci_sistem = find_edit_col(['sistem'], 6)
                ci_alt_sistem = find_edit_col(['alt', 'sistem'], 7)
                ci_tedarikci = find_edit_col(['tedarikçi'], 8)
                ci_ariza_tanimi = find_edit_col(['arıza', 'tanım'], 9)
                ci_ariza_sinifi = find_edit_col(['arıza', 'sınıf'], 10)
                ci_ariza_kaynagi = find_edit_col(['arıza', 'kaynak'], 11)
                ci_yapilan_islem = find_edit_col(['yapılan', 'işlem'], 12)
                ci_aksiyon = find_edit_col(['aksiyon'], 13)
                ci_garanti = find_edit_col(['garanti'], 14)
                ci_tespit_yontemi = find_edit_col(['tespit'], 15)
                ci_tamir_baslama_tarih = find_edit_col(['tamir', 'başlama', 'tarih'], 17)
                ci_tamir_baslama_saat = find_edit_col(['tamir', 'başlama', 'saat'], 18)
                ci_tamir_bitis_tarih = find_edit_col(['tamir', 'bitiş', 'tarih'], 19) if find_edit_col(['tamir', 'bitiş', 'tarih'], None) is not None else find_edit_col(['tamir', 'bitişi', 'tarih'], 19)
                ci_tamir_bitis_saat = find_edit_col(['tamir', 'bitiş', 'saat'], 20) if find_edit_col(['tamir', 'bitiş', 'saat'], None) is not None else find_edit_col(['tamir', 'bitişi', 'saat'], 20)
                ci_ariza_tipi = find_edit_col(['arıza', 'tip'], 24)
                ci_detayli = find_edit_col(['detaylı'], 25)
                
                row_found = False
                for row_num, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
                    # FRACAS ID sütunu
                    cell_value = str(row[ci_fracas_id].value or '').strip()
                    
                    if cell_value == fracas_id:
                        logger.info(f"[EDIT-FRACAS] Satır bulundu: {row_num}")
                        
                        # Değerleri güncelle (indeksler 0-başlangıçlı)
                        # B=1: Araç No, C=2: Sistem, D=3: Hata Tarihi
                        # J=9: Arıza Tanımı, K=10: Arıza Sınıfı, M=12: Yapılan İşlem
                        
                        try:
                            # Araç Bilgileri
                            row[ci_arac_no].value = arac_no if arac_no else row[ci_arac_no].value
                            row[ci_arac_modul].value = arac_modul if arac_modul else row[ci_arac_modul].value
                            row[ci_km].value = kilometre if kilometre else row[ci_km].value
                            
                            # Arıza Zamanı
                            if hata_tarihi:
                                row[ci_hata_tarih].value = hata_tarihi
                            
                            # Sistem Bilgileri
                            row[ci_sistem].value = sistem if sistem else row[ci_sistem].value
                            row[ci_alt_sistem].value = alt_sistem if alt_sistem else row[ci_alt_sistem].value
                            row[ci_tedarikci].value = tedarikci if tedarikci else row[ci_tedarikci].value
                            
                            # Arıza Detayları
                            row[ci_ariza_tanimi].value = ariza_tanimi if ariza_tanimi else row[ci_ariza_tanimi].value
                            row[ci_ariza_sinifi].value = ariza_sinifi if ariza_sinifi else row[ci_ariza_sinifi].value
                            row[ci_ariza_kaynagi].value = ariza_kaynagi if ariza_kaynagi else row[ci_ariza_kaynagi].value
                            row[ci_yapilan_islem].value = yapilan_islem if yapilan_islem else row[ci_yapilan_islem].value
                            row[ci_aksiyon].value = aksiyon if aksiyon else row[ci_aksiyon].value
                            row[ci_garanti].value = garanti_kapsami if garanti_kapsami else row[ci_garanti].value
                            row[ci_tespit_yontemi].value = ariza_tespit_yontemi if ariza_tespit_yontemi else row[ci_tespit_yontemi].value
                            
                            # Tamir Zamanları
                            row[ci_tamir_baslama_tarih].value = tamir_baslama_tarih if tamir_baslama_tarih else row[ci_tamir_baslama_tarih].value
                            row[ci_tamir_baslama_saat].value = tamir_baslama_saati if tamir_baslama_saati else row[ci_tamir_baslama_saat].value
                            row[ci_tamir_bitis_tarih].value = tamir_bitisi_tarih if tamir_bitisi_tarih else row[ci_tamir_bitis_tarih].value
                            row[ci_tamir_bitis_saat].value = tamir_bitisi_saati if tamir_bitisi_saati else row[ci_tamir_bitis_saat].value
                            
                            # Diğer Arıza Bilgileri
                            row[ci_ariza_tipi].value = ariza_tipi if ariza_tipi else row[ci_ariza_tipi].value
                            row[ci_detayli].value = detayli_bilgi if detayli_bilgi else row[ci_detayli].value
                            
                            row_found = True
                            logger.info(f"[EDIT-FRACAS] Satır {row_num} güncellendi - Tamir zamanları: {tamir_baslama_tarih} {tamir_baslama_saati}")
                        except Exception as cell_error:
                            logger.error(f"[EDIT-FRACAS] Cell yazma hatası satır {row_num}: {cell_error}")
                        
                        break
                
                if not row_found:
                    wb.close()
                    return jsonify({'success': False, 'error': f'FRACAS ID bulunamadı: {fracas_id}'}), 404
                
                # Dosyayı kaydet
                try:
                    wb.save(fracas_file)
                    wb.close()
                    logger.info(f"[EDIT-FRACAS] Dosya kaydedildi: {fracas_file}")
                except Exception as save_error:
                    logger.error(f"[EDIT-FRACAS] Dosya kaydetme hatası: {save_error}")
                    return jsonify({'success': False, 'error': f'Dosya kaydetme hatası: {save_error}'}), 500
                
                return jsonify({
                    'success': True,
                    'message': 'FRACAS verisi başarıyla güncellendi',
                    'fracas_id': fracas_id
                })
            
            except Exception as e:
                logger.error(f"[API] edit-fracas-inline hatası: {e}")
                return jsonify({
                    'success': False,
                    'error': str(e)
                }), 500

        @app.route('/api/parts-lookup', methods=['GET'])
        @login_required
        def parts_lookup():
            """Bileşen numarası - Nesne kısa metni arasında hızlı lookup yapıyor (proje-bazlı)"""
            
            query = request.args.get('q', '').strip()
            project = request.args.get('project') or session.get('current_project', 'belgrad')
            
            if not query or len(query) < 2:
                return jsonify([])
            
            # Cache'i proje için yükle
            parts_cache = load_parts_cache(project)
            
            if not parts_cache:
                return jsonify([])
            
            query_lower = query.lower()
            query_upper = query.upper()
            results = []
            
            # Hızlı arama: tam eşleşme başta, sonra kısmi
            # 1. Bileşen numarası ile tam başlangıç eşleşmesi
            for part in parts_cache:
                if part['bilesen_no'].startswith(query_upper):
                    results.append({
                        'bilesen_no': part['bilesen_no'],
                        'nesne_metni': part['nesne_metni']
                    })
                    if len(results) >= 15:
                        return jsonify(results)
            
            # 2. Nesne metni ile tam başlangıç eşleşmesi
            for part in parts_cache:
                if part['nesne_metni'].startswith(query_upper):
                    # Duplike kontrol
                    if not any(r['bilesen_no'] == part['bilesen_no'] for r in results):
                        results.append({
                            'bilesen_no': part['bilesen_no'],
                            'nesne_metni': part['nesne_metni']
                        })
                    if len(results) >= 15:
                        return jsonify(results)
            
            # 3. Bileşen numarası veya Nesne metni içinde kısmi eşleşme
            for part in parts_cache:
                if query_upper in part['bilesen_no'] or query_upper in part['nesne_metni']:
                    # Duplike kontrol
                    if not any(r['bilesen_no'] == part['bilesen_no'] for r in results):
                        results.append({
                            'bilesen_no': part['bilesen_no'],
                            'nesne_metni': part['nesne_metni']
                        })
                    if len(results) >= 15:
                        return jsonify(results)
            
            return jsonify(results)

        @app.route('/ekipmanlar')
        @login_required
        def ekipmanlar():
            current_project = session.get('current_project', 'belgrad')
            equipment = Equipment.query.filter_by(project_code=current_project).all()
            return render_template('ekipmanlar.html', equipment=equipment)

        @app.route('/arizalar')
        @login_required
        def arizalar():
            """Redirect to ariza_listesi_veriler page"""
            return redirect(url_for('ariza_listesi_veriler'))

        @app.route('/is-emirleri')
        @login_required
        def is_emirleri():
            orders = WorkOrder.query.all()
            stats = {
                'toplam': WorkOrder.query.count(),
                'beklemede': WorkOrder.query.filter_by(status='pending').count() if hasattr(WorkOrder, 'status') else 0,
                'devam_ediyor': WorkOrder.query.filter_by(status='in_progress').count() if hasattr(WorkOrder, 'status') else 0,
                'tamamlandi': WorkOrder.query.filter_by(status='completed').count() if hasattr(WorkOrder, 'status') else 0,
            }
            return render_template('is_emirleri.html', orders=orders, stats=stats)

        @app.route('/bakim-planlari')
        @login_required
        def bakim_planlari():
            import json
            
            # Proje bazlı başlat
            project_code = session.get('current_project', 'belgrad').lower()
            
            # NOTE: Equipment tablosu tek kaynak - Excel sync devre dışı
            # KM verileri sadece /tramvay-km sayfasından girilmelidir
            
            # Equipment verilerini yükle (KM bilgileri)
            try:
                from utils_project_excel_store import get_tramvay_list_with_km
                equipments = get_tramvay_list_with_km(project_code)
            except Exception as e:
                logger.error(f"bakim_planlari equipment error: {e}")
                equipments = Equipment.query.filter_by(project_code=project_code).all()
            
            # Bakım verilerini yükle (proje bazlı SADECE - fallback YOK!)
            maintenance_file = os.path.join(os.path.dirname(__file__), 'data', project_code, 'maintenance.json')
            maintenance_data = {}
            
            # SADECE proje-spesifik veriyi yükle, fallback yapma
            if os.path.exists(maintenance_file):
                with open(maintenance_file, 'r', encoding='utf-8') as f:
                    maintenance_data = json.load(f)
            
            return render_template('bakim_planlari.html', 
                                 maintenance_data=maintenance_data,
                                 equipments=equipments,
                                 project_name=session.get('project_name', 'Belgrad'))
        
        @app.route('/api/bakim-plani-tablosu')
        @login_required
        def bakim_plani_tablosu():
            """Tüm KM noktalarında yapılması gereken bakımları döndür"""
            import json
            
            current_project = session.get('current_project', 'belgrad').lower()
            
            # SYNC: Excel ile Database'i eşitle
            sync_equipment_with_excel(current_project)
            
            # maintenance.json'u proje bazlı yükle
            maintenance_file = os.path.join(os.path.dirname(__file__), 'data', current_project, 'maintenance.json')
            maintenance_data = {}
            
            # SADECE proje-spesifik veriyi yükle, fallback yapma
            if os.path.exists(maintenance_file):
                with open(maintenance_file, 'r', encoding='utf-8') as f:
                    maintenance_data = json.load(f)
            
            # Bakım seviyelerini sırala (70K ve 140K hariç - bunlar sadece specific KM'lerde)
            sorted_levels = sorted([(k, v['km']) for k, v in maintenance_data.items() 
                                   if k not in ['70K', '140K']], key=lambda x: x[1])
            
            # Tüm KM noktaları için bakımları hesapla
            schedule = []
            max_km = 300000
            
            for km in range(0, max_km + 1000, 1000):
                applicable = []
                
                for level_name, level_km in sorted_levels:
                    if km > 0 and km % level_km == 0:
                        works = maintenance_data[level_name].get('works', [])
                        works_count = len([w for w in works if w.startswith('BOZ')])
                        applicable.append({
                            'level': level_name,
                            'km': level_km,
                            'works': works_count,
                            'ratio': km // level_km
                        })
                
                # SPECIAL: 72000 KM'de 70K bakımını da ekle
                if km == 72000:
                    works = maintenance_data.get('70K', {}).get('works', [])
                    works_count = len([w for w in works if w.startswith('BOZ')])
                    applicable.append({
                        'level': '70K',
                        'km': 70000,
                        'works': works_count,
                        'ratio': 1  # Special case olduğu için 1
                    })
                
                if applicable:
                    total_works = sum(m['works'] for m in applicable)
                    
                    # Kapsamı belirle
                    if total_works >= 30:
                        scope = 'urgent'
                        scope_label = 'ÇOK KAPSAMLI'
                    elif total_works >= 20:
                        scope = 'heavy'
                        scope_label = 'KAPSAMLI'
                    else:
                        scope = 'normal'
                        scope_label = 'ORTA'
                    
                    schedule.append({
                        'km': km,
                        'maintenances': [f"{m['level']} (×{m['ratio']})" for m in applicable],
                        'maintenance_detail': applicable,
                        'total_works': total_works,
                        'scope': scope,
                        'scope_label': scope_label
                    })
            
            return jsonify(schedule)
        
        @app.route('/api/bakim-verileri')
        @login_required
        def bakim_verileri():
            """Tüm araçların bakım durumunu tablo formatında döndür (Equipment tablosundan direkt okuma)"""
            import json
            from datetime import datetime
            
            current_project = session.get('current_project', 'belgrad').lower()
            
            # SYNC: Excel ile Database'i eşitle
            sync_equipment_with_excel(current_project)
            
            # SINGLE SOURCE OF TRUTH: Equipment tablosu
            # Excel senkronizasyonu devre dışı - sadece Equipment'dan okuyoruz
            
            # Bakım verilerini yükle (proje bazlı SADECE)
            maintenance_file = os.path.join(os.path.dirname(__file__), 'data', current_project, 'maintenance.json')
            maintenance_data = {}
            
            if os.path.exists(maintenance_file):
                with open(maintenance_file, 'r', encoding='utf-8') as f:
                    maintenance_data = json.load(f)
            
            # Bakım KM'lerini sırayla al
            maintenance_levels = sorted([k for k in maintenance_data.keys()], 
                                       key=lambda x: int(x.replace('K', '')) * 1000)
            
            # Equipment'dan tüm araçları al (tek kaynak)
            # Excel'den filtrele yapma - Excel'i Equipment'dan populate ediyor,
            # o yüzden direkt Equipment'dan araçları al
            equipment_list = Equipment.query.filter(
                Equipment.parent_id == None,
                Equipment.project_code == current_project
            ).order_by(Equipment.equipment_code).all()
            
            logger.info(f"\1")
            
            # Kontrol: tüm Equipment'lerin project_code'larını göster (debug)
            all_equipment_projects = Equipment.query.filter(Equipment.parent_id == None).with_entities(Equipment.project_code, Equipment.equipment_code).limit(50).all()
            logger.info(f"\1")
            
            # Sonuç listesi
            result = []
            
            for eq in equipment_list:
                tram_id = str(eq.equipment_code)  # equipment_code kullan (1531, 1532...)
                # SINGLE SOURCE: Equipment tablosundan KM al (hiçbir yerde değişiklik olmayacak)
                current_km = getattr(eq, 'current_km', 0) or 0
                
                # Tüm bakım seviyelerini hesapla ve en yakınını bul
                all_maintenances = {}
                nearest_maintenance = None
                min_km_left = float('inf')
                
                for level in maintenance_levels:
                    level_km = int(level.replace('K', '')) * 1000
                    
                    # Bu bakımın katları: 6K = 6, 12, 18, 24... (limit 300K)
                    max_km = 300000  # 300K
                    multiples = []
                    
                    for i in range(1, (max_km // level_km) + 2):
                        km_value = level_km * i
                        multiples.append(km_value)
                    
                    # Sonraki bakım tarihi bul
                    next_due = None
                    for km_value in multiples:
                        if km_value > current_km:
                            next_due = km_value
                            break
                    
                    # Status belirle
                    if next_due is None:
                        # Tüm katları geçmiş
                        status = 'overdue'
                        km_left = current_km - multiples[-1]
                    else:
                        km_left = next_due - current_km
                        if km_left <= 0:
                            status = 'overdue'
                        elif km_left <= 500:
                            status = 'urgent'
                        elif km_left <= 2000:
                            status = 'warning'
                        else:
                            status = 'normal'
                    
                    maint_info = {
                        'level': level,
                        'level_km': level_km,
                        'next_km': next_due or multiples[-1],
                        'km_left': km_left if km_left > 0 else 0,
                        'status': status,
                        'multiples': multiples,
                        'works': maintenance_data.get(level, {}).get('works', [])
                    }
                    
                    all_maintenances[level] = maint_info
                    
                    # En yakını bul (pozitif km_left ile en küçük olanı; eşitse en yüksek level olanı)
                    if km_left > 0:
                        if km_left < min_km_left:
                            min_km_left = km_left
                            nearest_maintenance = maint_info
                        elif km_left == min_km_left and level_km > nearest_maintenance.get('level_km', 0):
                            # Eşit km_left'te daha yüksek level bakımını seç (18K > 6K vs)
                            nearest_maintenance = maint_info
                
                # Eğer hiç pozitif km_left yoksa (tümü geçmiş), en son bakımı al
                if nearest_maintenance is None:
                    nearest_maintenance = all_maintenances[maintenance_levels[-1]]
                
                # SPECIAL: Eğer next_km 72K'nin katıysa (72, 144, 216, 288K), 70K işlerini ekle
                next_km = nearest_maintenance.get('next_km', 0)
                is_70k_due = False
                if next_km > 0 and next_km % 72000 == 0 and 'all_maintenances' in locals():
                    # 70K işlerini ekle
                    is_70k_due = True
                    # 70K works sayısını al
                    seventy_k_works = len([w for w in maintenance_data.get('70K', {}).get('works', []) if w.startswith('BOZ')])
                    
                    # Eğer nearest_maintenance'da 70K yoksa, works sayısını artır
                    if 'level' in nearest_maintenance and nearest_maintenance['level'] != '70K':
                        # 70K'ı works'e not et (UI'da gösterilecek)
                        original_works = nearest_maintenance.get('works', [])
                        nearest_maintenance['has_70k'] = True
                        nearest_maintenance['additional_70k_works'] = seventy_k_works
                
                result.append({
                    'tram_id': tram_id,
                    'tram_name': eq.name if hasattr(eq, 'name') else tram_id,
                    'current_km': current_km,
                    'nearest_maintenance': nearest_maintenance,
                    'all_maintenances': all_maintenances
                })
            
            return jsonify({
                'tramps': result,
                'levels': maintenance_levels,
                'headers': maintenance_levels
            })

        @app.route('/yedek-parca')
        @login_required
        def yedek_parca():
            parts = SparePartInventory.query.all()
            stats = {
                'toplam': SparePartInventory.query.count(),
                'kritik_stok': SparePartInventory.query.filter(SparePartInventory.quantity < SparePartInventory.min_quantity).count() if hasattr(SparePartInventory, 'quantity') else 0,
            }
            return render_template('yedek_parca.html', parts=parts, stats=stats)

        @app.route('/uyarilar')
        @login_required
        def uyarilar():
            return render_template('uyarilar.html')

        @app.route('/analiz')
        @login_required
        def analiz():
            return render_template('analysis.html')

        @app.route('/raporlar')
        @login_required
        def raporlar():
            return render_template('reports.html')

        @app.route('/ayarlar')
        @login_required
        def ayarlar():
            return render_template('settings.html')

        # Additional route aliases for template compatibility
        @app.route('/arac-listesi')
        @login_required
        def arac_listesi():
            """Alias for ekipmanlar"""
            return ekipmanlar()

        @app.route('/ariza-ekle', methods=['GET', 'POST'])
        @login_required
        def ariza_ekle():
            """Alias for yeni_ariza_bildir"""
            return yeni_ariza_bildir()

        @app.route('/ariza/<int:id>')
        @login_required
        def ariza_detay(id):
            """Failure detail"""
            failure = Failure.query.get_or_404(id)
            return render_template('ariza_detay.html', failure=failure)

        @app.route('/ariza/<int:id>/guncelle', methods=['POST'])
        @login_required
        def ariza_guncelle(id):
            """Update failure"""
            failure = Failure.query.get_or_404(id)
            failure.status = request.form.get('status', failure.status)
            db.session.commit()
            flash('Failure updated', 'success')
            return redirect(url_for('ariza_detay', id=id))

        @app.route('/ekipman/<int:id>')
        @login_required
        def ekipman_detay(id):
            """Equipment detail"""
            equipment = Equipment.query.get_or_404(id)
            return render_template('ekipman_detay.html', equipment=equipment)

        @app.route('/is-emri/ekle', methods=['GET', 'POST'])
        @login_required
        def is_emri_ekle():
            """Create work order"""
            if request.method == 'POST':
                order = WorkOrder(
                    title=request.form.get('title'),
                    description=request.form.get('description'),
                    status='acik'
                )
                db.session.add(order)
                db.session.commit()
                flash('Work order created', 'success')
                return redirect(url_for('is_emirleri'))
            return render_template('is_emri_ekle.html')

        @app.route('/is-emri/<int:id>')
        @login_required
        def is_emri_detay(id):
            """Work order detail"""
            order = WorkOrder.query.get_or_404(id)
            return render_template('is_emri_detay.html', order=order)

        @app.route('/is-emri/<int:id>/guncelle', methods=['POST'])
        @login_required
        def is_emri_guncelle(id):
            """Update work order"""
            order = WorkOrder.query.get_or_404(id)
            order.status = request.form.get('status', order.status)
            db.session.commit()
            flash('Work order updated', 'success')
            return redirect(url_for('is_emri_detay', id=id))

        @app.route('/bakim-plani/ekle', methods=['GET', 'POST'])
        @login_required
        def bakim_plani_ekle():
            """Create maintenance plan"""
            if request.method == 'POST':
                plan = MaintenancePlan(
                    name=request.form.get('name'),
                    description=request.form.get('description'),
                    is_active=True
                )
                db.session.add(plan)
                db.session.commit()
                flash('Maintenance plan created', 'success')
                return redirect(url_for('bakim_planlari'))
            return render_template('bakim_plani_ekle.html')

        @app.route('/api/bakim-upload', methods=['POST'])
        @login_required
        def bakim_upload():
            """Bakım kartı dosyası upload (imza ile)"""
            try:
                file = request.files.get('file')
                tram_id = request.form.get('tram_id')
                level = request.form.get('level')
                signature = request.form.get('signature', '')  # Base64 imza
                
                if not file or not tram_id or not level:
                    return jsonify({'error': 'Eksik parametreler'}), 400
                
                # Klasör oluştur
                from datetime import datetime
                current_project = session.get('current_project', 'belgrad')
                bakim_dir = os.path.join(os.path.dirname(__file__), 'logs', current_project, 'Bakım')
                os.makedirs(bakim_dir, exist_ok=True)
                
                # Dosya adı: YYYYMMDD_TramID_Level_filename
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                filename = f"{timestamp}_{tram_id}_{level}_{file.filename}"
                filepath = os.path.join(bakim_dir, filename)
                
                # Dosyayı kaydet
                file.save(filepath)
                
                # Imza varsa ayrı dosya olarak kaydet
                if signature and signature.startswith('data:image'):
                    # Base64'ten PNG'ye dönüştür
                    import base64
                    header, encoded = signature.split(',')
                    signature_bytes = base64.b64decode(encoded)
                    signature_filename = f"{timestamp}_{tram_id}_{level}_signature.png"
                    signature_filepath = os.path.join(bakim_dir, signature_filename)
                    with open(signature_filepath, 'wb') as f:
                        f.write(signature_bytes)
                
                logger.info(f'[BAKIM] Dosya yüklendi: {filename}')
                
                return jsonify({
                    'success': True,
                    'message': 'Bakım kartı başarıyla yüklendi',
                    'file': filename
                })
            except Exception as e:
                logger.error(f'[BAKIM] Upload hatası: {e}')
                return jsonify({'error': str(e)}), 500

        @app.route('/api/bakim-excel-sheets')
        @login_required
        def bakim_excel_sheets():
            """Belgrad-Bakım.xlsx'deki sheet'leri listele"""
            try:
                import pandas as pd
                bakim_excel = os.path.join(os.path.dirname(__file__), 'data', 'belgrad', 'Belgrad-Bakım.xlsx')
                
                if not os.path.exists(bakim_excel):
                    return jsonify({'error': 'Belgrad-Bakım.xlsx bulunamadı'}), 404
                
                # Sheet adlarını al
                xl_file = pd.ExcelFile(bakim_excel)
                sheets = xl_file.sheet_names
                
                return jsonify({'sheets': sheets})
            except Exception as e:
                logger.error(f'[BAKIM] Excel sheet listesi hatası: {e}')
                return jsonify({'error': str(e)}), 500

        @app.route('/api/bakim-sheet-items/<sheet_name>')
        @login_required
        def bakim_sheet_items(sheet_name):
            """Belgrad-Bakım.xlsx'deki belirli sheet'in maddeleri"""
            try:
                import pandas as pd
                bakim_excel = os.path.join(os.path.dirname(__file__), 'data', 'belgrad', 'Belgrad-Bakım.xlsx')
                
                if not os.path.exists(bakim_excel):
                    return jsonify({'error': 'Belgrad-Bakım.xlsx bulunamadı'}), 404
                
                # Sheet'i oku (tüm rows veriye baştan doldurulmuş olabilir)
                df = pd.read_excel(bakim_excel, sheet_name=sheet_name)
                
                # Boş olmayan satırları al
                df = df.dropna(how='all')
                
                # JSON formatına çev
                items = []
                for idx, row in df.iterrows():
                    # Sütununları dictionary'ye dönüştür
                    item = {}
                    for col in df.columns:
                        val = row.get(col)
                        if pd.notna(val):
                            item[col] = str(val)
                    if item:  # Boş olmayan satırları ekle
                        items.append(item)
                
                return jsonify({
                    'sheet': sheet_name,
                    'items': items,
                    'count': len(items)
                })
            except Exception as e:
                logger.error(f'[BAKIM] Sheet items hatası: {e}')
                return jsonify({'error': str(e)}), 500

        @app.route('/api/bakim-tablosu-transpose')
        @login_required
        def bakim_tablosu_transpose():
            """Transpoze bakım tablosu: Satırda araçlar, sütunda KM noktaları"""
            try:
                import json
                from models import Equipment
                
                def get_color_by_progress(progress_percent):
                    """Progress'e göre renk döndür (0% yeşil, 100% kırmızı)"""
                    if progress_percent < 0:
                        progress_percent = 0
                    if progress_percent > 100:
                        progress_percent = 100
                    
                    # Gradient: yeşil -> sarı -> turuncu -> kırmızı
                    if progress_percent < 50:
                        # Yeşil (#d4edda) -> Sarı (#fff3cd)
                        # 0% -> 50%
                        p = progress_percent / 50
                        r = int(0xd4 + (0xff - 0xd4) * p)
                        g = int(0xed + (0xf3 - 0xed) * p)
                        b = int(0xda + (0xcd - 0xda) * p)
                    elif progress_percent < 70:
                        # Sarı (#fff3cd) -> Turuncu (#ffc107)
                        # 50% -> 70%
                        p = (progress_percent - 50) / 20
                        r = int(0xff)
                        g = int(0xf3 + (0xc1 - 0xf3) * p)
                        b = int(0xcd + (0x07 - 0xcd) * p)
                    elif progress_percent < 90:
                        # Turuncu (#ffc107) -> Koyu Turuncu (#ff8c00)
                        # 70% -> 90%
                        p = (progress_percent - 70) / 20
                        r = int(0xff)
                        g = int(0xc1 + (0x8c - 0xc1) * p)
                        b = int(0x07 + (0x00 - 0x07) * p)
                    else:
                        # Koyu Turuncu -> Kırmızı (#f8d7da)
                        # 90% -> 100%
                        p = (progress_percent - 90) / 10
                        r = int(0xff)
                        g = int(0x8c + (0xd7 - 0x8c) * p)
                        b = int(0x00 + (0xda - 0x00) * p)
                    
                    return f'#{r:02x}{g:02x}{b:02x}'
                
                logger.info('[BAKIM] Transpoze tablo yükleniyor...')
                
                # Bakım seviyeleri ve KM değerleri
                maintenance_km = {
                    '6K': 6000,
                    '18K': 18000,
                    '24K': 24000,
                    '36K': 36000,
                    '60K': 60000,
                    '70K': 70000,
                    '100K': 100000,
                    '140K': 140000,
                    '210K': 210000,
                    '300K': 300000
                }
                
                # KM noktaları: 6000'ün katları (6000, 12000, 18000... 300000'a kadar)
                km_points = list(range(6000, 300001, 6000))
                logger.info(f'[BAKIM] KM noktaları: {len(km_points)} adet')
                
                # Equipment tablosundan araçları al
                current_project = session.get('current_project', 'belgrad')
                
                # SYNC: Excel ile Database'i eşitle
                sync_equipment_with_excel(current_project)
                
                # Sadece mevcut KM verisi olan araçları al (0'dan büyük)
                equipments = Equipment.query.filter_by(project_code=current_project)\
                    .filter(Equipment.current_km > 0)\
                    .order_by(Equipment.equipment_code).all()
                logger.info(f'[BAKIM] {len(equipments)} araç bulundu (current_km > 0)')
                
                result = {
                    'km_points': km_points,
                    'tramps': []
                }
                
                # Araçlar için veri hazırla
                for equipment in equipments:
                    try:
                        current_km = int(equipment.current_km or 0)
                        
                        row_data = {
                            'tram_id': equipment.equipment_code,
                            'current_km': current_km,
                            'maintenance_at_km': {}  # Her KM noktası için yapılacak bakımlar
                        }
                        
                        # Her KM noktası için yapılacak bakımları hesapla
                        for km_point in km_points:
                            bakimlar = []
                            
                            # Her bakım seviyesi için kontrol et
                            for level, level_km in maintenance_km.items():
                                if km_point > 0 and km_point % level_km == 0:
                                    bakimlar.append(level)
                            
                            # Renk belirle: current_km'e göre progress hesapla
                            if km_point > current_km:
                                # Gelecek bakım - yüzde 0
                                progress_percent = 0
                                status = 'pending'
                            else:
                                # Geçmiş bakım - ne kadar ilerledik?
                                # Önceki KM noktasını bul (KM noktaları 6000'ün katları)
                                prev_km_point = km_point - 6000
                                if prev_km_point < 0:
                                    prev_km_point = 0
                                
                                # current_km, prev_km_point ile km_point arasında kaç %?
                                if km_point > prev_km_point:
                                    progress_percent = ((current_km - prev_km_point) / (km_point - prev_km_point)) * 100
                                    progress_percent = max(0, min(100, progress_percent))
                                else:
                                    progress_percent = 100
                                
                                if progress_percent >= 100:
                                    status = 'completed'
                                elif progress_percent >= 90:
                                    status = 'urgent'
                                elif progress_percent >= 70:
                                    status = 'warning'
                                else:
                                    status = 'normal'
                            
                            # Renk belirle
                            color = get_color_by_progress(progress_percent)
                            
                            row_data['maintenance_at_km'][km_point] = {
                                'bakimlar': bakimlar,  # Hangi bakımlar yapılacak: ['6K', '18K']
                                'color': color,
                                'status': status,
                                'progress': round(progress_percent, 1)
                            }
                        
                        result['tramps'].append(row_data)
                    except Exception as eq_err:
                        logger.error(f'[BAKIM] Araç işleme hatası {equipment.equipment_code}: {eq_err}')
                        continue
                
                logger.info(f'[BAKIM] {len(result["tramps"])} araç işlendi')
                return jsonify(result)
            except Exception as e:
                logger.error(f'[BAKIM] Transpoze tablo hatası: {type(e).__name__}: {e}')
                import traceback
                logger.error(traceback.format_exc())
                return jsonify({'error': str(e)}), 500

        @app.route('/api/bakim-pdf/<int:km>')
        @login_required
        def bakim_pdf(km):
            """Excel sekmesini PDF olarak döndür"""
            try:
                import openpyxl
                from io import BytesIO
                from reportlab.lib.pagesizes import A4, landscape
                from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
                from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
                from reportlab.lib.units import inch
                from reportlab.lib import colors
                
                # KM'ye göre sekme adını belirle
                km_mapping = {
                    6000: '6K',
                    18000: '18K',
                    24000: '24K',
                    36000: '36K',
                    60000: '60K',
                    72000: '72K',
                    102000: '102K',
                    138000: '138K',
                    144000: '144K',
                    204000: '204K',
                    210000: '210K',
                    216000: '216K',
                    276000: '276K',
                    300000: '300K'
                }
                
                sheet_name = km_mapping.get(km)
                if not sheet_name:
                    return jsonify({'error': f'{km} KM için bakım tablosu bulunamadı'}), 404
                
                current_project = session.get('current_project', 'belgrad')
                excel_file = os.path.join(os.path.dirname(__file__), 'data', current_project, f'{current_project.capitalize()}-Bakım.xlsx')
                
                if not os.path.exists(excel_file):
                    logger.error(f'[BAKIM] Dosya bulunamadı: {excel_file}')
                    return jsonify({'error': 'Excel dosyası bulunamadı'}), 404
                
                # Excel dosyasını oku
                wb = openpyxl.load_workbook(excel_file)
                if sheet_name not in wb.sheetnames:
                    return jsonify({'error': f'{sheet_name} sekmesi bulunamadı'}), 404
                
                ws = wb[sheet_name]
                
                # Verileri oku
                data = []
                for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False):
                    row_data = []
                    for cell in row:
                        value = cell.value
                        if value is None:
                            row_data.append('')
                        else:
                            row_data.append(str(value))
                    data.append(row_data)
                
                # PDF oluştur
                pdf_buffer = BytesIO()
                doc = SimpleDocTemplate(pdf_buffer, pagesize=landscape(A4), topMargin=0.5*inch, bottomMargin=0.5*inch)
                
                styles = getSampleStyleSheet()
                story = []
                
                # Başlık ekle
                title_style = ParagraphStyle(
                    'CustomTitle',
                    parent=styles['Heading1'],
                    fontSize=16,
                    textColor=colors.HexColor('#1a2332'),
                    spaceAfter=12
                )
                title = Paragraph(f'{km:,} KM Bakım Tablosu', title_style)
                story.append(title)
                story.append(Spacer(1, 0.2*inch))
                
                # Tabloyu oluştur
                if data:
                    # İlk 50 satır max (PDF'yi çok uzun yapmaması için)
                    table_data = data[:50]
                    
                    # Boş satırları filtrele
                    table_data = [row for row in table_data if any(cell.strip() for cell in row)]
                    
                    if table_data:
                        table = Table(table_data, colWidths=[0.5*inch]*len(table_data[0]))
                        table.setStyle(TableStyle([
                            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0d6efd')),
                            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                            ('FONTSIZE', (0, 0), (-1, 0), 8),
                            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                            ('FONTSIZE', (0, 1), (-1, -1), 7),
                            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f9f9f9')])
                        ]))
                        story.append(table)
                
                # PDF oluştur
                doc.build(story)
                pdf_buffer.seek(0)
                
                logger.info(f'[BAKIM] {km} KM PDF oluşturuldu')
                
                # PDF'yi download olarak döndür
                return send_file(
                    pdf_buffer,
                    mimetype='application/pdf',
                    as_attachment=True,
                    download_name=f'{km}km_bakim.pdf'
                )
            except Exception as e:
                logger.error(f'[BAKIM] PDF oluşturma hatası: {type(e).__name__}: {e}')
                import traceback
                logger.error(traceback.format_exc())
                return jsonify({'error': str(e)}), 500

        @app.route('/api/bakim-upload/<project>/<int:km>', methods=['POST'])
        @login_required
        def bakim_km_upload(project, km):
            """Bakım dosyası yükle"""
            try:
                # Kullanıcı yetkisi kontrol et
                if not current_user.can_access_project(project):
                    return jsonify({'error': 'Yetkisiz'}), 403
                
                if 'file' not in request.files:
                    return jsonify({'error': 'Dosya gerekli'}), 400
                
                file = request.files['file']
                if file.filename == '':
                    return jsonify({'error': 'Dosya seç'}), 400
                
                if not allowed_file(file.filename):
                    return jsonify({'error': f'Dosya türü desteklenmiyor. İzin verilen: {ALLOWED_EXTENSIONS}'}), 400
                
                # Bakım klasörünü oluştur
                bakim_folder = os.path.join(os.path.dirname(__file__), 'data', project, 'Bakim')
                os.makedirs(bakim_folder, exist_ok=True)
                
                # Dosya adını belirle: {km}km_bakim_{timestamp}.{ext}
                from datetime import datetime
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                filename = secure_filename(f'{km}km_bakim_{timestamp}_{file.filename}')
                filepath = os.path.join(bakim_folder, filename)
                
                # Dosyayı kaydet
                file.save(filepath)
                
                logger.info(f'[BAKIM] Dosya yüklendi: {project}/{km}km - {filename}')
                
                return jsonify({
                    'success': True,
                    'filename': filename,
                    'path': filepath,
                    'km': km
                })
            except Exception as e:
                logger.error(f'[BAKIM] Dosya yükleme hatası: {type(e).__name__}: {e}')
                import traceback
                logger.error(traceback.format_exc())
                return jsonify({'error': str(e)}), 500

        @app.route('/bakim-dosyalar/<project>/<int:km>')
        @login_required
        def bakim_dosyalar(project, km):
            """O KM'de yüklenen dosyaları listele"""
            try:
                if not current_user.can_access_project(project):
                    return jsonify({'error': 'Yetkisiz'}), 403
                
                bakim_folder = os.path.join(os.path.dirname(__file__), 'data', project, 'Bakim')
                
                if not os.path.exists(bakim_folder):
                    return jsonify({'files': []})
                
                # O KM'nin dosyalarını bul
                files = []
                for filename in os.listdir(bakim_folder):
                    if filename.startswith(f'{km}km_bakim_'):
                        filepath = os.path.join(bakim_folder, filename)
                        file_size = os.path.getsize(filepath)
                        file_date = os.path.getmtime(filepath)
                        
                        files.append({
                            'filename': filename,
                            'size': file_size,
                            'date': file_date,
                            'display_name': filename.replace(f'{km}km_bakim_', '')
                        })
                
                # Son yüklenen sırada göster
                files.sort(key=lambda x: x['date'], reverse=True)
                
                logger.info(f'[BAKIM] {project}/{km}km dosyaları: {len(files)} adet')
                return jsonify({'files': files})
            except Exception as e:
                logger.error(f'[BAKIM] Dosya listesi hatası: {type(e).__name__}: {e}')
                return jsonify({'error': str(e)}), 500

        @app.route('/bakim-dosya-indir/<project>/<int:km>/<filename>')
        @login_required
        def bakim_dosya_indir(project, km, filename):
            """Yüklenen dosyayı indir"""
            try:
                if not current_user.can_access_project(project):
                    return jsonify({'error': 'Yetkisiz'}), 403
                
                bakim_folder = os.path.join(os.path.dirname(__file__), 'data', project, 'Bakim')
                filepath = os.path.join(bakim_folder, filename)
                
                # Güvenlik: dosyanın belirtilen klasörde olup olmadığını kontrol et
                if not os.path.abspath(filepath).startswith(os.path.abspath(bakim_folder)):
                    return jsonify({'error': 'Dosya bulunamadı'}), 404
                
                if not os.path.exists(filepath):
                    return jsonify({'error': 'Dosya bulunamadı'}), 404
                
                logger.info(f'[BAKIM] Dosya indirildi: {project}/{km}km - {filename}')
                
                return send_file(
                    filepath,
                    as_attachment=True,
                    download_name=filename
                )
            except Exception as e:
                logger.error(f'[BAKIM] Dosya indirme hatası: {type(e).__name__}: {e}')
                return jsonify({'error': str(e)}), 500

        @app.route('/dokuman-listesi')
        @login_required
        def dokuman_listesi():
            """Dokümanlar - Proje spesifik filtreleme"""
            from models import TechnicalDocument
            import os
            
            current_project = session.get('current_project', 'belgrad')
            
            # Kategori ve tip filtreleri
            category = request.args.get('category', 'all')
            doc_type = request.args.get('type', 'all')
            
            # Doküman sorguksı
            query = TechnicalDocument.query.filter_by(project_code=current_project, is_active=True)
            
            if category != 'all':
                query = query.filter_by(category=category)
            if doc_type != 'all':
                query = query.filter_by(document_type=doc_type)
            
            dokumanlar = query.order_by(TechnicalDocument.created_at.desc()).all()
            
            # İstatistikler
            stats = {
                'toplam': len(dokumanlar),
                'kilavuz': len([d for d in dokumanlar if d.document_type == 'manual']),
                'sema': len([d for d in dokumanlar if d.document_type == 'schematic']),
                'prosedur': len([d for d in dokumanlar if d.document_type == 'procedure'])
            }
            
            return render_template('dokumanlar.html', 
                                 dokumanlar=dokumanlar, 
                                 stats=stats,
                                 category=category,
                                 doc_type=doc_type)

        @app.route('/dokuman/ekle', methods=['GET', 'POST'])
        @login_required
        def dokuman_ekle():
            """Yeni doküman ekle (Proje spesifik)"""
            from models import TechnicalDocument, Equipment
            
            current_project = session.get('current_project', 'belgrad')
            
            if request.method == 'POST':
                title = request.form.get('title')
                document_type = request.form.get('document_type')
                category = request.form.get('category')
                description = request.form.get('description')
                version = request.form.get('version', '1.0')
                
                # Yeni doküman kodu oluştur
                from datetime import datetime
                document_code = f"{current_project.upper()}-DOC-{datetime.now().strftime('%Y%m%d%H%M%S')}"
                
                dokuman = TechnicalDocument(
                    document_code=document_code,
                    project_code=current_project,
                    title=title,
                    document_type=document_type,
                    category=category,
                    description=description,
                    version=version,
                    is_active=True
                )
                
                db.session.add(dokuman)
                db.session.commit()
                
                flash(f'✅ Doküman başarıyla eklendi: {document_code}', 'success')
                return redirect(url_for('dokuman_listesi'))
            
            # GET - Form göster - mevcut projeden ekipmanlar
            current_project = session.get('current_project', 'belgrad')
            ekipmanlar = Equipment.query.filter_by(project_code=current_project).all()
            return render_template('dokuman_ekle.html', ekipmanlar=ekipmanlar)

        @app.route('/dokuman/<int:id>')
        @login_required
        def dokuman_detay(id):
            """Doküman detayı (Proje spesifik)"""
            from models import TechnicalDocument
            
            current_project = session.get('current_project', 'belgrad')
            dokuman = TechnicalDocument.query.filter_by(id=id, project_code=current_project).first()
            
            if not dokuman:
                flash('❌ Doküman bulunamadı', 'danger')
                return redirect(url_for('dokuman_listesi'))
            
            # View count artır
            dokuman.view_count += 1
            dokuman.last_accessed = datetime.now()
            db.session.commit()
            
            return render_template('dokuman_detay.html', dokuman=dokuman)

        @app.route('/dokuman/<int:id>/indir')
        @login_required
        def dokuman_indir(id):
            """Doküman indir (Proje spesifik)"""
            from models import TechnicalDocument
            
            current_project = session.get('current_project', 'belgrad')
            dokuman = TechnicalDocument.query.filter_by(id=id, project_code=current_project).first()
            
            if not dokuman or not dokuman.file_path:
                flash('❌ Dosya bulunamadı', 'danger')
                return redirect(url_for('dokuman_listesi'))
            
            dokuman.download_count += 1
            db.session.commit()
            
            try:
                return send_file(dokuman.file_path, as_attachment=True, download_name=dokuman.file_name)
            except Exception as e:
                flash(f'❌ Dosya indirme hatası: {str(e)}', 'danger')
                return redirect(url_for('dokuman_detay', id=id))

        @app.route('/kullanicilar')
        @login_required
        def kullanicilar():
            """Users list"""
            users = User.query.all()
            return render_template('kullanicilar.html', users=users)

        @app.route('/kullanici/ekle', methods=['GET', 'POST'])
        @login_required
        def kullanici_ekle():
            """Add user"""
            if request.method == 'POST':
                user = User(
                    username=request.form.get('username'),
                    email=request.form.get('email'),
                    full_name=request.form.get('full_name'),
                    role='user'
                )
                user.set_password(request.form.get('password'))
                db.session.add(user)
                db.session.commit()
                flash('User created', 'success')
                return redirect(url_for('kullanicilar'))
            return render_template('kullanici-ekle.html')

        @app.route('/proje-sec')
        def proje_sec():
            """Select project"""
            current_project = session.get('current_project', 'belgrad')
            # Add metadata to projects
            projects_with_data = []
            for project in PROJECTS:
                project_data = project.copy()
                project_data['has_fracas'] = False
                project_data['vehicle_count'] = 0
                project_data['failure_count'] = 0
                project_data['open_failures'] = 0
                projects_with_data.append(project_data)
            
            return render_template('proje_sec.html', projects=projects_with_data, current_project=current_project)

        @app.route('/proje-degistir/<project_code>')
        @login_required
        def proje_degistir(project_code):
            """Change project"""
            logger.info(f"\1")
            project = next((p for p in PROJECTS if p['code'] == project_code), None)
            
            if project:
                session['current_project'] = project_code
                session['project_code'] = project_code.lower()
                session['project_name'] = f"{project['flag']} {project['name']}"
                logger.info(f"\1")
                flash(f"Proje değiştirildi: {project['flag']} {project['name']}", 'success')
            else:
                flash('Geçersiz proje!', 'error')
            
            return redirect(url_for('dashboard.index'))

        @app.route('/profil-guncelle', methods=['POST'])
        @login_required
        def profil_guncelle():
            """Update profile"""
            return update_profile()

        @app.route('/audit-log')
        @login_required
        def audit_log():
            """Audit log"""
            return render_template('audit-log.html')
        @app.route('/tramvay-km')
        @login_required
        def tramvay_km():
            """Tram kilometer tracking - Single source of truth: Equipment DB"""
            
            project_code = session.get('current_project', 'belgrad').lower()
            
            # Use centralized helper function - handles all KM logic
            try:
                from utils_project_excel_store import get_tramvay_list_with_km
                equipments = get_tramvay_list_with_km(project_code)
            except Exception as e:
                logger.error(f"tramvay_km error: {e}")
                equipments = Equipment.query.filter_by(project_code=project_code).all()
            
            # Calculate stats
            stats = {
                'toplam_tramvay': len(equipments),
                'toplam_km': sum(getattr(e, 'current_km', 0) or 0 for e in equipments),
                'ortalama_km': sum(getattr(e, 'current_km', 0) or 0 for e in equipments) // len(equipments) if equipments else 0,
                'max_km': max([getattr(e, 'current_km', 0) or 0 for e in equipments]) if equipments else 0,
            }
            
            return render_template('tramvay_km.html', 
                                 stats=stats,
                                 equipments=equipments,
                                 project_name=session.get('project_name', 'Belgrad'))

        @app.route('/tramvay-km/guncelle', methods=['POST'])
        @login_required
        def tramvay_km_guncelle():
            """Update tram km - Single source: Equipment table
            Sync (bootstrap/sync) handles all propagation to Excel/JSON
            """
            logger.info(f"\1")
            try:
                tram_id = request.form.get('tram_id')
                current_km = request.form.get('current_km', 0)
                notes = request.form.get('notes', '')
                project_code = session.get('current_project', 'belgrad').lower()
                logger.info(f"\1")
                
                # Find equipment - try equipment_code first, then id
                equipment = Equipment.query.filter_by(equipment_code=str(tram_id), project_code=project_code).first()
                if not equipment and str(tram_id).isdigit():
                    equipment = Equipment.query.filter_by(id=int(tram_id), project_code=project_code).first()
                
                tram_code = str(equipment.equipment_code) if equipment else str(tram_id)
                old_km = equipment.current_km if equipment else 0
                
                # Create if not exists
                if not equipment:
                    equipment = Equipment(
                        equipment_code=tram_code,
                        name=f'Tramvay {tram_code}',
                        equipment_type='Tramvay',
                        current_km=0,
                        monthly_km=0,
                        notes='',
                        project_code=project_code
                    )
                    db.session.add(equipment)
                
                # Update KM in database (single source of truth)
                new_km = int(current_km) if current_km else 0
                equipment.current_km = new_km
                equipment.notes = notes
                db.session.commit()
                
                # Log the change
                log_km_change(
                    tram_id=tram_code,
                    old_km=old_km,
                    new_km=new_km,
                    user=current_user.username if current_user else 'system',
                    project_code=project_code,
                    notes=notes
                )
                
                # Log to Excel for permanent audit trail
                try:
                    km_excel_logger = KMExcelLogger(project_code)
                    km_excel_logger.log_km_to_excel(
                        tram_id=tram_code,
                        previous_km=old_km,
                        new_km=new_km,
                        reason=notes or 'Gün sonu sayımı',
                        user=current_user.username if current_user else 'Sistem',
                        system_type='Manuel'
                    )
                except Exception as excel_err:
                    logger.warning(f'KM Excel logging failed for {tram_code}: {excel_err}')
                
                # NOTE: Excel senkronizasyonu kaldırıldı - Equipment tablosu tek kaynak
                # Excel'in Database'i override etmesini önlemek için burada sync yapılmıyor
                
                flash(f'✅ {tram_code} KM bilgileri kaydedildi', 'success')
                
            except Exception as e:
                db.session.rollback()
                flash(f'❌ Kaydedilme hatası: {str(e)}', 'danger')
            
            return redirect(url_for('tramvay_km'))

        @app.route('/tramvay-km/toplu-guncelle', methods=['POST'])
        @login_required
        def tramvay_km_toplu_guncelle():
            """Bulk KM update - clean single path via Equipment table"""
            try:
                updates = request.get_json() or {}
                count = 0
                errors = []
                project_code = session.get('current_project', 'belgrad').lower()
                
                for tram_id, data in updates.items():
                    try:
                        tram_code = str(tram_id)
                        
                        # Find or create equipment
                        equipment = Equipment.query.filter_by(
                            equipment_code=tram_code, 
                            project_code=project_code
                        ).first()
                        
                        if not equipment and tram_code.isdigit():
                            equipment = Equipment.query.filter_by(
                                id=int(tram_code), 
                                project_code=project_code
                            ).first()
                            if equipment:
                                tram_code = str(equipment.equipment_code)
                        
                        if not equipment:
                            equipment = Equipment(
                                equipment_code=tram_code,
                                name=f'Tramvay {tram_code}',
                                equipment_type='Tramvay',
                                current_km=0,
                                monthly_km=0,
                                notes='',
                                project_code=project_code
                            )
                            db.session.add(equipment)
                        
                        # Update KM if provided
                        if 'current_km' in data and data['current_km']:
                            try:
                                new_km = int(float(data['current_km']))
                                old_km = equipment.current_km if equipment else 0
                                equipment.current_km = new_km
                                
                                # Sync to Excel/JSON
                                upsert_km(
                                    project_code=project_code,
                                    tram_id=tram_code,
                                    current_km=new_km,
                                    notes=str(data.get('notes', '') or ''),
                                    updated_by=current_user.username if current_user else 'admin'
                                )
                                
                                # Also log to Excel audit trail
                                try:
                                    km_excel_logger = KMExcelLogger(project_code)
                                    km_excel_logger.log_km_to_excel(
                                        tram_id=tram_code,
                                        previous_km=old_km,
                                        new_km=new_km,
                                        reason=data.get('notes', '') or 'Toplu güncelleme',
                                        user=current_user.username if current_user else 'Sistem',
                                        system_type='Manuel'
                                    )
                                except Exception as excel_err:
                                    logger.warning(f'KM Excel logging failed for {tram_code}: {excel_err}')
                                
                            except Exception as km_err:
                                errors.append(f"{tram_id}: Geçersiz KM değeri ({km_err})")
                                continue
                        
                        # Update notes if provided
                        if 'notes' in data:
                            equipment.notes = str(data['notes']).strip()
                        
                        count += 1
                        
                    except Exception as e:
                        errors.append(f"Tramvay {tram_id}: {str(e)}")
                
                db.session.commit()
                
                # No need to sync back - upsert_km already wrote to Excel 
                # (sync would read old Excel data and potentially overwrite DB)
                
                message = f'✅ {count} araç başarıyla kaydedildi'
                if errors:
                    message += f' ({len(errors)} hata)'
                
                return jsonify({'success': True, 'message': message}), 200
                
            except Exception as e:
                db.session.rollback()
                logger.error(f"Toplu guncelle error: {e}")
                return jsonify({'success': False, 'message': f'❌ Hata: {str(e)}'}), 500

        @app.route('/servis-durumu', methods=['GET', 'POST'])
        @login_required
        def servis_durumu():
            """Service status"""
            from datetime import timedelta, datetime
            from openpyxl import load_workbook
            import os

            current_project = session.get('current_project', 'belgrad').lower()
            
            # POST isteği - Yeni durum kayıt et
            if request.method == 'POST':
                try:
                    tarih = request.form.get('tarih')
                    tramvay_id = request.form.get('tramvay_id')
                    durum = request.form.get('durum')
                    sistem = request.form.get('sistem', '')
                    aciklama = request.form.get('aciklama', '')
                    
                    # ServiceStatus modeli varsa kayıt et
                    try:
                        from models import ServiceStatus
                        # Mevcut kaydı kontrol et
                        existing = ServiceStatus.query.filter_by(
                            tram_id=tramvay_id,
                            date=tarih,
                            project_code=current_project
                        ).first()
                        
                        alt_sistem = request.form.get('alt_sistem', '')
                        
                        if existing:
                            existing.status = durum
                            existing.sistem = sistem
                            existing.alt_sistem = alt_sistem
                            existing.aciklama = aciklama
                        else:
                            new_status = ServiceStatus(
                                tram_id=tramvay_id,
                                date=tarih,
                                status=durum,
                                sistem=sistem,
                                alt_sistem=alt_sistem,
                                aciklama=aciklama,
                                project_code=current_project
                            )
                            db.session.add(new_status)
                        
                        db.session.commit()

                        # DB -> Excel sync (proje bazlı)
                        upsert_service_status(
                            project_code=current_project,
                            status_date=tarih,
                            tram_id=str(tramvay_id),
                            status=durum,
                            sistem=sistem,
                            alt_sistem=alt_sistem,
                            aciklama=aciklama,
                            updated_by=current_user.username if current_user else 'system'
                        )
                        
                        # Log the status change
                        service_logger = ServiceStatusLogger(session.get('current_project', 'belgrad'))
                        service_logger.log_status_change(
                            tram_id=tramvay_id,
                            date=tarih,
                            status=durum,
                            sistem=sistem,
                            alt_sistem=alt_sistem,
                            aciklama=aciklama,
                            user_id=current_user.id
                        )
                        
                        flash(f'Durum başarıyla kaydedildi: {tramvay_id} - {tarih}', 'success')
                    except Exception as e:
                        flash(f'Durum kaydı hatası: {str(e)}', 'warning')
                    
                    return redirect(request.url)
                except Exception as e:
                    flash(f'Bir hata oluştu: {str(e)}', 'danger')
                    return redirect(request.url)
            
            # Excel -> DB sync (dışarıdan Excel güncellenmişse uygulamaya yansıt)
            try:
                sync_service_excel_to_db(current_project)
            except Exception as e:
                logger.warning(f"Service Excel sync warning ({current_project}): {e}")

            equipments = Equipment.query.filter_by(project_code=current_project).all()
            today_str = datetime.now().strftime('%Y-%m-%d')
            
            # Tramvaylar - Excel'den yükle (İstatistik hesaplaması için gerekli)
            excel_path = None
            project = session.get('current_project', 'belgrad')
            data_dir = os.path.join(os.path.dirname(__file__), 'data', project)
            
            if os.path.exists(data_dir):
                for file in os.listdir(data_dir):
                    if file.endswith('.xlsx') and 'Veriler' in file:
                        excel_path = os.path.join(data_dir, file)
                        break
                # Veriler.xlsx bulamazsa ilk xlsx'i al
                if not excel_path:
                    for file in os.listdir(data_dir):
                        if file.endswith('.xlsx'):
                            excel_path = os.path.join(data_dir, file)
                            break
            
            tramvaylar = []
            if excel_path and os.path.exists(excel_path):
                try:
                    wb = load_workbook(excel_path)
                    
                    # Sayfa2'den tram_id'leri oku (öncelikli)
                    if 'Sayfa2' in wb.sheetnames:
                        ws = wb['Sayfa2']
                        headers = []
                        
                        # Header'ı oku (1. satır)
                        header_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
                        tram_id_col = None
                        
                        # tram_id sütununu bul
                        for idx, header in enumerate(header_row):
                            if header and 'tram' in str(header).lower():
                                tram_id_col = idx
                                break
                        
                        if tram_id_col is not None:
                            # Tramvay ID'lerini oku (tekrar edenleri görmezden gel)
                            seen_trams = set()
                            for row in ws.iter_rows(min_row=2, values_only=True):
                                tram_id = row[tram_id_col] if row[tram_id_col] else None
                                if tram_id:
                                    tram_str = str(int(tram_id)) if isinstance(tram_id, (int, float)) else str(tram_id)
                                    if tram_str not in seen_trams:
                                        seen_trams.add(tram_str)
                                        tram_obj = type('Tram', (), {
                                            'id': tram_str,
                                            'weekly_status': {}
                                        })()
                                        tramvaylar.append(tram_obj)
                    
                    # Eğer Sayfa2'den tramvay bulunamadıysa, tüm sheet'leri kontrol et
                    if not tramvaylar:
                        for sheet_name in wb.sheetnames:
                            ws = wb[sheet_name]
                            headers = []
                            
                            # Header'ı oku (1. satır)
                            header_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
                            tram_id_col = None
                            
                            # tram_id sütununu bul
                            for idx, header in enumerate(header_row):
                                if header and 'tram' in str(header).lower():
                                    tram_id_col = idx
                                    break
                            
                            if tram_id_col is not None:
                                # Tramvay ID'lerini oku
                                for row in ws.iter_rows(min_row=2, values_only=True):
                                    tram_id = row[tram_id_col] if row[tram_id_col] else None
                                    if tram_id:
                                        tram_obj = type('Tram', (), {
                                            'id': str(int(tram_id)) if isinstance(tram_id, (int, float)) else str(tram_id),
                                            'weekly_status': {}
                                        })()
                                        tramvaylar.append(tram_obj)
                                break  # İlk sheet'i bulduktan sonra çık
                except Exception as e:
                    logger.info(f"\1")
                    pass
            
            # Eğer Excel'den yüklenemezse, veritabanından yükle
            if not tramvaylar:
                for eq in equipments:
                    tram_obj = type('Tram', (), {
                        'id': str(eq.id),
                        'weekly_status': {}
                    })()
                    tramvaylar.append(tram_obj)
            
            # SİSTEM, TEDARİKÇİ ve ALT SİSTEM verilerini Veriler.xlsx'ten çek (renk bazlı)
            sistemler = {}  # {sistem_adi: {tedarikçiler: [...], alt_sistemler: [...]}}
            
            veriler_path = None
            if os.path.exists(data_dir):
                for file in os.listdir(data_dir):
                    if 'veriler' in file.lower() and file.endswith('.xlsx'):
                        veriler_path = os.path.join(data_dir, file)
                        break
            
            if veriler_path and os.path.exists(veriler_path):
                try:
                    wb = load_workbook(veriler_path)
                    ws = wb['Sayfa1']  # Sayfa1'i aç
                    
                    # Renk tanımları
                    KIRMIZI = 'FFFF0000'  # SİSTEM
                    SARI = 'FFFFFF00'     # TEDARİKÇİ
                    MAVI = 'FF0070C0'     # ALT SİSTEM
                    
                    # Her sütunu kontrol et (sütun başına bir sistem)
                    for col in range(1, ws.max_column + 1):
                        sistem_adi = None
                        
                        # Sütundaki tüm satırları tara
                        for row in range(1, ws.max_row + 1):
                            cell = ws.cell(row=row, column=col)
                            value = cell.value
                            fill = cell.fill
                            
                            color_hex = None
                            if fill and fill.start_color:
                                color_hex = str(fill.start_color.rgb) if fill.start_color.rgb else None
                            
                            # Kırmızı renkli ise sistem bulundu (row 1'de)
                            if color_hex == KIRMIZI and value:
                                sistem_adi = str(value).strip()
                                if sistem_adi not in sistemler:
                                    sistemler[sistem_adi] = {
                                        'tedarikçiler': set(),
                                        'alt_sistemler': set()
                                    }
                            
                            # Sarı renkli ise tedarikçi
                            elif color_hex == SARI and value and sistem_adi:
                                sistemler[sistem_adi]['tedarikçiler'].add(str(value).strip())
                            
                            # Mavi renkli ise alt sistem
                            elif color_hex == MAVI and value and sistem_adi:
                                sistemler[sistem_adi]['alt_sistemler'].add(str(value).strip())
                    
                    # Set'leri list'e çevir ve sort et
                    sistemler = {
                        k: {
                            'tedarikçiler': sorted(list(v['tedarikçiler'])),
                            'alt_sistemler': sorted(list(v['alt_sistemler']))
                        }
                        for k, v in sistemler.items()
                    }
                except Exception as e:
                    logger.info(f"\1")
                    sistemler = {}
            
            import json
            sistemler_json = json.dumps(sistemler)
            logger.info(f"\1")
            logger.info(f"\1")
            logger.info(f"\1")
            logger.info(f"\1")
            for sistem_adi, data in sistemler.items():
                logger.info(f"\1")
                if data.get('tedarikçiler'):
                    logger.info(f"\1")
                if data.get('alt_sistemler'):
                    logger.info(f"\1")
            logger.info(f"\1")
            
            # ========== İSTATİSTİKLER - Tramvaylar yüklendikten sonra hesapla ==========
            stats = {
                'Servis': 0,
                'Servis Dışı': 0,
                'İşletme Kaynaklı Servis Dışı': 0,
                'erisebilirlik': '0%'
            }
            
            try:
                from models import ServiceStatus
                today_records = ServiceStatus.query.filter_by(date=today_str, project_code=current_project).all()
                
                servis_count = 0
                servis_disi_count = 0
                isletme_kaynak_count = 0
                
                for record in today_records:
                    if record.status == 'Servis':
                        servis_count += 1
                    elif record.status == 'Servis Dışı':
                        servis_disi_count += 1
                    elif record.status == 'İşletme Kaynaklı Servis Dışı':
                        isletme_kaynak_count += 1
                
                stats['Servis'] = servis_count
                stats['Servis Dışı'] = servis_disi_count
                stats['İşletme Kaynaklı Servis Dışı'] = isletme_kaynak_count
                
                # Erişilebilirlik hesapla: (Servis + İşletme Kaynaklı Servis Dışı) / Toplam * 100
                # (Bakımda olanlar da saysın - aynı dashboard hesaplaması gibi)
                total_tramvaylar = len(tramvaylar) if tramvaylar else 0
                if total_tramvaylar > 0:
                    available = servis_count + isletme_kaynak_count
                    erisebilirlik_percent = (available / total_tramvaylar) * 100
                    stats['erisebilirlik'] = f"{erisebilirlik_percent:.1f}%"
                
                logger.info(f"\1")
                logger.info(f"\1")
                logger.info(f"\1")
                logger.info(f"\1")
                logger.info(f"\1")
                logger.info(f"\1")
                
            except Exception as e:
                logger.info(f"\1")
                stats = {
                    'Servis': 0,
                    'Servis Dışı': 0,
                    'İşletme Kaynaklı Servis Dışı': 0,
                    'erisebilirlik': '0%'
                }
            # ========== İSTATİSTİKLER SONU ==========
            
            # Son 7 günü hesapla
            last_7_days = []
            for i in range(6, -1, -1):
                date = datetime.now() - timedelta(days=i)
                last_7_days.append(date.strftime('%Y-%m-%d'))
            
            # 7 günlük status matrix - veritabanından çek (varsa)
            status_matrix = {}
            tram_ids_list = [t.id if hasattr(t, 'id') else str(t) for t in tramvaylar]
            
            # Eğer ServiceStatus modeli varsa sorgu yap
            if hasattr(db, 'session'):
                try:
                    from models import ServiceStatus
                    for tram_id in tram_ids_list:
                        status_matrix[tram_id] = {}
                        for date in last_7_days:
                            status_record = ServiceStatus.query.filter_by(
                                tram_id=tram_id,
                                date=date,
                                project_code=current_project
                            ).first()
                            if status_record:
                                status_matrix[tram_id][date] = (
                                    status_record.status if hasattr(status_record, 'status') else 'Unknown',
                                    status_record.sistem if hasattr(status_record, 'sistem') else '',
                                    status_record.alt_sistem if hasattr(status_record, 'alt_sistem') else '',
                                    status_record.aciklama if hasattr(status_record, 'aciklama') else ''
                                )
                except:
                    pass
            
            # Bir önceki günün verileri al (form önceden doldurma için)
            yesterday_str = (datetime.now() - timedelta(days=1)).strftime('%Y-%m-%d')
            yesterday_data = {}
            try:
                from models import ServiceStatus
                yesterday_records = ServiceStatus.query.filter_by(date=yesterday_str, project_code=current_project).all()
                for record in yesterday_records:
                    yesterday_data[record.tram_id] = {
                        'status': record.status if hasattr(record, 'status') else '',
                        'sistem': record.sistem if hasattr(record, 'sistem') else '',
                        'alt_sistem': record.alt_sistem if hasattr(record, 'alt_sistem') else '',
                        'aciklama': record.aciklama if hasattr(record, 'aciklama') else ''
                    }
            except:
                pass
            
            return render_template('servis_durumu.html', 
                                 stats=stats, 
                                 tramvaylar=tramvaylar,
                                 tram_ids=tram_ids_list,
                                 last_7_days=last_7_days,
                                 dates=last_7_days,
                                 status_matrix=status_matrix,
                                 today=datetime.now().strftime('%Y-%m-%d'),
                                 sistemler=sistemler,
                                 sistemler_json=sistemler_json,
                                 project_name=session.get('project_name', 'Belgrad'),
                                 yesterday_data=yesterday_data)

        @app.route('/servis-durumu/guncelle', methods=['POST'])
        @login_required
        def servis_durumu_guncelle():
            """Update service status"""
            flash('Service status updated', 'success')
            return redirect(url_for('servis_durumu'))

        @app.route('/servis-durumu/toplu-servise-al', methods=['POST'])
        @login_required
        def servis_durumu_toplu_servise_al():
            """Bulk send to service - belirtilen tarihe göre"""
            try:
                from datetime import datetime
                from models import ServiceStatus
                
                data = request.get_json(force=True, silent=True)
                
                if not data:
                    return jsonify({'success': False, 'message': 'Geçersiz istek (JSON gerekli)'}), 400
                
                tram_ids = data.get('tram_ids', [])
                tarih = data.get('tarih', datetime.now().strftime('%Y-%m-%d'))  # Parametre'den tarih al, yoksa bugün
                
                if not tram_ids:
                    return jsonify({'success': False, 'message': 'Araç listesi boş'}), 400
                
                # Tarihi YYYY-MM-DD format'ında valide et
                try:
                    datetime.strptime(tarih, '%Y-%m-%d')
                except ValueError:
                    return jsonify({'success': False, 'message': 'Geçersiz tarih format'}), 400
                
                for tram_id in tram_ids:
                    # Mevcut kaydı kontrol et
                    existing = ServiceStatus.query.filter_by(
                        tram_id=str(tram_id),
                        date=tarih,
                        project_code=session.get('current_project', 'belgrad').lower()
                    ).first()
                    
                    if existing:
                        existing.status = 'Servis'
                        existing.alt_sistem = ''
                    else:
                        new_status = ServiceStatus(
                            tram_id=str(tram_id),
                            date=tarih,  # Belirtilen tarihi kullan
                            status='Servis',
                            sistem='',
                            alt_sistem='',
                            aciklama='Toplu servise alındı',
                            project_code=session.get('current_project', 'belgrad').lower()
                        )
                        db.session.add(new_status)
                
                db.session.commit()
                
                # Log bulk service operation
                for tram_id in tram_ids:
                    upsert_service_status(
                        project_code=session.get('current_project', 'belgrad').lower(),
                        status_date=tarih,
                        tram_id=str(tram_id),
                        status='Servis',
                        sistem='',
                        alt_sistem='',
                        aciklama='Toplu servise alındı',
                        updated_by=current_user.username if current_user else 'system'
                    )

                    service_logger = ServiceStatusLogger(session.get('current_project', 'belgrad'))
                    service_logger.log_status_change(
                        tram_id=tram_id,
                        date=tarih,
                        status='Servis',
                        sistem='',
                        alt_sistem='',
                        aciklama='Toplu servise alındı',
                        user_id=current_user.id
                    )
                
                return jsonify({'success': True, 'message': f'{len(tram_ids)} araç {tarih} tarihinde servise alındı'}), 200
            except Exception as e:
                db.session.rollback()
                logger.error(f'toplu-servise-al error: {str(e)}')
                return jsonify({'success': False, 'message': str(e)}), 400

        @app.route('/api/copy_previous_day', methods=['POST'])
        @login_required
        def api_copy_previous_day():
            """Önceki günün verilerini bugüne kopyala"""
            try:
                data = request.get_json()
                yesterday_data = data.get('yesterday_data', {})
                
                if not yesterday_data:
                    return jsonify({'success': False, 'message': 'Dün\'ün veri yok'}), 400
                
                # Bugünkü tarihi al
                from datetime import datetime, timedelta, date
                today = date.today()
                yesterday = today - timedelta(days=1)
                
                today_str = today.strftime('%Y-%m-%d')
                yesterday_str = yesterday.strftime('%Y-%m-%d')
                
                # Tüm araçlar için dünün durumunu bugüne kopyala
                from models import ServiceStatus
                
                count = 0
                for tram_id, yesterday_record in yesterday_data.items():
                    tram_str = str(tram_id)
                    
                    # Bugün için zaten kayıt varsa kontrol et
                    existing = ServiceStatus.query.filter_by(
                        tram_id=tram_str,
                        date=today_str
                    ).first()
                    
                    if existing:
                        # Mevcut kaydı güncelle
                        existing.status = yesterday_record.get('status', '')
                        existing.sistem = yesterday_record.get('sistem', '')
                        existing.aciklama = yesterday_record.get('aciklama', '')
                    else:
                        # Yeni kayıt oluştur
                        new_record = ServiceStatus(
                            tram_id=tram_str,
                            date=today_str,
                            status=yesterday_record.get('status', ''),
                            sistem=yesterday_record.get('sistem', ''),
                            aciklama=yesterday_record.get('aciklama', '')
                        )
                        db.session.add(new_record)
                    count += 1
                
                db.session.commit()
                
                # Log copy operation
                service_logger = ServiceStatusLogger(session.get('current_project', 'belgrad'))
                for tram_id, yesterday_record in yesterday_data.items():
                    service_logger.log_status_change(
                        tram_id=str(tram_id),
                        date=today_str,
                        status=yesterday_record.get('status', ''),
                        sistem=yesterday_record.get('sistem', ''),
                        alt_sistem=yesterday_record.get('alt_sistem', ''),
                        aciklama=f'Dünün durumundan kopyalandı ({yesterday_str})',
                        user_id=current_user.id
                    )
                
                return jsonify({
                    'success': True,
                    'message': f'✅ {count} araç dünün durumuna göre güncellendi'
                }), 200
                
            except Exception as e:
                db.session.rollback()
                return jsonify({'success': False, 'message': str(e)}), 400

        @app.route('/servis-durumu/indir')
        @login_required
        def servis_durumu_indir():
            """Download service status"""
            flash('Service status downloaded', 'info')
            return redirect(url_for('servis_durumu'))

        @app.route('/servis-durumu/root-cause-analysis', methods=['GET', 'POST'])
        @login_required
        def servis_durumu_rca():
            """Root Cause Analysis raporunu indir"""
            try:
                from flask import send_file
                
                # Tarih aralığını al
                start_date = request.args.get('start_date') or request.form.get('start_date')
                end_date = request.args.get('end_date') or request.form.get('end_date')
                tram_id = request.args.get('tram_id') or request.form.get('tram_id')
                
                # RCA analizi yap
                analysis = RootCauseAnalyzer.analyze_service_disruptions(
                    start_date=start_date,
                    end_date=end_date,
                    tram_id=tram_id
                )
                
                # Excel raporunu oluştur
                filepath = RootCauseAnalyzer.generate_rca_excel(analysis)
                
                return send_file(
                    filepath,
                    as_attachment=True,
                    download_name=f"RCA_Raporu_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
                
            except Exception as e:
                logger.error(f'RCA report error: {str(e)}')
                flash(f'Rapor oluşturulurken hata: {str(e)}', 'error')
                return redirect(url_for('servis_durumu'))

        @app.route('/yedek-parca/ekle', methods=['GET', 'POST'])
        @login_required
        def yedek_parca_ekle():
            """Add spare part"""
            if request.method == 'POST':
                part = SparePartInventory(
                    part_name=request.form.get('part_name'),
                    part_code=request.form.get('part_code')
                )
                db.session.add(part)
                db.session.commit()
                flash('Spare part added', 'success')
                return redirect(url_for('yedek_parca'))
            return render_template('yedek-parca-ekle.html')

        # Error handlers
        @app.errorhandler(404)
        def not_found(error):
            return render_template('404.html'), 404

        @app.errorhandler(500)
        def internal_error(error):
            import traceback
            # Log the full error
            print("\n" + "="*80)
            logger.info("\1")
            print("="*80)
            logger.info(f"\1")
            logger.info(f"\1")
            logger.info(f"\1")
            print(traceback.format_exc())
            print("="*80 + "\n")
            
            db.session.rollback()
            
            # Try to render the error template, if that fails, show plain text
            try:
                return render_template('500.html'), 500
            except:
                return f"""
                <html>
                    <body>
                        <h1>500 - Internal Server Error</h1>
                        <p>An error occurred. Check the server logs for details.</p>
                        <p><a href="/">Go back</a></p>
                    </body>
                </html>
                """, 500

        @app.before_request
        def log_request():
            """Log all incoming requests"""
            logger.info(f"\1")

        @app.after_request
        def optimize_response(response):
            """Optimize responses with caching and compression"""
            # Log response
            logger.info(f"\1")
            
            # Determine cache strategy based on content type and path
            if request.path.startswith('/static/'):
                # Cache static assets for 1 year
                response.headers['Cache-Control'] = 'public, max-age=31536000, immutable'
                response.headers['ETag'] = 'W/' + str(hash(response.get_data())) if response.is_sequence else ''
            elif response.content_type and 'application/json' in response.content_type:
                # API responses - no cache
                response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
                response.headers['Pragma'] = 'no-cache'
            elif response.content_type and 'text/html' in response.content_type:
                # HTML pages - short cache (5 minutes)
                response.headers['Cache-Control'] = 'public, max-age=300'
                response.headers['X-Content-Type-Options'] = 'nosniff'
            else:
                # Other content - no cache
                response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
            
            # Security headers
            response.headers['X-Frame-Options'] = 'SAMEORIGIN'
            response.headers['X-XSS-Protection'] = '1; mode=block'
            response.headers['Strict-Transport-Security'] = 'max-age=31536000; includeSubDomains'
            
            return response

        # Initialize parts cache on app startup
        try:
            load_parts_cache()
        except Exception as e:
            logger.info(f"\1")

        print('create_app finished')
        print(f'App object type: {type(app)}')
        print(f'App object: {app}')
        return app
        
    except Exception as e:
        # Critical error occurred during app initialization
        import traceback
        print('CRITICAL ERROR in create_app():')
        print(f'{type(e).__name__}: {e}')
        traceback.print_exc()
        return None


def init_sample_data(app):
    """Initialize sample data - DISABLED (test data removed)"""
    pass


# ==================== MODULE LEVEL APP INITIALIZATION ====================
# Module olarak import edilirse app variable'ı vardır
app = create_app()

if __name__ == '__main__':
    # app = create_app() - already created at module level
    if app:
        with app.app_context():
            db.create_all()  # Ensure all tables exist
            # Cache'i app başlangıcında yükle
            load_parts_cache()
        init_sample_data(app)  # Initialize sample data
    # Cloud ve local deployment için PORT ayarı
    port = int(os.environ.get('PORT', 5000))
    # **PRODUCTION'A AYARLI (debug=False default)**
    # DEBUG MODE'U AÇMAK İÇİN: set FLASK_ENV=development
    debug_mode = os.environ.get('FLASK_ENV', 'production') == 'development'
    logger.info(f"\1")
    logger.info(f"\1")
    logger.info(f"\1")
    logger.info(f"\1")
    logger.info(f"\1")
    logger.info(f"\1")
    logger.info(f"\1")
    app.run(host='0.0.0.0', port=port, debug=debug_mode, threaded=True)
