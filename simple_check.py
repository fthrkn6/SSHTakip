#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""Veriler.xlsx Say fa2 tram_id kontrolu"""

from openpyxl import load_workbook
from pathlib import Path

base = Path('.')

print("\nVERILER.XLSX SAYFA2 KONTROL\n")
print("="*70)

for proj in ['belgrad', 'kayseri']:
    path = base / 'data' / proj / 'Veriler.xlsx'
    print(f"\n{proj.upper()}: {path}")
    
    if not path.exists():
        print("  DOSYA BULUNAMADI")
        continue
    
    wb = load_workbook(path)
    if 'Sayfa2' not in wb.sheetnames:
        print(f"  Sayfalar: {wb.sheetnames}")
        continue
    
    ws = wb['Sayfa2']
    
    # Header'lar
    headers = [cell.value for cell in ws[1] if cell.value]
    print(f"  Sutunlar: {headers[:6]}")
    
    # ilk 10 tram_id
    trams = []
    for row in ws.iter_rows(min_row=2, max_row=50, max_col=1, values_only=True):
        if row[0]:
            tid = str(row[0]).strip()
            if tid:
                trams.append(tid)
    
    print(f"  Toplam tram: {len(trams)}")
    print(f"  Ornekler: {trams[:10]}")

print("\n" + "="*70)
print("\nBU VERILER EQUIPMENT TABLOSUNA YUKLENSIN MI?")
print("Eger tamam ise yapim...\n")
