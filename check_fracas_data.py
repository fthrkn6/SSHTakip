#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""Check FRACAS template data and display"""

import sys
import os
import pandas as pd
from openpyxl import load_workbook

def check_fracas_file():
    """Read and display FRACAS template data"""
    
    fracas_path = "logs/belgrad/ariza_listesi/Fracas_BELGRAD.xlsx"
    
    print("\n" + "="*80)
    print("FRACAS TEMPLATE DATA CHECK")
    print("="*80)
    
    if not os.path.exists(fracas_path):
        print(f"❌ File not found: {fracas_path}")
        return False
    
    print(f"✅ File found: {fracas_path}")
    
    # Read with openpyxl to check headers
    print("\n1. Reading FRACAS Sheet structure...")
    wb = load_workbook(fracas_path)
    ws = wb['FRACAS']
    
    # Check headers (row 4, index 3)
    print(f"   Headers in row 4:")
    headers = []
    for col_idx in range(1, 37):  # Columns A-AJ (36 columns)
        from openpyxl.utils import get_column_letter
        col_letter = get_column_letter(col_idx)
        header = ws[f'{col_letter}4'].value
        headers.append(str(header) if header else f"[Empty]")
        if col_idx <= 10 or col_idx >= 26:  # Show first 10 and last 10
            print(f"      {col_letter}4: {header}")
    
    # Read data with pandas
    print("\n2. Reading data rows with pandas...")
    df = pd.read_excel(fracas_path, sheet_name='FRACAS', header=3)  # Row 4 (index 3) has headers
    
    print(f"   Total rows: {len(df)}")
    print(f"   Total columns: {len(df.columns)}")
    print(f"\n   Column names: {list(df.columns)[:15]}...")  # Show first 15
    
    # Show last 5 rows with focus on our test data
    print("\n3. Last 5 data rows:")
    if len(df) > 0:
        display_df = df.tail(5).copy()
        # Select key columns for display
        key_cols = ['Araç Numarası', 'Araç Module', 'Sistem', 'Arıza Sınıfı']
        if 'Detaylı Bilgi' in df.columns:
            key_cols.append('Detaylı Bilgi')
        
        available_cols = [col for col in key_cols if col in df.columns]
        if available_cols:
            print(display_df[available_cols].to_string())
        
        # Check if our test data is there
        print("\n4. Checking for recent 'Servise Engel' entries:")
        if 'Detaylı Bilgi' in df.columns:
            mask = df['Detaylı Bilgi'].astype(str).str.contains('Servise Engel', na=False, case=False)
            servise_engel_rows = df[mask]
            
            if len(servise_engel_rows) > 0:
                print(f"   ✅ Found {len(servise_engel_rows)} rows with 'Servise Engel'!")
                last_rows = servise_engel_rows.tail(3)
                for idx, (i, row) in enumerate(last_rows.iterrows()):
                    arac_col = [col for col in df.columns if 'Araç Numarası' in col][0] if any('Araç Numarası' in col for col in df.columns) else None
                    if arac_col:
                        arac = row[arac_col]
                    else:
                        arac = "N/A"
                    print(f"      Row {i+1}: Araç={arac}, Detaylı Bilgi={row['Detaylı Bilgi']}")
        else:
            print("   ❌ 'Detaylı Bilgi' column not found")
    
    wb.close()
    return True

if __name__ == '__main__':
    success = check_fracas_file()
    sys.exit(0 if success else 1)
