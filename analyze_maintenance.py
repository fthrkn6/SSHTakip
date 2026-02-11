from openpyxl import load_workbook

# Excel dosyasını yükle
excel_file = 'data/belgrad/Belgrad-Bakım.xlsx'

# Openpyxl ile yükle
wb = load_workbook(excel_file)
print("Sheet'ler:", wb.sheetnames)

# 6K sheet'ini analiz et
sheet = wb['6K']
print(f"\nSheet boyutu: {sheet.dimensions}")

# Tüm satırları tarayıp WORKS kelimesini bul
print("\nİlk 20 satır:")
for row_idx in range(1, 21):
    row_data = []
    for col_idx in range(1, 6):  # İlk 5 sütun
        cell = sheet.cell(row=row_idx, column=col_idx)
        value = cell.value
        if value:
            row_data.append(f"{cell.coordinate}: {value}")
    if row_data:
        print(f"Row {row_idx}: {', '.join(row_data)}")

# WORKS'ü bul
print("\nWORKS satırını aranız...")
for row in sheet.iter_rows(min_row=1, max_row=50, values_only=False):
    for cell in row:
        if cell.value and 'WORK' in str(cell.value).upper():
            print(f"Bulundu: {cell.coordinate} = {cell.value}")
            # Bu satırından sonraki verileri göster
            work_row = cell.row
            print(f"\n{cell.value} satırından sonra:")
            for i in range(work_row+1, min(work_row+15, sheet.max_row+1)):
                work_cell = sheet.cell(row=i, column=cell.column)
                if work_cell.value:
                    print(f"  {work_cell.value}")
            break
